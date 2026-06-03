"""iter133 — Team Score module.

COMPLETELY ISOLATED from the hiring pipeline. Uses ONLY the two
collections:
    ts_rounds      — Team Rounds (round_name, total_score)
    ts_employees   — Team Employee Scores (full employee profile +
                     round_scores dict)

Zero reads/writes to: pipeline_data, naukri_applies,
bb_applicant_updates, bb_rounds, bb_job_roles, job_titles_master, or
any hiring collection.

All endpoints under /api/team-score/* require admin auth.
"""

import csv
import io
import logging
import os
import re
from datetime import datetime, timezone
from typing import Optional

from bson import ObjectId
from fastapi import APIRouter, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorDatabase
from pydantic import BaseModel

_logger = logging.getLogger("team_score")

# Wired by team_score.attach(app, db, require_auth) from server.py at
# startup. Keeps this module decoupled from server.py module-level state.
_db: Optional[AsyncIOMotorDatabase] = None
_require_auth = None

ts_router = APIRouter(prefix="/api/team-score", tags=["team-score"])


# ─────────────────────── Helpers ─────────────────────────────────────────

def _oid(s: str) -> ObjectId:
    try:
        return ObjectId(s)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _strip_oid(doc: dict) -> dict:
    if not doc:
        return doc
    if "_id" in doc:
        doc["id"] = str(doc.pop("_id"))
    return doc


# ─────────────────────── Pydantic models ─────────────────────────────────

class RoundIn(BaseModel):
    round_name: str
    total_score: float


class EmployeeIn(BaseModel):
    name: str
    email: Optional[str] = ""
    linkedin_id: Optional[str] = ""
    role: Optional[str] = ""
    joining_date: Optional[str] = ""
    college: Optional[str] = ""
    nirf_rank: Optional[str] = ""
    degree: Optional[str] = ""
    passing_year: Optional[str] = ""
    round_scores: Optional[dict] = None  # {round_name: score}
    employee_status: Optional[str] = "active"  # active | inactive


# ─────────────────────── Rounds CRUD ─────────────────────────────────────

@ts_router.get("/rounds")
async def list_rounds(request: Request):
    await _require_auth(request)
    rows = await _db.ts_rounds.find({}).sort("round_name", 1).to_list(None)
    return {"rounds": [_strip_oid(r) for r in rows]}


@ts_router.post("/rounds")
async def create_round(body: RoundIn, request: Request):
    await _require_auth(request)
    name = (body.round_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Round name required")
    existing = await _db.ts_rounds.find_one(
        {"round_name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}
    )
    if existing:
        raise HTTPException(status_code=409, detail="Round already exists")
    doc = {
        "round_name": name,
        "total_score": float(body.total_score or 0),
        "created_at": _now(),
        "updated_at": _now(),
    }
    res = await _db.ts_rounds.insert_one(doc)
    doc["id"] = str(res.inserted_id)
    doc.pop("_id", None)
    return doc


@ts_router.put("/rounds/{round_id}")
async def update_round(round_id: str, body: RoundIn, request: Request):
    await _require_auth(request)
    doc = await _db.ts_rounds.find_one({"_id": _oid(round_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Round not found")
    old_name = doc.get("round_name")
    new_name = (body.round_name or "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="Round name required")
    await _db.ts_rounds.update_one(
        {"_id": _oid(round_id)},
        {"$set": {
            "round_name": new_name,
            "total_score": float(body.total_score or 0),
            "updated_at": _now(),
        }},
    )
    # If renamed → rename the key in every employee's round_scores
    if old_name and new_name != old_name:
        async for emp in _db.ts_employees.find(
            {f"round_scores.{old_name}": {"$exists": True}}
        ):
            scores = emp.get("round_scores", {}) or {}
            scores[new_name] = scores.pop(old_name, None)
            await _db.ts_employees.update_one(
                {"_id": emp["_id"]}, {"$set": {"round_scores": scores}}
            )
    return {"success": True}


@ts_router.delete("/rounds/{round_id}")
async def delete_round(round_id: str, request: Request):
    await _require_auth(request)
    doc = await _db.ts_rounds.find_one({"_id": _oid(round_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Round not found")
    name = doc.get("round_name")
    await _db.ts_rounds.delete_one({"_id": _oid(round_id)})
    # Also remove the matching key from every employee's round_scores.
    if name:
        await _db.ts_employees.update_many(
            {f"round_scores.{name}": {"$exists": True}},
            {"$unset": {f"round_scores.{name}": ""}},
        )
    return {"success": True}


# ─────────────────────── Employees CRUD + filter ─────────────────────────

async def _ensure_round_exists(round_name: str, total_score: float = 0) -> str:
    """Create a round on the fly if missing. Returns the canonical name
    (case-preserving)."""
    name = (round_name or "").strip()
    if not name:
        return ""
    existing = await _db.ts_rounds.find_one(
        {"round_name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}
    )
    if existing:
        return existing["round_name"]
    await _db.ts_rounds.insert_one({
        "round_name": name,
        "total_score": float(total_score or 0),
        "created_at": _now(),
        "updated_at": _now(),
    })
    return name


def _build_filter_query(params: dict) -> dict:
    """Map UI filter params to a Mongo filter against ts_employees."""
    q = {}
    if params.get("employee_status"):
        q["employee_status"] = params["employee_status"]
    for f in ("name", "email", "role"):
        v = (params.get(f) or "").strip()
        if v:
            q[f] = {"$regex": re.escape(v), "$options": "i"}
    nirf = (params.get("nirf_rank") or "").strip()
    if nirf:
        q["nirf_rank"] = nirf
    return q


@ts_router.get("/employees")
async def list_employees(
    request: Request,
    employee_status: Optional[str] = None,
    name: Optional[str] = None,
    email: Optional[str] = None,
    role: Optional[str] = None,
    nirf_rank: Optional[str] = None,
):
    await _require_auth(request)
    q = _build_filter_query({
        "employee_status": employee_status, "name": name, "email": email,
        "role": role, "nirf_rank": nirf_rank,
    })
    # Sort: active first, then inactive; within each, name ASC.
    rows = await _db.ts_employees.find(q).to_list(None)
    rows.sort(key=lambda r: (
        (r.get("employee_status") or "active").lower() != "active",
        (r.get("name") or "").lower(),
    ))
    return {"employees": [_strip_oid(r) for r in rows]}


@ts_router.get("/filters")
async def filter_options(request: Request):
    """DISTINCT values pulled exclusively from ts_employees."""
    await _require_auth(request)
    async def _d(field):
        vals = await _db.ts_employees.distinct(field)
        return sorted([str(v) for v in vals if v])
    return {
        "employee_status": ["active", "inactive"],
        "name": await _d("name"),
        "email": await _d("email"),
        "role": await _d("role"),
        "nirf_rank": await _d("nirf_rank"),
    }


@ts_router.post("/employees")
async def create_employee(body: EmployeeIn, request: Request):
    await _require_auth(request)
    if not (body.name or "").strip():
        raise HTTPException(status_code=400, detail="Name required")
    scores = body.round_scores or {}
    # Auto-create missing rounds (rare for manual create; harmless).
    for rn in list(scores.keys()):
        await _ensure_round_exists(rn)
    doc = {
        "employee_status": (body.employee_status or "active").lower(),
        "name": body.name.strip(),
        "email": (body.email or "").strip(),
        "linkedin_id": (body.linkedin_id or "").strip(),
        "role": (body.role or "").strip(),
        "joining_date": (body.joining_date or "").strip(),
        "college": (body.college or "").strip(),
        "nirf_rank": (body.nirf_rank or "").strip(),
        "degree": (body.degree or "").strip(),
        "passing_year": (body.passing_year or "").strip(),
        "round_scores": {k: float(v) for k, v in scores.items() if v not in (None, "")},
        "created_at": _now(),
        "updated_at": _now(),
    }
    res = await _db.ts_employees.insert_one(doc)
    doc["id"] = str(res.inserted_id)
    doc.pop("_id", None)
    return doc


@ts_router.put("/employees/{emp_id}")
async def update_employee(emp_id: str, body: EmployeeIn, request: Request):
    await _require_auth(request)
    exists = await _db.ts_employees.find_one({"_id": _oid(emp_id)})
    if not exists:
        raise HTTPException(status_code=404, detail="Employee not found")
    scores = body.round_scores or {}
    for rn in list(scores.keys()):
        await _ensure_round_exists(rn)
    await _db.ts_employees.update_one(
        {"_id": _oid(emp_id)},
        {"$set": {
            "name": body.name.strip(),
            "email": (body.email or "").strip(),
            "linkedin_id": (body.linkedin_id or "").strip(),
            "role": (body.role or "").strip(),
            "joining_date": (body.joining_date or "").strip(),
            "college": (body.college or "").strip(),
            "nirf_rank": (body.nirf_rank or "").strip(),
            "degree": (body.degree or "").strip(),
            "passing_year": (body.passing_year or "").strip(),
            "round_scores": {k: float(v) for k, v in scores.items() if v not in (None, "")},
            "updated_at": _now(),
        }},
    )
    return {"success": True}


@ts_router.delete("/employees/{emp_id}")
async def delete_employee(emp_id: str, request: Request):
    await _require_auth(request)
    r = await _db.ts_employees.delete_one({"_id": _oid(emp_id)})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"success": True}


@ts_router.post("/employees/{emp_id}/activate")
async def activate_employee(emp_id: str, request: Request):
    await _require_auth(request)
    r = await _db.ts_employees.update_one(
        {"_id": _oid(emp_id)},
        {"$set": {"employee_status": "active", "updated_at": _now()}},
    )
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"success": True, "employee_status": "active"}


@ts_router.post("/employees/{emp_id}/deactivate")
async def deactivate_employee(emp_id: str, request: Request):
    await _require_auth(request)
    r = await _db.ts_employees.update_one(
        {"_id": _oid(emp_id)},
        {"$set": {"employee_status": "inactive", "updated_at": _now()}},
    )
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"success": True, "employee_status": "inactive"}


# ─────────────────────── Export ──────────────────────────────────────────

async def _collect_export_rows(filter_q: dict):
    rounds = await _db.ts_rounds.find({}).sort("round_name", 1).to_list(None)
    round_headers = [(r["round_name"], r.get("total_score", 0)) for r in rounds]
    emps = await _db.ts_employees.find(filter_q).to_list(None)
    # Active first, separator row, then inactive — per spec.
    actives = [e for e in emps if (e.get("employee_status") or "active").lower() == "active"]
    inactives = [e for e in emps if (e.get("employee_status") or "active").lower() == "inactive"]
    actives.sort(key=lambda e: (e.get("name") or "").lower())
    inactives.sort(key=lambda e: (e.get("name") or "").lower())
    base_cols = ["Name", "Email ID", "LinkedIn ID", "Role", "Joining Date",
                 "College", "NIRF Rank", "Degree", "Passing Year"]
    headers = base_cols + [f"{rn}({int(ts) if float(ts).is_integer() else ts})"
                           for rn, ts in round_headers]

    def emp_row(e):
        scores = e.get("round_scores") or {}
        row = [
            e.get("name") or "",
            e.get("email") or "",
            e.get("linkedin_id") or "",
            e.get("role") or "",
            e.get("joining_date") or "",
            e.get("college") or "",
            e.get("nirf_rank") or "",
            e.get("degree") or "",
            e.get("passing_year") or "",
        ]
        for rn, _ in round_headers:
            v = scores.get(rn)
            row.append("" if v in (None, "") else v)
        return row

    rows = [emp_row(e) for e in actives]
    if inactives:
        sep = ["INACTIVE EMPLOYEES"] + [""] * (len(headers) - 1)
        rows.append(sep)
        rows.extend([emp_row(e) for e in inactives])
    return headers, rows


@ts_router.get("/export")
async def export_team_scores(
    request: Request,
    fmt: str = Query("csv", regex="^(csv|xlsx)$"),
    employee_status: Optional[str] = None,
    name: Optional[str] = None,
    email: Optional[str] = None,
    role: Optional[str] = None,
    nirf_rank: Optional[str] = None,
):
    await _require_auth(request)
    q = _build_filter_query({
        "employee_status": employee_status, "name": name, "email": email,
        "role": role, "nirf_rank": nirf_rank,
    })
    headers, rows = await _collect_export_rows(q)
    if fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        w.writerows(rows)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="team_scores.csv"'},
        )
    # xlsx
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Team Scores"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="team_scores.xlsx"'},
    )


# ─────────────────────── Import ──────────────────────────────────────────

_ROUND_COL_RE = re.compile(r"^\s*([^()]+?)\s*\(\s*([-\d.]+)\s*\)\s*$")
_BASE_COL_MAP = {
    "name": "name", "employee name": "name", "full name": "name",
    "email": "email", "email id": "email", "email_id": "email",
    "linkedin": "linkedin_id", "linkedin id": "linkedin_id", "linkedin_id": "linkedin_id",
    "role": "role", "designation": "role",
    "joining date": "joining_date", "joining_date": "joining_date", "doj": "joining_date",
    "college": "college", "institute": "college",
    "nirf rank": "nirf_rank", "nirf_rank": "nirf_rank", "nirf": "nirf_rank",
    "degree": "degree", "qualification": "degree",
    "passing year": "passing_year", "passing_year": "passing_year", "year of passing": "passing_year",
}


def _parse_round_header(h: str):
    """If h is 'A(20)' style → ('A', 20.0). Else None."""
    m = _ROUND_COL_RE.match(str(h or ""))
    if not m:
        return None
    return (m.group(1).strip(), float(m.group(2)))


def _row_has_inactive_marker(row) -> bool:
    return any(
        isinstance(v, str) and "inactive employees" in v.strip().lower()
        for v in row
    )


@ts_router.post("/import")
async def import_team_scores(request: Request, file: UploadFile = File(...)):
    await _require_auth(request)
    raw = await file.read()
    filename = (file.filename or "").lower()
    rows: list = []
    if filename.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [r for r in reader]
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else c for c in r])
    else:
        raise HTTPException(status_code=400, detail="Use .csv or .xlsx")

    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")

    headers = [str(h or "").strip() for h in rows[0]]
    body = rows[1:]

    # Map columns: base fields + round-header columns.
    base_idx = {}
    round_idx = []  # list of (col_index, round_name, total_score)
    for i, h in enumerate(headers):
        key = (h or "").strip().lower()
        if key in _BASE_COL_MAP:
            base_idx[_BASE_COL_MAP[key]] = i
            continue
        parsed = _parse_round_header(h)
        if parsed:
            round_idx.append((i, parsed[0], parsed[1]))

    if "name" not in base_idx:
        raise HTTPException(status_code=400, detail="`Name` column missing")

    # Ensure every detected round exists in ts_rounds (auto-create).
    created_rounds = []
    for _, rn, ts in round_idx:
        existing = await _db.ts_rounds.find_one(
            {"round_name": {"$regex": f"^{re.escape(rn)}$", "$options": "i"}}
        )
        if not existing:
            await _db.ts_rounds.insert_one({
                "round_name": rn, "total_score": float(ts),
                "created_at": _now(), "updated_at": _now(),
            })
            created_rounds.append(rn)

    # Walk rows, switching status at "INACTIVE EMPLOYEES" markers.
    status = "active"
    inserted, updated, separators = 0, 0, 0
    for r in body:
        # Pad row to header length.
        if len(r) < len(headers):
            r = list(r) + [""] * (len(headers) - len(r))
        # Skip fully-empty rows.
        if all((str(v).strip() == "") for v in r):
            continue
        # Status marker?
        if _row_has_inactive_marker(r):
            status = "inactive"
            separators += 1
            continue

        def _at(field):
            i = base_idx.get(field)
            return "" if i is None else str(r[i] or "").strip()

        nm = _at("name")
        if not nm:
            continue
        scores = {}
        for i, rn, _ in round_idx:
            v = r[i]
            if v in (None, ""):
                continue
            try:
                scores[rn] = float(v)
            except (TypeError, ValueError):
                continue
        doc = {
            "employee_status": status,
            "name": nm,
            "email": _at("email"),
            "linkedin_id": _at("linkedin_id"),
            "role": _at("role"),
            "joining_date": _at("joining_date"),
            "college": _at("college"),
            "nirf_rank": _at("nirf_rank"),
            "degree": _at("degree"),
            "passing_year": _at("passing_year"),
            "round_scores": scores,
            "updated_at": _now(),
        }
        # Upsert by (name, email) — same combo means same person.
        existing = await _db.ts_employees.find_one({
            "name": nm,
            "email": doc["email"],
        })
        if existing:
            await _db.ts_employees.update_one(
                {"_id": existing["_id"]}, {"$set": doc}
            )
            updated += 1
        else:
            doc["created_at"] = _now()
            await _db.ts_employees.insert_one(doc)
            inserted += 1

    return {
        "success": True,
        "inserted": inserted,
        "updated": updated,
        "separators": separators,
        "rounds_created": created_rounds,
    }


# ─────────────────────── Attach ──────────────────────────────────────────

def attach(app, db, require_auth):
    """Called once from server.startup_event. Wires the router and the
    db/auth handles into this module without creating import cycles."""
    global _db, _require_auth
    _db = db
    _require_auth = require_auth
    app.include_router(ts_router)
    _logger.info("Team Score module attached at /api/team-score")
