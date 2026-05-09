"""
WhatsApp Missed Export (iter67)
============================================
Resend interview schedule details + meeting links via WhatsApp for candidates
who missed/deleted their original message.

Pipeline:
  upload(xlsx/csv) → auto-map columns → 5-priority match against pipeline_data
  + bb_registrations → fetch latest active schedule → preview → bulk send
  via existing AiSensy "Candidate FollowUp" template (5 params:
  [name, role, date, time, schedule_link]) → log to bb_resend_history.

Strict allowlist from messaging.send_whatsapp() remains in place.
"""
import io
import os
import re
import uuid
import logging
import pandas as pd
from typing import List, Optional, Dict, Any, Tuple
from datetime import datetime, timezone, timedelta

from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from motor.motor_asyncio import AsyncIOMotorDatabase

from messaging import send_whatsapp, can_send_message, FRONTEND_URL

_logger = logging.getLogger("bb_resend")

resend_router = APIRouter(prefix="/api/bb/resend", tags=["WhatsAppResend"])

_db: Optional[AsyncIOMotorDatabase] = None
_get_user = None  # async callable injected from server.py (get_current_user)


def init_resend(db: AsyncIOMotorDatabase, get_user_dep):
    global _db, _get_user
    _db = db
    _get_user = get_user_dep


# ---------------------------------------------------------------------------
# Column auto-mapping
# ---------------------------------------------------------------------------
_COLUMN_ALIASES = {
    "name":  ["name", "candidate_name", "full_name", "fullname", "candidate", "applicant_name", "applicant"],
    "email": ["email", "email_id", "emailid", "mail", "e_mail", "email_address", "emailaddress"],
    "phone": ["phone", "mobile", "phone_number", "phonenumber", "mobile_number", "mobilenumber", "contact", "contact_number", "contactno", "phone_no", "mobile_no"],
}


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(s or "").strip().lower()).strip("_")


def _auto_map_columns(df: pd.DataFrame) -> Dict[str, str]:
    """Return mapping {logical_field: actual_df_column} for name/email/phone."""
    mapping: Dict[str, str] = {}
    slug_to_col = {_slug(c): c for c in df.columns}
    for field, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            slug = _slug(alias)
            if slug in slug_to_col:
                mapping[field] = slug_to_col[slug]
                break
    return mapping


# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------
def _norm_email(v) -> str:
    return str(v or "").strip().lower()


def _norm_phone(v) -> str:
    digits = "".join(ch for ch in str(v or "") if ch.isdigit())
    return digits[-10:] if len(digits) >= 10 else digits


def _norm_name(v) -> str:
    s = re.sub(r"\s+", " ", str(v or "").strip().lower())
    return s


def _is_valid_phone(p: str) -> bool:
    n = _norm_phone(p)
    return len(n) == 10 and n[0] in "6789"


# ---------------------------------------------------------------------------
# Schedule date/time formatting
# ---------------------------------------------------------------------------
def _fmt_date(iso) -> str:
    if not iso:
        return ""
    s = str(iso)
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        y, m, d = s[:4], s[5:7], s[8:10]
        return f"{d}-{m}-{y}"
    return s


def _fmt_time(t) -> str:
    if not t:
        return ""
    s = str(t)
    # Already 12h?
    if "AM" in s.upper() or "PM" in s.upper():
        return s
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            h = int(parts[0])
            m = int(parts[1])
            ampm = "AM" if h < 12 else "PM"
            h12 = h if 1 <= h <= 12 else (h - 12 if h > 12 else 12)
            return f"{h12}:{m:02d} {ampm}"
        except Exception:
            return s
    return s


# ---------------------------------------------------------------------------
# Matching engine — 5 priorities
# ---------------------------------------------------------------------------
async def _find_in_collection(coll: str, query: dict) -> Optional[dict]:
    cursor = _db[coll].find(query, {"_id": 0})
    docs = await cursor.to_list(length=5)
    if not docs:
        return None
    # If multiple, choose the one with the most-recent schedule_date / last_update
    def _score(d):
        return (str(d.get("schedule_date") or ""), str(d.get("last_update") or ""))
    docs.sort(key=_score, reverse=True)
    return docs[0]


async def _match_candidate(name: str, email: str, phone: str) -> Tuple[Optional[dict], str, int, int, str]:
    """
    Returns (matched_doc, status, confidence, priority_used, source_collection)
    status ∈ {Exact Match, Partial Match, Multiple Match, No Match}
    """
    e = _norm_email(email)
    p = _norm_phone(phone)
    n = _norm_name(name)

    # Pre-build regex-safe last-10 phone matcher
    phone_regex_clauses = []
    if p:
        phone_regex_clauses.append({"phone": {"$regex": f"{re.escape(p)}$"}})

    # ---- Priority 1: name + email (exact email, name fuzzy substring)
    if e and n:
        for coll in ("pipeline_data", "bb_registrations"):
            cursor = _db[coll].find({"email": e}, {"_id": 0})
            docs = await cursor.to_list(length=10)
            for d in docs:
                if _norm_name(d.get("name")) == n or n in _norm_name(d.get("name")) or _norm_name(d.get("name")) in n:
                    return d, "Exact Match", 100, 1, coll

    # ---- Priority 2: name + phone
    if p and n:
        for coll in ("pipeline_data", "bb_registrations"):
            cursor = _db[coll].find({"$or": phone_regex_clauses} if phone_regex_clauses else {}, {"_id": 0})
            docs = await cursor.to_list(length=10)
            for d in docs:
                dn = _norm_name(d.get("name"))
                if dn == n or n in dn or dn in n:
                    return d, "Exact Match", 95, 2, coll

    # ---- Priority 3: email only
    if e:
        for coll in ("pipeline_data", "bb_registrations"):
            cursor = _db[coll].find({"email": e}, {"_id": 0})
            docs = await cursor.to_list(length=5)
            if len(docs) > 1:
                # Multiple match — pick latest by schedule_date, but flag
                docs.sort(key=lambda d: str(d.get("schedule_date") or ""), reverse=True)
                return docs[0], "Multiple Match", 75, 3, coll
            if docs:
                return docs[0], "Partial Match", 80, 3, coll

    # ---- Priority 4: phone only
    if p:
        for coll in ("pipeline_data", "bb_registrations"):
            cursor = _db[coll].find({"$or": phone_regex_clauses}, {"_id": 0})
            docs = await cursor.to_list(length=5)
            if len(docs) > 1:
                docs.sort(key=lambda d: str(d.get("schedule_date") or ""), reverse=True)
                return docs[0], "Multiple Match", 70, 4, coll
            if docs:
                return docs[0], "Partial Match", 75, 4, coll

    return None, "No Match", 0, 0, ""


# ---------------------------------------------------------------------------
# Schedule data fetcher — pulls latest ACTIVE schedule + token
# ---------------------------------------------------------------------------
async def _fetch_schedule(matched: dict, source_coll: str) -> Dict[str, Any]:
    """
    Combine schedule data from bb_registrations + pipeline_data.
    Returns a dict with: schedule_date, schedule_time, job_role, interview_round,
    schedule_token, schedule_link, hr_name, status, otp_verified, has_active_schedule.
    """
    email = _norm_email(matched.get("email"))
    phone = _norm_phone(matched.get("phone"))

    # Prefer bb_registrations (it carries schedule_token); fall back to pipeline_data
    reg = None
    if email or phone:
        q = {"$or": []}
        if email:
            q["$or"].append({"email": email})
        if phone:
            q["$or"].append({"phone": {"$regex": f"{re.escape(phone)}$"}})
        if q["$or"]:
            cursor = _db.bb_registrations.find(q, {"_id": 0}).sort("schedule_date", -1)
            regs = await cursor.to_list(length=5)
            # Pick the latest non-cancelled one
            for r in regs:
                if str(r.get("status") or "").lower() not in ("cancelled", "canceled"):
                    reg = r
                    break

    pipe = matched if source_coll == "pipeline_data" else None
    if not pipe and (email or phone):
        q = {"$or": []}
        if email:
            q["$or"].append({"email": email})
        if phone:
            q["$or"].append({"phone": {"$regex": f"{re.escape(phone)}$"}})
        if q["$or"]:
            pipe = await _db.pipeline_data.find_one(q, {"_id": 0})

    pipe = pipe or {}
    reg = reg or {}

    # Choose latest schedule between the two sources
    schedule_date = reg.get("schedule_date") or pipe.get("schedule_date") or ""
    schedule_time = reg.get("schedule_time") or pipe.get("schedule_time") or ""
    job_role = (reg.get("job_role") or pipe.get("job_role") or "").strip() or "Interview"
    if isinstance(reg.get("job_roles"), list) and reg["job_roles"]:
        job_role = ", ".join(reg["job_roles"])

    schedule_token = reg.get("schedule_token") or ""
    reschedule_count = int(reg.get("reschedule_count") or pipe.get("reschedule_count") or 0)
    interview_round = f"Round {reschedule_count + 1}" if reschedule_count else "Round 1"
    status = reg.get("status") or pipe.get("status") or pipe.get("result_status") or ""
    otp_verified = bool(reg.get("otp_verified") or pipe.get("otp_verified"))

    has_active_schedule = bool(schedule_date and schedule_time)

    schedule_link = ""
    if schedule_token and FRONTEND_URL:
        schedule_link = f"{FRONTEND_URL}/schedule-interview/{schedule_token}"

    return {
        "schedule_date": str(schedule_date or ""),
        "schedule_time": str(schedule_time or ""),
        "job_role": job_role,
        "interview_round": interview_round,
        "schedule_token": schedule_token,
        "schedule_link": schedule_link,
        "hr_name": "BluBridge HR Team",
        "status": status,
        "otp_verified": otp_verified,
        "has_active_schedule": has_active_schedule,
        "source_collection": source_coll,
    }


# ---------------------------------------------------------------------------
# Token generation (creates lightweight bb_registrations record if missing)
# ---------------------------------------------------------------------------
async def _ensure_schedule_token(matched: dict, sched: dict) -> str:
    """If no token exists, create a lightweight bb_registrations doc with one."""
    if sched.get("schedule_token"):
        return sched["schedule_token"]
    if not sched.get("has_active_schedule"):
        return ""

    email = _norm_email(matched.get("email"))
    phone = _norm_phone(matched.get("phone"))
    if not (email or phone):
        return ""

    new_token = uuid.uuid4().hex
    doc = {
        "name": matched.get("name") or "",
        "email": email,
        "phone": phone,
        "job_role": sched.get("job_role"),
        "schedule_date": sched.get("schedule_date"),
        "schedule_time": sched.get("schedule_time"),
        "schedule_token": new_token,
        "status": "Scheduled",
        "reschedule_count": 0,
        "created_via": "resend_module",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await _db.bb_registrations.insert_one(doc)
    return new_token


# ===========================================================================
# Schemas
# ===========================================================================
class SendRequest(BaseModel):
    upload_id: str
    row_ids: Optional[List[str]] = None  # None or empty => all matched rows
    only_failed: bool = False


class TestSendRequest(BaseModel):
    name: Optional[str] = "Test Candidate"
    job_role: Optional[str] = "AI Engineer"
    schedule_date: Optional[str] = "2026-02-15"
    schedule_time: Optional[str] = "10:30 AM"


# ===========================================================================
# Endpoints
# ===========================================================================
COOLDOWN_SECONDS = 5 * 60  # 5-minute cooldown per candidate
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB


@resend_router.post("/upload")
async def upload_resend_file(
    file: UploadFile = File(...),
    request: Request = None,
):
    user = await _get_user(request)
    if not (file.filename or "").lower().endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(400, "File must be .csv or .xlsx")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(400, "File exceeds 5 MB limit")

    try:
        if file.filename.lower().endswith(".csv"):
            for enc in ("utf-8", "latin-1", "cp1252"):
                try:
                    df = pd.read_csv(io.BytesIO(content), encoding=enc, dtype=str)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise HTTPException(400, "Unable to decode CSV")
        else:
            df = pd.read_excel(io.BytesIO(content), dtype=str)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse file: {e}")

    df.columns = [str(c).strip() for c in df.columns]
    df = df.fillna("")

    col_map = _auto_map_columns(df)
    if "email" not in col_map and "phone" not in col_map:
        raise HTTPException(
            400,
            "Could not auto-detect Email or Phone column. Expected one of: "
            f"{', '.join(_COLUMN_ALIASES['email'] + _COLUMN_ALIASES['phone'])}"
        )

    rows: List[dict] = []
    matched_count = 0
    for _, r in df.iterrows():
        name_in = str(r[col_map.get("name", "")] if "name" in col_map else "").strip()
        email_in = str(r[col_map.get("email", "")] if "email" in col_map else "").strip()
        phone_in = str(r[col_map.get("phone", "")] if "phone" in col_map else "").strip()

        if not (email_in or phone_in):
            continue  # skip blank rows

        matched, status, conf, priority, source = await _match_candidate(name_in, email_in, phone_in)
        sched: Dict[str, Any] = {}
        cand: Dict[str, Any] = {
            "input_name": name_in,
            "input_email": email_in,
            "input_phone": phone_in,
        }
        if matched:
            sched = await _fetch_schedule(matched, source)
            cand.update({
                "name": matched.get("name") or name_in,
                "email": matched.get("email") or email_in,
                "phone": matched.get("phone") or phone_in,
                "candidate_id": matched.get("id") or "",
                "current_status": sched.get("status"),
                "otp_verified": sched.get("otp_verified"),
            })
            matched_count += 1
        else:
            cand.update({"name": name_in, "email": email_in, "phone": phone_in})

        rows.append({
            "row_id": uuid.uuid4().hex,
            "match_status": status,
            "match_confidence": conf,
            "priority_used": priority,
            "source_collection": source,
            "candidate": cand,
            "schedule": sched,
            "whatsapp": {
                "last_status": "pending" if matched and sched.get("has_active_schedule") else None,
                "last_sent_at": None,
                "retry_count": 0,
                "failure_reason": None,
            },
        })

    upload_id = uuid.uuid4().hex
    doc = {
        "upload_id": upload_id,
        "filename": file.filename,
        "uploaded_by": user,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "total_rows": len(rows),
        "matched_rows": matched_count,
        "column_mapping": col_map,
        "rows": rows,
    }
    await _db.bb_resend_uploads.insert_one(doc)

    return {
        "upload_id": upload_id,
        "filename": file.filename,
        "total_rows": len(rows),
        "matched_rows": matched_count,
        "column_mapping": col_map,
    }


@resend_router.get("/preview/{upload_id}")
async def get_preview(
    upload_id: str,
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    match_status: Optional[str] = None,
    whatsapp_status: Optional[str] = None,
    search: Optional[str] = None,
):
    await _get_user(request)
    doc = await _db.bb_resend_uploads.find_one({"upload_id": upload_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Upload not found")

    rows: List[dict] = doc.get("rows", [])

    # Filtering
    if match_status:
        rows = [r for r in rows if r.get("match_status") == match_status]
    if whatsapp_status:
        rows = [r for r in rows if (r.get("whatsapp") or {}).get("last_status") == whatsapp_status]
    if search:
        q = search.strip().lower()
        rows = [r for r in rows if any(
            q in str((r.get("candidate") or {}).get(k, "")).lower()
            for k in ("name", "email", "phone", "input_name", "input_email", "input_phone")
        )]

    total = len(rows)
    start = (page - 1) * page_size
    end = start + page_size
    page_rows = rows[start:end]

    return {
        "upload_id": upload_id,
        "filename": doc.get("filename"),
        "uploaded_at": doc.get("uploaded_at"),
        "total_rows": doc.get("total_rows"),
        "matched_rows": doc.get("matched_rows"),
        "column_mapping": doc.get("column_mapping"),
        "page": page,
        "page_size": page_size,
        "filtered_total": total,
        "rows": page_rows,
    }


@resend_router.get("/uploads")
async def list_uploads(request: Request, limit: int = Query(20, ge=1, le=100)):
    await _get_user(request)
    cursor = _db.bb_resend_uploads.find({}, {"_id": 0, "rows": 0}).sort("uploaded_at", -1).limit(limit)
    return {"uploads": await cursor.to_list(length=limit)}


# ---------------------------------------------------------------------------
# Sender helpers
# ---------------------------------------------------------------------------
async def _send_one(row: dict, user: str, upload_id: str) -> Tuple[str, Optional[str]]:
    """Send WhatsApp for a single row. Returns (status, failure_reason)."""
    cand = row.get("candidate") or {}
    sched = row.get("schedule") or {}
    name = cand.get("name") or cand.get("input_name") or ""
    email = cand.get("email") or cand.get("input_email") or ""
    phone = cand.get("phone") or cand.get("input_phone") or ""

    # ---- Validations ----
    if row.get("match_status") == "No Match":
        return "skipped", "Candidate not matched"
    if not _is_valid_phone(phone):
        return "failed", f"Invalid phone: {phone}"
    if not sched.get("has_active_schedule"):
        return "skipped", "No active schedule available"

    # ---- Cooldown ----
    last = (row.get("whatsapp") or {}).get("last_sent_at")
    if last:
        try:
            last_dt = datetime.fromisoformat(str(last).replace("Z", "+00:00"))
            if (datetime.now(timezone.utc) - last_dt).total_seconds() < COOLDOWN_SECONDS:
                return "skipped", f"Cooldown active ({COOLDOWN_SECONDS//60} min)"
        except Exception:
            pass

    # ---- Ensure schedule token ----
    token = sched.get("schedule_token")
    if not token:
        token = await _ensure_schedule_token(cand, sched)
        if not token:
            return "failed", "Could not generate schedule link"
        sched["schedule_token"] = token
        sched["schedule_link"] = f"{FRONTEND_URL}/schedule-interview/{token}" if FRONTEND_URL else f"/schedule-interview/{token}"

    schedule_link = sched.get("schedule_link") or (f"{FRONTEND_URL}/schedule-interview/{token}" if FRONTEND_URL else "")
    if not schedule_link:
        return "failed", "Schedule link missing"

    # ---- TEST_MODE gate (informational; send_whatsapp also enforces) ----
    allowed, reason = await can_send_message(email, phone)
    if not allowed:
        return "blocked", f"Recipient blocked by gate ({reason})"

    formatted_date = _fmt_date(sched.get("schedule_date"))
    formatted_time = _fmt_time(sched.get("schedule_time"))
    job_role = sched.get("job_role") or "Interview"

    ok = await send_whatsapp(
        "Candidate FollowUp", phone, email,
        # iter69e (#11) — AiSensy "Candidate FollowUp" template expects
        # exactly 4 params (verified against AiSensy on 2026-05-08). Sending
        # the 5th `schedule_link` produced HTTP 400 → silent drop. Aligned
        # with `messaging.notify_missed_reminder`.
        [name, job_role, formatted_date, formatted_time],
        is_test=False,
    )

    # ---- History log ----
    await _db.bb_resend_history.insert_one({
        "history_id": uuid.uuid4().hex,
        "upload_id": upload_id,
        "row_id": row.get("row_id"),
        "sent_by": user,
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "candidate": {"name": name, "email": email, "phone": phone},
        "template": "Candidate FollowUp",
        "params": [name, job_role, formatted_date, formatted_time],
        "status": "success" if ok else "failed",
        "failure_reason": None if ok else "AiSensy send failed",
        "retry_count": int((row.get("whatsapp") or {}).get("retry_count") or 0) + 1,
    })

    return ("success" if ok else "failed", None if ok else "AiSensy send failed")


async def _persist_row_status(upload_id: str, row_id: str, status: str, reason: Optional[str]):
    now = datetime.now(timezone.utc).isoformat()
    update = {
        "$set": {
            "rows.$[r].whatsapp.last_status": status,
            "rows.$[r].whatsapp.last_sent_at": now,
            "rows.$[r].whatsapp.failure_reason": reason,
        },
        "$inc": {"rows.$[r].whatsapp.retry_count": 1},
    }
    await _db.bb_resend_uploads.update_one(
        {"upload_id": upload_id},
        update,
        array_filters=[{"r.row_id": row_id}],
    )


@resend_router.post("/send")
async def send_resend(payload: SendRequest, request: Request):
    user = await _get_user(request)
    doc = await _db.bb_resend_uploads.find_one({"upload_id": payload.upload_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Upload not found")

    rows = doc.get("rows", [])
    target_ids = set(payload.row_ids or [])

    selected = []
    for r in rows:
        if target_ids and r.get("row_id") not in target_ids:
            continue
        if payload.only_failed and (r.get("whatsapp") or {}).get("last_status") != "failed":
            continue
        selected.append(r)

    if not selected:
        raise HTTPException(400, "No rows to send")

    summary = {"success": 0, "failed": 0, "blocked": 0, "skipped": 0, "results": []}
    for r in selected:
        status, reason = await _send_one(r, user, payload.upload_id)
        summary[status] = summary.get(status, 0) + 1
        summary["results"].append({
            "row_id": r.get("row_id"),
            "name": (r.get("candidate") or {}).get("name"),
            "phone": (r.get("candidate") or {}).get("phone"),
            "status": status,
            "reason": reason,
        })
        await _persist_row_status(payload.upload_id, r.get("row_id"), status, reason)

    return summary


@resend_router.post("/test")
async def send_test_message(payload: TestSendRequest, request: Request):
    """Send a test WhatsApp message to the first allowlisted number."""
    user = await _get_user(request)
    target_email, target_phone = "rishi.nayak@blubridge.com", "9443109903"
    # iter69e (#11) — 4-param Candidate FollowUp template (no schedule_link).
    fmt_date = _fmt_date(payload.schedule_date)
    ok = await send_whatsapp(
        "Candidate FollowUp", target_phone, target_email,
        [payload.name, payload.job_role, fmt_date, payload.schedule_time],
        is_test=False,
    )
    await _db.bb_resend_history.insert_one({
        "history_id": uuid.uuid4().hex,
        "upload_id": "test",
        "row_id": "test",
        "sent_by": user,
        "sent_at": datetime.now(timezone.utc).isoformat(),
        "candidate": {"name": payload.name, "email": target_email, "phone": target_phone},
        "template": "Candidate FollowUp",
        "params": [payload.name, payload.job_role, fmt_date, payload.schedule_time],
        "status": "success" if ok else "failed",
        "failure_reason": None if ok else "AiSensy send failed",
        "retry_count": 1,
    })
    return {"success": bool(ok), "to": target_phone}


@resend_router.get("/history")
async def get_history(
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    upload_id: Optional[str] = None,
    status: Optional[str] = None,
):
    await _get_user(request)
    q = {}
    if upload_id:
        q["upload_id"] = upload_id
    if status:
        q["status"] = status
    total = await _db.bb_resend_history.count_documents(q)
    cursor = _db.bb_resend_history.find(q, {"_id": 0}).sort("sent_at", -1).skip((page - 1) * page_size).limit(page_size)
    return {
        "page": page, "page_size": page_size, "total": total,
        "rows": await cursor.to_list(length=page_size),
    }


@resend_router.get("/template-preview")
async def template_preview():
    """Return the WhatsApp template body the UI should display as preview."""
    return {
        "template": "Candidate FollowUp",
        "body": (
            "Hello {{name}},\n\n"
            "This is a reminder regarding your interview schedule for the role of {{job_role}}.\n\n"
            "Interview Details:\n"
            "📅 Date: {{schedule_date}}\n"
            "⏰ Time: {{schedule_time}}\n"
            "🎯 Round: {{interview_round}}\n\n"
            "🔗 Interview Link:\n{{schedule_link}}\n\n"
            "Please join on time.\n\n"
            "For support contact:\n{{hr_name}}\n\n"
            "Thank You,\nBluBridge Hiring Team"
        ),
        "params": ["name", "job_role", "schedule_date", "schedule_time", "schedule_link"],
    }


# ---------------------------------------------------------------------------
# Export — Download preview as CSV / XLSX
# ---------------------------------------------------------------------------
_EXPORT_COLUMNS = [
    "Candidate Name", "Email", "Phone",
    "Match Status", "Match Confidence", "Priority Used",
    "Job Role", "Interview Round",
    "Schedule Date", "Schedule Time",
    "Schedule Link", "Schedule Token",
    "WhatsApp Status", "Last Sent At", "Retry Count", "Failure Reason",
    "Source Collection",
]


def _row_to_export_record(r: dict) -> dict:
    c = r.get("candidate") or {}
    s = r.get("schedule") or {}
    w = r.get("whatsapp") or {}
    return {
        "Candidate Name":   c.get("name") or c.get("input_name") or "",
        "Email":            c.get("email") or c.get("input_email") or "",
        "Phone":            c.get("phone") or c.get("input_phone") or "",
        "Match Status":     r.get("match_status") or "",
        "Match Confidence": r.get("match_confidence") or 0,
        "Priority Used":    r.get("priority_used") or 0,
        "Job Role":         s.get("job_role") or "",
        "Interview Round":  s.get("interview_round") or "",
        "Schedule Date":    _fmt_date(s.get("schedule_date")) or "",
        "Schedule Time":    s.get("schedule_time") or "",
        "Schedule Link":    s.get("schedule_link") or "",
        "Schedule Token":   s.get("schedule_token") or "",
        "WhatsApp Status":  w.get("last_status") or "pending",
        "Last Sent At":     w.get("last_sent_at") or "",
        "Retry Count":      w.get("retry_count") or 0,
        "Failure Reason":   w.get("failure_reason") or "",
        "Source Collection":r.get("source_collection") or "",
    }


@resend_router.get("/export/{upload_id}")
async def export_preview(
    upload_id: str,
    request: Request,
    fmt: str = Query("xlsx", regex="^(xlsx|csv)$"),
    match_status: Optional[str] = None,
    whatsapp_status: Optional[str] = None,
):
    """Download the preview table (with current filters applied) as CSV or XLSX."""
    await _get_user(request)
    doc = await _db.bb_resend_uploads.find_one({"upload_id": upload_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Upload not found")

    rows = doc.get("rows", [])
    if match_status:
        rows = [r for r in rows if r.get("match_status") == match_status]
    if whatsapp_status:
        rows = [r for r in rows if (r.get("whatsapp") or {}).get("last_status") == whatsapp_status]

    records = [_row_to_export_record(r) for r in rows]
    df = pd.DataFrame(records, columns=_EXPORT_COLUMNS)

    base_name = (doc.get("filename") or "whatsapp-resend").rsplit(".", 1)[0]
    fname = f"{base_name}-results.{fmt}"

    buf = io.BytesIO()
    if fmt == "csv":
        df.to_csv(buf, index=False, encoding="utf-8-sig")
        media = "text/csv; charset=utf-8"
    else:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Resend Results", index=False)
            ws = writer.sheets["Resend Results"]
            from openpyxl.styles import Font, PatternFill, Alignment
            hdr_fill = PatternFill(start_color="1D3A8A", end_color="1D3A8A", fill_type="solid")
            hdr_font = Font(bold=True, color="FFFFFF", size=11)
            for col_idx in range(1, len(_EXPORT_COLUMNS) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
            # Auto-fit columns (approximate)
            widths = [22, 30, 16, 16, 14, 12, 22, 16, 14, 12, 60, 32, 14, 22, 10, 32, 18]
            from openpyxl.utils import get_column_letter
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    buf.seek(0)
    return StreamingResponse(
        buf, media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
