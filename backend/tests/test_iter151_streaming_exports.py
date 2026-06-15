"""iter151 — Streaming export endpoints (chunked-streaming refactor).

Verifies:
  1. /api/applicants/export and /api/attended/export return correct content.
  2. The main applicants cursor is NEVER materialised via to_list (proves
     the streaming refactor stays in place; a regression that re-introduces
     `.to_list(None)` on `pipeline_data` would fail this test).
  3. No `bb_users` mutation — auth is bypassed via dependency override,
     not by writing to the user collection (per the strict no-password-
     mutation rule documented in /app/memory/test_credentials.md).
"""

from __future__ import annotations

import os
import sys

import pytest
import pytest_asyncio

os.environ.setdefault("FRONTEND_URL", "http://localhost")
os.environ.setdefault("RESEND_API_KEY", "test")
os.environ.setdefault("MONGO_URL", os.environ.get("MONGO_URL", "mongodb://localhost:27017"))
os.environ.setdefault("DB_NAME", os.environ.get("DB_NAME", "test_iter151"))

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

pytestmark = pytest.mark.asyncio(loop_scope="module")


@pytest_asyncio.fixture(loop_scope="module", scope="module")
async def app_module():
    import server  # noqa: E402
    return server


@pytest_asyncio.fixture(loop_scope="module")
async def auth_override(app_module):
    from server import app, get_current_user
    app.dependency_overrides[get_current_user] = lambda: "pytest-user"
    yield
    app.dependency_overrides.pop(get_current_user, None)


@pytest_asyncio.fixture(loop_scope="module")
async def seeded_db(app_module):
    db = app_module.db
    await db.pipeline_data.delete_many({"_iter151_marker": True})
    docs = []
    for i in range(50):
        docs.append({
            "_iter151_marker": True,
            "isTest": False,
            "name": f"Iter151Cand{i:03d}",
            "email": f"iter151cand{i:03d}@example.test",
            "phone": f"90000{i:05d}",
            "_normalized_job_role": "Engineer",
            "_college_status": "NIRF",
            "_college_resolved": "Test College",
            "degree": "B.Tech",
            "course": "CSE",
            "year_of_graduation": "2025",
            "submitted_at": "2026-01-15T10:00:00",
            "last_update": "2026-01-15T10:00:00",
            "schedule_date": "2026-02-01",
            "schedule_time": "10:00",
            "email_type": "shortlist",
            "otp_verified": "2026-02-01T10:30:00",
            "result_status": "Pass",
        })
    await db.pipeline_data.insert_many(docs)
    yield db
    await db.pipeline_data.delete_many({"_iter151_marker": True})


async def test_applicants_export_csv_works(app_module, auth_override, seeded_db):
    from httpx import AsyncClient, ASGITransport
    transport = ASGITransport(app=app_module.app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        r = await ac.get("/api/applicants/export", params={
            "format": "csv",
            "search": "iter151cand",
        })
    assert r.status_code == 200
    body = r.text
    assert "Name,Email,Phone" in body
    assert body.count("\n") >= 50
    assert "iter151cand000@example.test" in body


async def test_attended_export_xlsx_works(app_module, auth_override, seeded_db):
    from httpx import AsyncClient, ASGITransport
    transport = ASGITransport(app=app_module.app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        r = await ac.get("/api/attended/export", params={
            "format": "xlsx",
            "search": "iter151cand",
        })
    assert r.status_code == 200
    assert "attachment" in r.headers.get("content-disposition", "")
    assert len(r.content) > 0
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb.active
    first_row = next(ws.iter_rows(values_only=True))
    assert first_row[0] == "Name"


async def test_pipeline_data_is_NOT_materialised_via_to_list(app_module, auth_override, seeded_db, monkeypatch):
    """Regression guard: if a future change re-introduces
    `pipeline_data.find(...).to_list(None)` inside either export endpoint,
    this test fails immediately."""
    from motor.motor_asyncio import AsyncIOMotorCursor
    calls: list = []
    orig_to_list = AsyncIOMotorCursor.to_list

    async def _spy_to_list(self, length=None):
        try:
            coll_name = self.collection.name
        except Exception:
            coll_name = "?"
        calls.append(coll_name)
        return await orig_to_list(self, length)

    monkeypatch.setattr(AsyncIOMotorCursor, "to_list", _spy_to_list)

    from httpx import AsyncClient, ASGITransport
    transport = ASGITransport(app=app_module.app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        r1 = await ac.get("/api/applicants/export", params={"format": "csv", "search": "iter151cand"})
        r2 = await ac.get("/api/attended/export", params={"format": "csv", "search": "iter151cand"})
    assert r1.status_code == 200
    assert r2.status_code == 200
    assert "pipeline_data" not in calls, (
        f"pipeline_data was materialised via to_list during export — regression! Calls: {calls}"
    )


async def test_no_user_collection_mutation(app_module, auth_override, seeded_db):
    before = await app_module.db.bb_users.count_documents({})
    from httpx import AsyncClient, ASGITransport
    transport = ASGITransport(app=app_module.app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        await ac.get("/api/applicants/export", params={"format": "csv", "search": "iter151cand"})
        await ac.get("/api/attended/export", params={"format": "csv", "search": "iter151cand"})
    after = await app_module.db.bb_users.count_documents({})
    assert before == after, "Export endpoints must not touch bb_users"
