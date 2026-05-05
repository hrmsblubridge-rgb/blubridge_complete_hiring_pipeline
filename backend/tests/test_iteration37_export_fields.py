"""Iteration 37: Dynamic field selection for Interview Reports export."""
import os
import io
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://applicant-details.preview.emergentagent.com").rstrip("/")
ADMIN_USER = "Admin User"
ADMIN_PASS = "Admin User"


@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    r = s.post(f"{BASE_URL}/api/login", json={"username": ADMIN_USER, "password": ADMIN_PASS}, timeout=30)
    if r.status_code != 200:
        pytest.skip(f"Login failed status={r.status_code}")
    return s


# ---------- /export-fields catalog endpoint ----------
class TestExportFieldsCatalog:
    def test_export_fields_returns_sections(self, client):
        r = client.get(f"{BASE_URL}/api/bb/interview-reports/export-fields", timeout=60)
        assert r.status_code == 200
        data = r.json()
        assert "sections" in data and "total_matching" in data
        section_ids = [s["id"] for s in data["sections"]]
        assert "registration" in section_ids
        assert "interview" in section_ids
        for s in data["sections"]:
            assert "label" in s and isinstance(s["fields"], list)
            for f in s["fields"]:
                assert "key" in f and "label" in f

    def test_missing_db_keys_omitted_from_catalog(self, client):
        """date_of_birth, current_location, preferred_location, date_of_application
        are NOT present in pipeline_data so should be silently omitted."""
        r = client.get(f"{BASE_URL}/api/bb/interview-reports/export-fields", timeout=60)
        assert r.status_code == 200
        data = r.json()
        all_keys = []
        for s in data["sections"]:
            all_keys.extend([f["key"] for f in s["fields"]])
        # Per spec these should NOT appear when pipeline_data lacks them
        for missing in ["date_of_birth", "current_location", "preferred_location"]:
            assert missing not in all_keys, f"{missing} should be hidden (not in pipeline_data sample)"
        # Core fields MUST appear
        for present in ["name", "email", "phone", "job_role", "schedule_date", "attendance"]:
            assert present in all_keys, f"{present} should be in catalog"

    def test_total_matching_is_int(self, client):
        r = client.get(f"{BASE_URL}/api/bb/interview-reports/export-fields", timeout=60)
        data = r.json()
        assert isinstance(data["total_matching"], int)
        assert data["total_matching"] >= 0


# ---------- /export selective columns ----------
class TestExportSelective:
    def test_export_with_4_fields_returns_4_columns(self, client):
        params = {
            "fields": "name,email,job_role,schedule_date",
            "startDate": "2026-02-15",
            "endDate": "2026-02-28",
        }
        r = client.get(f"{BASE_URL}/api/bb/interview-reports/export", params=params, timeout=120)
        assert r.status_code in (200, 404), f"Got {r.status_code}: {r.text[:300]}"
        if r.status_code == 404:
            pytest.skip("No data in 2026-02-15..2026-02-28 range — env-dependent")
        ct = r.headers.get("content-type", "")
        assert "spreadsheet" in ct or "excel" in ct or "xlsx" in ct, f"Unexpected content-type: {ct}"
        # Read XLSX header row
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header == ["Name", "Email", "Job Role", "Schedule Date"], f"Got header: {header}"

    def test_export_drops_bogus_field_silently(self, client):
        params = {
            "fields": "name,xxxBogus,email",
            "startDate": "2026-02-20",
            "endDate": "2026-02-22",
        }
        r = client.get(f"{BASE_URL}/api/bb/interview-reports/export", params=params, timeout=120)
        if r.status_code == 404:
            pytest.skip("No data in date range — env-dependent")
        assert r.status_code == 200, f"Got {r.status_code}: {r.text[:300]}"
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header == ["Name", "Email"], f"Bogus field not dropped, header: {header}"

    def test_export_only_invalid_field_returns_400(self, client):
        r = client.get(
            f"{BASE_URL}/api/bb/interview-reports/export",
            params={"fields": "invalid_only"},
            timeout=60,
        )
        assert r.status_code == 400
        body = r.json()
        msg = body.get("detail") or body.get("message") or ""
        assert "at least one field" in msg.lower(), f"Got: {body}"

    def test_export_empty_fields_param_falls_back_to_full_catalog(self, client):
        """fields= (empty) should be backward-compat: full catalog (all 16+ columns)."""
        params = {"fields": "", "startDate": "2026-02-20", "endDate": "2026-02-22"}
        r = client.get(f"{BASE_URL}/api/bb/interview-reports/export", params=params, timeout=180)
        if r.status_code == 404:
            pytest.skip("No data in range")
        assert r.status_code == 200, f"Got {r.status_code}: {r.text[:300]}"
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # Full catalog has 20 entries but only those with required db keys present should appear
        # Spec says "all 16 columns" backward compat
        assert "Name" in header and "Email" in header
        assert len(header) >= 14, f"Expected ~16 columns, got {len(header)}: {header}"

    def test_export_no_data_returns_404(self, client):
        params = {
            "startDate": "2099-01-01",
            "endDate": "2099-01-02",
            "fields": "name",
        }
        r = client.get(f"{BASE_URL}/api/bb/interview-reports/export", params=params, timeout=60)
        assert r.status_code == 404
        body = r.json()
        msg = body.get("detail") or body.get("message") or ""
        assert "no data" in msg.lower(), f"Got: {body}"


# ---------- Regression: list endpoint pagination + sort ----------
class TestRegressionInterviewReports:
    def test_list_pagination_works(self, client):
        r = client.get(
            f"{BASE_URL}/api/bb/interview-reports",
            params={"page": 1, "limit": 10},
            timeout=60,
        )
        assert r.status_code == 200
        d = r.json()
        assert "data" in d and "total" in d
        assert len(d["data"]) <= 10

    def test_list_sort_by_name_asc(self, client):
        r = client.get(
            f"{BASE_URL}/api/bb/interview-reports",
            params={"page": 1, "limit": 10, "sort_by": "name", "sort_dir": "asc"},
            timeout=60,
        )
        assert r.status_code == 200
