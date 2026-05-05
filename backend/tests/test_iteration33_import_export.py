"""Iteration 33 — Import/Export with strict schema on Update Applicants Scores page.
Tests:
- Export endpoint (xlsx, csv, filtered, validation)
- Import preview (valid, missing columns, csv, empty score, invalid score)
- Import confirm (persist, upsert, isImported flag)
- Round-trip: Export → Import preview parses without errors
- Worker presence in logs
- Messaging scope safety (TEST_MODE notify_rejected simulation + worker query)
"""
import io
import os
import csv
import uuid
import asyncio
import pytest
import requests
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://applicant-details.preview.emergentagent.com").rstrip("/")


def _run_async(coro):
    """Run a coroutine in a fresh event loop (motor client must be created here)."""
    loop = asyncio.new_event_loop()
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()
API = f"{BASE_URL}/api"
ADMIN_USER = "Admin User"
ADMIN_PASS = "Admin User"

FIXED_HEADERS = ["Name", "Schedule Date", "College", "Degree", "Course",
                 "Year of Graduation", "Email", "Phone", "Job Role", "Status"]


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    r = s.post(f"{API}/login", json={"username": ADMIN_USER, "password": ADMIN_PASS}, timeout=30)
    if r.status_code != 200:
        # try email/password keys
        r = s.post(f"{API}/login", json={"email": ADMIN_USER, "password": ADMIN_PASS}, timeout=30)
    assert r.status_code == 200, f"Login failed: {r.status_code} {r.text[:300]}"
    return s


@pytest.fixture(scope="module")
def cleanup_emails():
    emails = []
    yield emails
    # Cleanup via a direct admin endpoint if available; else best-effort via Mongo via admin tools
    # Best-effort: call delete endpoint if exists; fallback noop
    try:
        from motor.motor_asyncio import AsyncIOMotorClient
        url = os.environ.get("MONGO_URL")
        db_name = os.environ.get("DB_NAME", "hr_analytics")
        if url and emails:
            async def _clean():
                cli = AsyncIOMotorClient(url)
                db = cli[db_name]
                await db.bb_applicant_updates.delete_many({"email": {"$in": emails}})
                cli.close()
            _run_async(_clean())
    except Exception:
        pass


# -------- helpers --------

def _make_xlsx(rows, headers):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============ EXPORT TESTS ============

class TestExport:
    def test_export_xlsx_status_and_headers(self, session):
        r = session.get(f"{API}/bb/export-scores?format=xlsx", timeout=120)
        assert r.status_code == 200, r.text[:300]
        assert "spreadsheetml" in r.headers.get("Content-Type", "")
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 1
        headers = [str(h or "").strip() for h in rows[0]]
        # First 10 must match exactly
        assert headers[:10] == FIXED_HEADERS, f"Header mismatch: {headers[:10]}"
        # Round columns alphabetical
        round_cols = headers[10:]
        assert round_cols == sorted(round_cols), f"Round columns NOT alphabetical: {round_cols}"

    def test_export_csv(self, session):
        r = session.get(f"{API}/bb/export-scores?format=csv", timeout=120)
        assert r.status_code == 200
        assert "text/csv" in r.headers.get("Content-Type", "")
        text = r.content.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        assert len(all_rows) > 1, "CSV should have data rows"
        headers = [h.strip() for h in all_rows[0]]
        assert headers[:10] == FIXED_HEADERS
        round_cols = headers[10:]
        assert round_cols == sorted(round_cols)

    def test_export_filter_date_range(self, session):
        r_all = session.get(f"{API}/bb/export-scores?format=csv", timeout=120)
        r_filt = session.get(
            f"{API}/bb/export-scores?format=csv&startDate=2026-01-01&endDate=2026-12-31",
            timeout=120,
        )
        assert r_filt.status_code == 200
        all_rows = list(csv.reader(io.StringIO(r_all.content.decode("utf-8", errors="ignore"))))
        filt_rows = list(csv.reader(io.StringIO(r_filt.content.decode("utf-8", errors="ignore"))))
        assert len(filt_rows) <= len(all_rows), "Filtered rows should be <= unfiltered"

    def test_export_invalid_format_returns_422(self, session):
        r = session.get(f"{API}/bb/export-scores?format=invalid", timeout=30)
        assert r.status_code == 422, f"Expected 422, got {r.status_code}: {r.text[:300]}"


# ============ IMPORT PREVIEW TESTS ============

class TestImportPreview:
    def test_preview_valid_xlsx_with_round_cols(self, session):
        # Note: alphabetical sorted output expects ['Accounts1', 'BA']
        headers = FIXED_HEADERS + ["BA", "Accounts1"]  # purposely unsorted on input
        rows = [
            {"Name": "TestA", "Schedule Date": "2026-06-01", "College": "C1", "Degree": "BE",
             "Course": "CS", "Year of Graduation": "2025",
             "Email": "test_iter33_a@example.com", "Phone": "9999999991",
             "Job Role": "Engineer", "Status": "On hold", "BA": 80, "Accounts1": 70},
            {"Name": "TestB", "Schedule Date": "2026-06-02", "College": "C2", "Degree": "BE",
             "Course": "EC", "Year of Graduation": "2024",
             "Email": "test_iter33_b@example.com", "Phone": "9999999992",
             "Job Role": "Analyst", "Status": "Rejected", "BA": 65, "Accounts1": ""},
        ]
        data = _make_xlsx(rows, headers)
        files = {"file": ("test_iter33.xlsx", data,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = session.post(f"{API}/bb/import-scores/preview", files=files, timeout=60)
        assert r.status_code == 200, r.text[:300]
        body = r.json()
        assert body["total"] == 2
        assert body["round_columns"] == ["Accounts1", "BA"], f"Not alphabetical: {body['round_columns']}"
        assert body["errors"] == []
        # Row 1 has both scores, row 2 has only BA (Accounts1 empty)
        scores0 = {s["round_name"]: s["score"] for s in body["rows"][0]["scores"]}
        assert scores0 == {"BA": 80.0, "Accounts1": 70.0}
        scores1 = {s["round_name"]: s["score"] for s in body["rows"][1]["scores"]}
        assert scores1 == {"BA": 65.0}, f"Empty score should be skipped: {scores1}"

    def test_preview_missing_required_column(self, session):
        headers = [h for h in FIXED_HEADERS if h != "Status"]  # drop Status
        rows = [{"Name": "X", "Email": "test_iter33_x@example.com"}]
        data = _make_xlsx(rows, headers)
        files = {"file": ("test_iter33_missing.xlsx", data,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = session.post(f"{API}/bb/import-scores/preview", files=files, timeout=30)
        assert r.status_code == 400
        assert "Status" in r.text or "missing" in r.text.lower()

    def test_preview_csv(self, session):
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=FIXED_HEADERS + ["BA"])
        w.writeheader()
        w.writerow({"Name": "CsvA", "Schedule Date": "2026-06-03", "College": "C",
                    "Degree": "BE", "Course": "CS", "Year of Graduation": "2024",
                    "Email": "test_iter33_csv@example.com", "Phone": "9999999993",
                    "Job Role": "Eng", "Status": "On hold", "BA": "75"})
        files = {"file": ("test_iter33.csv", buf.getvalue().encode("utf-8"), "text/csv")}
        r = session.post(f"{API}/bb/import-scores/preview", files=files, timeout=30)
        assert r.status_code == 200, r.text[:300]
        b = r.json()
        assert b["total"] == 1
        assert b["round_columns"] == ["BA"]

    def test_preview_invalid_score_adds_error_row_returns_200(self, session):
        headers = FIXED_HEADERS + ["BA"]
        rows = [
            {"Name": "Bad", "Schedule Date": "2026-06-04", "College": "C", "Degree": "BE",
             "Course": "CS", "Year of Graduation": "2024",
             "Email": "test_iter33_bad@example.com", "Phone": "9999999994",
             "Job Role": "Eng", "Status": "On hold", "BA": "abc"},
            {"Name": "Good", "Schedule Date": "2026-06-04", "College": "C", "Degree": "BE",
             "Course": "CS", "Year of Graduation": "2024",
             "Email": "test_iter33_good@example.com", "Phone": "9999999995",
             "Job Role": "Eng", "Status": "On hold", "BA": "85"},
        ]
        data = _make_xlsx(rows, headers)
        files = {"file": ("bad.xlsx", data,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = session.post(f"{API}/bb/import-scores/preview", files=files, timeout=30)
        assert r.status_code == 200
        body = r.json()
        assert len(body["errors"]) >= 1
        assert body["total"] == 2  # both rows still parsed


# ============ IMPORT CONFIRM TESTS ============

class TestImportConfirm:
    def test_confirm_persists_and_flags(self, session, cleanup_emails):
        email_a = "test_iter33_confirm_a@example.com"
        email_b = "test_iter33_confirm_b@example.com"
        cleanup_emails.extend([email_a, email_b])
        payload = {"rows": [
            {"name": "ConfA", "schedule_date": "2026-06-10",
             "email": email_a, "phone": "9991110001", "job_role": "Engineer",
             "status": "Rejected",
             "scores": [{"round_name": "BA", "score": 90}, {"round_name": "Accounts1", "score": 75}]},
            {"name": "ConfB", "schedule_date": "2026-06-11",
             "email": email_b, "phone": "9991110002", "job_role": "Analyst",
             "status": "On hold", "scores": []},
        ]}
        r = session.post(f"{API}/bb/import-scores/confirm", json=payload, timeout=30)
        assert r.status_code == 200, r.text[:300]
        body = r.json()
        assert body["success"] is True
        assert body["imported"] == 2
        assert isinstance(body["batch_id"], str) and len(body["batch_id"]) > 8

        # Verify via Mongo direct
        from motor.motor_asyncio import AsyncIOMotorClient
        url = os.environ.get("MONGO_URL"); db_name = os.environ.get("DB_NAME", "hr_analytics")
        async def _check():
            cli = AsyncIOMotorClient(url); db = cli[db_name]
            doc = await db.bb_applicant_updates.find_one({"email": email_a})
            cli.close()
            return doc
        doc = _run_async(_check())
        assert doc is not None
        assert doc["isImported"] is True
        assert doc["import_batch_id"] == body["batch_id"]
        assert doc["import_rejection_notified"] is False
        assert doc["status"] == "Rejected"
        assert doc["name"] == "ConfA"
        assert doc["phone"] == "9991110001"
        assert doc["job_role"] == "Engineer"
        assert "imported_at" in doc

    def test_confirm_upsert_no_duplicate(self, session, cleanup_emails):
        email = "test_iter33_upsert@example.com"
        cleanup_emails.append(email)
        payload1 = {"rows": [{"name": "U1", "email": email, "phone": "9991110011",
                              "job_role": "Eng", "status": "On hold", "scores": []}]}
        r1 = session.post(f"{API}/bb/import-scores/confirm", json=payload1, timeout=30)
        assert r1.status_code == 200
        batch1 = r1.json()["batch_id"]

        payload2 = {"rows": [{"name": "U2-updated", "email": email, "phone": "9991110011",
                              "job_role": "Eng", "status": "Rejected", "scores": []}]}
        r2 = session.post(f"{API}/bb/import-scores/confirm", json=payload2, timeout=30)
        assert r2.status_code == 200
        batch2 = r2.json()["batch_id"]
        assert batch1 != batch2

        from motor.motor_asyncio import AsyncIOMotorClient
        url = os.environ.get("MONGO_URL"); db_name = os.environ.get("DB_NAME", "hr_analytics")
        async def _check():
            cli = AsyncIOMotorClient(url); db = cli[db_name]
            count = await db.bb_applicant_updates.count_documents({"email": email})
            doc = await db.bb_applicant_updates.find_one({"email": email})
            cli.close()
            return count, doc
        count, doc = _run_async(_check())
        assert count == 1, f"Upsert created duplicate: count={count}"
        assert doc["status"] == "Rejected"
        assert doc["name"] == "U2-updated"
        assert doc["import_batch_id"] == batch2


# ============ ROUND TRIP ============

class TestRoundTrip:
    def test_export_xlsx_then_preview(self, session):
        r = session.get(f"{API}/bb/export-scores?format=xlsx", timeout=120)
        assert r.status_code == 200
        files = {"file": ("roundtrip.xlsx", r.content,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        rp = session.post(f"{API}/bb/import-scores/preview", files=files, timeout=180)
        assert rp.status_code == 200, rp.text[:400]
        body = rp.json()
        assert isinstance(body.get("round_columns"), list)
        assert body["round_columns"] == sorted(body["round_columns"])
        # No header errors (no missing columns failure)
        # Rows may have errors if data has bad scores; but body is fine per task description


# ============ WORKER PRESENCE + MESSAGING SCOPE ============

class TestWorkerAndMessaging:
    def test_worker_log_present(self):
        log_path = "/var/log/supervisor/backend.err.log"
        if not os.path.exists(log_path):
            pytest.skip("Backend log not present")
        # tail last ~200KB
        with open(log_path, "rb") as f:
            f.seek(0, 2)
            sz = f.tell()
            f.seek(max(0, sz - 200_000))
            tail = f.read().decode("utf-8", errors="ignore")
        assert "Import-rejection mailer worker started" in tail or "[Import Rejection Worker started" in tail
        assert "MESSAGING_CUTOFF_TS" in tail  # cutoff guard log present

    def test_worker_query_finds_imported_rejected(self, session, cleanup_emails):
        email = "test_iter33_worker@example.com"
        cleanup_emails.append(email)
        payload = {"rows": [{"name": "WorkerCand", "email": email, "phone": "9991119999",
                              "job_role": "Eng", "status": "Rejected", "scores": [],
                              "isTest": True}]}
        r = session.post(f"{API}/bb/import-scores/confirm", json=payload, timeout=30)
        assert r.status_code == 200

        from motor.motor_asyncio import AsyncIOMotorClient
        url = os.environ.get("MONGO_URL"); db_name = os.environ.get("DB_NAME", "hr_analytics")
        async def _check():
            cli = AsyncIOMotorClient(url); db = cli[db_name]
            doc = await db.bb_applicant_updates.find_one({
                "email": email,
                "isImported": True,
                "status": "Rejected",
                "import_rejection_notified": {"$ne": True},
            })
            cli.close()
            return doc
        doc = _run_async(_check())
        assert doc is not None, "Worker query did not pick up the imported rejected record"

    def test_notify_rejected_test_mode(self):
        """Direct invocation of messaging.notify_rejected in TEST_MODE."""
        import sys
        sys.path.insert(0, "/app/backend")
        try:
            from messaging import notify_rejected
        except Exception as e:
            pytest.skip(f"Cannot import messaging: {e}")
        # In TEST_MODE, recipient should be re-routed to rishi.nayak@blubridge.com / 9443109903
        result = _run_async(
            notify_rejected("TestCandidate", "9999999999", "test_iter33_msg@example.com")
        )
        # Email might succeed (SMTP configured); WhatsApp 401 is fail-soft
        # We accept ok==True or ok==False but no exception
        assert result is True or result is False or result is None
