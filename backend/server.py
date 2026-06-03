from dotenv import load_dotenv
from pathlib import Path
import sys as _sys

# Startup debug — surfaces in Render logs to pinpoint where boot stops.
print("SERVER STARTING...", flush=True)

# Ensure the backend directory is on sys.path so bare imports like
# `from bb_modules import ...`, `from messaging import ...`, `from _fmt import ...`
# resolve regardless of how uvicorn loads this module:
#   • `cd backend && uvicorn server:app`           (dev / supervisor)
#   • `uvicorn backend.server:app` from repo root  (Render start command)
_BACKEND_DIR = str(Path(__file__).resolve().parent)
if _BACKEND_DIR not in _sys.path:
    _sys.path.insert(0, _BACKEND_DIR)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, UploadFile, File, Depends, Query
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
from pymongo import ReturnDocument
import logging
import shutil
import glob as glob_module
import asyncio
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import jwt
import pandas as pd
import io
import re
import gc
import resource as _resource
from bson import ObjectId

# MongoDB connection
try:
    mongo_url = os.environ['MONGO_URL']
    _db_name = os.environ['DB_NAME']
except KeyError as _e:
    print(f"[startup] FATAL: required env var {_e} is missing.", flush=True)
    raise
client = AsyncIOMotorClient(mongo_url)
db = client[_db_name]
print("MONGO CLIENT INITIALIZED", flush=True)

# JWT Configuration
JWT_SECRET = os.environ.get("JWT_SECRET", "recruitment-analytics-secret-key")
JWT_ALGORITHM = "HS256"

# Create the main app
app = FastAPI()
print("FASTAPI APP CREATED", flush=True)
api_router = APIRouter(prefix="/api")

# ============ AUTH HELPERS ============

def create_token(username: str) -> str:
    payload = {
        "sub": username,
        "exp": datetime.now(timezone.utc) + timedelta(hours=24),
        "type": "access"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> str:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload["sub"]
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ============ DATA HELPERS ============

def normalize_phone(phone) -> str:
    if pd.isna(phone) or phone is None:
        return ""
    # Handle numeric types — pandas reads phone columns as float when NaN present
    if isinstance(phone, (int, float)):
        try:
            phone_str = str(int(phone))
        except (ValueError, OverflowError):
            return ""
    else:
        phone_str = str(phone).strip()
    # Remove all spaces
    phone_str = phone_str.replace(" ", "")
    # Handle comma-separated: take first number
    if "," in phone_str:
        phone_str = phone_str.split(",")[0].strip()
    # Strip non-digit characters (+, -, etc.)
    phone_str = re.sub(r'[^\d]', '', phone_str)
    # Normalize to 10 digits based on length
    if len(phone_str) == 10:
        return phone_str
    if len(phone_str) > 10:
        return phone_str[-10:]
    return phone_str

def normalize_email(email) -> str:
    if pd.isna(email) or email is None:
        return ""
    return str(email).strip().lower()

import time as _time_module

MONTH_MAP = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12"
}


def normalize_date(val) -> str:
    """Convert any date value/string to canonical YYYY-MM-DD format.
    Handles: pd.Timestamp, datetime, DD-MMM-YYYY, DD-MM-YYYY, YYYY-MM-DD, etc."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (pd.Timestamp, datetime)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # Already ISO? (YYYY-MM-DD)
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return s[:10]
    parts = re.split(r'[-/]', s)
    if len(parts) == 3:
        # DD-MMM-YYYY (24-Mar-2026)
        if parts[1].lower() in MONTH_MAP:
            mm = MONTH_MAP[parts[1].lower()]
            return f"{parts[2]}-{mm}-{parts[0].zfill(2)}"
        # DD-MM-YYYY (24-03-2026) — day first if parts[0] <= 31
        if len(parts[0]) <= 2 and len(parts[2]) == 4:
            return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
        # MM-DD-YYYY fallback
        if len(parts[2]) == 4:
            return f"{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"
    return s


def clean_value(val):
    if pd.isna(val) or val is None:
        return None
    if isinstance(val, (pd.Timestamp, datetime)):
        # Date-only timestamps → ISO YYYY-MM-DD
        if hasattr(val, 'hour') and val.hour == 0 and val.minute == 0 and val.second == 0:
            return val.strftime("%Y-%m-%d")
        return val.isoformat()
    if isinstance(val, _time_module.struct_time):
        return str(val)
    # Handle datetime.time objects (from Excel time columns like schedule_time)
    if hasattr(val, 'hour') and hasattr(val, 'minute') and not isinstance(val, datetime):
        return val.strftime("%I:%M %p") if hasattr(val, 'strftime') else str(val)
    return val


def build_sort(sort_by: Optional[str], sort_dir: Optional[str], allowed: dict, default: dict) -> dict:
    """Translate API-facing sort_by/sort_dir into a Mongo $sort dict.
    `allowed` maps the public field name to its concrete DB field path.
    Falls back to `default` when input is empty or not whitelisted.
    sort_dir: 'asc' (default) or 'desc'.
    """
    if not sort_by:
        return default
    db_field = allowed.get(sort_by)
    if not db_field:
        return default
    direction = -1 if (sort_dir or "").lower() == "desc" else 1
    return {db_field: direction}


def _rss_mb() -> float:
    """Return current process resident memory in MB. Linux: ru_maxrss is in KB.
    macOS: ru_maxrss is in bytes. iter116 — added so bulk-upload worker can
    log RAM before/after each XLSX parse + after GC, making Render OOM
    incidents (512 MB cap) traceable to a specific upload."""
    try:
        kb = _resource.getrusage(_resource.RUSAGE_SELF).ru_maxrss
        # On macOS this is bytes; on Linux it's KB. Detect via magnitude.
        return round(kb / 1024.0, 1) if kb > 10_000_000 else round(kb / 1024.0, 1)
    except Exception:
        return -1.0


def parse_file(file_content: bytes, filename: str) -> pd.DataFrame:
    if filename.lower().endswith('.csv'):
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                return pd.read_csv(io.BytesIO(file_content), encoding=encoding)
            except UnicodeDecodeError:
                continue
        raise ValueError("Unable to decode CSV file")
    elif filename.lower().endswith(('.xlsx', '.xls')):
        return pd.read_excel(io.BytesIO(file_content))
    else:
        raise ValueError("Unsupported file format. Use .csv or .xlsx")

# ============ SCHEMA MAPPING LAYER ============

# Naukri Applies: display column name → normalized DB field name
NAUKRI_COLUMN_MAP = {
    "Job Title": "job_title",
    "Date of application": "date_of_application",
    "Name": "name",
    "Email ID": "email",
    "Phone Number": "phone",
    "Current Location": "current_location",
    "Preferred Locations": "preferred_locations",
    "Total Experience": "total_experience",
    "Curr. Company name": "curr_company_name",
    "Curr. Company Designation": "curr_company_designation",
    "Department": "department",
    "Role": "role",
    "Industry": "industry",
    "Key Skills": "key_skills",
    "Annual Salary": "annual_salary",
    "Notice period/ Availability to join": "notice_period",
    "Resume Headline": "resume_headline",
    "Summary": "summary",
    "Under Graduation degree": "ug_degree",
    "UG Specialization": "ug_specialization",
    "UG University/institute Name": "ug_university",
    "UG Graduation year": "ug_graduation_year",
    "Post graduation degree": "pg_degree",
    "PG specialization": "pg_specialization",
    "PG university/institute name": "pg_university",
    "PG graduation year": "pg_graduation_year",
    "Doctorate degree": "doctorate_degree",
    "Doctorate specialization": "doctorate_specialization",
    "Doctorate university/institute name": "doctorate_university",
    "Doctorate graduation year": "doctorate_graduation_year",
    "Gender": "gender",
    "Marital Status": "marital_status",
    "Home Town/City": "home_town",
    "Pin Code": "pin_code",
    "Work permit for USA": "work_permit_usa",
    "Date of Birth": "date_of_birth",
    "Permanent Address": "permanent_address",
    "Last Workflow activity": "last_workflow_activity",
    "Last Workflow activity by": "last_workflow_activity_by",
    "Time of Last Workflow activity Update": "time_last_workflow_update",
    "Latest Pipeline Stage": "latest_pipeline_stage",
    "Pipeline Status Updated By": "pipeline_status_updated_by",
    "Time when Stage updated": "time_stage_updated",
    "Download": "download",
    "Downloaded By": "downloaded_by",
    "Time Of Download": "time_of_download",
    "Viewed": "viewed",
    "Viewed By": "viewed_by",
    "Time Of View": "time_of_view",
    "Emailed": "emailed",
    "Emailed By": "emailed_by",
    "Time Of Email": "time_of_email",
    "Calling Status": "calling_status",
    "Calling Status updated by": "calling_status_updated_by",
    "Time of Calling activity update": "time_calling_update",
    "Comment 1": "comment_1",
    "Comment 1 BY": "comment_1_by",
    "Time Comment 1 posted": "time_comment_1",
    "Comment 2": "comment_2",
    "Comment 2 BY": "comment_2_by",
    "Time Comment 2 posted": "time_comment_2",
    "Comment 3": "comment_3",
    "Comment 3 BY": "comment_3_by",
    "Time Comment 3 posted": "time_comment_3",
    "Comment 4": "comment_4",
    "Comment 4 BY": "comment_4_by",
    "Time Comment 4 posted": "time_comment_4",
    "Comment 5": "comment_5",
    "Comment 5 BY": "comment_5_by",
    "Time Comment 5 posted": "time_comment_5",
    "Source": "source",
    "Candidate profile": "candidate_profile",
}

# Case-insensitive reverse lookup for Naukri column matching
_NAUKRI_CI_LOOKUP = {k.strip().lower(): v for k, v in NAUKRI_COLUMN_MAP.items()}

# HR Internal Pipeline: canonical column set (already snake_case)
PIPELINE_EXPECTED_COLUMNS = {
    "id", "name", "email", "phone", "age", "gender", "hr_team",
    "year_of_graduation", "college", "college_type", "degree", "course",
    "submitted_at", "last_update", "email_send", "current_location",
    "location", "state", "loca_change", "attend_inperson", "job_role",
    "email_type", "confirm_box", "resend", "whatsapp_reminder_sent",
    "schedule_date", "schedule_time", "reschedule_count", "otp",
    "otp_verified", "otp_expired", "otp_send", "message_id",
    "result_mail", "result_update", "result_status", "reschedule",
    "reschedule_mail", "rescheduled", "reschedule_date",
}

# Case-insensitive lookup for pipeline columns
_PIPELINE_CI_LOOKUP = {c.lower(): c for c in PIPELINE_EXPECTED_COLUMNS}

# ============ AUTH ENDPOINTS ============

import bcrypt as _bcrypt


def _hash_pw(plain: str) -> str:
    return _bcrypt.hashpw(plain.encode("utf-8"), _bcrypt.gensalt()).decode("utf-8")


def _verify_pw(plain: str, hashed: str) -> bool:
    try:
        return _bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


async def _seed_admin_user():
    """iter77 — Idempotent admin seeding. On first boot, creates the
    `Admin User` document in `bb_users` with bcrypt-hashed default
    password. Subsequent boots do nothing if the user already exists —
    so users can change the password without it being reset on restart."""
    existing = await db.bb_users.find_one({"username": "Admin User"})
    if existing is None:
        await db.bb_users.insert_one({
            "username": "Admin User",
            "password_hash": _hash_pw("Admin User"),
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })


class LoginRequest(BaseModel):
    username: str
    password: str


@api_router.post("/login")
async def login(response: Response, data: LoginRequest):
    # iter77 — Verify against bb_users (bcrypt). On miss, fall back to the
    # legacy hardcoded default once and seed the user — keeps existing
    # deployments unaffected during the migration window.
    user_doc = await db.bb_users.find_one({"username": data.username})
    if user_doc and _verify_pw(data.password, user_doc.get("password_hash", "")):
        ok = True
    elif data.username == "Admin User" and data.password == "Admin User" and user_doc is None:
        await _seed_admin_user()
        ok = True
    else:
        ok = False

    if not ok:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    token = create_token(data.username)
    response.set_cookie(
        key="access_token", value=token, httponly=True, secure=False,
        samesite="lax", max_age=86400, path="/",
    )
    return {"success": True, "message": "Login successful", "username": data.username}


@api_router.post("/logout")
async def logout(response: Response):
    response.delete_cookie(key="access_token", path="/")
    return {"success": True, "message": "Logged out"}


@api_router.get("/auth/check")
async def check_auth(user: str = Depends(get_current_user)):
    return {"authenticated": True, "username": user}


@api_router.get("/me")
async def get_me(user: str = Depends(get_current_user)):
    """iter77 — Profile info for the currently signed-in user."""
    doc = await db.bb_users.find_one({"username": user}, {"_id": 0, "password_hash": 0})
    if not doc:
        return {"username": user, "role": "admin", "created_at": None}
    return doc


class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str


@api_router.post("/auth/change-password")
async def change_password(data: ChangePasswordRequest, user: str = Depends(get_current_user)):
    """iter77 — Verify old password via bcrypt; replace stored hash."""
    if not data.new_password or len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    if data.old_password == data.new_password:
        raise HTTPException(status_code=400, detail="New password must be different from old password")

    user_doc = await db.bb_users.find_one({"username": user})
    if user_doc is None:
        # First-time path — accept legacy default once, then create the
        # user with the new password. Defensive against accounts that
        # somehow never got seeded.
        if data.old_password != "Admin User":
            raise HTTPException(status_code=401, detail="Old password is incorrect")
        await db.bb_users.insert_one({
            "username": user,
            "password_hash": _hash_pw(data.new_password),
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        return {"success": True, "message": "Password updated"}

    if not _verify_pw(data.old_password, user_doc.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Old password is incorrect")
    await db.bb_users.update_one(
        {"username": user},
        {"$set": {
            "password_hash": _hash_pw(data.new_password),
            "password_updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    return {"success": True, "message": "Password updated"}

# ============ UPLOAD ENDPOINTS ============

@api_router.post("/upload/naukri")
async def upload_naukri(file: UploadFile = File(...), user: str = Depends(get_current_user)):
    """Upload Naukri Applies data — mapping-driven schema alignment"""
    try:
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File size exceeds 10MB limit")

        df = parse_file(content, file.filename)
        df.columns = df.columns.str.strip()

        # Map CSV columns → DB field names using the mapping dictionary
        col_map = {}  # csv_column_name → db_field_name
        for csv_col in df.columns:
            if csv_col in NAUKRI_COLUMN_MAP:
                col_map[csv_col] = NAUKRI_COLUMN_MAP[csv_col]
            elif csv_col.strip().lower() in _NAUKRI_CI_LOOKUP:
                col_map[csv_col] = _NAUKRI_CI_LOOKUP[csv_col.strip().lower()]

        mapped_fields = set(col_map.values())
        if "email" not in mapped_fields and "phone" not in mapped_fields:
            raise HTTPException(status_code=400, detail="Could not detect 'Email ID' or 'Phone Number' column")

        total = len(df)
        inserted = 0
        updated = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                doc = {}
                # Store mapped columns
                for csv_col, db_field in col_map.items():
                    doc[db_field] = clean_value(row.get(csv_col))

                # Store unmapped columns (data loss prevention)
                for csv_col in df.columns:
                    if csv_col not in col_map:
                        safe = re.sub(r'[^\w]', '_', csv_col.strip().lower()).strip('_')
                        doc[f"_extra_{safe}"] = clean_value(row.get(csv_col))

                # Normalize identifiers
                email = normalize_email(doc.get("email"))
                phone = normalize_phone(doc.get("phone"))
                doc["email"] = email
                doc["phone"] = phone

                # Normalize date fields to ISO YYYY-MM-DD
                for date_field in ("date_of_application", "date_of_birth"):
                    if doc.get(date_field):
                        doc[date_field] = normalize_date(doc[date_field])

                doc["updated_at"] = datetime.now(timezone.utc)

                if not email and not phone:
                    errors.append(f"Row {idx + 2}: Missing Email ID and Phone Number")
                    continue

                # UPSERT on email OR phone
                query = {"$or": []}
                if email:
                    query["$or"].append({"email": email})
                if phone:
                    query["$or"].append({"phone": phone})

                existing = await db.naukri_applies.find_one(query)
                if existing:
                    await db.naukri_applies.update_one({"_id": existing["_id"]}, {"$set": doc})
                    updated += 1
                else:
                    doc["created_at"] = datetime.now(timezone.utc)
                    await db.naukri_applies.insert_one(doc)
                    inserted += 1

            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        # iter123 — Defer heavy post-upload reprocessing to background task.
        # `reprocess_matching()` + `_sync_job_titles_master()` traverse all
        # rows; for production-sized collections they easily exceed Render's
        # 30s HTTP request timeout, producing 502s. The bulk-upload route
        # already runs async; individual uploads now match that pattern.
        async def _bg_post_upload_naukri():
            try:
                await reprocess_matching()
                await _sync_job_titles_master()
                logger.info("[upload/naukri:bg_post] reprocess + sync complete")
            except Exception as _be:
                logger.exception(f"[upload/naukri:bg_post] failed: {_be!r}")
        asyncio.create_task(_bg_post_upload_naukri())

        return {
            "success": True,
            "message": f"Naukri data uploaded. Inserted: {inserted}, Updated: {updated}",
            "total": total,
            "inserted": inserted,
            "updated": updated,
            "errors": errors[:10],
            "mapped_columns": len(col_map),
            "unmapped_columns": len(df.columns) - len(col_map),
            "background_processing": True,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/upload/pipeline")
async def upload_pipeline(file: UploadFile = File(...), user: str = Depends(get_current_user)):
    """Upload HR Internal Pipeline data — handles duplicate columns, strict schema"""
    try:
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File size exceeds 10MB limit")

        df = parse_file(content, file.filename)
        df.columns = df.columns.str.strip()

        # Deduplicate columns: keep first occurrence only
        cols = df.columns.tolist()
        seen_bases = set()
        keep_indices = []
        for i, col in enumerate(cols):
            base = re.sub(r'\.\d+$', '', col.strip()).lower()
            if base not in seen_bases:
                seen_bases.add(base)
                keep_indices.append(i)
        df = df.iloc[:, keep_indices]
        df.columns = [re.sub(r'\.\d+$', '', c.strip()) for c in df.columns]

        # Map CSV columns → DB field names
        col_map = {}
        for csv_col in df.columns:
            normalized = csv_col.strip().lower()
            if normalized in _PIPELINE_CI_LOOKUP:
                db_field = _PIPELINE_CI_LOOKUP[normalized]
                col_map[csv_col] = "pipeline_id" if db_field == "id" else db_field

        mapped_fields = set(col_map.values())
        if "email" not in mapped_fields and "phone" not in mapped_fields:
            raise HTTPException(status_code=400, detail="Could not detect 'email' or 'phone' column")

        total = len(df)
        inserted = 0
        updated = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                doc = {}
                for csv_col, db_field in col_map.items():
                    doc[db_field] = clean_value(row.get(csv_col))

                # Store unmapped columns (data loss prevention)
                for csv_col in df.columns:
                    if csv_col not in col_map:
                        safe = re.sub(r'[^\w]', '_', csv_col.strip().lower()).strip('_')
                        doc[f"_extra_{safe}"] = clean_value(row.get(csv_col))

                email = normalize_email(doc.get("email"))
                phone = normalize_phone(doc.get("phone"))
                doc["email"] = email
                doc["phone"] = phone
                doc["updated_at"] = datetime.now(timezone.utc)

                if not email and not phone:
                    errors.append(f"Row {idx + 2}: Missing email and phone")
                    continue

                # UPSERT on email OR phone
                query = {"$or": []}
                if email:
                    query["$or"].append({"email": email})
                if phone:
                    query["$or"].append({"phone": phone})

                existing = await db.pipeline_data.find_one(query)
                if existing:
                    await db.pipeline_data.update_one({"_id": existing["_id"]}, {"$set": doc})
                    updated += 1
                else:
                    doc["created_at"] = datetime.now(timezone.utc)
                    await db.pipeline_data.insert_one(doc)
                    inserted += 1

            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        # iter123 — Defer heavy post-upload reprocessing to background task
        # to prevent Render 502 timeouts on large pipeline files.
        async def _bg_post_upload_pipeline():
            try:
                await reprocess_matching()
                await _sync_job_titles_master()
                logger.info("[upload/pipeline:bg_post] reprocess + sync complete")
            except Exception as _be:
                logger.exception(f"[upload/pipeline:bg_post] failed: {_be!r}")
        asyncio.create_task(_bg_post_upload_pipeline())

        return {
            "success": True,
            "message": f"Pipeline data uploaded. Inserted: {inserted}, Updated: {updated}",
            "total": total,
            "inserted": inserted,
            "updated": updated,
            "errors": errors[:10],
            "mapped_columns": len(col_map),
            "unmapped_columns": len(df.columns) - len(col_map),
            "background_processing": True,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============ MATCHING & PROCESSING ============

def is_null_or_empty(val):
    """Check if a value is None, empty string, or missing"""
    return val is None or val == ""

async def renormalize_collection(collection_name: str):
    """Re-normalize email, phone, and date fields in an existing collection."""
    coll = db[collection_name]
    cursor = coll.find({})
    fixed = 0
    async for doc in cursor:
        updates = {}
        old_email = doc.get("email", "")
        old_phone = doc.get("phone", "")
        new_email = normalize_email(old_email)
        new_phone = normalize_phone(old_phone)
        if new_email != old_email:
            updates["email"] = new_email
        if new_phone != old_phone:
            updates["phone"] = new_phone
        # Normalize date fields
        for df_field in ("date_of_application", "date_of_birth"):
            old_val = doc.get(df_field)
            if old_val:
                new_val = normalize_date(old_val)
                if new_val and new_val != str(old_val):
                    updates[df_field] = new_val
        if updates:
            await coll.update_one({"_id": doc["_id"]}, {"$set": updates})
            fixed += 1
    return fixed


async def _persist_derived_fields(collection_name: str):
    """Compute & persist `_college_status`, `_nirf_category`, `_college_resolved`,
    `_match_confidence`, `_normalized_job_role` for every document in `collection_name`.
    Used after `reprocess_matching` and bulk uploads to keep endpoints fast."""
    from pymongo import UpdateOne
    coll = db[collection_name]
    rank_lookup = await _build_college_rank_lookup()
    mappings = await _get_job_keyword_mappings()

    ops = []
    cursor = coll.find({}, {
        "_id": 1, "ug_university": 1, "pg_university": 1,
        "college": 1, "job_title": 1, "job_role": 1,
    })
    async for doc in cursor:
        cc = _classify_college(doc, rank_lookup)
        cs = cc["college_status"]
        # iter110 — `_nirf_category` keeps the binary premium gate ("NIRF" =
        # rank 1..100) while exposing the full bucketed label on every row.
        cat = "NIRF" if cs.startswith("NIRF - #") else cs
        # Pipeline records use `job_role`; naukri uses `job_title`
        raw_role = doc.get("job_title") or doc.get("job_role") or ""
        normalized_role = _resolve_normalized_job_role(raw_role, mappings)
        ops.append(UpdateOne(
            {"_id": doc["_id"]},
            {"$set": {
                "_college_status": cs,
                "_nirf_category": cat,
                "_college_resolved": cc.get("college") or "-",
                "_match_confidence": cc.get("match_confidence") or None,
                "_normalized_job_role": normalized_role or "Unknown",
            }}
        ))
        if len(ops) >= 1000:
            await coll.bulk_write(ops, ordered=False)
            ops = []
    if ops:
        await coll.bulk_write(ops, ordered=False)

    await coll.create_index("_college_status")
    await coll.create_index("_nirf_category")
    await coll.create_index("_normalized_job_role")


async def reprocess_matching():
    """Rebuild `registered_candidates` as the INNER-JOIN enriched view.

    CLASSIFICATION RULE (May 2026, view-based):
      - Registered   = every record in `pipeline_data` (HR internal).
      - Unregistered = naukri_applies records with no pipeline match.
    For disk-efficiency on Atlas free tier we do NOT duplicate all pipeline
    records into `registered_candidates`. Endpoints read counts directly from
    raw collections (`/api/data/registered`, `/api/data/unregistered`,
    `/api/data/classification`). `registered_candidates` stays as the
    enriched (naukri+pipeline) join view used by scoring/scheduling pages.

    Test records (`isTest: true`) in raw collections are skipped.
    """

    # Step 1: Re-normalize identifiers in raw collections so matching works
    await renormalize_collection("naukri_applies")
    await renormalize_collection("pipeline_data")

    naukri_list = await db.naukri_applies.find({"isTest": {"$ne": True}}).to_list(None)
    pipeline_list = await db.pipeline_data.find({"isTest": {"$ne": True}}).to_list(None)

    pipeline_by_email = {}
    pipeline_by_phone = {}
    for p in pipeline_list:
        em = normalize_email(p.get('email'))
        ph = normalize_phone(p.get('phone'))
        if em:
            pipeline_by_email[em] = p
        if ph:
            pipeline_by_phone[ph] = p

    await db.registered_candidates.drop()

    registered_docs = []
    naukri_updates = []
    _skip_keys = {"_id", "_is_registered", "created_at", "updated_at"}

    for naukri in naukri_list:
        email = normalize_email(naukri.get('email'))
        phone = normalize_phone(naukri.get('phone'))

        pipeline_match = None
        if email and email in pipeline_by_email:
            pipeline_match = pipeline_by_email[email]
        elif phone and phone in pipeline_by_phone:
            pipeline_match = pipeline_by_phone[phone]

        is_registered = pipeline_match is not None

        # Batched update for speed (was per-row await update_one — minutes on Atlas)
        from pymongo import UpdateOne as _UpdateOne
        naukri_updates.append(_UpdateOne(
            {"_id": naukri["_id"]},
            {"$set": {"_is_registered": is_registered}}
        ))

        if is_registered:
            doc = {k: v for k, v in pipeline_match.items() if k not in _skip_keys}
            for k, v in naukri.items():
                if k in _skip_keys:
                    continue
                if v is not None and v != "":
                    doc[k] = v
                elif k not in doc:
                    doc[k] = v
            if not doc.get("job_title"):
                doc["job_title"] = doc.get("job_role") or ""
            doc["_has_naukri_match"] = True
            registered_docs.append(doc)

    # Flush naukri _is_registered updates in bulk (1 round-trip per chunk vs N)
    chunk = 1000
    for i in range(0, len(naukri_updates), chunk):
        await db.naukri_applies.bulk_write(naukri_updates[i:i + chunk], ordered=False)

    chunk = 2000
    for i in range(0, len(registered_docs), chunk):
        await db.registered_candidates.insert_many(registered_docs[i:i + chunk])

    await db.registered_candidates.create_index([("email", 1), ("phone", 1)])
    await db.registered_candidates.create_index("email_type")
    await db.registered_candidates.create_index("result_status")
    await db.registered_candidates.create_index("schedule_date")
    await db.registered_candidates.create_index("otp_verified")
    await db.naukri_applies.create_index("_is_registered")

    await _persist_derived_fields("registered_candidates")
    await _persist_derived_fields("naukri_applies")
    # iter125 — Persist `_normalized_job_role` on pipeline_data too. Without
    # this pass, freshly-uploaded HR pipeline rows had no normalized field
    # set and were excluded from `/api/job-roles` (which filters on
    # `_normalized_job_role NOT IN [None,"","Unknown"]`). New job roles
    # entering through pipeline uploads therefore stayed invisible on the
    # Job Roles page even though `_sync_job_titles_master` correctly
    # mirrored them into `bb_job_roles` and `job_titles_master`.
    await _persist_derived_fields("pipeline_data")

# ============ JOB ROLE NORMALIZATION ============

def _normalize_text_for_matching(text: str) -> str:
    """Normalize text for keyword matching: lowercase, trim, remove punctuation."""
    if not text:
        return ""
    return re.sub(r'[^\w\s]', '', text.lower().strip())


async def _sync_job_titles_master():
    """Extract distinct job titles from BOTH naukri_applies AND pipeline_data,
    and upsert into job_titles_master + bb_job_roles (case-insensitive
    deduplication). Spec #10A — keeps the mapping picker AND the Job Roles
    page in sync after every dataset upload.

    iter125 — Hardened with structured logging for production debugging:
      [JobRoleSync] DETECTED new_role="<raw>" source=<naukri|pipeline>
      [JobRoleSync] INSERT job_titles_master normalized="<norm>"
      [JobRoleSync] INSERT bb_job_roles name="<raw>"
      [JobRoleSync] SUMMARY scanned=<N> jtm_inserts=<X> bb_inserts=<Y>

    iter127 — Coverage extended:
      * Also scans `_normalized_job_role` (the resolved canonical value
        shown in Analytics) — previously omitted, so canonical roles like
        "AI & ML Engineer" derived from raw "AI And ML Engineer - C++ or
        Java Developer" never made it into the catalog.
      * Also scans `registered_candidates` — covers the college-drive
        intake that bypasses pipeline_data/naukri_applies.
      * Eliminates the gap where analytics shows a role but Job Roles /
        Job Role dropdown / Unmapped Job Keywords don't.
    """
    sources = []
    # iter127 — Scan ALL fields that can surface a role in any user-facing
    # view. Order matters only for the source tag in logs; the dedupe key
    # (normalized title) collapses duplicates across sources.
    scan_targets = [
        # (collection, field, source_tag)
        (db.naukri_applies, "job_title", "naukri"),
        (db.naukri_applies, "_normalized_job_role", "naukri_canonical"),
        (db.pipeline_data, "job_role", "pipeline"),
        (db.pipeline_data, "job_title", "pipeline_legacy"),
        (db.pipeline_data, "_normalized_job_role", "pipeline_canonical"),
        (db.registered_candidates, "job_role", "registered"),
        (db.registered_candidates, "job_title", "registered_legacy"),
        (db.registered_candidates, "_normalized_job_role", "registered_canonical"),
    ]
    for coll, field, tag in scan_targets:
        try:
            cursor = coll.aggregate([
                {"$match": {field: {"$nin": [None, ""]}}},
                {"$group": {"_id": f"${field}"}},
            ])
            async for r in cursor:
                v = r.get("_id")
                if v:
                    sources.append((v, tag))
        except Exception as _scan_err:
            logger.warning(
                f"[JobRoleSync] scan skipped coll={coll.name} field={field}: {_scan_err!r}"
            )

    seen = set()
    scanned = 0
    jtm_inserts = 0
    bb_inserts = 0
    for raw_value, source_tag in sources:
        scanned += 1
        raw = str(raw_value or "").strip()
        if not raw:
            continue
        # iter127 — Skip the literal "Unknown" bucket; it's not a real role.
        if raw.lower() == "unknown":
            continue
        normalized = _normalize_text_for_matching(raw)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)

        existing = await db.job_titles_master.find_one({"normalized_job_title": normalized})
        if not existing:
            try:
                await db.job_titles_master.insert_one({
                    "raw_job_title": raw,
                    "normalized_job_title": normalized,
                    "is_mapped": False,
                })
                jtm_inserts += 1
                logger.info(
                    f"[JobRoleSync] DETECTED new_role={raw!r} source={source_tag} | "
                    f"INSERT job_titles_master normalized={normalized!r}"
                )
            except Exception as _e:
                # Likely a race on the unique index; safe to swallow.
                logger.warning(
                    f"[JobRoleSync] job_titles_master insert skipped for {raw!r}: {_e}"
                )

        # iter69f (#10A) — auto-upsert into bb_job_roles so the Job Roles page
        # always lists every distinct title from imports + manual creates.
        # Case-insensitive match against existing rows.
        bb_existing = await db.bb_job_roles.find_one(
            {"name": {"$regex": f"^{re.escape(raw)}$", "$options": "i"}}
        )
        if not bb_existing:
            try:
                await db.bb_job_roles.insert_one({
                    "name": raw,
                    "source": "imported",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                })
                bb_inserts += 1
                logger.info(
                    f"[JobRoleSync] INSERT bb_job_roles name={raw!r} source={source_tag}"
                )
            except Exception as _e:
                logger.warning(
                    f"[JobRoleSync] bb_job_roles insert skipped for {raw!r}: {_e}"
                )

    logger.info(
        f"[JobRoleSync] SUMMARY scanned={scanned} unique={len(seen)} "
        f"jtm_inserts={jtm_inserts} bb_inserts={bb_inserts}"
    )
    return {"scanned": scanned, "unique": len(seen), "jtm_inserts": jtm_inserts, "bb_inserts": bb_inserts}


# iter127 — Periodic safety-net sync. Runs `_sync_job_titles_master` every
# `JOB_ROLE_SYNC_INTERVAL_SECONDS` seconds (default 900 = 15 min) so the
# catalog converges even if a post-upload background task silently dies
# (Render redeploy mid-task, OOM kill, unhandled exception in the
# `_bg_post_upload_*` task wrapper). The sync function itself is
# idempotent (dedupes via case-insensitive lookups + the unique index on
# job_titles_master.normalized_job_title), so re-running it is harmless
# and cheap when nothing changed.
async def _periodic_job_titles_sync():
    interval = int(os.environ.get("JOB_ROLE_SYNC_INTERVAL_SECONDS", "900"))
    if interval < 60:
        interval = 60  # floor at 1 minute
    logger.info(f"[JobRoleSync] periodic safety-net started interval={interval}s")
    # First run is immediate (covers the case where the startup one-shot
    # was scheduled but hasn't run yet); after that, wait `interval`.
    while True:
        try:
            await asyncio.sleep(interval)
            await _sync_job_titles_master()
        except asyncio.CancelledError:
            raise
        except Exception as e:
            logger.warning(f"[JobRoleSync] periodic tick failed: {e!r}")


async def _get_job_keyword_mappings() -> list:
    """Fetch all job keyword mappings from DB."""
    return await db.job_keyword_mapping.find({}, {"_id": 0}).to_list(None)


# iter99 — Single source of truth for keyword → canonical-title lookup.
# Returns a dict[normalized_keyword] -> canonical_job_role and a set of all
# normalized canonical titles. Callers pass this into per-row resolution
# helpers so we hit Mongo only once per request, not once per row.
async def _build_canonical_index() -> tuple:
    """Returns (kw_to_canonical: dict, canonical_set: set). All keys are
    normalized via `_normalize_text_for_matching`. The canonical_set also
    includes each canonical role's own normalized name so an applicant
    already on a canonical title resolves to itself."""
    mappings = await _get_job_keyword_mappings()
    kw_to_canonical = {}
    canonical_set = set()
    for m in mappings:
        canonical = (m.get("job_role") or "").strip()
        if not canonical:
            continue
        canonical_set.add(_normalize_text_for_matching(canonical))
        for kw in m.get("keywords", []) or []:
            nk = _normalize_text_for_matching(kw)
            if nk:
                kw_to_canonical[nk] = canonical
        # Also map the canonical-name → itself so reverse lookups are cheap.
        kw_to_canonical[_normalize_text_for_matching(canonical)] = canonical
    return kw_to_canonical, canonical_set


def _canonicalize_job_role(raw: str, kw_to_canonical: dict) -> str:
    """Read-time canonicalization: given a raw job-role string and the
    pre-built keyword index, return the canonical title if mapped, else the
    raw input unchanged. Empty/None inputs return ''. Does NOT mutate the DB.
    """
    if not raw:
        return ""
    nk = _normalize_text_for_matching(raw)
    return kw_to_canonical.get(nk, raw)


def _resolve_normalized_job_role(job_title: str, mappings: list) -> str:
    """Given a raw job_title and keyword mappings, return the canonical job role.
    Matches by exact normalized comparison (keywords are full job titles).
    Returns first match or falls back to raw job_title."""
    if not job_title:
        return job_title or "Unknown"
    normalized_title = _normalize_text_for_matching(job_title)
    for mapping in mappings:
        for keyword in mapping.get("keywords", []):
            kw_normalized = _normalize_text_for_matching(keyword)
            if kw_normalized and kw_normalized == normalized_title:
                return mapping["job_role"]
    return job_title


class JobKeywordMappingCreate(BaseModel):
    job_role: str
    keywords: List[str]


class JobKeywordMappingUpdate(BaseModel):
    job_role: Optional[str] = None
    keywords: Optional[List[str]] = None


@api_router.get("/job-titles/unmatched")
async def get_unmatched_job_titles(user: str = Depends(get_current_user)):
    """Return job titles not yet mapped to a canonical job role.

    iter99 — Two-layer dedupe:
      1. Pull EVERY raw title from job_titles_master AND every imported name
         from bb_job_roles (so manually-created job roles also appear in the
         unmapped list until they're mapped).
      2. Collapse by `_normalize_text_for_matching(title)` so case/punctuation
         variants ('ABC' vs 'abc' vs 'A.B.C.') merge into ONE row.
      3. Exclude anything already in `job_keyword_mapping.keywords[]` OR
         used as a canonical `job_role` (regardless of the stale `is_mapped`
         flag on `job_titles_master`).
    """
    kw_to_canonical, _ = await _build_canonical_index()
    mapped_norm_set = set(kw_to_canonical.keys())

    candidates: dict = {}  # normalized -> first raw seen

    async for t in db.job_titles_master.find({}, {"_id": 0, "raw_job_title": 1, "normalized_job_title": 1}):
        raw = (t.get("raw_job_title") or "").strip()
        if not raw:
            continue
        norm = t.get("normalized_job_title") or _normalize_text_for_matching(raw)
        if not norm or norm in mapped_norm_set:
            continue
        candidates.setdefault(norm, raw)

    # bb_job_roles also feeds the unmapped picker (per spec: manually-created
    # job roles should appear until mapped).
    async for r in db.bb_job_roles.find({}, {"_id": 0, "name": 1}):
        raw = (r.get("name") or "").strip()
        if not raw:
            continue
        norm = _normalize_text_for_matching(raw)
        if not norm or norm in mapped_norm_set:
            continue
        candidates.setdefault(norm, raw)

    titles = sorted(candidates.values(), key=lambda s: s.lower())
    return {"titles": titles}


@api_router.get("/job-keyword-mappings")
async def list_job_keyword_mappings(user: str = Depends(get_current_user)):
    """List all job keyword mappings."""
    mappings = await db.job_keyword_mapping.find({}).to_list(None)
    for m in mappings:
        m["id"] = str(m.pop("_id"))
    return {"mappings": mappings}


@api_router.post("/job-keyword-mappings")
async def create_job_keyword_mapping(data: JobKeywordMappingCreate, user: str = Depends(get_current_user)):
    """Create a new job keyword mapping and mark keywords as mapped."""
    keywords = [k.strip() for k in data.keywords if k.strip()]
    if not keywords:
        raise HTTPException(status_code=400, detail="At least one keyword is required")
    doc = {
        "job_role": data.job_role.strip(),
        "keywords": keywords,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.job_keyword_mapping.insert_one(doc)
    # Mark keywords as mapped in job_titles_master
    for kw in keywords:
        norm = _normalize_text_for_matching(kw)
        if norm:
            await db.job_titles_master.update_many(
                {"normalized_job_title": norm},
                {"$set": {"is_mapped": True}}
            )
    # iter108 — Trigger background reprocess so existing applicants whose raw
    # job_title matches one of the new keywords get reclassified out of "Unknown".
    asyncio.create_task(_trigger_deferred_reprocess(reason=f"mapping_create:{doc['job_role']}"))
    return {"success": True, "id": str(result.inserted_id), "job_role": doc["job_role"], "keywords": doc["keywords"]}


@api_router.put("/job-keyword-mappings/{mapping_id}")
async def update_job_keyword_mapping(mapping_id: str, data: JobKeywordMappingUpdate, user: str = Depends(get_current_user)):
    """Update an existing job keyword mapping. Handles keyword add/remove and is_mapped flags."""
    try:
        oid = ObjectId(mapping_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid mapping ID")

    old_mapping = await db.job_keyword_mapping.find_one({"_id": oid})
    if not old_mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")

    updates = {}
    if data.job_role is not None:
        updates["job_role"] = data.job_role.strip()

    new_keywords = None
    if data.keywords is not None:
        new_keywords = [k.strip() for k in data.keywords if k.strip()]
        updates["keywords"] = new_keywords

    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")

    await db.job_keyword_mapping.update_one({"_id": oid}, {"$set": updates})

    # Handle is_mapped flag changes for keywords
    if new_keywords is not None:
        old_keywords = set(old_mapping.get("keywords", []))
        new_keywords_set = set(new_keywords)
        removed = old_keywords - new_keywords_set
        added = new_keywords_set - old_keywords

        # Unmap removed keywords (only if not used by another mapping)
        for kw in removed:
            norm = _normalize_text_for_matching(kw)
            if norm:
                other = await db.job_keyword_mapping.find_one(
                    {"keywords": kw, "_id": {"$ne": oid}}
                )
                if not other:
                    await db.job_titles_master.update_many(
                        {"normalized_job_title": norm},
                        {"$set": {"is_mapped": False}}
                    )

        # Map newly added keywords
        for kw in added:
            norm = _normalize_text_for_matching(kw)
            if norm:
                await db.job_titles_master.update_many(
                    {"normalized_job_title": norm},
                    {"$set": {"is_mapped": True}}
                )

    # iter108 — Reclassify existing applicants whenever keywords or the
    # canonical job_role label changes.
    asyncio.create_task(_trigger_deferred_reprocess(reason=f"mapping_update:{mapping_id}"))
    return {"success": True}


@api_router.delete("/job-keyword-mappings/{mapping_id}")
async def delete_job_keyword_mapping(mapping_id: str, user: str = Depends(get_current_user)):
    """Delete a job keyword mapping and release its keywords."""
    try:
        oid = ObjectId(mapping_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid mapping ID")

    mapping = await db.job_keyword_mapping.find_one({"_id": oid})
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")

    await db.job_keyword_mapping.delete_one({"_id": oid})

    # Unmap keywords (only if not used by another mapping)
    for kw in mapping.get("keywords", []):
        norm = _normalize_text_for_matching(kw)
        if norm:
            other = await db.job_keyword_mapping.find_one({"keywords": kw})
            if not other:
                await db.job_titles_master.update_many(
                    {"normalized_job_title": norm},
                    {"$set": {"is_mapped": False}}
                )

    # iter108 — Deleting a mapping un-maps its keywords; existing applicants
    # whose raw title matched ONLY this mapping will now correctly fall back
    # to the raw title (or "Unknown" if raw is empty).
    asyncio.create_task(_trigger_deferred_reprocess(reason=f"mapping_delete:{mapping_id}"))
    return {"success": True}


# ============ DASHBOARD COUNTS ENDPOINT ============

# Helpers for NULL-safe MongoDB queries
_null_filter = {"$in": [None, ""]}
_not_null_filter = {"$nin": [None, ""], "$exists": True}


# ============ DRILL-DOWN DATA ENDPOINTS ============

@api_router.get("/data/unregistered")
async def get_unregistered(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    user: str = Depends(get_current_user)
):
    """Unregistered (May 2026 rule): present in `naukri_applies` but NOT in `pipeline_data`.
    Uses persisted `_is_registered` flag on naukri_applies."""
    skip = (page - 1) * limit
    query = {"_is_registered": {"$ne": True}, "isTest": {"$ne": True}}
    total = await db.naukri_applies.count_documents(query)
    cursor = db.naukri_applies.find(query, {
        "_id": 0, "name": 1, "email": 1, "phone": 1,
        "job_title": 1, "date_of_application": 1, "gender": 1, "date_of_birth": 1
    }).skip(skip).limit(limit)
    data = await cursor.to_list(None)
    return {
        "data": data, "total": total, "page": page, "limit": limit,
        "columns": ["name", "email", "phone", "job_title", "date_of_application", "gender", "date_of_birth"]
    }

@api_router.get("/data/registered")
async def get_registered(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    user: str = Depends(get_current_user)
):
    """Registered (May 2026 rule): present in `pipeline_data` (HR internal dataset),
    regardless of Naukri presence. Reads directly from pipeline_data (no duplication)."""
    skip = (page - 1) * limit
    query = {"isTest": {"$ne": True}}
    total = await db.pipeline_data.count_documents(query)
    cursor = db.pipeline_data.find(query, {
        "_id": 0, "name": 1, "email": 1, "phone": 1,
        "job_role": 1, "schedule_date": 1, "gender": 1, "college": 1, "degree": 1,
        "last_update": 1
    }).skip(skip).limit(limit)
    raw = await cursor.to_list(None)
    # Map to the expected column set (job_title <- job_role, date_of_application <- last_update)
    data = [{
        "name": d.get("name") or "-",
        "email": d.get("email") or "-",
        "phone": d.get("phone") or "-",
        "job_title": d.get("job_role") or "-",
        "date_of_application": d.get("last_update") or "-",
        "gender": d.get("gender") or "-",
        "date_of_birth": d.get("degree") or "-",  # pipeline has no DOB; show degree instead
    } for d in raw]
    return {
        "data": data, "total": total, "page": page, "limit": limit,
        "columns": ["name", "email", "phone", "job_title", "date_of_application", "gender", "date_of_birth"]
    }


@api_router.get("/data/classification")
async def get_classification_counts(user: str = Depends(get_current_user)):
    """Explicit classification summary per new rule (May 2026).
      - total_registered   = pipeline_data (HR internal) count, excluding isTest rows
      - total_unregistered = naukri_applies with `_is_registered != True`
      - total_naukri       = naukri_applies count
      - matched            = naukri records with a pipeline match (intersection)
    """
    exclude_test = {"isTest": {"$ne": True}}
    total_registered = await db.pipeline_data.count_documents(exclude_test)
    total_naukri = await db.naukri_applies.count_documents(exclude_test)
    total_unregistered = await db.naukri_applies.count_documents(
        {**exclude_test, "_is_registered": {"$ne": True}}
    )
    matched = await db.naukri_applies.count_documents(
        {**exclude_test, "_is_registered": True}
    )
    return {
        "total_registered": total_registered,
        "total_unregistered": total_unregistered,
        "total_naukri": total_naukri,
        "matched": matched,
        "note": "Registered = all HR pipeline records; Unregistered = naukri without pipeline match.",
    }

@api_router.get("/data/shortlisted")
async def get_shortlisted(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    user: str = Depends(get_current_user)
):
    """Shortlisted: registered_candidates WHERE email_type matches shortlist"""
    skip = (page - 1) * limit
    query = {"email_type": {"$regex": "shortlist", "$options": "i"}}
    total = await db.registered_candidates.count_documents(query)
    cursor = db.registered_candidates.find(query, {
        "_id": 0, "name": 1, "email": 1, "phone": 1, "job_title": 1,
        "date_of_application": 1, "gender": 1, "location": 1, "email_type": 1
    }).skip(skip).limit(limit)
    data = await cursor.to_list(None)
    return {
        "data": data, "total": total, "page": page, "limit": limit,
        "columns": ["name", "email", "phone", "job_title", "date_of_application", "gender", "location", "email_type"]
    }

@api_router.get("/data/rejected")
async def get_rejected(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    user: str = Depends(get_current_user)
):
    """Rejected: registered_candidates WHERE email_type IN (reject, rejected)"""
    skip = (page - 1) * limit
    query = {"email_type": {"$regex": "^reject", "$options": "i"}}
    total = await db.registered_candidates.count_documents(query)
    cursor = db.registered_candidates.find(query, {
        "_id": 0, "name": 1, "email": 1, "phone": 1, "job_title": 1,
        "date_of_application": 1, "gender": 1, "date_of_birth": 1,
        "location": 1, "email_type": 1
    }).skip(skip).limit(limit)
    data = await cursor.to_list(None)
    return {
        "data": data, "total": total, "page": page, "limit": limit,
        "columns": ["name", "email", "phone", "job_title", "date_of_application", "gender", "date_of_birth", "location", "email_type"]
    }

@api_router.get("/data/scheduled")
async def get_scheduled(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    user: str = Depends(get_current_user)
):
    """Scheduled: Shortlisted AND schedule_date/time NOT NULL (strict hierarchy)"""
    skip = (page - 1) * limit
    query = {
        "email_type": {"$regex": "shortlist", "$options": "i"},
        "schedule_date": _not_null_filter,
        "schedule_time": _not_null_filter
    }
    total = await db.registered_candidates.count_documents(query)
    cursor = db.registered_candidates.find(query, {
        "_id": 0, "name": 1, "email": 1, "phone": 1, "job_title": 1,
        "date_of_application": 1, "gender": 1,
        "schedule_date": 1, "schedule_time": 1, "reschedule_count": 1
    }).skip(skip).limit(limit)
    data = await cursor.to_list(None)
    return {
        "data": data, "total": total, "page": page, "limit": limit,
        "columns": ["name", "email", "phone", "job_title", "date_of_application", "gender", "schedule_date", "schedule_time", "reschedule_count"]
    }

@api_router.get("/data/not-scheduled")
async def get_not_scheduled(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    user: str = Depends(get_current_user)
):
    """Not Scheduled: Shortlisted AND schedule_date/time IS NULL (strict hierarchy)"""
    skip = (page - 1) * limit
    query = {
        "email_type": {"$regex": "shortlist", "$options": "i"},
        "$and": [
            {"$or": [{"schedule_date": None}, {"schedule_date": ""}, {"schedule_date": {"$exists": False}}]},
            {"$or": [{"schedule_time": None}, {"schedule_time": ""}, {"schedule_time": {"$exists": False}}]}
        ]
    }
    total = await db.registered_candidates.count_documents(query)
    cursor = db.registered_candidates.find(query, {
        "_id": 0, "name": 1, "email": 1, "phone": 1, "job_title": 1,
        "date_of_application": 1, "gender": 1, "date_of_birth": 1,
        "location": 1, "loca_change": 1, "attend_inperson": 1,
        "email_type": 1, "confirm_box": 1
    }).skip(skip).limit(limit)
    data = await cursor.to_list(None)
    return {
        "data": data, "total": total, "page": page, "limit": limit,
        "columns": ["name", "email", "phone", "job_title", "date_of_application", "gender", "date_of_birth", "location", "loca_change", "attend_inperson", "email_type", "confirm_box"]
    }

@api_router.get("/data/attended")
async def get_attended(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    user: str = Depends(get_current_user)
):
    """Attended: Shortlisted AND Scheduled AND otp_verified NOT NULL (strict hierarchy)"""
    skip = (page - 1) * limit
    query = {
        "email_type": {"$regex": "shortlist", "$options": "i"},
        "schedule_date": _not_null_filter,
        "schedule_time": _not_null_filter,
        "otp_verified": _not_null_filter
    }
    total = await db.registered_candidates.count_documents(query)
    cursor = db.registered_candidates.find(query, {
        "_id": 0, "name": 1, "email": 1, "phone": 1, "job_title": 1,
        "date_of_application": 1, "gender": 1, "date_of_birth": 1,
        "schedule_date": 1, "schedule_time": 1, "reschedule_count": 1,
        "otp_verified": 1, "result_mail": 1, "result_update": 1, "result_status": 1
    }).skip(skip).limit(limit)
    data = await cursor.to_list(None)
    return {
        "data": data, "total": total, "page": page, "limit": limit,
        "columns": ["name", "email", "phone", "job_title", "date_of_application", "gender", "date_of_birth", "schedule_date", "schedule_time", "reschedule_count", "otp_verified", "result_mail", "result_update", "result_status"]
    }

@api_router.get("/data/not-attended")
async def get_not_attended(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    user: str = Depends(get_current_user)
):
    """Not Attended: Shortlisted AND Scheduled AND otp_verified IS NULL (strict hierarchy)"""
    skip = (page - 1) * limit
    query = {
        "email_type": {"$regex": "shortlist", "$options": "i"},
        "schedule_date": _not_null_filter,
        "schedule_time": _not_null_filter,
        "$or": [
            {"otp_verified": None},
            {"otp_verified": ""},
            {"otp_verified": {"$exists": False}}
        ]
    }
    total = await db.registered_candidates.count_documents(query)
    cursor = db.registered_candidates.find(query, {
        "_id": 0, "name": 1, "email": 1, "phone": 1, "job_title": 1,
        "date_of_application": 1, "gender": 1, "date_of_birth": 1,
        "schedule_date": 1, "schedule_time": 1, "reschedule_count": 1,
        "otp_verified": 1, "otp_expired": 1
    }).skip(skip).limit(limit)
    data = await cursor.to_list(None)
    return {
        "data": data, "total": total, "page": page, "limit": limit,
        "columns": ["name", "email", "phone", "job_title", "date_of_application", "gender", "date_of_birth", "schedule_date", "schedule_time", "reschedule_count", "otp_verified", "otp_expired"]
    }

# ============ SUMMARY & ROLE ANALYTICS ENDPOINTS ============

async def _aggregate_funnel_stats(match_filter: dict) -> list:
    """Aggregate funnel statistics grouped by job_title from registered_candidates.
    Uses STRICT HIERARCHY: Shortlisted/Rejected from email_type,
    Scheduled = subset of Shortlisted, Attended = subset of Scheduled."""

    # Helper expressions for reuse
    _is_shortlisted = {"$regexMatch": {"input": {"$ifNull": ["$email_type", ""]}, "regex": "shortlist", "options": "i"}}
    _has_schedule = {"$and": [
        {"$ne": [{"$ifNull": ["$schedule_date", ""]}, ""]},
        {"$ne": [{"$ifNull": ["$schedule_date", None]}, None]},
        {"$ne": [{"$ifNull": ["$schedule_time", ""]}, ""]},
        {"$ne": [{"$ifNull": ["$schedule_time", None]}, None]}
    ]}
    _no_schedule = {"$and": [
        {"$or": [{"$eq": [{"$ifNull": ["$schedule_date", None]}, None]}, {"$eq": ["$schedule_date", ""]}]},
        {"$or": [{"$eq": [{"$ifNull": ["$schedule_time", None]}, None]}, {"$eq": ["$schedule_time", ""]}]}
    ]}
    _has_otp = {"$and": [
        {"$ne": [{"$ifNull": ["$otp_verified", ""]}, ""]},
        {"$ne": [{"$ifNull": ["$otp_verified", None]}, None]}
    ]}
    _no_otp = {"$or": [
        {"$eq": [{"$ifNull": ["$otp_verified", None]}, None]},
        {"$eq": ["$otp_verified", ""]}
    ]}

    pipeline = [
        {"$match": match_filter},
        {"$group": {
            "_id": "$job_title",
            "total_applicants": {"$sum": 1},
            # Shortlisted: email_type matches shortlist
            "shortlisted": {"$sum": {"$cond": [_is_shortlisted, 1, 0]}},
            # Rejected: email_type matches reject
            "rejected": {"$sum": {"$cond": [
                {"$regexMatch": {"input": {"$ifNull": ["$email_type", ""]}, "regex": "^reject", "options": "i"}},
                1, 0
            ]}},
            # Scheduled = Shortlisted AND has schedule
            "scheduled": {"$sum": {"$cond": [
                {"$and": [_is_shortlisted, _has_schedule]},
                1, 0
            ]}},
            # Not Scheduled = Shortlisted AND no schedule
            "not_scheduled": {"$sum": {"$cond": [
                {"$and": [_is_shortlisted, _no_schedule]},
                1, 0
            ]}},
            # Attended = Shortlisted AND scheduled AND otp_verified
            "attended": {"$sum": {"$cond": [
                {"$and": [_is_shortlisted, _has_schedule, _has_otp]},
                1, 0
            ]}},
            # Not Attended = Shortlisted AND scheduled AND no otp_verified
            "not_attended": {"$sum": {"$cond": [
                {"$and": [_is_shortlisted, _has_schedule, _no_otp]},
                1, 0
            ]}}
        }},
        {"$sort": {"_id": 1}},
        {"$project": {
            "_id": 0,
            "job_role": "$_id",
            "total_applicants": 1,
            "shortlisted": 1,
            "rejected": 1,
            "scheduled": 1,
            "not_scheduled": 1,
            "attended": 1,
            "not_attended": 1
        }}
    ]
    return await db.registered_candidates.aggregate(pipeline).to_list(None)


@api_router.get("/summary")
async def get_summary(
    startDate: str = Query(None),
    endDate: str = Query(None),
    search: str = Query(None),
    user: str = Depends(get_current_user)
):
    """Job role-wise funnel statistics split by NIRF / Non-NIRF.

    iter118 — Aggregation rebuilt to match the user-supplied business rules
    exactly. Key corrections vs prior version:
      * `isTest` filter REMOVED — test rows ARE counted when their dates fall
        inside the selected range (per user spec).
      * Date upper-bound now includes `\\uffff` suffix so ISO timestamps with
        the time portion (e.g. `2026-05-22T09:27:20+00:00`) still match a
        single-day filter like `endDate=2026-05-22`.
      * `Rejected` = email_type NOT shortlist (the user's exact rule). Empty /
        typo'd email_types (e.g. 'raject') now count as Rejected, matching
        the production data shapes we audited.
      * `Interview Scheduled / Not Scheduled / Attended / Not Attended` no
        longer require the shortlist precondition — they evaluate
        `schedule_date` / `schedule_time` / `otp_verified` directly.
      * `Attended` requires `otp_verified` to be NOT NULL AND NOT in
        {0, "0", false, "", null}; production stores only {None, 1.0, True, ""}
        so this is equivalent to "truthy non-empty value".

    OPTIMIZED: All grouping at DB level via persisted `_normalized_job_role`
    + `_nirf_category` indexes; no full-collection scans.
    """
    base_match: dict = {}
    if search:
        base_match["_normalized_job_role"] = {"$regex": re.escape(search), "$options": "i"}

    pipe_match = dict(base_match)
    naukri_match = dict(base_match)
    if startDate or endDate:
        # iter118 — separate date filters because pipeline_data carries
        # ISO timestamps with time portion (`2026-05-22T09:27:20+00:00`)
        # in `last_update` while naukri_applies stores a plain date string
        # in `date_of_application`. Both upper bounds use `\\uffff` so the
        # entire day is included regardless of any trailing time.
        if startDate:
            pipe_match["last_update"] = {"$gte": startDate}
            naukri_match["date_of_application"] = {"$gte": startDate}
        if endDate:
            pipe_match.setdefault("last_update", {})["$lte"] = endDate + "\uffff"
            naukri_match.setdefault("date_of_application", {})["$lte"] = endDate + "\uffff"

    # ---- Helper expressions for funnel stages (iter118 — user-spec exact) ----
    # email_type matches /shortlist/i  (covers shortlist, Shortlist, shortlisted,
    # Shortlisted, and trailing-whitespace variants like 'shortlist ').
    is_shortlisted = {
        "$regexMatch": {
            "input": {"$ifNull": ["$email_type", ""]},
            "regex": "shortlist",
            "options": "i",
        }
    }
    # NOT shortlist  → user's literal rule for "Rejected"
    is_rejected = {"$not": is_shortlisted}
    has_schedule = {
        "$and": [
            {"$ne": [{"$ifNull": ["$schedule_date", ""]}, ""]},
            {"$ne": [{"$ifNull": ["$schedule_time", ""]}, ""]},
        ]
    }
    not_has_schedule = {"$not": has_schedule}
    # otp_verified considered "attended" when value exists AND is not falsy.
    # Production distinct values: {None, 1.0, True, ""}. We exclude
    # null/empty/0/false; everything else is treated as truthy.
    otp_truthy = {
        "$and": [
            {"$ne": [{"$ifNull": ["$otp_verified", None]}, None]},
            {"$ne": [{"$ifNull": ["$otp_verified", ""]}, ""]},
            {"$ne": [{"$ifNull": ["$otp_verified", 0]}, 0]},
            {"$ne": [{"$ifNull": ["$otp_verified", "0"]}, "0"]},
            {"$ne": [{"$ifNull": ["$otp_verified", False]}, False]},
        ]
    }

    # iter125c — Robust role/category resolution.
    # Mirrors the data-table fallback chain `_normalized_job_role → job_role
    # → job_title` inside the `$group._id` so candidates whose derived field
    # hasn't been persisted yet still bucket under their REAL role instead
    # of collapsing into "Unknown". Same idea as iter125b for Interview
    # Reports — keeps Summary Statistics consistent with the canonical
    # role on every row regardless of background-reprocess timing.
    _role_id_expr = {
        "$let": {
            "vars": {
                "norm": {"$ifNull": ["$_normalized_job_role", ""]},
                "jr": {"$ifNull": ["$job_role", ""]},
                "jt": {"$ifNull": ["$job_title", ""]},
            },
            "in": {
                "$cond": [
                    {"$and": [
                        {"$ne": ["$$norm", ""]},
                        {"$ne": ["$$norm", "Unknown"]},
                    ]},
                    "$$norm",
                    {"$cond": [
                        {"$ne": ["$$jr", ""]}, "$$jr",
                        {"$cond": [{"$ne": ["$$jt", ""]}, "$$jt", "Unknown"]},
                    ]},
                ],
            },
        },
    }

    # PIPELINE-FIRST: aggregate funnel + counts from pipeline_data
    pipe_pipeline = [
        {"$match": pipe_match},
        {"$group": {
            "_id": {
                "role": _role_id_expr,
                "cat": {"$ifNull": ["$_nirf_category", "Non NIRF"]},
            },
            "total_registered": {"$sum": 1},
            "shortlisted": {"$sum": {"$cond": [is_shortlisted, 1, 0]}},
            "rejected": {"$sum": {"$cond": [is_rejected, 1, 0]}},
            "scheduled": {"$sum": {"$cond": [has_schedule, 1, 0]}},
            "not_scheduled": {"$sum": {"$cond": [not_has_schedule, 1, 0]}},
            "attended": {"$sum": {"$cond": [{"$and": [has_schedule, otp_truthy]}, 1, 0]}},
            "not_attended": {"$sum": {"$cond": [{"$and": [has_schedule, {"$not": otp_truthy}]}, 1, 0]}},
        }},
    ]
    pipe_results = await db.pipeline_data.aggregate(pipe_pipeline, allowDiskUse=False).to_list(None)
    pipe_buckets = {(r["_id"]["role"], r["_id"]["cat"]): r for r in pipe_results}

    # Naukri counts per (role, cat) — total + matched (registered)
    naukri_pipeline = [
        {"$match": naukri_match},
        {"$group": {
            "_id": {
                "role": _role_id_expr,
                "cat": {"$ifNull": ["$_nirf_category", "Non NIRF"]},
            },
            "naukri_total": {"$sum": 1},
            "naukri_unregistered": {"$sum": {"$cond": [
                {"$ne": [{"$ifNull": ["$_is_registered", False]}, True]}, 1, 0
            ]}},
        }},
    ]
    naukri_results = await db.naukri_applies.aggregate(naukri_pipeline, allowDiskUse=False).to_list(None)
    naukri_buckets = {(r["_id"]["role"], r["_id"]["cat"]): r for r in naukri_results}

    # Combine
    results = []
    all_keys = set(pipe_buckets.keys()) | set(naukri_buckets.keys())
    for (role, cat) in sorted(all_keys):
        p = pipe_buckets.get((role, cat), {})
        n = naukri_buckets.get((role, cat), {})
        results.append({
            "job_role": f"{role} - {cat}",
            "total_naukri": n.get("naukri_total", 0),
            "total_registered": p.get("total_registered", 0),
            "total_unregistered": n.get("naukri_unregistered", 0),
            "shortlisted": p.get("shortlisted", 0),
            "rejected": p.get("rejected", 0),
            "scheduled": p.get("scheduled", 0),
            "not_scheduled": p.get("not_scheduled", 0),
            "attended": p.get("attended", 0),
            "not_attended": p.get("not_attended", 0),
        })

    total_registered = sum(r["total_registered"] for r in results)
    total_naukri = sum(r["total_naukri"] for r in results)
    total_unregistered = sum(r["total_unregistered"] for r in results)

    return {
        "data": results,
        "total_registered": total_registered,           # NEW RULE: pipeline count
        "total_registered_hr": total_registered,        # alias kept for backward compat (iter28)
        "total_naukri": total_naukri,
        "total_unregistered_naukri": total_unregistered,
    }


@api_router.get("/job-roles")
async def get_job_roles(user: str = Depends(get_current_user)):
    """Unique job roles with HR-internal applicant counts (pipeline_data).
    iter99 — Roll up by READ-TIME canonical title using job_keyword_mapping,
    so dashboard charts and filters see one row per canonical job role even
    if pipeline_data still carries raw imported variants.
    iter125c — Aggregation now falls back through
    `_normalized_job_role → job_role → job_title` inside `$group._id`, so
    freshly-uploaded rows whose derived field isn't persisted yet still
    surface under their real role (instead of being silently dropped by
    the legacy `{"_normalized_job_role": {"$nin": [None,"","Unknown"]}}`
    pre-filter)."""
    role_id_expr = {
        "$let": {
            "vars": {
                "norm": {"$ifNull": ["$_normalized_job_role", ""]},
                "jr": {"$ifNull": ["$job_role", ""]},
                "jt": {"$ifNull": ["$job_title", ""]},
            },
            "in": {
                "$cond": [
                    {"$and": [
                        {"$ne": ["$$norm", ""]},
                        {"$ne": ["$$norm", "Unknown"]},
                    ]},
                    "$$norm",
                    {"$cond": [
                        {"$ne": ["$$jr", ""]}, "$$jr",
                        {"$cond": [{"$ne": ["$$jt", ""]}, "$$jt", ""]},
                    ]},
                ],
            },
        },
    }
    pipeline = [
        {"$match": {"isTest": {"$ne": True}}},
        {"$group": {"_id": role_id_expr, "count": {"$sum": 1}}},
        {"$match": {"_id": {"$nin": ["", "Unknown"]}}},
    ]
    raw_results = await db.pipeline_data.aggregate(pipeline, allowDiskUse=False).to_list(None)
    kw_to_canonical, _ = await _build_canonical_index()
    rolled: dict = {}
    for r in raw_results:
        canon = _canonicalize_job_role(r["_id"], kw_to_canonical)
        if not canon or canon.strip().lower() in ("", "unknown"):
            continue
        rolled[canon] = rolled.get(canon, 0) + r["count"]
    results = [{"job_role": k, "count": v} for k, v in rolled.items()]
    results.sort(key=lambda x: (-x["count"], x["job_role"].lower()))
    return {"job_roles": results}


def derive_status(doc: dict) -> str:
    """Derive candidate status using STRICT HIERARCHY based on email_type field.
    Hierarchy: Registered → Shortlisted → Interview Scheduled → Attended/Not Attended
               Registered → Rejected
    """
    email_type = str(doc.get("email_type") or "").strip()
    otp_verified = str(doc.get("otp_verified") or "").strip()
    schedule_date = str(doc.get("schedule_date") or "").strip()
    schedule_time = str(doc.get("schedule_time") or "").strip()

    # Rejected: email_type IN ('reject', 'rejected')
    if re.match(r"^reject", email_type, re.IGNORECASE):
        return "Rejected"

    # Shortlisted hierarchy: email_type IN ('shortlist', 'shortlisted')
    if re.search(r"shortlist", email_type, re.IGNORECASE):
        # Interview Scheduled = Shortlisted AND has schedule
        if schedule_date and schedule_time:
            # Attended = Interview Scheduled AND otp_verified
            if otp_verified:
                return "Attended"
            # Not Attended = Interview Scheduled AND no otp_verified
            return "Not Attended"
        # Shortlisted but not yet scheduled
        return "Shortlisted"

    return "Registered"


@api_router.get("/role")
async def get_role_applicants(
    jobRole: str = Query(..., description="Job role to analyze"),
    startDate: str = Query(None),
    endDate: str = Query(None),
    search: str = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(100, ge=1, le=500),
    sort_by: Optional[str] = Query(None),
    sort_dir: Optional[str] = Query(None),
    user: str = Depends(get_current_user)
):
    """Lightweight applicant table — no score fetching. Supports search by name/email/phone."""
    role = jobRole.strip()
    match = {"job_title": {"$regex": f"^{re.escape(role)}$", "$options": "i"}}
    if startDate or endDate:
        date_filter = {}
        if startDate:
            date_filter["$gte"] = startDate
        if endDate:
            date_filter["$lte"] = endDate
        match["date_of_application"] = date_filter
    if search:
        search_re = {"$regex": re.escape(search), "$options": "i"}
        match["$or"] = [{"name": search_re}, {"email": search_re}, {"phone": search_re}]

    total = await db.registered_candidates.count_documents(match)
    skip = (page - 1) * limit
    sort_spec = build_sort(sort_by, sort_dir, allowed={
        "name": "name", "email": "email", "phone": "phone", "gender": "gender",
        "date_of_birth": "date_of_birth", "date_of_application": "date_of_application",
    }, default={"name": 1})
    cursor = db.registered_candidates.find(match, {
        "_id": 0, "name": 1, "email": 1, "phone": 1, "gender": 1,
        "date_of_birth": 1, "date_of_application": 1,
        "email_type": 1, "otp_verified": 1,
        "schedule_date": 1, "schedule_time": 1
    }).sort(list(sort_spec.items())).skip(skip).limit(limit)
    docs = await cursor.to_list(None)

    applicants = [{
        "name": doc.get("name") or "-",
        "email": doc.get("email") or "-",
        "phone": doc.get("phone") or "-",
        "gender": doc.get("gender") or "-",
        "date_of_birth": doc.get("date_of_birth") or "-",
        "date_of_application": doc.get("date_of_application") or "-",
        "status": derive_status(doc),
    } for doc in docs]

    return {
        "data": applicants,
        "total": total,
        "page": page,
        "limit": limit,
        "columns": ["name", "email", "phone", "gender", "date_of_birth", "date_of_application", "status"]
    }


# ============ GLOBAL APPLICANTS TABLE ============

@api_router.get("/applicants")
async def get_global_applicants(
    jobRole: str = Query(None),
    dateType: str = Query("Registered"),
    startDate: str = Query(None),
    endDate: str = Query(None),
    search: str = Query(None),
    name: str = Query(None),
    email: str = Query(None),
    phone: str = Query(None),
    collegeStatus: str = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(100, ge=1, le=500),
    sort_by: Optional[str] = Query(None),
    sort_dir: Optional[str] = Query(None),
    user: str = Depends(get_current_user)
):
    """Global registered applicants table (HR-internal source, May 2026 rule).

    Reads directly from `pipeline_data` (every HR-internal record is "Registered"
    regardless of Naukri presence). All filters pushed to MongoDB via persisted
    `_normalized_job_role` and `_nirf_category` fields. DB-level pagination.
    """
    match = {"isTest": {"$ne": True}}

    # Date filter — iter116: "Registered" filter MUST use immutable
    # `submitted_at` (set once at registration time and never overwritten).
    # The previous `last_update` choice was overwritten by every downstream
    # action (schedule / OTP-verify / status change), so a candidate who
    # registered on 22/05 IST and then scheduled later that same day could
    # be pushed off the "Registered=22/05" filter when the schedule write's
    # UTC timestamp crossed midnight. `schedule_date` continues to drive
    # the "Scheduled" filter.
    if startDate and endDate:
        date_field = "submitted_at" if dateType == "Registered" else "schedule_date"
        match[date_field] = {"$gte": startDate, "$lte": endDate + "\uffff"}

    # iter110 — College status filter accepts the 5 canonical values directly
    # OR the legacy "Non NIRF" alias (matches any non-premium bucket).
    if collegeStatus and collegeStatus.strip() and collegeStatus.strip().lower() != "all":
        fval = collegeStatus.strip()
        if fval == "NIRF":
            match["_nirf_category"] = "NIRF"
        elif fval in ("Non NIRF", "Non-NIRF"):
            match["_nirf_category"] = {"$ne": "NIRF"}
        elif fval in ("Non-NIRF 101-150", "Non-NIRF 151-200", "Non-NIRF 201-300", "Non-NIRF - No Rank"):
            match["_nirf_category"] = fval
        else:
            match["_college_status"] = fval

    # Job role filter (uses persisted _normalized_job_role)
    if jobRole and jobRole.strip() and jobRole.strip().lower() != "all jobs":
        match["_normalized_job_role"] = {"$regex": f"^{re.escape(jobRole.strip())}$", "$options": "i"}

    # Search filter
    if search:
        search_re = {"$regex": re.escape(search), "$options": "i"}
        match["$or"] = [
            {"name": search_re}, {"email": search_re},
            {"phone": search_re}, {"_normalized_job_role": search_re},
        ]

    # iter111 — Per-field Name / Email / Phone filters (regex partial match).
    _npe_clauses = []
    if name and name.strip():
        _npe_clauses.append({"name": {"$regex": re.escape(name.strip()), "$options": "i"}})
    if email and email.strip():
        _npe_clauses.append({"email": {"$regex": re.escape(email.strip()), "$options": "i"}})
    if phone and phone.strip():
        _npe_clauses.append({"phone": {"$regex": re.escape(phone.strip())}})
    if _npe_clauses:
        match["$and"] = (match.get("$and") or []) + _npe_clauses

    total = await db.pipeline_data.count_documents(match)
    skip = (page - 1) * limit

    projection = {
        "_id": 0, "name": 1, "email": 1, "phone": 1,
        "_college_status": 1, "_college_resolved": 1, "_match_confidence": 1,
        "_normalized_job_role": 1,
        "degree": 1, "email_type": 1, "otp_verified": 1,
        "schedule_date": 1, "schedule_time": 1, "result_status": 1,
        "last_update": 1, "submitted_at": 1, "job_role": 1, "job_title": 1,
    }

    pipeline = [
        {"$match": match},
        {"$sort": build_sort(sort_by, sort_dir, allowed={
            "name": "name", "email": "email", "phone": "phone",
            "college_status": "_college_status", "college": "_college_resolved",
            "degree": "degree", "job_role": "_normalized_job_role",
            "registered_date": "submitted_at", "schedule_date": "schedule_date",
            "schedule_time": "schedule_time",
        }, default={"name": 1})},
        {"$skip": skip},
        {"$limit": limit},
        {"$project": projection},
    ]
    docs = await db.pipeline_data.aggregate(pipeline, allowDiskUse=False).to_list(None)

    applicants = []
    for doc in docs:
        cs = doc.get("_college_status") or "Non-NIRF - No Rank"
        normalized_role = doc.get("_normalized_job_role") or doc.get("job_role") or doc.get("job_title") or "Unknown"

        email_type = str(doc.get("email_type") or "").strip().lower()
        otp_verified = str(doc.get("otp_verified") or "").strip()
        schedule_date = str(doc.get("schedule_date") or "").strip()
        schedule_time = str(doc.get("schedule_time") or "").strip()
        result_status_raw = str(doc.get("result_status") or "").strip()

        if (email_type in ("shortlist", "shortlisted") and schedule_date and schedule_time
                and otp_verified and otp_verified != "0"):
            reg_status = "Attended"
        elif (email_type in ("shortlist", "shortlisted") and schedule_date and schedule_time
                and (not otp_verified or otp_verified == "0")):
            reg_status = "Not Attended"
        elif email_type in ("shortlist", "shortlisted") and schedule_date and schedule_time:
            reg_status = "Interview Scheduled"
        elif email_type in ("shortlist", "shortlisted") and (not schedule_date) and (not schedule_time):
            reg_status = "Interview Not Scheduled"
        elif email_type in ("reject", "rejected"):
            reg_status = "Rejected"
        elif email_type in ("shortlist", "shortlisted"):
            reg_status = "Shortlisted"
        else:
            reg_status = "Registered"

        if reg_status == "Attended":
            res_status = result_status_raw if result_status_raw and result_status_raw not in ("-", "") else "NA"
        else:
            res_status = "NA"

        applicants.append({
            "name": doc.get("name") or "-",
            "email": doc.get("email") or "-",
            "phone": doc.get("phone") or "-",
            "college_status": cs,
            "college": doc.get("_college_resolved") or "-",
            "match_confidence": doc.get("_match_confidence") or "-",
            "degree": doc.get("degree") or "-",
            "job_role": normalized_role or "-",
            "registered_status": reg_status,
            "registered_date": doc.get("submitted_at") or doc.get("last_update") or "-",
            "schedule_date": doc.get("schedule_date") or "-",
            "schedule_time": doc.get("schedule_time") or "-",
            "attended_or_not": "Attended" if reg_status == "Attended" else "Not Attended",
            "result_status": res_status,
        })

    return {"data": applicants, "total": total, "page": page, "limit": limit}


# iter123 — View Applicants Export. Re-runs the same filters as
# `/api/applicants` but un-paginated, supports CSV + XLSX.
def _build_global_applicants_match(jobRole, dateType, startDate, endDate, search, name, email, phone, collegeStatus):
    match = {"isTest": {"$ne": True}}
    if startDate and endDate:
        date_field = "submitted_at" if dateType == "Registered" else "schedule_date"
        match[date_field] = {"$gte": startDate, "$lte": endDate + "\uffff"}
    if collegeStatus and collegeStatus.strip() and collegeStatus.strip().lower() != "all":
        fval = collegeStatus.strip()
        if fval == "NIRF":
            match["_nirf_category"] = "NIRF"
        elif fval in ("Non NIRF", "Non-NIRF"):
            match["_nirf_category"] = {"$ne": "NIRF"}
        elif fval in ("Non-NIRF 101-150", "Non-NIRF 151-200", "Non-NIRF 201-300", "Non-NIRF - No Rank"):
            match["_nirf_category"] = fval
        else:
            match["_college_status"] = fval
    if jobRole and jobRole.strip() and jobRole.strip().lower() != "all jobs":
        match["_normalized_job_role"] = {"$regex": f"^{re.escape(jobRole.strip())}$", "$options": "i"}
    if search:
        search_re = {"$regex": re.escape(search), "$options": "i"}
        match["$or"] = [
            {"name": search_re}, {"email": search_re},
            {"phone": search_re}, {"_normalized_job_role": search_re},
        ]
    _npe_clauses = []
    if name and name.strip():
        _npe_clauses.append({"name": {"$regex": re.escape(name.strip()), "$options": "i"}})
    if email and email.strip():
        _npe_clauses.append({"email": {"$regex": re.escape(email.strip()), "$options": "i"}})
    if phone and phone.strip():
        _npe_clauses.append({"phone": {"$regex": re.escape(phone.strip())}})
    if _npe_clauses:
        match["$and"] = (match.get("$and") or []) + _npe_clauses
    return match


def _derive_registered_status(doc):
    """Replicate the registered_status classifier from /api/applicants."""
    email_type = str(doc.get("email_type") or "").strip().lower()
    otp_verified = str(doc.get("otp_verified") or "").strip()
    sch_d = str(doc.get("schedule_date") or "").strip()
    sch_t = str(doc.get("schedule_time") or "").strip()
    if email_type in ("shortlist", "shortlisted") and sch_d and sch_t and otp_verified and otp_verified != "0":
        return "Attended"
    if email_type in ("shortlist", "shortlisted") and sch_d and sch_t and (not otp_verified or otp_verified == "0"):
        return "Not Attended"
    if email_type in ("shortlist", "shortlisted") and sch_d and sch_t:
        return "Interview Scheduled"
    if email_type in ("shortlist", "shortlisted") and (not sch_d) and (not sch_t):
        return "Interview Not Scheduled"
    if email_type in ("reject", "rejected"):
        return "Rejected"
    if email_type in ("shortlist", "shortlisted"):
        return "Shortlisted"
    return "Registered"


@api_router.get("/applicants/export")
async def export_global_applicants(
    jobRole: str = Query(None),
    dateType: str = Query("Registered"),
    startDate: str = Query(None),
    endDate: str = Query(None),
    search: str = Query(None),
    name: str = Query(None),
    email: str = Query(None),
    phone: str = Query(None),
    collegeStatus: str = Query(None),
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
    user: str = Depends(get_current_user),
):
    """Export filtered View Applicants rows to CSV/XLSX.

    iter123 — Honours all filters currently applied on the page. Returns
    the same 17 fields the user requested, in the same order shown to
    the user, so the export mirrors the table they see.
    """
    from fastapi.responses import StreamingResponse
    import io
    import csv as _csv

    match = _build_global_applicants_match(jobRole, dateType, startDate, endDate, search, name, email, phone, collegeStatus)
    projection = {
        "_id": 0, "name": 1, "email": 1, "phone": 1, "age": 1, "gender": 1,
        "_college_status": 1, "_college_resolved": 1, "college": 1,
        "degree": 1, "course": 1, "year_of_graduation": 1,
        "_normalized_job_role": 1, "job_role": 1, "job_title": 1,
        "email_type": 1, "otp_verified": 1,
        "schedule_date": 1, "schedule_time": 1,
        "submitted_at": 1, "last_update": 1, "result_status": 1,
    }
    docs = await db.pipeline_data.find(match, projection).sort([("name", 1)]).to_list(None)
    if not docs:
        raise HTTPException(status_code=404, detail="No data available to export")

    headers = [
        "Name", "Email", "Phone", "Age", "Gender",
        "College Status", "College", "Degree", "Course", "Year of Graduation",
        "Job Role", "Registered Status", "Registered Date",
        "Schedule Date", "Schedule Time", "Attended or Not", "Result Status",
    ]

    def _row(d):
        reg_status = _derive_registered_status(d)
        attended = "Attended" if reg_status == "Attended" else "Not Attended"
        res_status = (d.get("result_status") or "").strip() if reg_status == "Attended" else "NA"
        if reg_status == "Attended" and (not res_status or res_status == "-"):
            res_status = "NA"
        reg_date = (d.get("submitted_at") or d.get("last_update") or "")[:10]
        return [
            d.get("name") or "",
            d.get("email") or "",
            d.get("phone") or "",
            d.get("age") or "",
            d.get("gender") or "",
            d.get("_college_status") or "",
            d.get("_college_resolved") or d.get("college") or "",
            d.get("degree") or "",
            d.get("course") or "",
            d.get("year_of_graduation") or "",
            d.get("_normalized_job_role") or d.get("job_role") or d.get("job_title") or "Unknown",
            reg_status,
            reg_date,
            d.get("schedule_date") or "",
            d.get("schedule_time") or "",
            attended,
            res_status,
        ]

    today_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    fname = f"View_Applicants_{today_iso}"

    if format == "csv":
        buf = io.StringIO()
        w = _csv.writer(buf)
        w.writerow(headers)
        for d in docs:
            w.writerow(_row(d))
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{fname}.csv"'},
        )

    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("View Applicants")
    ws.append(headers)
    for d in docs:
        ws.append(_row(d))
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        iter([bio.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'},
    )


# ============ ATTENDED APPLICANTS MODULE ============

@api_router.get("/attended-roles")
async def get_attended_roles(user: str = Depends(get_current_user)):
    """Job role boxes with attended applicant counts (HR-internal pipeline_data).
    OPTIMIZED: aggregation on persisted `_normalized_job_role`."""
    pipeline = [
        {"$match": {
            "isTest": {"$ne": True},
            "email_type": {"$regex": "shortlist", "$options": "i"},
            "schedule_date": _not_null_filter,
            "schedule_time": _not_null_filter,
            "otp_verified": _not_null_filter,
        }},
        {"$group": {
            "_id": {"$ifNull": ["$_normalized_job_role", "Unknown"]},
            "count": {"$sum": 1},
        }},
        {"$sort": {"_id": 1}},
        {"$project": {"_id": 0, "job_role": "$_id", "count": 1}},
    ]
    roles = await db.pipeline_data.aggregate(pipeline, allowDiskUse=False).to_list(None)
    return {"job_roles": roles}


@api_router.get("/attended")
async def get_attended_applicants(
    jobRole: str = Query(None),
    startDate: str = Query(None),
    endDate: str = Query(None),
    search: str = Query(None),
    name: str = Query(None),
    email: str = Query(None),
    phone: str = Query(None),
    round: str = Query(None),
    collegeStatus: str = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(100, ge=1, le=500),
    sort_by: Optional[str] = Query(None),
    sort_dir: Optional[str] = Query(None),
    user: str = Depends(get_current_user)
):
    """Global attended applicants table (HR-internal pipeline_data) with scores.

    iter70 — Round columns are now built DYNAMICALLY from `bb_rounds` (DISTINCT
    `name`, alphabetical, displayed AFTER `result_status`) and per-applicant
    scores are fetched from `bb_applicant_updates.scores[]` (matched by
    email OR phone). This keeps the table in lock-step with Update Applicant
    Scores + Score Sheet Import.
    """

    # ---- iter70 — Build dynamic round columns from bb_rounds ----
    # iter79 — Stricter dedup: collapse whitespace + lowercase so legacy
    # variants ("Accounts1" vs "Accounts 1") do NOT produce duplicate columns.
    round_cursor = db.bb_rounds.find(
        {"$or": [{"active": {"$ne": False}}, {"active": {"$exists": False}}]},
        {"_id": 0, "name": 1},
    )
    seen_norm = set()
    dynamic_rounds: list = []
    async for r in round_cursor:
        rn = (r.get("name") or "").strip()
        if not rn:
            continue
        norm = re.sub(r"\s+", "", rn).lower()
        if norm in seen_norm:
            continue
        seen_norm.add(norm)
        dynamic_rounds.append(rn)
    dynamic_rounds.sort(key=lambda x: x.lower())

    match = {"isTest": {"$ne": True}, "otp_verified": _not_null_filter}

    if startDate and endDate:
        match["schedule_date"] = {**_not_null_filter, "$gte": startDate, "$lte": endDate}

    # iter110 — College status filter (5 buckets + legacy "Non NIRF" alias).
    if collegeStatus and collegeStatus.strip() and collegeStatus.strip().lower() != "all":
        fval = collegeStatus.strip()
        if fval == "NIRF":
            match["_nirf_category"] = "NIRF"
        elif fval in ("Non NIRF", "Non-NIRF"):
            match["_nirf_category"] = {"$ne": "NIRF"}
        elif fval in ("Non-NIRF 101-150", "Non-NIRF 151-200", "Non-NIRF 201-300", "Non-NIRF - No Rank"):
            match["_nirf_category"] = fval
        else:
            match["_college_status"] = fval

    # Job role filter via persisted _normalized_job_role
    if jobRole and jobRole.strip() and jobRole.strip().lower() != "all jobs":
        match["_normalized_job_role"] = {"$regex": f"^{re.escape(jobRole.strip())}$", "$options": "i"}

    if search:
        search_re = {"$regex": re.escape(search), "$options": "i"}
        match["$or"] = [
            {"name": search_re}, {"email": search_re},
            {"phone": search_re}, {"_normalized_job_role": search_re},
        ]

    # iter111 — Per-field Name / Email / Phone filters (regex partial match).
    _npe_clauses = []
    if name and name.strip():
        _npe_clauses.append({"name": {"$regex": re.escape(name.strip()), "$options": "i"}})
    if email and email.strip():
        _npe_clauses.append({"email": {"$regex": re.escape(email.strip()), "$options": "i"}})
    if phone and phone.strip():
        _npe_clauses.append({"phone": {"$regex": re.escape(phone.strip())}})
    if _npe_clauses:
        match["$and"] = (match.get("$and") or []) + _npe_clauses

    # Round filter — match against bb_applicant_updates.scores[].round_name
    if round:
        target_lower = round.strip().lower()
        score_emails = set()
        score_phones = set()
        async for sr in db.bb_applicant_updates.find(
            {"scores.round_name": {"$regex": f"^{re.escape(round.strip())}$", "$options": "i"}},
            {"_id": 0, "email": 1, "phone": 1, "scores": 1},
        ):
            # Confirm at least one score entry matches the requested round (case-insensitive)
            has_round = any(
                (s.get("round_name") or "").strip().lower() == target_lower
                for s in (sr.get("scores") or [])
            )
            if not has_round:
                continue
            se = normalize_email(sr.get("email"))
            sp = normalize_phone(sr.get("phone"))
            if se: score_emails.add(se)
            if sp: score_phones.add(sp)
        id_filter = []
        if score_emails:
            id_filter.append({"email": {"$in": list(score_emails)}})
        if score_phones:
            id_filter.append({"phone": {"$in": list(score_phones)}})
        if id_filter:
            round_or = {"$or": id_filter} if len(id_filter) > 1 else id_filter[0]
            match = {"$and": [match, round_or]}
        else:
            base_cols = ["name", "email", "phone", "college_status", "college",
                         "degree", "course", "year_of_graduation", "job_role",
                         "schedule_date", "result_status"]
            return {"data": [], "total": 0, "page": page, "limit": limit,
                    "columns": base_cols + dynamic_rounds, "round_columns": dynamic_rounds}

    total = await db.pipeline_data.count_documents(match)
    skip = (page - 1) * limit

    pipeline = [
        {"$match": match},
        {"$sort": build_sort(sort_by, sort_dir, allowed={
            "name": "name", "email": "email", "phone": "phone",
            "college_status": "_college_status", "college": "_college_resolved",
            "degree": "degree", "course": "course",
            "year_of_graduation": "year_of_graduation",
            "job_role": "_normalized_job_role",
            "schedule_date": "schedule_date", "schedule_time": "schedule_time",
            "result_status": "result_status",
        }, default={"name": 1})},
        {"$skip": skip},
        {"$limit": limit},
    ]
    docs = await db.pipeline_data.aggregate(pipeline, allowDiskUse=False).to_list(None)

    # ---- iter70 — Fetch scores from bb_applicant_updates (NOT score_sheet) ----
    page_emails = list({normalize_email(d.get("email")) for d in docs if d.get("email")})
    page_phones = list({normalize_phone(d.get("phone")) for d in docs if d.get("phone")})
    upd_query = []
    if page_emails: upd_query.append({"email": {"$in": page_emails}})
    if page_phones: upd_query.append({"phone": {"$in": page_phones}})
    upd_records = []
    if upd_query:
        upd_records = await db.bb_applicant_updates.find(
            {"$or": upd_query} if len(upd_query) > 1 else upd_query[0],
            {"_id": 0, "email": 1, "phone": 1, "scores": 1},
        ).to_list(None)

    # Index applicant scores by normalized email + phone
    # iter79 — Match the column-dedup key (whitespace-collapsed + lowercased)
    # so a score saved as "Accounts1" still surfaces in the "Accounts 1" column.
    def _round_key(rn: str) -> str:
        return re.sub(r"\s+", "", (rn or "")).lower()

    scores_by_email: Dict[str, Dict[str, Any]] = {}
    scores_by_phone: Dict[str, Dict[str, Any]] = {}
    for upd in upd_records:
        ue = normalize_email(upd.get("email"))
        up_ = normalize_phone(upd.get("phone"))
        # Build {round_key: score} map for this applicant
        s_map: Dict[str, Any] = {}
        for s in (upd.get("scores") or []):
            rn = (s.get("round_name") or "").strip()
            if rn:
                s_map[_round_key(rn)] = s.get("score")
        if ue:
            scores_by_email.setdefault(ue, {}).update(s_map)
        if up_:
            scores_by_phone.setdefault(up_, {}).update(s_map)

    applicants = []
    for doc in docs:
        cs = doc.get("_college_status") or "Non-NIRF - No Rank"
        normalized_role = doc.get("_normalized_job_role") or doc.get("job_role") or doc.get("job_title") or "Unknown"

        doc_email = normalize_email(doc.get("email"))
        doc_phone = normalize_phone(doc.get("phone"))

        # Resolve per-applicant round → score lookup
        round_lookup: Dict[str, Any] = {}
        if doc_email and doc_email in scores_by_email:
            round_lookup.update(scores_by_email[doc_email])
        if doc_phone and doc_phone in scores_by_phone:
            for k, v in scores_by_phone[doc_phone].items():
                round_lookup.setdefault(k, v)

        row = {
            "name": doc.get("name") or "-",
            "email": doc.get("email") or "-",
            "phone": doc.get("phone") or "-",
            "college_status": cs,
            "college": doc.get("_college_resolved") or "-",
            "match_confidence": doc.get("_match_confidence") or "-",
            "degree": doc.get("degree") or "-",
            "course": doc.get("course") or "-",
            "year_of_graduation": doc.get("year_of_graduation") or "-",
            "job_role": normalized_role or "-",
            "schedule_date": doc.get("schedule_date") or "-",
            "result_status": doc.get("result_status") or "-",
        }
        # Populate dynamic round columns (display "-" when no score recorded)
        for rn in dynamic_rounds:
            v = round_lookup.get(_round_key(rn))
            row[rn] = v if v not in (None, "", "-") else "-"
        applicants.append(row)

    base_cols = ["name", "email", "phone", "college_status", "college", "degree", "course",
                 "year_of_graduation", "job_role", "schedule_date", "result_status"]
    columns = base_cols + dynamic_rounds

    return {
        "data": applicants,
        "total": total,
        "page": page,
        "limit": limit,
        "columns": columns,
        "round_columns": dynamic_rounds,
    }


# iter123 — View Attended Applicants Export. Dynamic round columns from
# bb_rounds (alphabetical) appended after Result Status. Honours all
# filters currently applied on the page.
@api_router.get("/attended/export")
async def export_attended_applicants(
    jobRole: str = Query(None),
    startDate: str = Query(None),
    endDate: str = Query(None),
    search: str = Query(None),
    name: str = Query(None),
    email: str = Query(None),
    phone: str = Query(None),
    round: str = Query(None),
    collegeStatus: str = Query(None),
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
    user: str = Depends(get_current_user),
):
    from fastapi.responses import StreamingResponse
    import io
    import csv as _csv

    # Resolve dynamic round columns first (mirrors /api/attended logic).
    round_cursor = db.bb_rounds.find(
        {"$or": [{"active": {"$ne": False}}, {"active": {"$exists": False}}]},
        {"_id": 0, "name": 1},
    )
    seen_norm = set()
    dynamic_rounds: list = []
    async for r in round_cursor:
        rn = (r.get("name") or "").strip()
        if not rn:
            continue
        norm = re.sub(r"\s+", "", rn).lower()
        if norm in seen_norm:
            continue
        seen_norm.add(norm)
        dynamic_rounds.append(rn)
    dynamic_rounds.sort(key=lambda x: x.lower())

    # Build the same match the /attended endpoint uses (attended = OTP verified).
    match = {"isTest": {"$ne": True}, "otp_verified": _not_null_filter}
    if startDate and endDate:
        match["schedule_date"] = {**_not_null_filter, "$gte": startDate, "$lte": endDate}
    if collegeStatus and collegeStatus.strip() and collegeStatus.strip().lower() != "all":
        fval = collegeStatus.strip()
        if fval == "NIRF":
            match["_nirf_category"] = "NIRF"
        elif fval in ("Non NIRF", "Non-NIRF"):
            match["_nirf_category"] = {"$ne": "NIRF"}
        elif fval in ("Non-NIRF 101-150", "Non-NIRF 151-200", "Non-NIRF 201-300", "Non-NIRF - No Rank"):
            match["_nirf_category"] = fval
        else:
            match["_college_status"] = fval
    if jobRole and jobRole.strip() and jobRole.strip().lower() != "all jobs":
        match["_normalized_job_role"] = {"$regex": f"^{re.escape(jobRole.strip())}$", "$options": "i"}
    if search:
        search_re = {"$regex": re.escape(search), "$options": "i"}
        match["$or"] = [
            {"name": search_re}, {"email": search_re},
            {"phone": search_re}, {"_normalized_job_role": search_re},
        ]
    _npe_clauses = []
    if name and name.strip():
        _npe_clauses.append({"name": {"$regex": re.escape(name.strip()), "$options": "i"}})
    if email and email.strip():
        _npe_clauses.append({"email": {"$regex": re.escape(email.strip()), "$options": "i"}})
    if phone and phone.strip():
        _npe_clauses.append({"phone": {"$regex": re.escape(phone.strip())}})
    if _npe_clauses:
        match["$and"] = (match.get("$and") or []) + _npe_clauses

    projection = {
        "_id": 0, "name": 1, "email": 1, "phone": 1, "age": 1, "gender": 1,
        "_college_status": 1, "_college_resolved": 1, "college": 1,
        "degree": 1, "course": 1, "year_of_graduation": 1,
        "_normalized_job_role": 1, "job_role": 1, "job_title": 1,
        "schedule_date": 1, "result_status": 1,
    }
    docs = await db.pipeline_data.find(match, projection).sort([("name", 1)]).to_list(None)
    if not docs:
        raise HTTPException(status_code=404, detail="No data available to export")

    # Fetch all scores for these applicants in one shot.
    page_emails = list({normalize_email(d.get("email")) for d in docs if d.get("email")})
    page_phones = list({normalize_phone(d.get("phone")) for d in docs if d.get("phone")})
    upd_query = []
    if page_emails: upd_query.append({"email": {"$in": page_emails}})
    if page_phones: upd_query.append({"phone": {"$in": page_phones}})
    upd_records = []
    if upd_query:
        upd_records = await db.bb_applicant_updates.find(
            {"$or": upd_query} if len(upd_query) > 1 else upd_query[0],
            {"_id": 0, "email": 1, "phone": 1, "scores": 1},
        ).to_list(None)

    def _rkey(rn):
        return re.sub(r"\s+", "", (rn or "")).lower()
    scores_by_email: Dict[str, Dict[str, Any]] = {}
    scores_by_phone: Dict[str, Dict[str, Any]] = {}
    for upd in upd_records:
        ue = normalize_email(upd.get("email"))
        up_ = normalize_phone(upd.get("phone"))
        s_map = {}
        for s in (upd.get("scores") or []):
            rn = (s.get("round_name") or "").strip()
            if rn:
                s_map[_rkey(rn)] = s.get("score")
        if ue:
            scores_by_email.setdefault(ue, {}).update(s_map)
        if up_:
            scores_by_phone.setdefault(up_, {}).update(s_map)

    headers = [
        "Name", "Email", "Phone", "Age", "Gender",
        "College Status", "College", "Degree", "Course", "Year of Graduation",
        "Job Role", "Scheduled Date", "Result Status",
    ] + dynamic_rounds

    def _row(d):
        de = normalize_email(d.get("email"))
        dp = normalize_phone(d.get("phone"))
        rl: Dict[str, Any] = {}
        if de and de in scores_by_email:
            rl.update(scores_by_email[de])
        if dp and dp in scores_by_phone:
            for k, v in scores_by_phone[dp].items():
                rl.setdefault(k, v)
        row = [
            d.get("name") or "",
            d.get("email") or "",
            d.get("phone") or "",
            d.get("age") or "",
            d.get("gender") or "",
            d.get("_college_status") or "",
            d.get("_college_resolved") or d.get("college") or "",
            d.get("degree") or "",
            d.get("course") or "",
            d.get("year_of_graduation") or "",
            d.get("_normalized_job_role") or d.get("job_role") or d.get("job_title") or "Unknown",
            d.get("schedule_date") or "",
            d.get("result_status") or "",
        ]
        for rn in dynamic_rounds:
            row.append(rl.get(_rkey(rn), ""))
        return row

    today_iso = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    fname = f"View_Attended_Applicants_{today_iso}"

    if format == "csv":
        buf = io.StringIO()
        w = _csv.writer(buf)
        w.writerow(headers)
        for d in docs:
            w.writerow(_row(d))
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{fname}.csv"'},
        )

    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("View Attended Applicants")
    ws.append(headers)
    for d in docs:
        ws.append(_row(d))
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        iter([bio.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'},
    )


# ============ SCORE SHEET UPLOAD ============

SCORE_ROUND_COLUMNS = ["ZA", "C++", "Java", "BA", "LA", "Mensa Org", "Accounts2", "Accounts1", "BE", "Mensa", "BP"]
# Lowercase lookup map for case-insensitive matching
ROUND_NAME_MAP = {r.strip().lower(): r for r in SCORE_ROUND_COLUMNS}


@api_router.post("/upload/scoresheet")
async def upload_score_sheet(
    file: UploadFile = File(...),
    user: str = Depends(get_current_user)
):
    """Upload score sheet (CSV/XLSX). Fields: name, email, phone, score, round_name.

    Non-destructive smart upsert (Feb 2026 / iter45):
      * Email + Phone matching (email primary, phone secondary).
      * Per (email-or-phone, canonical round_name): the upload only overwrites
        when the existing record is older OR missing. Newer existing records
        are preserved + skipped.
      * Same phone with a different email → flagged as a conflict and skipped.
      * Round names are whitespace-collapsed; common aliases (Technical 1 →
        Round 1, Accounts1 → Accounts 1, etc.) are canonicalised before match.
    """
    from bb_modules import _norm_round, _norm_email, _norm_phone, _detect_score_phone_conflict

    content = await file.read()
    df = parse_file(content, file.filename)
    df.columns = df.columns.str.strip().str.lower()

    required = {"name", "email", "phone", "score", "round_name"}
    if not required.issubset(set(df.columns)):
        missing = required - set(df.columns)
        raise HTTPException(status_code=400, detail=f"Missing columns: {missing}")

    inserted = 0
    updated = 0
    skipped_newer = 0
    skipped_conflict = 0
    errors = []
    now_iso = datetime.now(timezone.utc).isoformat()

    for idx, row in df.iterrows():
        try:
            email = _norm_email(row.get("email"))
            phone = _norm_phone(row.get("phone"))
            name = str(row.get("name") or "").strip()
            raw_round = str(row.get("round_name") or "").strip()
            canon_round = _norm_round(raw_round)
            score_val = row.get("score")

            if not email and not phone:
                errors.append(f"Row {idx + 2}: Missing email and phone")
                continue
            if not canon_round:
                errors.append(f"Row {idx + 2}: Missing round_name")
                continue

            try:
                score = float(score_val) if not pd.isna(score_val) else 0.0
            except (ValueError, TypeError):
                score = 0.0

            # Conflict check: same phone, different email
            conflict = await _detect_score_phone_conflict(email, phone)
            if conflict:
                skipped_conflict += 1
                errors.append(f"Row {idx + 2}: SKIP {conflict}")
                continue

            # Build identity match across email/phone for the same canonical
            # round (canon comparison is case-insensitive via $regex).
            or_clauses = []
            if email:
                or_clauses.append({"email": email})
            if phone:
                or_clauses.append({"phone": phone})
            existing = None
            cursor = db.score_sheet.find(
                {"$or": or_clauses} if len(or_clauses) > 1 else or_clauses[0],
                {"_id": 1, "round_name": 1, "score": 1, "created_at": 1, "updated_at": 1},
            )
            async for r in cursor:
                if _norm_round(r.get("round_name")) == canon_round:
                    existing = r
                    break

            doc = {
                "name": name,
                "email": email,
                "phone": phone,
                "score": score,
                "round_name": raw_round or canon_round,
                "round_canonical": canon_round,
                "created_at": now_iso,
                "updated_at": now_iso,
            }

            if not existing:
                await db.score_sheet.insert_one(doc)
                inserted += 1
            else:
                existing_ts = str(existing.get("updated_at") or existing.get("created_at") or "")
                if existing_ts and existing_ts > now_iso:
                    skipped_newer += 1
                    continue
                # Replace older / equal-timestamp record with the new one
                await db.score_sheet.update_one(
                    {"_id": existing["_id"]},
                    {"$set": {**doc, "created_at": existing.get("created_at") or now_iso}},
                )
                updated += 1

            # ---- Iter48: sync into bb_applicant_updates (append-only) ----
            # Reuses the score-sheet match identity so the same upload makes
            # the round visible on the "Update Applicants Scores" page.
            sync_or = []
            if email: sync_or.append({"email": email})
            if phone: sync_or.append({"phone": phone})
            au = await db.bb_applicant_updates.find_one(
                {"$or": sync_or} if len(sync_or) > 1 else sync_or[0],
                {"_id": 0, "email": 1, "scores": 1, "name": 1, "phone": 1},
            )
            au_scores = list((au or {}).get("scores") or [])
            au_round_lc = {
                str(s.get("round_name") or "").strip().lower()
                for s in au_scores if s.get("round_name")
            }
            if canon_round.lower() not in au_round_lc:
                au_scores.append({"round_name": raw_round or canon_round, "score": score})
                target_email = (au or {}).get("email") or email or phone
                await db.bb_applicant_updates.update_one(
                    {"email": target_email},
                    {"$set": {
                        "email": target_email,
                        "scores": au_scores,
                        "name": name or (au or {}).get("name") or "",
                        "phone": phone or (au or {}).get("phone") or "",
                        "updated_at": now_iso,
                    },
                    "$setOnInsert": {"status": "On hold", "isImported": True}},
                    upsert=True,
                )

            # ---- Iter48: register the round into bb_rounds ----
            ex_round = await db.bb_rounds.find_one(
                {"name": {"$regex": f"^{re.escape(raw_round or canon_round)}$",
                          "$options": "i"}},
                {"_id": 1, "active": 1},
            )
            if not ex_round:
                await db.bb_rounds.insert_one({
                    "name": raw_round or canon_round,
                    "active": True, "order": 0,
                    "source": "score_sheet",
                    "created_at": now_iso,
                })
            elif ex_round.get("active") is False:
                await db.bb_rounds.update_one(
                    {"_id": ex_round["_id"]}, {"$set": {"active": True}}
                )

        except Exception as e:
            errors.append(f"Row {idx + 2}: {str(e)}")

    return {
        "success": True,
        "message": f"Score sheet processed. Inserted: {inserted}, Updated: {updated}, "
                   f"Skipped (newer): {skipped_newer}, Skipped (conflict): {skipped_conflict}",
        "inserted": inserted,
        "updated": updated,
        "skipped_newer": skipped_newer,
        "skipped_conflict": skipped_conflict,
        "errors": errors[:20]
    }


# ============ COLLEGE RANK UPLOAD ============

COLLEGE_RANK_COLUMNS = {"rank", "college_name", "short_name", "city", "state"}

@api_router.post("/upload/college-rank")
async def upload_college_rank(
    file: UploadFile = File(...),
    user: str = Depends(get_current_user)
):
    """Upload college rank list (CSV/XLSX). Columns: Rank, College Name, Short Name, City, State."""
    content = await file.read()
    df = parse_file(content, file.filename)
    df.columns = df.columns.str.strip()
    # Map column names to snake_case
    col_map = {}
    for c in df.columns:
        cl = c.strip().lower().replace(" ", "_")
        if cl in COLLEGE_RANK_COLUMNS:
            col_map[c] = cl
    mapped = set(col_map.values())
    if "college_name" not in mapped:
        raise HTTPException(status_code=400, detail=f"Missing 'College Name' column. Found: {list(df.columns)}")

    # Clear existing and re-insert
    await db.college_rank_list.delete_many({})
    inserted = 0
    for _, row in df.iterrows():
        try:
            doc = {}
            for csv_col, db_field in col_map.items():
                val = row.get(csv_col)
                if pd.isna(val):
                    doc[db_field] = None
                elif db_field == "rank":
                    # iter112 — Accept BOTH integer ranks (1..300) and range
                    # strings ("101-150", "151-200", "201-300"). The classifier
                    # in `_rank_to_college_status` handles both shapes. Without
                    # this, range strings silently became None and ~200 valid
                    # NIRF colleges fell into "Non-NIRF - No Rank".
                    s = str(val).strip()
                    try:
                        doc[db_field] = int(float(s))
                    except (ValueError, TypeError):
                        # Preserve range strings; normalize dash variants.
                        norm = s.replace(" ", "").replace("–", "-").replace("—", "-")
                        if norm in ("101-150", "151-200", "201-300"):
                            doc[db_field] = norm
                        else:
                            doc[db_field] = None
                else:
                    doc[db_field] = str(val).strip()
            if not doc.get("college_name"):
                continue
            await db.college_rank_list.insert_one(doc)
            inserted += 1
        except Exception:
            pass
    return {"success": True, "inserted": inserted}


async def _build_college_rank_lookup() -> dict:
    """Build structured lookup for multi-criteria college matching.
    Groups rank entries by base_name for efficient lookup.
    Returns: {entries_by_base: {base: [entries]}, cities: set, states: set}"""
    docs = await db.college_rank_list.find({}, {"_id": 0}).to_list(None)

    # First pass: collect all known cities and states (normalized)
    city_set = set()
    state_set = set()
    for doc in docs:
        city = _normalize_college_text(doc.get("city") or "")
        state = _normalize_college_text(doc.get("state") or "")
        if city:
            city_set.add(city)
        if state:
            state_set.add(state)

    # Second pass: build lookup keyed by base_name
    entries_by_base = {}

    for doc in docs:
        rank = doc.get("rank")
        college_name = (doc.get("college_name") or "").strip()
        short_name = (doc.get("short_name") or "").strip()
        city = _normalize_college_text(doc.get("city") or "")
        state = _normalize_college_text(doc.get("state") or "")

        if not college_name:
            continue

        normalized = _normalize_college_text(college_name)

        # For rank entries, remove generic words + own location tokens to get base
        own_locations = set()
        for token in normalized.split():
            if token in city_set or token in state_set:
                own_locations.add(token)
        # Also check multi-word city/state phrases
        if city:
            for t in city.split():
                if t in normalized:
                    own_locations.add(t)
        if state:
            for t in state.split():
                if t in normalized:
                    own_locations.add(t)

        base = _extract_college_base(normalized, own_locations)

        entry = {
            "college_name": college_name,
            "short_name": short_name,
            "normalized": normalized,
            "base": base,
            "city": city,
            "state": state,
            "rank": rank,
        }

        if base:
            entries_by_base.setdefault(base, []).append(entry)

        # Also index by short_name base
        if short_name:
            short_norm = _normalize_college_text(short_name)
            short_base = _extract_college_base(short_norm, own_locations)
            if short_base and short_base != base:
                entries_by_base.setdefault(short_base, []).append(entry)

    return {"entries_by_base": entries_by_base, "cities": city_set, "states": state_set}


_GENERIC_COLLEGE_WORDS = frozenset({
    "university", "institute", "college", "of", "the", "and", "for",
})


def _normalize_college_text(text: str) -> str:
    """Normalize: lowercase, trim, remove punctuation, single spaces."""
    if not text:
        return ""
    text = text.lower().strip()
    text = re.sub(r'[,.\-&()\[\]/]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _extract_college_base(normalized_text: str, location_tokens: set = None) -> str:
    """Extract base name by removing generic words and location tokens."""
    tokens = normalized_text.split()
    remove = _GENERIC_COLLEGE_WORDS
    if location_tokens:
        remove = remove | location_tokens
    base_tokens = [t for t in tokens if t not in remove]
    return " ".join(base_tokens).strip()


def _match_college_entry(university_text: str, rank_lookup: dict) -> dict:
    """Match a university text against the rank lookup using multi-criteria matching.
    Returns: {rank, college_name, confidence: HIGH|MEDIUM|LOW|None}"""
    if not university_text or not university_text.strip():
        return {"rank": None, "college_name": "", "confidence": None}

    normalized = _normalize_college_text(university_text)
    cities = rank_lookup["cities"]
    states = rank_lookup["states"]
    entries_by_base = rank_lookup["entries_by_base"]

    # Extract location tokens from the university text
    location_tokens = set()
    extracted_city = None
    extracted_state = None

    # Single-word token matching
    for token in normalized.split():
        if token in cities:
            location_tokens.add(token)
            if not extracted_city:
                extracted_city = token
        if token in states:
            location_tokens.add(token)
            if not extracted_state:
                extracted_state = token

    # Multi-word city/state phrase matching
    for city in cities:
        if ' ' in city and city in normalized:
            extracted_city = city
            for t in city.split():
                location_tokens.add(t)
    for state in states:
        if ' ' in state and state in normalized:
            extracted_state = state
            for t in state.split():
                location_tokens.add(t)

    # Extract base name (remove generic words + location tokens)
    base = _extract_college_base(normalized, location_tokens)

    if not base:
        return {"rank": None, "college_name": university_text, "confidence": None}

    # Step 1: Exact base match
    candidates = list(entries_by_base.get(base, []))

    # Step 1a: Token-subset fallback if no exact match
    if not candidates:
        base_tokens = set(base.split())
        for key, entries in entries_by_base.items():
            key_tokens = set(key.split())
            if base_tokens and key_tokens:
                if base_tokens.issubset(key_tokens) or key_tokens.issubset(base_tokens):
                    candidates.extend(entries)
        # Deduplicate
        seen = set()
        deduped = []
        for c in candidates:
            uid = (c["college_name"], c.get("rank"))
            if uid not in seen:
                seen.add(uid)
                deduped.append(c)
        candidates = deduped

    if not candidates:
        return {"rank": None, "college_name": university_text, "confidence": None}

    # Step 2: Strong match (base + location)
    if extracted_city or extracted_state:
        for entry in candidates:
            city_match = (extracted_city and entry["city"]
                          and (extracted_city in entry["city"] or entry["city"] in extracted_city))
            state_match = (extracted_state and entry["state"]
                           and (extracted_state in entry["state"] or entry["state"] in extracted_state))
            if city_match or state_match:
                return {"rank": entry["rank"], "college_name": entry["college_name"], "confidence": "HIGH"}

    # Step 3: Base-only — single match
    if len(candidates) == 1:
        return {"rank": candidates[0]["rank"], "college_name": candidates[0]["college_name"], "confidence": "MEDIUM"}

    # Step 4: Multiple matches — disambiguate via NIRF (top-100 only).
    # iter112 — Rank may now be a range string like "101-150"; only ints
    # 1..100 count as top-NIRF for disambiguation.
    def _is_top_nirf_rank(r):
        return isinstance(r, int) and 1 <= r <= 100
    nirf_candidates = [e for e in candidates if _is_top_nirf_rank(e["rank"])]
    if len(nirf_candidates) == 1:
        return {"rank": nirf_candidates[0]["rank"], "college_name": nirf_candidates[0]["college_name"], "confidence": "MEDIUM"}

    # Ambiguous — multiple NIRF or none
    return {"rank": None, "college_name": university_text, "confidence": "LOW"}


def _rank_to_college_status(rank) -> str:
    """iter110 — Map a NIRF dataset rank value (int OR range-string like
    '101-150') to the canonical college_status label.

    Returns one of:
      • 'NIRF - #<rank>'        — rank 1..100 (premium)
      • 'Non-NIRF 101-150'      — rank 101..150 or string '101-150'
      • 'Non-NIRF 151-200'      — rank 151..200 or string '151-200'
      • 'Non-NIRF 201-300'      — rank 201..300 or string '201-300'
      • 'Non-NIRF - No Rank'    — None / empty / unrecognised
    """
    if rank is None or rank == "":
        return "Non-NIRF - No Rank"
    # Numeric rank
    try:
        r = int(rank)
        if 1 <= r <= 100:
            return f"NIRF - #{r}"
        if 101 <= r <= 150:
            return "Non-NIRF 101-150"
        if 151 <= r <= 200:
            return "Non-NIRF 151-200"
        if 201 <= r <= 300:
            return "Non-NIRF 201-300"
        return "Non-NIRF - No Rank"
    except (ValueError, TypeError):
        pass
    # String range — accept en/em dashes and extra whitespace
    s = str(rank).strip().replace(" ", "").replace("–", "-").replace("—", "-")
    if s == "101-150":
        return "Non-NIRF 101-150"
    if s == "151-200":
        return "Non-NIRF 151-200"
    if s == "201-300":
        return "Non-NIRF 201-300"
    return "Non-NIRF - No Rank"


def _classify_college(doc: dict, rank_lookup: dict) -> dict:
    """iter110 — Five-bucket college classification.

    Returns: {college_status, college, match_confidence}.

    Match priority unchanged from prior logic:
      1. Both UG and PG resolve to top-NIRF (rank 1..100) → PG wins.
      2. PG top-NIRF only → PG.
      3. UG top-NIRF only → UG.
      4. Neither UG nor PG text supplied AND fallback `college` exists → fallback.
      5. Otherwise pick whichever side has ANY rank → bucketed.
      6. No rank anywhere → 'Non-NIRF - No Rank'.
    """
    ug_text = (doc.get("ug_university") or "").strip()
    pg_text = (doc.get("pg_university") or "").strip()
    fallback_text = (doc.get("college") or "").strip()

    ug_match = _match_college_entry(ug_text, rank_lookup)
    pg_match = _match_college_entry(pg_text, rank_lookup)

    def _is_top_nirf(m: dict) -> bool:
        r = m.get("rank")
        try:
            return r is not None and 1 <= int(r) <= 100
        except (ValueError, TypeError):
            return False

    ug_top = _is_top_nirf(ug_match)
    pg_top = _is_top_nirf(pg_match)

    if pg_top:
        return {"college_status": _rank_to_college_status(pg_match["rank"]),
                "college": pg_text or "-", "match_confidence": pg_match["confidence"]}
    if ug_top:
        return {"college_status": _rank_to_college_status(ug_match["rank"]),
                "college": ug_text or "-", "match_confidence": ug_match["confidence"]}

    # Neither UG nor PG present — fall back to the pipeline-only `college` field
    if not ug_text and not pg_text and fallback_text:
        fb_match = _match_college_entry(fallback_text, rank_lookup)
        return {"college_status": _rank_to_college_status(fb_match["rank"]),
                "college": fallback_text, "match_confidence": fb_match.get("confidence")}

    # Neither top-NIRF: take whichever side has a non-None rank (101-300 buckets)
    if pg_match.get("rank") is not None:
        return {"college_status": _rank_to_college_status(pg_match["rank"]),
                "college": pg_text or "-", "match_confidence": pg_match["confidence"]}
    if ug_match.get("rank") is not None:
        return {"college_status": _rank_to_college_status(ug_match["rank"]),
                "college": ug_text or "-", "match_confidence": ug_match["confidence"]}

    return {"college_status": "Non-NIRF - No Rank",
            "college": ug_text or pg_text or fallback_text or "-",
            "match_confidence": ug_match.get("confidence") or pg_match.get("confidence")}

UPLOAD_BASE = Path(os.getenv("UPLOAD_BASE", "/tmp/uploads"))
PROCESSED_BASE = Path(os.getenv("PROCESSED_BASE", "/tmp/processed_files"))
BULK_TYPES_LIST = ["naukri", "pipeline", "score"]

# Create directories on module load (best-effort — on hardened FS just skip)
try:
    for _bt in BULK_TYPES_LIST:
        (UPLOAD_BASE / _bt).mkdir(parents=True, exist_ok=True)
        (PROCESSED_BASE / _bt).mkdir(parents=True, exist_ok=True)
except (PermissionError, OSError) as _e:
    logging.warning(f"[startup] could not create upload dirs at {UPLOAD_BASE} / {PROCESSED_BASE}: {_e}")



# ============ DB-DRIVEN BACKGROUND QUEUE WORKER ============

# iter129 — Per-file-type worker registry (replaces the legacy single
# `_worker_running` boolean). Each file_type (naukri, pipeline, score)
# gets its OWN concurrent worker task so a slow pipeline batch can never
# starve the naukri queue (or vice-versa). The set tracks file_types
# that currently have a live worker so start_all_workers is idempotent.
_worker_running: set = set()
# Single-flight guard for deferred reprocess_matching. Prevents overlapping
# `registered_candidates.drop()` + insert_many races when multiple drain
# events happen in quick succession.
_reprocess_lock = asyncio.Lock()
_reprocess_pending = False


async def _trigger_deferred_reprocess(reason: str = ""):
    """iter108 — Shared deferred reprocess trigger with single-flight guard.

    Used by:
      - Bulk-upload queue worker (after queue is drained)
      - Job-keyword-mapping create/update/delete endpoints (so newly mapped
        keywords reclassify existing applicants out of "Unknown" without a
        manual rebuild)

    Single-flight: if a reprocess is already running, sets the pending flag
    so one follow-up run executes after the current one finishes. Multiple
    rapid edits coalesce into ONE follow-up run.
    """
    global _reprocess_pending
    if _reprocess_lock.locked():
        _reprocess_pending = True
        logger.info(f"[Reprocess:COALESCED] reason={reason!r} — follow-up scheduled")
        return
    async with _reprocess_lock:
        while True:
            try:
                logger.info(f"[Reprocess:START] reason={reason!r}")
                t0 = datetime.now(timezone.utc)
                await reprocess_matching()
                await _sync_job_titles_master()
                elapsed = (datetime.now(timezone.utc) - t0).total_seconds()
                logger.info(f"[Reprocess:DONE] reason={reason!r} elapsed={elapsed:.1f}s")
            except Exception as e:
                logger.exception(f"[Reprocess:FAIL] reason={reason!r}: {e}")
            if _reprocess_pending:
                _reprocess_pending = False
                logger.info("[Reprocess:FOLLOWUP] running coalesced re-run")
                continue
            break


async def _backfill_unknown_classifications_once():
    """iter108 — One-shot backfill: reclassify legacy non-test rows currently
    stuck at `_normalized_job_role` in {None, '', 'Unknown'} using the
    LATEST `job_keyword_mapping` entries.

    Safety guarantees:
      - Only touches rows where `isTest != True` AND current value is
        Unknown/null/missing (never overwrites a successfully mapped row).
      - Only persists the new value when canonicalization yields a NON-Unknown
        result — otherwise the row stays unchanged.
      - Idempotent: a `bb_meta._id='iter108_unknown_backfill'` flag is set
        after a successful run so subsequent reboots skip the work.
      - Runs as a fire-and-forget startup task so app boot is not blocked.
    """
    try:
        meta = await db.bb_meta.find_one({"_id": "iter108_unknown_backfill"})
        if meta and meta.get("done"):
            return
        from pymongo import UpdateOne
        mappings = await _get_job_keyword_mappings()
        if not mappings:
            logger.info("[Iter108:UnknownBackfill] SKIP — no mappings configured yet")
            return
        total_fixed = 0
        for coll_name in ("pipeline_data", "naukri_applies"):
            coll = db[coll_name]
            ops = []
            stuck_filter = {
                "isTest": {"$ne": True},
                "$or": [
                    {"_normalized_job_role": {"$in": [None, "", "Unknown"]}},
                    {"_normalized_job_role": {"$exists": False}},
                ],
            }
            async for doc in coll.find(stuck_filter, {"_id": 1, "job_title": 1, "job_role": 1}):
                raw = doc.get("job_title") or doc.get("job_role") or ""
                new_val = _resolve_normalized_job_role(raw, mappings)
                # iter122 — Repair NEW UNKNOWN BUG: the original condition
                # `new_val != raw` skipped rows whose resolution returned the
                # RAW title verbatim (e.g. "Ai Ml Engineer" with no exact
                # mapping). Those rows remained stuck at `_normalized_job_role
                # = 'Unknown'` even though the raw title was perfectly usable.
                # New condition: update whenever resolution produced ANY
                # non-empty, non-"Unknown" value AND the existing stored
                # value is still in the stuck set {None, "", "Unknown"}.
                if new_val and new_val != "Unknown":
                    ops.append(UpdateOne(
                        {"_id": doc["_id"]},
                        {"$set": {"_normalized_job_role": new_val}}
                    ))
                if len(ops) >= 1000:
                    await coll.bulk_write(ops, ordered=False)
                    total_fixed += len(ops)
                    ops = []
            if ops:
                await coll.bulk_write(ops, ordered=False)
                total_fixed += len(ops)
            logger.info(f"[Iter108:UnknownBackfill] {coll_name} processed — total_fixed_so_far={total_fixed}")
        await db.bb_meta.update_one(
            {"_id": "iter108_unknown_backfill"},
            {"$set": {
                "done": True,
                "fixed_count": total_fixed,
                "ran_at": datetime.now(timezone.utc).isoformat(),
            }},
            upsert=True,
        )
        logger.info(f"[Iter108:UnknownBackfill] COMPLETED — reclassified {total_fixed} legacy rows")
    except Exception as e:
        # Never let backfill crash app startup. Re-run on next boot.
        logger.exception(f"[Iter108:UnknownBackfill] FAILED: {e}")


# iter123 — Admin-triggered backfill re-runs. Useful when the iter122
# fix lands in prod after the `bb_meta.done` flag has already been set
# on an earlier (buggy) backfill run. Authenticated users only.
@api_router.post("/admin/reset-backfill/{name}")
async def reset_backfill_flag(name: str, user: str = Depends(get_current_user)):
    """Reset a `bb_meta._id={name}` `done` flag and (if it's a known
    backfill) re-launch it as a background task. Safe to call repeatedly —
    each backfill itself is idempotent.
    """
    known_backfills = {
        "iter108_unknown_backfill": _backfill_unknown_classifications_once,
        "iter110_college_status_backfill": _backfill_college_status_once,
    }
    if name not in known_backfills:
        raise HTTPException(status_code=404, detail=f"Unknown backfill: {name}")
    await db.bb_meta.update_one(
        {"_id": name},
        {"$set": {
            "done": False,
            "reset_by": user,
            "reset_at": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )
    asyncio.create_task(known_backfills[name]())
    return {"success": True, "backfill": name, "status": "relaunched"}




async def _backfill_college_status_once():
    """iter110 — One-shot backfill: reclassify `_college_status` for legacy
    non-test rows where the value is missing OR uses the old binary
    "Non NIRF" label. Top-NIRF rows ("NIRF - #N") are left untouched —
    their rank-suffixed label is already correct under the new scheme.

    Idempotent via `bb_meta._id='iter110_college_status_backfill'`.
    Read-only outside the targeted match filter; never touches test rows.
    """
    try:
        meta = await db.bb_meta.find_one({"_id": "iter110_college_status_backfill"})
        if meta and meta.get("done"):
            return
        from pymongo import UpdateOne
        rank_lookup = await _build_college_rank_lookup()
        if not rank_lookup or not rank_lookup.get("entries_by_base"):
            logger.info("[Iter110:CollegeBackfill] SKIP — empty rank lookup")
            return
        total_fixed = 0
        for coll_name in ("pipeline_data", "naukri_applies", "bb_registrations"):
            coll = db[coll_name]
            ops = []
            stuck_filter = {
                "isTest": {"$ne": True},
                # iter112 — Also re-evaluate rows currently labelled
                # "Non-NIRF - No Rank". The NIRF dataset has been enriched
                # with 200 newly-ranked colleges (101-300 buckets) so many of
                # these previously No-Rank rows now resolve into real buckets.
                "$or": [
                    {"_college_status": {"$in": [None, "", "Non NIRF", "Non-NIRF - No Rank"]}},
                    {"_college_status": {"$exists": False}},
                ],
            }
            projection = {"_id": 1, "ug_university": 1, "pg_university": 1, "college": 1, "_college_status": 1}
            async for doc in coll.find(stuck_filter, projection):
                cc = _classify_college(doc, rank_lookup)
                cs = cc["college_status"]
                # iter112 — Skip no-op writes (same value already persisted).
                if cs == doc.get("_college_status"):
                    continue
                cat = "NIRF" if cs.startswith("NIRF - #") else cs
                ops.append(UpdateOne(
                    {"_id": doc["_id"]},
                    {"$set": {"_college_status": cs, "_nirf_category": cat}}
                ))
                if len(ops) >= 1000:
                    await coll.bulk_write(ops, ordered=False)
                    total_fixed += len(ops)
                    ops = []
            if ops:
                await coll.bulk_write(ops, ordered=False)
                total_fixed += len(ops)
            logger.info(f"[Iter110:CollegeBackfill] {coll_name} processed — total_fixed_so_far={total_fixed}")
        await db.bb_meta.update_one(
            {"_id": "iter110_college_status_backfill"},
            {"$set": {"done": True, "fixed_count": total_fixed,
                      "ran_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True,
        )
        logger.info(f"[Iter110:CollegeBackfill] COMPLETED — reclassified {total_fixed} legacy rows")
    except Exception as e:
        logger.exception(f"[Iter110:CollegeBackfill] FAILED: {e}")

# ---- iter67 — Per-host queue isolation ----
# The Mongo queue is shared across deployments (preview + production), but each
# host has its OWN /app/uploads filesystem. If a worker from host A claims a job
# uploaded on host B, the file is "not found" on host A's disk. We tag every
# upload row with HOST_ID and only allow workers to claim rows for their host.
import socket as _socket
HOST_ID = os.environ.get("HOST_ID") or _socket.gethostname() or "unknown-host"


async def _bg_queue_worker(file_type_scope: str = None):
    """Persistent background worker: continuously polls bulk_upload_queue
    for pending jobs of `file_type_scope` (e.g. "naukri", "pipeline",
    "score"). Processes ONE file at a time per scope, sequentially. Runs
    independent of UI/browser. Hardened against malformed legacy queue
    documents — never dies on a bad row.

    iter129 — Per-file-type concurrency: by scoping each worker to a
    single `file_type`, naukri uploads no longer wait behind a pipeline
    backlog (or vice-versa). Each worker runs as its own asyncio task
    launched by `start_all_workers`. Backward-compatible: passing
    `file_type_scope=None` reverts to the legacy single-FIFO behaviour.

    ATOMIC CLAIM: Uses `find_one_and_update` to claim the next pending
    row in a single round-trip — preventing any concurrent writer
    (older deploy, atlas trigger, duplicate task) from sniping the row
    between our find and update.

    POST-BATCH MATCHING: To keep per-file processing fast, the heavy
    `reprocess_matching()` rebuild is deferred until after the queue is
    drained (no more pending jobs in THIS scope). It runs ONCE per
    drained batch.
    """
    # iter129 — Per-file-type registry. Allows up to one worker per
    # file_type to be alive at any time; legacy callers without a scope
    # use the sentinel key "_legacy".
    scope_key = file_type_scope or "_legacy"
    if scope_key in _worker_running:
        return
    _worker_running.add(scope_key)
    worker_pid = os.getpid()
    logger.info(
        f"Background queue worker started (pid={worker_pid}, scope={scope_key!r})"
    )
    drained_pending_match = False  # True when at least one naukri/pipeline file was processed since last reprocess
    try:
        while True:
            try:
                # iter67 — claim only our host-private rows. Legacy "queued"/
                # "pending" rows are also accepted ONLY when stamped with our
                # host_id (or unset, for very old rows we created before this
                # fix). This prevents a legacy worker on another deployment
                # (sharing the same Mongo) from sniping our rows.
                # iter129 — Optionally scope to a single file_type so each
                # worker only sees its own queue (parallel drain pattern).
                now_iso = datetime.now(timezone.utc).isoformat()
                base_or = [
                    {"status": "queued_local", "host_id": HOST_ID},
                    {
                        "status": {"$in": ["queued", "pending"]},
                        "owner": {"$in": [None, "e1_recruitment_app"]},
                        "$or": [{"host_id": HOST_ID}, {"host_id": {"$exists": False}}, {"host_id": None}],
                    },
                ]
                claim_filter = {"$or": base_or}
                if file_type_scope:
                    # iter129 — Honour both new `file_type` field and the
                    # legacy `upload_type` field used by very old queue
                    # rows. Without this fallback, a typed worker would
                    # silently ignore any legacy-schema row of its type.
                    claim_filter["$and"] = [
                        {"$or": [
                            {"file_type": file_type_scope},
                            {"upload_type": file_type_scope},
                        ]},
                    ]
                job = await db.bulk_upload_queue.find_one_and_update(
                    claim_filter,
                    {"$set": {
                        "status": "processing",
                        "owner": "e1_recruitment_app",
                        "host_id": HOST_ID,
                        "updated_at": now_iso,
                        "worker_pid": worker_pid,
                        "worker_scope": scope_key,
                        "claimed_at": now_iso,
                    }},
                    sort=[("created_at", 1)],
                    return_document=ReturnDocument.AFTER,
                )
                if not job:
                    # Queue drained — schedule deferred reprocess_matching ONCE
                    # (fire-and-forget so the worker immediately returns to the
                    # claim loop and can pick up new uploads while reprocess runs).
                    if drained_pending_match:
                        asyncio.create_task(_trigger_deferred_reprocess(reason=f"queue_drained:{scope_key}"))
                        drained_pending_match = False
                    await asyncio.sleep(3)
                    continue

                job_id = job["_id"]
                # Accept both new schema (file_type/file_name/file_path) and legacy
                # schema (upload_type/filename/filepath) without crashing the worker.
                file_type = job.get("file_type") or job.get("upload_type")
                file_name = job.get("file_name") or job.get("filename")
                file_path = job.get("file_path") or job.get("filepath")

                if not (file_type and file_name and file_path):
                    # Malformed row — mark failed and move on. Never let one bad
                    # doc kill the worker (root cause of "files stuck pending").
                    logger.error(f"Queue: malformed job {job_id} keys={list(job.keys())} → marking failed")
                    await db.bulk_upload_queue.update_one(
                        {"_id": job_id},
                        {"$set": {
                            "status": "failed",
                            "error_message": "Malformed queue record — missing file_type/file_name/file_path",
                            "updated_at": datetime.now(timezone.utc).isoformat(),
                        }},
                    )
                    continue

                logger.info(f"Queue: processing {file_type}/{file_name} (id={job_id})")

                try:
                    path = Path(file_path)
                    if not path.exists():
                        # Legacy rows sometimes store the file content in the doc itself
                        # ("file_content"). Restore it to disk if so.
                        if job.get("file_content"):
                            path.parent.mkdir(parents=True, exist_ok=True)
                            content_blob = job["file_content"]
                            if isinstance(content_blob, str):
                                import base64
                                try:
                                    content_blob = base64.b64decode(content_blob)
                                except Exception:
                                    content_blob = content_blob.encode("utf-8", errors="ignore")
                            path.write_bytes(content_blob)
                        elif (job.get("host_id") and job.get("host_id") != HOST_ID):
                            # Cross-host orphan — file lives on another deployment.
                            # Release the claim back to queued so the correct host
                            # can pick it up. Do NOT mark failed.
                            logger.info(f"Queue: releasing cross-host job {file_name} back to queued (owned by host_id={job.get('host_id')})")
                            await db.bulk_upload_queue.update_one(
                                {"_id": job_id},
                                {"$set": {
                                    "status": "queued_local",
                                    "updated_at": datetime.now(timezone.utc).isoformat(),
                                }, "$unset": {"worker_pid": "", "claimed_at": ""}},
                            )
                            continue
                        else:
                            raise FileNotFoundError(f"File not found on disk: {file_path}")

                    content = path.read_bytes()
                    if not content:
                        raise ValueError("File is empty")
                    # iter116 — Memory observability for the Render 512 MB OOM
                    # issue. Log RSS before the parse (baseline), after the
                    # parse (peak), and after GC (released). A single
                    # `[QueueMem]` log line lets us see exactly which upload
                    # spiked memory and how much was recovered.
                    _mem_before = _rss_mb()
                    logger.info(
                        f"Queue: parsed {file_type}/{file_name} "
                        f"({len(content)} bytes) rss_mb={_mem_before}"
                    )
                    process_fn = _PROCESS_FN.get(file_type)
                    if not process_fn:
                        raise ValueError(f"Unknown file type: {file_type}")

                    # ---- Live row-count progress writer (Iter46) ----
                    # process_fn calls back every 200 rows so the queue doc
                    # carries `progress = {processed, total, percent}` for the
                    # status endpoint / UI to surface live.
                    async def _write_progress(processed: int, total: int):
                        try:
                            pct = int(round((processed / total) * 100)) if total else 0
                            await db.bulk_upload_queue.update_one(
                                {"_id": job_id},
                                {"$set": {
                                    "progress": {"processed": int(processed),
                                                  "total": int(total),
                                                  "percent": pct},
                                    "updated_at": datetime.now(timezone.utc).isoformat(),
                                }},
                            )
                        except Exception:
                            pass  # progress write must never crash the worker
                    result = await process_fn(content, file_name, progress_cb=_write_progress)
                    # iter116 — Drop the raw file bytes ASAP so they don't
                    # linger alongside the next job's content. Then explicit
                    # GC + RSS log so we can correlate uploads to OOM events.
                    del content
                    _mem_peak = _rss_mb()
                    gc.collect()
                    _mem_after = _rss_mb()
                    logger.info(
                        f"[QueueMem] file={file_name} peak_rss_mb={_mem_peak} "
                        f"after_gc_rss_mb={_mem_after} freed_mb={round(_mem_peak - _mem_after, 1)}"
                    )

                    if result.get("success"):
                        # Move file to processed_files directory
                        dest = PROCESSED_BASE / file_type / path.name
                        try:
                            shutil.move(str(path), str(dest))
                        except Exception as me:
                            logger.warning(f"Queue: move failed for {path} → {dest}: {me}")
                        await db.bulk_upload_queue.update_one(
                            {"_id": job_id},
                            {"$set": {
                                "status": "completed",
                                "result": result,
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }}
                        )
                        logger.info(f"Queue: completed {file_type}/{file_name} → {result}")
                        # Mark for deferred reprocess (only naukri/pipeline affect matching)
                        if file_type in ("naukri", "pipeline"):
                            drained_pending_match = True
                    else:
                        await db.bulk_upload_queue.update_one(
                            {"_id": job_id},
                            {"$set": {
                                "status": "failed",
                                "error_message": result.get("error", "Processing returned failure"),
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }}
                        )
                        logger.warning(f"Queue: failed {file_type}/{file_name} — {result}")

                except Exception as e:
                    logger.exception(f"Queue: error {file_type}/{file_name} — {e}")
                    await db.bulk_upload_queue.update_one(
                        {"_id": job_id},
                        {"$set": {
                            "status": "failed",
                            "error_message": str(e),
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )

                # Small delay between jobs
                await asyncio.sleep(0.5)
            except asyncio.CancelledError:
                raise
            except Exception as outer:
                # Last-resort safety net: log & sleep so the worker never dies on
                # an unexpected row. (Original cause of "files stuck pending".)
                logger.exception(f"Queue: unexpected outer-loop error — {outer}")
                await asyncio.sleep(5)

    except asyncio.CancelledError:
        logger.info(f"Background queue worker stopped (scope={scope_key!r})")
    finally:
        _worker_running.discard(scope_key)


# ============ BULK UPLOAD ENDPOINTS ============

@api_router.get("/bulk-upload/status")
async def bulk_upload_status(user: str = Depends(get_current_user)):
    """Return queue status for all types: pending, processing, completed, failed."""
    result = {}
    for utype in BULK_TYPES_LIST:
        # Pending + Processing from DB (queued is our discriminator; pending kept
        # for backward-compat)
        active = await db.bulk_upload_queue.find(
            {"file_type": utype, "status": {"$in": ["queued", "queued_local", "pending", "processing"]}, "host_id": HOST_ID},
            {"_id": 1, "file_name": 1, "file_path": 1, "status": 1, "created_at": 1,
             "error_message": 1, "progress": 1}
        ).sort("created_at", 1).to_list(None)
        pending = []
        for j in active:
            size = 0
            try:
                p = Path(j["file_path"])
                if p.exists():
                    size = p.stat().st_size
            except Exception:
                pass
            pending.append({
                "id": str(j["_id"]),
                "name": j["file_name"],
                "status": j["status"],
                "size": size,
                "progress": j.get("progress") or None,
            })

        # Completed from DB
        completed = await db.bulk_upload_queue.find(
            {"file_type": utype, "status": "completed"},
            {"_id": 1, "file_name": 1, "file_path": 1, "result": 1, "updated_at": 1}
        ).sort("updated_at", -1).to_list(None)
        processed = []
        for j in completed:
            size = 0
            try:
                p = PROCESSED_BASE / utype
                for fp in p.iterdir():
                    if fp.name.endswith(j["file_name"]) or j["file_name"] in fp.name:
                        size = fp.stat().st_size
                        break
            except Exception:
                pass
            processed.append({
                "id": str(j["_id"]),
                "name": j["file_name"],
                "size": size,
                "result": j.get("result"),
            })

        # Failed from DB (read both new error_message and legacy error fields)
        failed_docs = await db.bulk_upload_queue.find(
            {"file_type": utype, "status": "failed"},
            {"_id": 1, "file_name": 1, "error_message": 1, "error": 1, "updated_at": 1}
        ).sort("updated_at", -1).to_list(None)
        failed = [{"id": str(j["_id"]), "name": j["file_name"],
                    "error": j.get("error_message") or j.get("error") or "Unknown error"} for j in failed_docs]

        result[utype] = {"pending": pending, "processed": processed, "failed": failed}
    return result


@api_router.post("/bulk-upload/process-now")
async def trigger_bulk_process(user: str = Depends(get_current_user)):
    """Manual trigger — worker is always running, this confirms status."""
    pending_count = await db.bulk_upload_queue.count_documents({"status": "pending"})
    processing_count = await db.bulk_upload_queue.count_documents({"status": "processing"})
    return {"success": True, "message": "Worker is running", "pending": pending_count, "processing": processing_count}


@api_router.post("/bulk-upload/{upload_type}")
async def bulk_upload_files(
    upload_type: str,
    files: List[UploadFile] = File(...),
    user: str = Depends(get_current_user)
):
    """Save files to disk and enqueue DB records. Worker processes them in background."""
    if upload_type not in BULK_TYPES_LIST:
        raise HTTPException(status_code=400, detail=f"Invalid type: {upload_type}. Must be naukri, pipeline, or score")
    upload_dir = UPLOAD_BASE / upload_type
    upload_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    skipped = []
    for f in files:
        if not f.filename.lower().endswith(('.csv', '.xlsx')):
            skipped.append({"name": f.filename, "reason": "Only .csv or .xlsx allowed"})
            logger.warning(f"BulkUpload: skipped {f.filename} (invalid extension)")
            continue
        content = await f.read()
        if not content:
            skipped.append({"name": f.filename, "reason": "Empty file"})
            logger.warning(f"BulkUpload: skipped {f.filename} (empty)")
            continue
        safe_name = f"{int(datetime.now(timezone.utc).timestamp() * 1000)}_{ObjectId()}_{f.filename}"
        dest = upload_dir / safe_name
        try:
            dest.write_bytes(content)
        except Exception as e:
            logger.exception(f"BulkUpload: disk write failed for {f.filename}: {e}")
            skipped.append({"name": f.filename, "reason": f"Disk write failed: {e}"})
            continue
        try:
            # Insert with our internal status `queued` (not `pending`) so any
            # external/legacy worker that scans for {status: 'pending'} cannot
            # snipe our row. Our own worker reads {status: 'queued'}.
            await db.bulk_upload_queue.insert_one({
                "file_name": f.filename,
                "file_path": str(dest),
                "file_type": upload_type,
                "file_size": len(content),
                # iter67 — use a host-private status so a legacy worker on
                # another deployment (which scans for {"queued","pending"})
                # cannot snipe our row and falsely fail it as "file not found"
                # (the file lives only on our local /app/uploads filesystem).
                "status": "queued_local",
                "owner": "e1_recruitment_app",
                "host_id": HOST_ID,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "error_message": None,
                "result": None,
            })
            saved.append(safe_name)
            logger.info(f"BulkUpload: queued {upload_type}/{f.filename} ({len(content)} bytes)")
        except Exception as e:
            logger.exception(f"BulkUpload: DB enqueue failed for {f.filename}: {e}")
            skipped.append({"name": f.filename, "reason": f"DB enqueue failed: {e}"})
    return {"success": True, "saved": saved, "skipped": skipped, "count": len(saved)}


@api_router.post("/bulk-upload/{upload_type}/clear-failed")
async def clear_failed_uploads(upload_type: str, user: str = Depends(get_current_user)):
    """Archive all `failed` queue rows for a given upload_type so the UI is clean."""
    if upload_type not in BULK_TYPES_LIST:
        raise HTTPException(status_code=400, detail=f"Invalid type: {upload_type}")
    res = await db.bulk_upload_queue.update_many(
        {"file_type": upload_type, "status": "failed"},
        {"$set": {"status": "archived", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"success": True, "archived": res.modified_count}


@api_router.delete("/bulk-upload/{upload_type}/{queue_id}")
async def delete_bulk_file(upload_type: str, queue_id: str, user: str = Depends(get_current_user)):
    """Delete a pending file from queue and disk."""
    try:
        oid = ObjectId(queue_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid queue ID")
    job = await db.bulk_upload_queue.find_one({"_id": oid})
    if not job:
        raise HTTPException(status_code=404, detail="Queue entry not found")
    if job["status"] == "processing":
        raise HTTPException(status_code=409, detail="Cannot delete file while processing")
    # Remove file from disk
    try:
        path = Path(job["file_path"])
        if path.exists():
            path.unlink()
    except Exception:
        pass
    await db.bulk_upload_queue.delete_one({"_id": oid})
    return {"success": True, "deleted": job["file_name"]}


async def _process_naukri_file(content: bytes, filename: str, progress_cb=None) -> dict:
    """Core naukri processing — extracted from upload_naukri for reuse.
    `progress_cb(processed:int, total:int)` is called every 200 rows so the
    worker can surface live row-count progress on the queue document."""
    df = parse_file(content, filename)
    df.columns = df.columns.str.strip()
    col_map = {}
    for csv_col in df.columns:
        if csv_col in NAUKRI_COLUMN_MAP:
            col_map[csv_col] = NAUKRI_COLUMN_MAP[csv_col]
        elif csv_col.strip().lower() in _NAUKRI_CI_LOOKUP:
            col_map[csv_col] = _NAUKRI_CI_LOOKUP[csv_col.strip().lower()]
    mapped_fields = set(col_map.values())
    if "email" not in mapped_fields and "phone" not in mapped_fields:
        return {"success": False, "error": "No email/phone column found"}
    total_rows = int(len(df))
    if progress_cb:
        await progress_cb(0, total_rows)
    inserted = 0
    updated = 0
    for idx, row in df.iterrows():
        try:
            doc = {}
            for csv_col, db_field in col_map.items():
                doc[db_field] = clean_value(row.get(csv_col))
            for csv_col in df.columns:
                if csv_col not in col_map:
                    safe = re.sub(r'[^\w]', '_', csv_col.strip().lower()).strip('_')
                    doc[f"_extra_{safe}"] = clean_value(row.get(csv_col))
            email = normalize_email(doc.get("email"))
            phone = normalize_phone(doc.get("phone"))
            doc["email"] = email
            doc["phone"] = phone
            for date_field in ("date_of_application", "date_of_birth"):
                if doc.get(date_field):
                    doc[date_field] = normalize_date(doc[date_field])
            doc["updated_at"] = datetime.now(timezone.utc)
            if not email and not phone:
                continue
            query = {"$or": []}
            if email: query["$or"].append({"email": email})
            if phone: query["$or"].append({"phone": phone})
            existing = await db.naukri_applies.find_one(query)
            if existing:
                await db.naukri_applies.update_one({"_id": existing["_id"]}, {"$set": doc})
                updated += 1
            else:
                doc["created_at"] = datetime.now(timezone.utc)
                await db.naukri_applies.insert_one(doc)
                inserted += 1
        except Exception:
            pass
        # Live progress every 200 rows
        if progress_cb and (idx + 1) % 200 == 0:
            await progress_cb(idx + 1, total_rows)
    if progress_cb:
        await progress_cb(total_rows, total_rows)
    # iter116 — release the DataFrame ASAP after iteration completes so the
    # next file's parse doesn't accumulate on top of this one (Render 512 MB
    # OOM mitigation). `df.iterrows()` materializes a Series per row which
    # the GC sometimes leaves around; explicit del + collect releases both.
    del df
    gc.collect()
    # NOTE: reprocess_matching() / _sync_job_titles_master() are now deferred
    # to the queue worker (run ONCE after batch drains) for performance.
    return {"success": True, "inserted": inserted, "updated": updated, "total": total_rows}


async def _process_pipeline_file(content: bytes, filename: str, progress_cb=None) -> dict:
    """Core pipeline processing — extracted from upload_pipeline for reuse."""
    df = parse_file(content, filename)
    df.columns = df.columns.str.strip()
    cols = df.columns.tolist()
    seen_bases = set()
    keep_indices = []
    for i, col in enumerate(cols):
        base = re.sub(r'\.\d+$', '', col.strip()).lower()
        if base not in seen_bases:
            seen_bases.add(base)
            keep_indices.append(i)
    df = df.iloc[:, keep_indices]
    df.columns = [re.sub(r'\.\d+$', '', c.strip()) for c in df.columns]
    col_map = {}
    for csv_col in df.columns:
        normalized = csv_col.strip().lower()
        if normalized in _PIPELINE_CI_LOOKUP:
            db_field = _PIPELINE_CI_LOOKUP[normalized]
            col_map[csv_col] = "pipeline_id" if db_field == "id" else db_field
    mapped_fields = set(col_map.values())
    if "email" not in mapped_fields and "phone" not in mapped_fields:
        return {"success": False, "error": "No email/phone column found"}
    total_rows = int(len(df))
    if progress_cb:
        await progress_cb(0, total_rows)
    inserted = 0
    updated = 0
    for idx, row in df.iterrows():
        try:
            doc = {}
            for csv_col, db_field in col_map.items():
                doc[db_field] = clean_value(row.get(csv_col))
            for csv_col in df.columns:
                if csv_col not in col_map:
                    safe = re.sub(r'[^\w]', '_', csv_col.strip().lower()).strip('_')
                    doc[f"_extra_{safe}"] = clean_value(row.get(csv_col))
            email = normalize_email(doc.get("email"))
            phone = normalize_phone(doc.get("phone"))
            doc["email"] = email
            doc["phone"] = phone
            doc["updated_at"] = datetime.now(timezone.utc)
            if not email and not phone:
                continue
            query = {"$or": []}
            if email: query["$or"].append({"email": email})
            if phone: query["$or"].append({"phone": phone})
            existing = await db.pipeline_data.find_one(query)
            if existing:
                await db.pipeline_data.update_one({"_id": existing["_id"]}, {"$set": doc})
                updated += 1
            else:
                doc["created_at"] = datetime.now(timezone.utc)
                await db.pipeline_data.insert_one(doc)
                inserted += 1
        except Exception:
            pass
        if progress_cb and (idx + 1) % 200 == 0:
            await progress_cb(idx + 1, total_rows)
    if progress_cb:
        await progress_cb(total_rows, total_rows)
    # iter116 — release the DataFrame ASAP after iteration completes (see
    # _process_naukri_file for the same Render-OOM rationale).
    del df
    gc.collect()
    # NOTE: reprocess_matching() is deferred to queue worker post-batch.
    return {"success": True, "inserted": inserted, "updated": updated, "total": total_rows}


async def _process_score_file(content: bytes, filename: str, progress_cb=None) -> dict:
    """Core score sheet processing — extracted from upload_scoresheet for reuse.

    Iter48 — in addition to the legacy `score_sheet` insert, this now also:
        * Appends the (round_name, score) into `bb_applicant_updates.scores[]`
          for the matched applicant (email primary, phone fallback).
        * Upserts the round_name into `bb_rounds` (case-insensitive dedupe).
      So the same score becomes visible on both "View Attended Applicants"
      and "Update Applicants Scores" without a separate import.
    """
    df = parse_file(content, filename)
    df.columns = df.columns.str.strip().str.lower()
    required = {"name", "email", "phone", "score", "round_name"}
    if not required.issubset(set(df.columns)):
        return {"success": False, "error": f"Missing columns: {required - set(df.columns)}"}
    total_rows = int(len(df))
    if progress_cb:
        await progress_cb(0, total_rows)
    inserted = 0
    seen_round_names: set = set()
    for idx, row in df.iterrows():
        try:
            email = normalize_email(row.get("email"))
            phone = normalize_phone(row.get("phone"))
            name = str(row.get("name") or "").strip()
            round_name = str(row.get("round_name") or "").strip()
            score_val = row.get("score")
            if not email and not phone:
                continue
            if not round_name:
                continue
            try:
                score = float(score_val) if not pd.isna(score_val) else 0.0
            except (ValueError, TypeError):
                score = 0.0
            now_iso = datetime.now(timezone.utc).isoformat()
            doc = {"name": name, "email": email, "phone": phone, "score": score,
                   "round_name": round_name, "created_at": now_iso}
            await db.score_sheet.insert_one(doc)
            inserted += 1
            seen_round_names.add(round_name)

            # ---- Iter48: sync into bb_applicant_updates (append-only) ----
            or_clauses = []
            if email: or_clauses.append({"email": email})
            if phone: or_clauses.append({"phone": phone})
            existing = await db.bb_applicant_updates.find_one(
                {"$or": or_clauses} if len(or_clauses) > 1 else or_clauses[0],
                {"_id": 0, "email": 1, "scores": 1},
            )
            existing_scores = list((existing or {}).get("scores") or [])
            existing_round_lc = {
                str(s.get("round_name") or "").strip().lower()
                for s in existing_scores if s.get("round_name")
            }
            if round_name.lower() in existing_round_lc:
                # Don't overwrite — preserve existing per-applicant score.
                continue
            existing_scores.append({"round_name": round_name, "score": score})
            target_email = (existing or {}).get("email") or email or phone
            await db.bb_applicant_updates.update_one(
                {"email": target_email},
                {"$set": {
                    "email": target_email,
                    "scores": existing_scores,
                    "name": name or (existing or {}).get("name") or "",
                    "phone": phone or (existing or {}).get("phone") or "",
                    "updated_at": now_iso,
                },
                "$setOnInsert": {"status": "On hold", "isImported": True}},
                upsert=True,
            )
        except Exception:
            pass
        if progress_cb and (idx + 1) % 200 == 0:
            await progress_cb(idx + 1, total_rows)
    if progress_cb:
        await progress_cb(total_rows, total_rows)

    # ---- Iter48: auto-register round names into bb_rounds (case-insensitive)
    rounds_added = 0
    for rn in sorted(seen_round_names):
        ex = await db.bb_rounds.find_one(
            {"name": {"$regex": f"^{re.escape(rn)}$", "$options": "i"}},
            {"_id": 1, "active": 1},
        )
        if ex:
            if ex.get("active") is False:
                await db.bb_rounds.update_one({"_id": ex["_id"]}, {"$set": {"active": True}})
            continue
        await db.bb_rounds.insert_one({
            "name": rn, "active": True, "order": 0,
            "source": "score_sheet",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        rounds_added += 1
    # iter116 — release DataFrame after all post-processing finishes.
    del df
    gc.collect()
    return {"success": True, "inserted": inserted, "total": total_rows,
            "rounds_registered": rounds_added}


_PROCESS_FN = {
    "naukri": _process_naukri_file,
    "pipeline": _process_pipeline_file,
    "score": _process_score_file,
}


@api_router.post("/reprocess")
async def trigger_reprocess(user: str = Depends(get_current_user)):
    """Re-normalize all data and rebuild registered_candidates from scratch."""
    naukri_before = await db.naukri_applies.count_documents({})
    pipeline_before = await db.pipeline_data.count_documents({})
    registered_before = await db.registered_candidates.count_documents({})

    await reprocess_matching()

    registered_after = await db.registered_candidates.count_documents({})
    return {
        "success": True,
        "message": "Reprocessing complete — data re-normalized and matching rebuilt",
        "naukri_count": naukri_before,
        "pipeline_count": pipeline_before,
        "registered_before": registered_before,
        "registered_after": registered_after,
        "change": registered_after - registered_before,
    }


@api_router.get("/debug/matching")
async def debug_matching(user: str = Depends(get_current_user)):
    """Debug endpoint: shows every naukri record, its normalized identifiers,
    and whether a pipeline match was found (and why/why not)."""
    naukri_list = await db.naukri_applies.find({}, {"_id": 0}).to_list(None)
    pipeline_list = await db.pipeline_data.find({}, {"_id": 0}).to_list(None)

    # Build normalized pipeline lookups
    pipeline_by_email = {}
    pipeline_by_phone = {}
    for p in pipeline_list:
        em = normalize_email(p.get('email'))
        ph = normalize_phone(p.get('phone'))
        if em:
            pipeline_by_email[em] = {"name": p.get("name"), "email": em, "phone": ph, "job_role": p.get("job_role")}
        if ph:
            pipeline_by_phone[ph] = {"name": p.get("name"), "email": em, "phone": ph, "job_role": p.get("job_role")}

    results = []
    matched_count = 0
    unmatched_count = 0

    for naukri in naukri_list:
        n_email = normalize_email(naukri.get('email'))
        n_phone = normalize_phone(naukri.get('phone'))

        match_type = None
        pipeline_info = None

        if n_email and n_email in pipeline_by_email:
            match_type = "email"
            pipeline_info = pipeline_by_email[n_email]
        elif n_phone and n_phone in pipeline_by_phone:
            match_type = "phone"
            pipeline_info = pipeline_by_phone[n_phone]

        is_matched = match_type is not None
        if is_matched:
            matched_count += 1
        else:
            unmatched_count += 1

        results.append({
            "naukri_name": naukri.get("name"),
            "naukri_email": n_email,
            "naukri_phone": n_phone,
            "naukri_job_title": naukri.get("job_title"),
            "matched": is_matched,
            "match_type": match_type,
            "pipeline_match": pipeline_info,
        })

    return {
        "total_naukri": len(naukri_list),
        "total_pipeline": len(pipeline_list),
        "matched": matched_count,
        "unmatched": unmatched_count,
        "details": results,
    }


# Health check
@api_router.get("/")
async def root():
    return {"message": "Recruitment Analytics API", "status": "healthy", "version": "4.0"}

# Include the router in the main app
app.include_router(api_router)

# Include BluBridge modules router (with traceback safety per deployment spec)
try:
    from bb_modules import bb_router, pub_router, init_bb, backfill_form_slugs
    init_bb(db, get_current_user, _build_college_rank_lookup, _classify_college)
    app.include_router(bb_router)
    app.include_router(pub_router)
except Exception:
    import traceback
    traceback.print_exc()
    raise

# Include WhatsApp Resend module (iter67)
try:
    from bb_resend import resend_router, init_resend
    init_resend(db, get_current_user)
    app.include_router(resend_router)
except Exception:
    import traceback
    traceback.print_exc()
    raise

# Include Help / Templates module (iter67)
try:
    from bb_help import help_router
    app.include_router(help_router)
except Exception:
    import traceback
    traceback.print_exc()
    raise

# Include Manual Operations module (iter67)
try:
    from bb_manual import manual_router, init_manual, _ensure_default_test_credentials
    init_manual(db, get_current_user)
    app.include_router(manual_router)
except Exception:
    import traceback
    traceback.print_exc()
    raise

# Wire centralized messaging gate (TEST_MODE → bb_test_credentials lookup)
try:
    from messaging import init_messaging, is_test_mode
    init_messaging(db)
except Exception:
    import traceback
    traceback.print_exc()
    raise

# iter133 — Team Score module (isolated; new collections ts_rounds + ts_employees)
try:
    from team_score import attach as _attach_team_score

    async def _ts_require_auth(request):
        # Reuse the existing admin auth gate; tolerates the session
        # cookie path used everywhere else in the app.
        try:
            return await get_current_user(request)
        except Exception:
            from fastapi import HTTPException as _HE
            raise _HE(status_code=401, detail="Authentication required")
    _attach_team_score(app, db, _ts_require_auth)
except Exception:
    import traceback
    traceback.print_exc()
    raise


@app.get("/api/messaging/status")
async def messaging_status(user: str = Depends(get_current_user)):
    """Surface the current TEST_MODE status to the UI banner."""
    return {"test_mode": is_test_mode()}

# CORS Configuration
try:
    _cors_origins = [os.environ['FRONTEND_URL']]
except KeyError:
    print("[startup] FATAL: FRONTEND_URL env var is required but missing.", flush=True)
    raise
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
print("MIDDLEWARE + ROUTERS REGISTERED", flush=True)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Startup event
@app.on_event("startup")
async def startup_event():
    # Create indexes — tolerate Atlas quota errors so the app still boots on
    # an over-quota free tier (indexes most likely already exist anyway).
    _index_specs = [
        (db.naukri_applies, [("email", 1), ("phone", 1)]),
        (db.pipeline_data, [("email", 1), ("phone", 1)]),
        (db.registered_candidates, [("email", 1), ("phone", 1)]),
        (db.registered_candidates, "email_type"),
        (db.registered_candidates, "result_status"),
        (db.registered_candidates, "schedule_date"),
        (db.registered_candidates, "otp_verified"),
        (db.registered_candidates, "_normalized_job_role"),
        (db.registered_candidates, "_nirf_category"),
        (db.registered_candidates, "_college_status"),
        (db.registered_candidates, "name"),
        (db.naukri_applies, "_normalized_job_role"),
        (db.naukri_applies, "_nirf_category"),
        (db.bulk_upload_queue, [("status", 1), ("created_at", 1)]),
    ]
    for coll, spec in _index_specs:
        try:
            await coll.create_index(spec)
        except Exception as e:
            logger.warning(f"[startup] create_index skipped on {coll.name} ({spec}): {e}")
    try:
        await db.job_titles_master.create_index("normalized_job_title", unique=True)
    except Exception as e:
        logger.warning(f"[startup] create_index skipped on job_titles_master (unique): {e}")
    try:
        await db.job_titles_master.create_index("is_mapped")
    except Exception as e:
        logger.warning(f"[startup] create_index skipped on job_titles_master.is_mapped: {e}")

    # iter108 — One-shot backfill: reclassify legacy rows stuck at
    # Unknown/null `_normalized_job_role` using current keyword mappings.
    # Idempotent via `bb_meta` flag so reboots don't repeat the work.
    asyncio.create_task(_backfill_unknown_classifications_once())

    # iter127 — One-shot historical sync of every analytics-visible role
    # into `bb_job_roles` + `job_titles_master`. Catches the gap where a
    # canonical resolved role (e.g. "AI & ML Engineer" derived from a raw
    # upload title) appears in Summary Statistics but never made it into
    # the catalog (and therefore the Job Roles page / dropdowns / Unmapped
    # Keywords). Runs only once per process boot; subsequent uploads use
    # the upgraded `_sync_job_titles_master` directly + the periodic
    # safety net below.
    asyncio.create_task(_sync_job_titles_master())

    # iter127 — Periodic safety-net sync. Even if a post-upload background
    # task dies (Render restart, worker exception), the catalog converges
    # within minutes because this loop re-runs the sync every 15 minutes.
    asyncio.create_task(_periodic_job_titles_sync())

    # iter110 — One-shot backfill: reclassify rows whose `_college_status`
    # was set under the old binary scheme ("NIRF - #N" or "Non NIRF") and
    # any non-test rows still missing the field. Idempotent via bb_meta.
    asyncio.create_task(_backfill_college_status_once())

    # iter130 — Lifecycle (Activate/Deactivate) status backfill.
    # Adds `status='active'` to every existing bb_job_roles /
    # bb_job_openings / bb_hiring_forms row that lacks the field, and
    # creates a `status` index on each collection. Idempotent — only
    # touches rows where the field is missing/null/empty.
    from bb_modules import _ensure_status_indexes_and_backfill
    asyncio.create_task(_ensure_status_indexes_and_backfill())

    # Backfill slugs for existing hiring forms + ensure unique index
    try:
        await backfill_form_slugs()
    except Exception as e:
        logger.warning(f"[startup] backfill_form_slugs skipped: {e}")
    
    # Resume: reset any stuck "processing" records (from our own worker) back
    # to "queued_local". Only touches rows we own (this host) to avoid resurrecting
    # legacy/phantom rows from other deployments.
    # All cleanup writes wrapped in try/except so Atlas free-tier quota errors
    # don't kill the startup — login + reads still work even when DB is over quota.
    try:
        stuck = await db.bulk_upload_queue.update_many(
            {"status": "processing", "owner": "e1_recruitment_app", "host_id": HOST_ID},
            {"$set": {"status": "queued_local", "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        if stuck.modified_count > 0:
            logger.info(f"Reset {stuck.modified_count} stuck processing jobs to queued")
    except Exception as e:
        logger.warning(f"[startup] stuck-job reset skipped: {e}")

    # One-time cleanup: archive legacy phantom records (pre-fix queue rows that
    # had `error: 'Invalid upload_type'` and a stray `started_at` field). They
    # are not actionable and clutter the UI's "failed" list.
    try:
        legacy = await db.bulk_upload_queue.update_many(
            {"status": "failed", "error": "Invalid upload_type"},
            {"$set": {"status": "archived", "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        if legacy.modified_count > 0:
            logger.info(f"Archived {legacy.modified_count} legacy phantom failed records")
    except Exception as e:
        logger.warning(f"[startup] legacy-phantom archive skipped: {e}")

    # Iter67 — Archive orphan failed rows whose file no longer exists on disk
    # AND that belong to this host. Failed rows from OTHER hosts must not be
    # touched (those files only exist on the other deployment's filesystem).
    try:
        orphan_count = 0
        async for row in db.bulk_upload_queue.find(
            {"status": "failed", "host_id": HOST_ID},
            {"_id": 1, "file_path": 1}
        ):
            fp = row.get("file_path")
            if fp and not Path(fp).exists():
                await db.bulk_upload_queue.update_one(
                    {"_id": row["_id"]},
                    {"$set": {
                        "status": "archived",
                        "archive_reason": "orphan_file_missing_on_disk",
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }},
                )
                orphan_count += 1
        if orphan_count > 0:
            logger.info(f"Archived {orphan_count} orphan failed rows (file missing on disk)")
    except Exception as e:
        logger.warning(f"[startup] orphan archive skipped: {e}")

    # Iter67 — Reclaim cross-host failed rows that belong to us. The other
    # deployment's legacy worker may have sniped one of our jobs and falsely
    # marked it failed with "File not found on disk". If the file still exists
    # on OUR disk, restore the row to queued_local so we can retry.
    try:
        reclaimed = 0
        async for row in db.bulk_upload_queue.find(
            {"status": "failed", "host_id": HOST_ID, "error_message": {"$regex": "^File not found on disk"}},
            {"_id": 1, "file_path": 1}
        ):
            fp = row.get("file_path")
            if fp and Path(fp).exists():
                await db.bulk_upload_queue.update_one(
                    {"_id": row["_id"]},
                    {"$set": {
                        "status": "queued_local",
                        "error_message": None,
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                    }, "$unset": {"worker_pid": "", "claimed_at": ""}},
                )
                reclaimed += 1
        if reclaimed > 0:
            logger.info(f"Reclaimed {reclaimed} cross-host falsely-failed rows back to queued_local")
    except Exception as e:
        logger.warning(f"[startup] cross-host reclaim skipped: {e}")

    # Write test credentials (dev-only aid). Render's filesystem is read-only
    # under /app, so skip silently in production. Override via TEST_CRED_DIR env
    # var if you ever want it written elsewhere (e.g. a persistent disk).
    _cred_dir = Path(os.getenv("TEST_CRED_DIR", "/app/memory"))
    try:
        _cred_dir.mkdir(parents=True, exist_ok=True)
        with open(_cred_dir / "test_credentials.md", "w") as f:
            f.write("# Test Credentials\n\n")
            f.write("## Admin Account (RecruitIQ)\n")
            f.write("- Username: `Admin User`\n")
            f.write("- Password: `Admin User`\n")
    except (PermissionError, OSError):
        # Read-only FS (Render, Heroku, etc.) — not fatal, just skip.
        logger.info(f"[startup] test_credentials.md write skipped — {_cred_dir} not writable")
    except Exception as e:
        logger.warning(f"Failed to write test credentials: {e}")

    # iter68 — Seed default tester credentials BEFORE messaging workers start
    # so the centralized TEST_MODE gate has a non-empty allow list on first run.
    try:
        await _ensure_default_test_credentials()
        count = await db.bb_test_credentials.count_documents({})
        logger.info(f"[TEST_MODE] is_on={is_test_mode()} testers_in_db={count}")
    except Exception as e:
        logger.error(f"Failed to seed default tester credentials: {e}")

    # iter88 — One-shot historical rejection-flag backfill.
    # Sets `rejection_sent=True` on every existing rejected applicant so the
    # new evening dispatcher will never message any historical record.
    # Idempotent: marker doc in `bb_migrations` short-circuits on subsequent boots.
    try:
        from maintenance.backfill_rejection_flags import run_backfill as _run_rej_backfill
        result = await _run_rej_backfill(db)
        logger.info(f"[Backfill:rejection_flags] result={result}")
    except Exception as e:
        logger.warning(f"[startup] rejection-flag backfill skipped: {e}")

    # iter129 — Per-file-type concurrent queue workers. Each file_type
    # (naukri, pipeline, score) gets its own worker so a slow pipeline
    # batch can never starve the naukri queue (the user-reported
    # "naukri stuck at queued_local" bug — pure FIFO bottleneck).
    for _scope in ("naukri", "pipeline", "score"):
        asyncio.create_task(_bg_queue_worker(file_type_scope=_scope))
    logger.info("DB-driven background queue workers launched (per file_type)")

    # Start messaging background workers
    from bg_workers import init_workers, start_all_workers
    init_workers(db)
    await start_all_workers()
    logger.info("Messaging background workers launched")


# iter125d — Lightweight liveness probe for external uptime monitors
# (Render keep-alive / UptimeRobot / Pingdom). No auth, no DB queries,
# no logging — stays a true zero-overhead endpoint. Mounted directly on
# `app` (not `api_router`) so the path is `/health`, not `/api/health`,
# which is the convention most hosted-monitor services expect.
@app.api_route("/health", methods=["GET", "HEAD"])
async def health():
    return Response(status_code=200)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


# ============================================================================
# REACT SPA FALLBACK (production deployment, e.g. Render)
# ----------------------------------------------------------------------------
# Only activates when the built frontend exists on disk. In the Emergent dev
# preview the frontend runs on its own webpack dev server (port 3000), so the
# build folder is absent and this catch-all silently does nothing.
# ============================================================================
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

_FRONTEND_BUILD_DIR = Path(__file__).resolve().parent.parent / "frontend" / "build"
_FRONTEND_INDEX = _FRONTEND_BUILD_DIR / "index.html"

if _FRONTEND_INDEX.exists():
    # Serve hashed static assets (JS/CSS/images) under /static
    _STATIC_DIR = _FRONTEND_BUILD_DIR / "static"
    if _STATIC_DIR.exists():
        app.mount("/static", StaticFiles(directory=str(_STATIC_DIR)), name="static")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_react_routes(full_path: str):
        # Never shadow API or OpenAPI surfaces.
        if full_path.startswith(("api/", "api", "docs", "openapi.json", "redoc")):
            raise HTTPException(status_code=404)
        # Serve any built static file directly if it exists (favicon, manifest, etc.).
        candidate = _FRONTEND_BUILD_DIR / full_path
        if full_path and candidate.is_file():
            return FileResponse(str(candidate))
        # Otherwise fall through to React Router via index.html.
        return FileResponse(str(_FRONTEND_INDEX))
