"""BluBridge Modules — New features for the Hiring Pipeline.
Separate router to avoid modifying existing server.py logic."""

from fastapi import APIRouter, HTTPException, Request, Query
from pydantic import BaseModel, field_validator
import re as _re


def _validate_phone_10digits(v):
    """Shared Pydantic validator — normalize and validate a 10-digit phone.

    iter94 — Tolerant normalization (accepts the same formats the frontend
    accepts so backend and frontend are aligned):
      • bare 10 digits                          → keep
      • leading "0"  (11 chars: 0 + 10 digits)  → strip the 0
      • leading "+91" (13 chars: +91 + 10)      → strip the +91
      • leading "91"  (12 chars: 91 + 10)       → strip the 91
      • leading "91" but len==10 already        → keep (numbers like 9123456789
                                                   that just happen to start
                                                   with 91 are valid 10-digit
                                                   numbers, NOT a prefix)
    Anything else → 422 with the canonical helper-text message.
    Storage contract: DB always holds the bare 10-digit form.
    """
    s = (v or "").strip().replace(" ", "")
    if s.startswith("+91") and len(s) == 13 and s[3:].isdigit():
        s = s[3:]
    elif s.startswith("0") and s.lstrip("0").isdigit():
        # iter104 — Always strip ALL leading zeros up-front, then validate
        # the remainder. The earlier `len==11` gate caused "0123456789" (one
        # zero + 9 digits = 10 chars total) to be evaluated as if it were a
        # bare 10-digit number, producing a misleading "invalid" rejection.
        # Stripping leading zeros first makes the final length check decide.
        s = s.lstrip("0")
    elif s.startswith("91") and len(s) == 12 and s.isdigit():
        s = s[2:]
    # else: leave as-is (len==10 numbers starting with 91 keep all 10 digits)
    if not _re.fullmatch(r"[0-9]{10}", s):
        raise ValueError("Phone must contain 10 digits (you may type it with +91, 91, or leading 0 — we normalize automatically).")
    return s


def _validate_nonempty(v, field_name="field"):
    """Shared Pydantic validator — string must be non-empty after trim."""
    s = (v or "").strip()
    if not s:
        raise ValueError(f"{field_name} is required")
    return s
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from bson import ObjectId
import hashlib, secrets, re, logging, uuid
from _fmt import to_24h_db

bb_router = APIRouter(prefix="/api/bb")
# Public router — no auth prefix
pub_router = APIRouter(prefix="/api/pub")

_logger = logging.getLogger("bb_modules")


# ============ MESSAGING — Delegated to messaging.py ============
# All recipient overrides (TEST_MODE) happen in messaging.py centrally.
# No direct messaging logic in this file.

# Shared dependencies — injected from server.py
_db = None
_auth_fn = None
_build_college_rank_lookup_fn = None
_classify_college_fn = None


def init_bb(database, auth_fn, college_lookup_fn, classify_fn):
    global _db, _auth_fn, _build_college_rank_lookup_fn, _classify_college_fn
    _db = database
    _auth_fn = auth_fn
    _build_college_rank_lookup_fn = college_lookup_fn
    _classify_college_fn = classify_fn


async def backfill_form_slugs():
    """Backfill `slug` for any hiring form missing it, and ensure unique index.
    Safe to call repeatedly. Should be invoked at app startup.
    Also ensures the unique compound index on bb_college_schedules.
    """
    if _db is None:
        return
    try:
        # Build slug for docs missing it
        cursor = _db.bb_hiring_forms.find({"$or": [{"slug": {"$exists": False}}, {"slug": ""}, {"slug": None}]})
        async for f in cursor:
            base = _slugify(f.get("name") or f.get("job_role") or "form")
            slug = await _unique_slug(base)
            await _db.bb_hiring_forms.update_one({"_id": f["_id"]}, {"$set": {"slug": slug}})
            _logger.info(f"backfilled slug '{slug}' for form {f['_id']}")
        # Ensure unique index on slug (sparse to allow legacy nulls during migration)
        try:
            await _db.bb_hiring_forms.create_index("slug", unique=True, sparse=True)
        except Exception as e:
            _logger.warning(f"slug index creation skipped: {e}")
        # College schedules: compound index on (college_name, job_role) — collation
        # gives case-insensitive uniqueness so HR can't sneak in "IIT" + "Iit".
        try:
            await _db.bb_college_schedules.create_index(
                [("college_name", 1), ("job_role", 1)],
                name="college_role_unique",
                unique=True,
                collation={"locale": "en", "strength": 2},
            )
        except Exception as e:
            _logger.warning(f"college_schedules index creation skipped: {e}")
    except Exception as e:
        _logger.error(f"backfill_form_slugs failed: {e}")


async def _require_auth(request: Request):
    return await _auth_fn(request)


# ============ CANDIDATE MATCHING & ENRICHMENT HELPERS ============
# Smart cross-source resolver used by:
#   - verify_applicant_otp  (runtime fallback for the success card)
#   - register_applicant / register_college_applicant (preserve existing values)
#   - backfill_pipeline_extras.py (one-time legacy cleanup)

def _norm_email(e) -> str:
    return (str(e or "").strip().lower())

def _norm_phone(p) -> str:
    digits = re.sub(r"[^\d]", "", str(p or ""))
    return digits[-10:] if len(digits) >= 10 else digits

def _is_blank(v) -> bool:
    """Treat None / '' / 'NULL' / 'N/A' as missing (case-insensitive)."""
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s.upper() in ("NULL", "N/A", "NONE")


async def _classify_college_type(college_str: str) -> str:
    """Return 'NIRF - #N' or 'Non NIRF' for a college name; '' if no college given."""
    college = (college_str or "").strip()
    if not college:
        return ""
    rank_lookup = await _build_college_rank_lookup_fn()
    cc = _classify_college_fn(
        {"ug_university": college, "pg_university": "", "college": college},
        rank_lookup,
    )
    status = cc.get("college_status") or ""
    return status if status.startswith("NIRF - #") else ("Non NIRF" if college else "")


async def _resolve_candidate_extras(email: str, phone: str) -> dict:
    """Look up `college_type`, `source`, and `college` for a candidate from the
    other authoritative sources (bb_registrations -> naukri_applies). Read-only,
    used as a runtime fallback. Returns only fields we could resolve.

    Priority order per the master matching rule:
        1. bb_registrations (latest registered_at)
        2. naukri_applies   (matched profile)
    Email is the primary identifier, phone is secondary. Empty strings are
    treated as missing.
    """
    out: dict = {}
    em = _norm_email(email)
    ph = _norm_phone(phone)
    if not em and not ph:
        return out
    or_clauses = []
    if em:
        or_clauses.append({"email": em})
    if ph:
        or_clauses.append({"phone": ph})
    match = {"$or": or_clauses}

    # ---- bb_registrations ----
    try:
        reg = await _db.bb_registrations.find(match, {"_id": 0}).sort("registered_at", -1).limit(1).to_list(1)
        reg = reg[0] if reg else None
    except Exception:
        reg = None
    if reg:
        # college (name) → derive college_type via NIRF lookup
        if "college_type" not in out:
            ct = await _classify_college_type(reg.get("college") or "")
            if ct:
                out["college_type"] = ct
        if "source" not in out:
            # bb_registrations is the public registration form pipeline
            form_name = (reg.get("form_name") or "").lower()
            if "college" in form_name:
                out["source"] = "college_form"
            else:
                out["source"] = "registration_form"
        if "college" not in out and not _is_blank(reg.get("college")):
            out["college"] = reg.get("college")

    # ---- naukri_applies ----
    if "college_type" not in out or "source" not in out or "college" not in out:
        try:
            nk = await _db.naukri_applies.find_one(match, {"_id": 0})
        except Exception:
            nk = None
        if nk:
            if "college_type" not in out:
                pg = (nk.get("pg_university") or "").strip()
                ug = (nk.get("ug_university") or "").strip()
                rank_lookup = await _build_college_rank_lookup_fn()
                cc = _classify_college_fn(
                    {"ug_university": ug, "pg_university": pg,
                     "college": pg or ug},
                    rank_lookup,
                )
                status = cc.get("college_status") or ""
                # iter110 — Emit the full bucketed status (not the binary alias).
                ct = status if status else ("Non-NIRF - No Rank" if (ug or pg) else "")
                if ct:
                    out["college_type"] = ct
            if "source" not in out:
                # Naukri ATS uploads. Preserve original sub-source if present.
                raw_src = (nk.get("source") or "").strip()
                out["source"] = f"naukri:{raw_src}" if raw_src else "naukri"
            if "college" not in out:
                pg = (nk.get("pg_university") or "").strip()
                ug = (nk.get("ug_university") or "").strip()
                col = pg or ug
                if col:
                    out["college"] = col

    return out


async def _detect_match_conflict(email: str, phone: str) -> Optional[str]:
    """If the same phone is associated with a *different* email (or vice versa)
    across pipeline_data / bb_registrations / naukri_applies, return a short
    reason string. Used by the backfill script to safely skip ambiguous rows.
    """
    em = _norm_email(email)
    ph = _norm_phone(phone)
    if not em or not ph:
        return None
    for col in ("pipeline_data", "bb_registrations", "naukri_applies"):
        try:
            doc = await _db[col].find_one({"phone": ph}, {"_id": 0, "email": 1, "phone": 1})
        except Exception:
            doc = None
        if doc:
            other_em = _norm_email(doc.get("email"))
            if other_em and other_em != em:
                return f"phone {ph} matches different email {other_em} in {col}"
    return None


# ============ EXACT SCORE MAPPING (Round-aware) ============
# Master rule: Email primary + Phone secondary. Round names are treated as
# the round IDENTITY — we don't force a fixed Round 1/2/HR/Final taxonomy.
# Friendly aliases (Technical 1 → Round 1, etc.) are normalised when present.

# Friendly alias map per the Feb 2026 spec. Keys are case-insensitive after
# whitespace collapse. Values are the canonical bucket the score will surface
# under in `round_wise_scores`. Anything not in this map keeps its original
# (whitespace-collapsed) round name — e.g. "BP" stays "BP".
ROUND_NAME_ALIASES = {
    "technical 1": "Round 1",
    "technical1": "Round 1",
    "tech 1": "Round 1",
    "round 1": "Round 1",
    "round1": "Round 1",
    "technical 2": "Round 2",
    "technical2": "Round 2",
    "tech 2": "Round 2",
    "round 2": "Round 2",
    "round2": "Round 2",
    "hr interview": "HR Round",
    "hr round": "HR Round",
    "hr": "HR Round",
    "final discussion": "Final Round",
    "final round": "Final Round",
    "final": "Final Round",
    # Common spacing variants seen in live score_sheet
    "accounts1": "Accounts 1",
    "accounts2": "Accounts 2",
    "mensa org": "Mensa Org",
    "mensaorg": "Mensa Org",
}


def _norm_round(name: Optional[str]) -> str:
    """Whitespace-collapse + alias-map a round name. Empty input → ''."""
    if not name:
        return ""
    s = re.sub(r"\s+", " ", str(name).strip())
    return ROUND_NAME_ALIASES.get(s.lower(), s)


def _score_record_ts(rec: dict) -> str:
    """Return a sortable timestamp for a score_sheet record (ISO string)."""
    return str(rec.get("created_at") or rec.get("updated_at") or "")


async def _build_round_wise_scores(
    email: str,
    phone: str,
    pick: str = "latest",
) -> dict:
    """Resolve all score records for a candidate (email primary, phone fallback)
    and group them by canonical round name.

    `pick`:
        - "latest"  → pick the entry with the most recent created_at per round
        - "highest" → pick the highest score per round
        - "lowest"  → pick the lowest score per round
    Returns:
        {
          "round_wise_scores": { canonical_round: {score, raw_round, created_at} },
          "latest_round": <round name with the most recent created_at across all>,
          "latest_score": <its numeric score>,
          "total_score":  <sum of selected scores per round>,
          "rounds":       <ordered list of canonical round names>,
        }
    """
    em = _norm_email(email)
    ph = _norm_phone(phone)
    if not em and not ph:
        return {"round_wise_scores": {}, "latest_round": None,
                "latest_score": None, "total_score": 0, "rounds": []}

    or_clauses = []
    if em:
        or_clauses.append({"email": em})
    if ph:
        or_clauses.append({"phone": ph})
    cursor = _db.score_sheet.find(
        {"$or": or_clauses} if len(or_clauses) > 1 else or_clauses[0],
        {"_id": 0},
    )
    records = await cursor.to_list(None) or []
    if not records:
        return {"round_wise_scores": {}, "latest_round": None,
                "latest_score": None, "total_score": 0, "rounds": []}

    # Group by canonical round name
    grouped: dict = {}
    for rec in records:
        canon = _norm_round(rec.get("round_name"))
        if not canon:
            continue
        grouped.setdefault(canon, []).append(rec)

    selected = {}
    for canon, recs in grouped.items():
        if pick == "highest":
            best = max(recs, key=lambda r: float(r.get("score") or 0))
        elif pick == "lowest":
            best = min(recs, key=lambda r: float(r.get("score") or 0))
        else:  # latest (default)
            best = max(recs, key=_score_record_ts)
        try:
            score_val = float(best.get("score") or 0)
        except (TypeError, ValueError):
            score_val = 0.0
        selected[canon] = {
            "score": score_val,
            "raw_round": (best.get("round_name") or canon),
            "created_at": best.get("created_at"),
        }

    # Latest round across the whole candidate
    latest_round = None
    latest_score = None
    if selected:
        latest_round = max(selected, key=lambda k: str(selected[k].get("created_at") or ""))
        latest_score = selected[latest_round]["score"]
    total_score = sum((s["score"] or 0) for s in selected.values())

    return {
        "round_wise_scores": selected,
        "latest_round": latest_round,
        "latest_score": latest_score,
        "total_score": total_score,
        "rounds": sorted(selected.keys()),
    }


async def _detect_score_phone_conflict(email: str, phone: str) -> Optional[str]:
    """For score uploads — return a conflict reason if `phone` is already
    associated with a *different* email in score_sheet."""
    em = _norm_email(email)
    ph = _norm_phone(phone)
    if not em or not ph:
        return None
    doc = await _db.score_sheet.find_one(
        {"phone": ph, "email": {"$nin": [None, "", em]}},
        {"_id": 0, "email": 1},
    )
    if doc:
        other = _norm_email(doc.get("email"))
        # Allow comma-joined dup emails (e.g. "x@y.com, x@y.com")
        parts = {p.strip() for p in (other or "").split(",") if p.strip()}
        if other and other != em and em not in parts:
            return f"phone {ph} ↔ different email '{other}' in score_sheet"
    return None




def _oid(id_str: str) -> ObjectId:
    try:
        return ObjectId(id_str)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid ID format")


def _slugify(text: str) -> str:
    """Convert a form name to a URL-friendly slug.
    Lowercase, alphanumerics + hyphens only, collapse repeats, trim hyphens.
    """
    if not text:
        return ""
    s = text.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s


async def _unique_slug(base: str, exclude_id: Optional[ObjectId] = None) -> str:
    """Return a unique slug, appending -1, -2, ... on collision."""
    if not base:
        base = "form"
    candidate = base
    n = 0
    while True:
        q = {"slug": candidate}
        if exclude_id is not None:
            q["_id"] = {"$ne": exclude_id}
        existing = await _db.bb_hiring_forms.find_one(q, {"_id": 1})
        if not existing:
            return candidate
        n += 1
        candidate = f"{base}-{n}"


async def _resolve_form_by_slug_or_id(key: str):
    """Try slug first, fallback to ObjectId. Returns the form doc or None."""
    if not key:
        return None
    # Try slug first
    form = await _db.bb_hiring_forms.find_one({"slug": key})
    if form:
        return form
    # Fallback: ObjectId
    try:
        return await _db.bb_hiring_forms.find_one({"_id": ObjectId(key)})
    except Exception:
        return None


def _classify_reason(rejected_reasons: list) -> str:
    """Map raw rejection-reason strings to a single primary category code.
    Priority: AGE > GRADUATION_YEAR > LOCATION > GENERAL.
    """
    if not rejected_reasons:
        return ""
    joined = " | ".join(r.lower() for r in rejected_reasons)
    if "age" in joined:
        return "AGE"
    if "graduation" in joined or "grad year" in joined:
        return "GRADUATION_YEAR"
    if "location" in joined or "attend in person" in joined:
        return "LOCATION"
    return "GENERAL"


_REASON_UI_MESSAGE = {
    "AGE": "Thank you for your interest. Unfortunately, your profile does not meet our current eligibility criteria.",
    "GRADUATION_YEAR": "We are currently looking for candidates from the {grad_min} to {grad_max} batch only.",
    "LOCATION": "Currently, we are proceeding only with candidates who are willing to attend in-person interviews in Chennai.",
    "GENERAL": "Thank you for applying. We will get back to you if your profile matches future requirements.",
    "": "Thank you for completing your registration. We will review your profile and get in touch shortly.",
}


def _build_ui_message(reason: str, grad_min=None, grad_max=None) -> str:
    msg = _REASON_UI_MESSAGE.get(reason, _REASON_UI_MESSAGE[""])
    if reason == "GRADUATION_YEAR":
        msg = msg.format(grad_min=grad_min or "the recent", grad_max=grad_max or "current")
    return msg


def _build_sort(sort_by, sort_dir, allowed: dict, default: dict) -> dict:
    """Whitelist-based sort spec builder. Returns Mongo $sort dict.
    `allowed` maps API field → DB field. Falls back to `default` on unknown input.
    """
    if not sort_by:
        return default
    db_field = allowed.get(sort_by)
    if not db_field:
        return default
    direction = -1 if (sort_dir or "").lower() == "desc" else 1
    return {db_field: direction}


# ============ JOB ROLES ============

class JobRoleBody(BaseModel):
    name: str

@bb_router.get("/job-roles")
async def list_job_roles(request: Request):
    """List job roles for dropdowns / management UI.

    iter99 — Read-time canonicalization:
      1. Every canonical `job_role` from `job_keyword_mapping` is included
         (these are the targets recruiters mapped raw keywords to).
      2. Every `bb_job_roles` row whose normalized name is NOT a mapped
         keyword and NOT already a canonical target is included (truly
         independent manual creates).
      3. Case-insensitive dedupe by normalized name — the canonical row
         wins so the displayed casing matches the mapping target.
    The returned shape is unchanged: {roles: [{id, name, source?, ...}]}.
    """
    await _require_auth(request)

    # Build the canonical-keyword index via server.py helper to keep one
    # source of truth.
    from server import _build_canonical_index as _bci, _normalize_text_for_matching as _nrm
    kw_to_canonical, canonical_set = await _bci()

    # All bb_job_roles rows keyed by normalized name (preserves _id for the
    # frontend's edit/delete actions).
    bb_rows: dict = {}
    async for r in _db.bb_job_roles.find({}).sort("name", 1):
        nm = (r.get("name") or "").strip()
        if not nm:
            continue
        bb_rows.setdefault(_nrm(nm), r)

    out: list = []
    seen_norms = set()

    # 1) Canonical mapping targets first.
    for canonical in {kw_to_canonical[k] for k in kw_to_canonical}:
        nrm = _nrm(canonical)
        if nrm in seen_norms:
            continue
        seen_norms.add(nrm)
        # Prefer existing bb_job_roles _id so edit/delete keeps working.
        if nrm in bb_rows:
            r = bb_rows[nrm]
            r["id"] = str(r.pop("_id"))
            r["name"] = canonical  # force canonical casing
            out.append(r)
        else:
            out.append({"id": f"canonical:{nrm}", "name": canonical, "source": "canonical"})

    # 2) bb_job_roles rows that aren't mapped keywords and not already
    #    surfaced as canonical → truly independent manual creates.
    for nrm, r in bb_rows.items():
        if nrm in seen_norms or nrm in kw_to_canonical:
            continue
        seen_norms.add(nrm)
        r["id"] = str(r.pop("_id"))
        out.append(r)

    out.sort(key=lambda x: (x.get("name") or "").lower())
    return {"roles": out}

@bb_router.post("/job-roles")
async def create_job_role(data: JobRoleBody, request: Request):
    await _require_auth(request)
    name = data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    # iter69f (#10C) — Reject case-insensitive duplicates so the Job Roles
    # page can never grow two cards for the same title.
    dup = await _db.bb_job_roles.find_one(
        {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}
    )
    if dup:
        raise HTTPException(status_code=409, detail="Job role already exists")
    doc = {"name": name, "source": "manual",
           "created_at": datetime.now(timezone.utc).isoformat()}
    result = await _db.bb_job_roles.insert_one(doc)
    # iter69f (#10C) — Mirror manual additions into job_titles_master so the
    # mapping picker offers them alongside imported titles.
    try:
        from server import _normalize_text_for_matching as _nrm
        normalized = _nrm(name)
        if normalized:
            existing_jtm = await _db.job_titles_master.find_one(
                {"normalized_job_title": normalized}
            )
            if not existing_jtm:
                await _db.job_titles_master.insert_one({
                    "raw_job_title": name,
                    "normalized_job_title": normalized,
                    "is_mapped": False,
                    "source": "manual",
                })
    except Exception as _e:
        _logger.warning(f"[job-roles] job_titles_master mirror skipped: {_e}")
    return {"success": True, "id": str(result.inserted_id), "name": name}

@bb_router.put("/job-roles/{role_id}")
async def update_job_role(role_id: str, data: JobRoleBody, request: Request):
    await _require_auth(request)
    result = await _db.bb_job_roles.update_one({"_id": _oid(role_id)}, {"$set": {"name": data.name.strip()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}

@bb_router.delete("/job-roles/{role_id}")
async def delete_job_role(role_id: str, request: Request):
    await _require_auth(request)
    result = await _db.bb_job_roles.delete_one({"_id": _oid(role_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}


# ============ FORM TYPES ============

class FormTypeBody(BaseModel):
    name: str

@bb_router.get("/form-types")
async def list_form_types(request: Request):
    await _require_auth(request)
    types = await _db.bb_form_types.find({}).sort("name", 1).to_list(None)
    for t in types:
        t["id"] = str(t.pop("_id"))
    return {"form_types": types}

@bb_router.post("/form-types")
async def create_form_type(data: FormTypeBody, request: Request):
    await _require_auth(request)
    name = data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    doc = {"name": name, "created_at": datetime.now(timezone.utc).isoformat()}
    result = await _db.bb_form_types.insert_one(doc)
    return {"success": True, "id": str(result.inserted_id), "name": name}

@bb_router.put("/form-types/{type_id}")
async def update_form_type(type_id: str, data: FormTypeBody, request: Request):
    await _require_auth(request)
    result = await _db.bb_form_types.update_one({"_id": _oid(type_id)}, {"$set": {"name": data.name.strip()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}

@bb_router.delete("/form-types/{type_id}")
async def delete_form_type(type_id: str, request: Request):
    await _require_auth(request)
    result = await _db.bb_form_types.delete_one({"_id": _oid(type_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}


# ============ HIRING FORMS ============

class ConditionsBody(BaseModel):
    age_min: Optional[int] = None
    age_max: Optional[int] = None
    grad_year_min: Optional[int] = None
    grad_year_max: Optional[int] = None
    locations: Optional[List[str]] = None
    location_change: Optional[str] = "NA"
    attend_in_person: Optional[str] = "NA"
    college_limit: Optional[str] = "Both"

class HiringFormCreate(BaseModel):
    name: str
    form_type_id: str
    job_role: str
    conditions: Optional[ConditionsBody] = None
    job_description_attached: Optional[bool] = False
    job_opening_id: Optional[str] = None
    show_instruction_page: Optional[bool] = False
    instruction_content: Optional[str] = ""

class HiringFormUpdate(BaseModel):
    name: Optional[str] = None
    form_type_id: Optional[str] = None
    job_role: Optional[str] = None
    conditions: Optional[ConditionsBody] = None
    job_description_attached: Optional[bool] = None
    job_opening_id: Optional[str] = None
    show_instruction_page: Optional[bool] = None
    instruction_content: Optional[str] = None

@bb_router.get("/hiring-forms")
async def list_hiring_forms(request: Request):
    await _require_auth(request)
    forms = await _db.bb_hiring_forms.find({}).sort("created_at", -1).to_list(None)
    for f in forms:
        f["id"] = str(f.pop("_id"))
        # Backfill slug on-the-fly for legacy docs missing it
        if not f.get("slug"):
            base = _slugify(f.get("name") or f.get("job_role") or "form")
            slug = await _unique_slug(base)
            await _db.bb_hiring_forms.update_one({"_id": _oid(f["id"])}, {"$set": {"slug": slug}})
            f["slug"] = slug
    return {"forms": forms}

@bb_router.post("/hiring-forms")
async def create_hiring_form(data: HiringFormCreate, request: Request):
    await _require_auth(request)
    ft = await _db.bb_form_types.find_one({"_id": _oid(data.form_type_id)})
    if not ft:
        raise HTTPException(status_code=400, detail="Form type not found")
    cond = data.conditions.dict() if data.conditions else {}
    if cond.get("locations"):
        cond["locations"] = [l.strip() for l in cond["locations"] if l.strip()]
    name = data.name.strip()
    slug = await _unique_slug(_slugify(name))
    doc = {
        "name": name, "form_type_id": data.form_type_id,
        "form_type_name": ft["name"], "job_role": data.job_role.strip(),
        "slug": slug,
        "conditions": cond,
        "job_description_attached": data.job_description_attached or False,
        "job_opening_id": data.job_opening_id or None,
        "show_instruction_page": data.show_instruction_page or False,
        "instruction_content": (data.instruction_content or "").strip(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await _db.bb_hiring_forms.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None)
    return {"success": True, "form": doc}

@bb_router.put("/hiring-forms/{form_id}")
async def update_hiring_form(form_id: str, data: HiringFormUpdate, request: Request):
    await _require_auth(request)
    oid = _oid(form_id)
    updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if data.name is not None:
        new_name = data.name.strip()
        updates["name"] = new_name
        # Auto-regenerate slug on rename
        updates["slug"] = await _unique_slug(_slugify(new_name), exclude_id=oid)
    if data.job_role is not None:
        updates["job_role"] = data.job_role.strip()
    if data.form_type_id is not None:
        ft = await _db.bb_form_types.find_one({"_id": _oid(data.form_type_id)})
        if not ft:
            raise HTTPException(status_code=400, detail="Form type not found")
        updates["form_type_id"] = data.form_type_id
        updates["form_type_name"] = ft["name"]
    if data.conditions is not None:
        cond = data.conditions.dict()
        if cond.get("locations"):
            cond["locations"] = [l.strip() for l in cond["locations"] if l.strip()]
        updates["conditions"] = cond
    if data.job_description_attached is not None:
        updates["job_description_attached"] = data.job_description_attached
    if data.job_opening_id is not None:
        updates["job_opening_id"] = data.job_opening_id
    if data.show_instruction_page is not None:
        updates["show_instruction_page"] = data.show_instruction_page
    if data.instruction_content is not None:
        updates["instruction_content"] = data.instruction_content.strip()
    result = await _db.bb_hiring_forms.update_one({"_id": oid}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}

@bb_router.delete("/hiring-forms/{form_id}")
async def delete_hiring_form(form_id: str, request: Request):
    await _require_auth(request)
    result = await _db.bb_hiring_forms.delete_one({"_id": _oid(form_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}


# ============ ROUNDS ============

class RoundBody(BaseModel):
    name: str
    order: Optional[int] = None


@bb_router.get("/rounds")
async def list_rounds(request: Request, includeInactive: bool = Query(False)):
    """List rounds. Active-only by default; pass includeInactive=true to see all.

    Iter51 — case-insensitive + whitespace-collapsed dedupe at render time.
    Should never trigger for a clean dataset (the create/import paths already
    enforce uniqueness), but acts as a safety net so the UI never shows
    "Accounts1" + "Accounts 1" as separate tabs even if legacy bad data exists.
    """
    await _require_auth(request)
    q = {} if includeInactive else {"active": {"$ne": False}}
    rounds = await _db.bb_rounds.find(q).sort([("order", 1), ("name", 1)]).to_list(None)
    deduped = []
    seen_canon = set()
    for r in rounds:
        r["id"] = str(r.pop("_id"))
        canon = _norm_round(r.get("name") or "").lower()
        if canon and canon in seen_canon:
            continue
        seen_canon.add(canon)
        if "active" not in r:
            r["active"] = True
        if "order" not in r:
            r["order"] = 0
        deduped.append(r)
    return {"rounds": deduped}


async def _round_in_use(name: str) -> bool:
    """Return True if any score record references this round name."""
    if not name:
        return False
    # bb_applicant_updates.scores[].round_name OR score_sheet.round_name
    if await _db.score_sheet.find_one({"round_name": name}, {"_id": 1}):
        return True
    if await _db.bb_applicant_updates.find_one({"scores.round_name": name}, {"_id": 1}):
        return True
    return False


@bb_router.post("/rounds")
async def create_round(data: RoundBody, request: Request):
    """Create a new round. Round name must be unique (case-insensitive)."""
    await _require_auth(request)
    name = (data.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    # Uniqueness check (case-insensitive) across active+inactive
    existing = await _db.bb_rounds.find_one(
        {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}},
        {"_id": 1, "active": 1, "name": 1},
    )
    if existing:
        if existing.get("active") is False:
            raise HTTPException(status_code=409, detail=f"A round named '{existing['name']}' already exists but is inactive. Restore it instead.")
        raise HTTPException(status_code=409, detail="Round name already exists")
    doc = {
        "name": name,
        "active": True,
        "order": int(data.order) if data.order is not None else 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await _db.bb_rounds.insert_one(doc)
    return {"success": True, "id": str(result.inserted_id), "name": name, "active": True, "order": doc["order"]}


@bb_router.put("/rounds/{round_id}")
async def update_round(round_id: str, data: RoundBody, request: Request):
    """Update a round's name and/or order. Renames cascade to all referenced
    score records (score_sheet + bb_applicant_updates.scores[]) so historical
    data stays linked. Name must remain unique."""
    await _require_auth(request)
    oid = _oid(round_id)
    current = await _db.bb_rounds.find_one({"_id": oid})
    if not current:
        raise HTTPException(status_code=404, detail="Not found")
    new_name = (data.name or "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="Name required")
    # Uniqueness (excluding self)
    dup = await _db.bb_rounds.find_one({
        "_id": {"$ne": oid},
        "name": {"$regex": f"^{re.escape(new_name)}$", "$options": "i"},
    }, {"_id": 1})
    if dup:
        raise HTTPException(status_code=409, detail="Round name already exists")
    updates = {"name": new_name, "updated_at": datetime.now(timezone.utc).isoformat()}
    if data.order is not None:
        updates["order"] = int(data.order)
    await _db.bb_rounds.update_one({"_id": oid}, {"$set": updates})

    # Cascade rename to historical score records — keeps mappings intact
    old_name = current.get("name")
    if old_name and old_name != new_name:
        await _db.score_sheet.update_many(
            {"round_name": old_name}, {"$set": {"round_name": new_name}},
        )
        await _db.bb_applicant_updates.update_many(
            {"scores.round_name": old_name},
            {"$set": {"scores.$[el].round_name": new_name}},
            array_filters=[{"el.round_name": old_name}],
        )
    return {"success": True}


@bb_router.delete("/rounds/{round_id}")
async def delete_round(round_id: str, request: Request, hard: bool = Query(False)):
    """Logical delete by default — sets `active=false` so historical scores
    remain queryable but the round is hidden from normal dropdowns. Hard delete
    is BLOCKED if any score record references this round (data-safety guarantee).
    """
    await _require_auth(request)
    oid = _oid(round_id)
    current = await _db.bb_rounds.find_one({"_id": oid})
    if not current:
        raise HTTPException(status_code=404, detail="Not found")
    name = current.get("name")
    in_use = await _round_in_use(name)

    if hard:
        if in_use:
            raise HTTPException(
                status_code=409,
                detail="Cannot hard-delete: round is referenced by existing applicant scores. Use logical delete instead.",
            )
        await _db.bb_rounds.delete_one({"_id": oid})
        return {"success": True, "deleted": "hard"}

    # Logical delete (default)
    await _db.bb_rounds.update_one(
        {"_id": oid},
        {"$set": {"active": False, "deleted_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"success": True, "deleted": "logical", "in_use": in_use}


@bb_router.post("/rounds/{round_id}/restore")
async def restore_round(round_id: str, request: Request):
    """Re-enable a logically-deleted round."""
    await _require_auth(request)
    oid = _oid(round_id)
    current = await _db.bb_rounds.find_one({"_id": oid})
    if not current:
        raise HTTPException(status_code=404, detail="Not found")
    # Guard against name collision with another active round
    name = current.get("name") or ""
    dup = await _db.bb_rounds.find_one({
        "_id": {"$ne": oid},
        "active": {"$ne": False},
        "name": {"$regex": f"^{re.escape(name)}$", "$options": "i"},
    }, {"_id": 1})
    if dup:
        raise HTTPException(status_code=409, detail="Cannot restore: another active round with this name exists. Rename one of them first.")
    await _db.bb_rounds.update_one(
        {"_id": oid},
        {"$set": {"active": True, "restored_at": datetime.now(timezone.utc).isoformat()},
         "$unset": {"deleted_at": ""}},
    )
    return {"success": True}


# ============ COLLEGE SCHEDULES (HR-Configured Drives) ============

class CollegeScheduleBody(BaseModel):
    college_name: str
    job_role: Optional[str] = None     # legacy — joined string accepted
    job_roles: Optional[List[str]] = None  # iter56 — preferred structured array
    schedule_date: str   # ISO YYYY-MM-DD
    schedule_time: str   # 24h HH:MM:SS or HH:MM
    notes: Optional[str] = ""


class CollegeScheduleUpdate(BaseModel):
    college_name: Optional[str] = None
    job_role: Optional[str] = None
    job_roles: Optional[List[str]] = None
    schedule_date: Optional[str] = None
    schedule_time: Optional[str] = None
    notes: Optional[str] = None


def _normalize_roles(job_role: Optional[str], job_roles: Optional[List[str]]) -> List[str]:
    """Iter56 — Coerce either input form into a clean, deduped, ordered list.
    Accepts:  ['AI/ML','HR']  OR  'AI/ML, HR, AI/ML' (legacy comma string).
    """
    raw = []
    if job_roles:
        raw = list(job_roles)
    elif job_role:
        raw = job_role.split(",")
    out: List[str] = []
    seen = set()
    for r in raw:
        s = (r or "").strip()
        if not s:
            continue
        k = s.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(s)
    return out


@bb_router.get("/college-schedules")
async def list_college_schedules(request: Request, includeInactive: bool = Query(False), college: Optional[str] = Query(None)):
    """List HR-configured college schedules. Active-only by default."""
    await _require_auth(request)
    q = {} if includeInactive else {"active": {"$ne": False}}
    if college:
        q["college_name"] = {"$regex": f"^{re.escape(college.strip())}$", "$options": "i"}
    docs = await _db.bb_college_schedules.find(q).sort([("college_name", 1), ("schedule_date", 1)]).to_list(None)
    for d in docs:
        d["id"] = str(d.pop("_id"))
        if "active" not in d:
            d["active"] = True
        # Iter56 — always expose `job_roles` array. Backfill from legacy joined `job_role` string.
        if "job_roles" not in d or not isinstance(d.get("job_roles"), list):
            d["job_roles"] = _normalize_roles(d.get("job_role"), None)
    return {"schedules": docs}


@bb_router.post("/college-schedules")
async def create_college_schedule(data: CollegeScheduleBody, request: Request):
    """Create a college+roles+date+time mapping. Iter56 — Job Role accepted as
    array (preferred) or legacy comma string. Stored as both `job_roles` array
    and `job_role` joined string for backward compat with downstream readers
    (register endpoint regex matches on `job_role`)."""
    await _require_auth(request)
    college = (data.college_name or "").strip()
    roles = _normalize_roles(data.job_role, data.job_roles)
    date = (data.schedule_date or "").strip()
    time = (data.schedule_time or "").strip()
    if not (college and roles and date and time):
        raise HTTPException(status_code=400, detail="college_name, job_roles, schedule_date and schedule_time are required")
    doc = {
        "college_name": college,
        "job_roles": roles,
        "job_role": ",".join(roles),  # legacy compat — used by register endpoint regex matcher
        "schedule_date": date,
        # iter83 — Strict 24-hour normalization. Replaces brittle "len==3 else +':00'".
        "schedule_time": to_24h_db(time),
        "notes": (data.notes or "").strip(),
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    res = await _db.bb_college_schedules.insert_one(doc)
    doc["id"] = str(res.inserted_id); doc.pop("_id", None)
    return {"success": True, "schedule": doc}


@bb_router.put("/college-schedules/{sched_id}")
async def update_college_schedule(sched_id: str, data: CollegeScheduleUpdate, request: Request):
    await _require_auth(request)
    oid = _oid(sched_id)
    cur = await _db.bb_college_schedules.find_one({"_id": oid})
    if not cur:
        raise HTTPException(status_code=404, detail="Not found")
    updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if data.college_name is not None:
        updates["college_name"] = data.college_name.strip()
    # Iter56 — when either form of role input is present, normalize and write both fields
    if data.job_role is not None or data.job_roles is not None:
        roles = _normalize_roles(data.job_role, data.job_roles)
        if not roles:
            raise HTTPException(status_code=400, detail="At least one job role is required")
        updates["job_roles"] = roles
        updates["job_role"] = ",".join(roles)
    if data.schedule_date is not None:
        updates["schedule_date"] = data.schedule_date.strip()
    if data.schedule_time is not None:
        # iter83 — Strict 24-hour normalization on update too.
        try:
            updates["schedule_time"] = to_24h_db(data.schedule_time)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"Invalid schedule_time: {e}")
    if data.notes is not None:
        updates["notes"] = data.notes.strip()
    await _db.bb_college_schedules.update_one({"_id": oid}, {"$set": updates})
    return {"success": True}


@bb_router.delete("/college-schedules/{sched_id}")
async def delete_college_schedule(sched_id: str, request: Request, hard: bool = Query(False)):
    """Logical delete by default (sets active=false)."""
    await _require_auth(request)
    oid = _oid(sched_id)
    cur = await _db.bb_college_schedules.find_one({"_id": oid})
    if not cur:
        raise HTTPException(status_code=404, detail="Not found")
    if hard:
        await _db.bb_college_schedules.delete_one({"_id": oid})
        return {"success": True, "deleted": "hard"}
    await _db.bb_college_schedules.update_one(
        {"_id": oid},
        {"$set": {"active": False, "deleted_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"success": True, "deleted": "logical"}


@bb_router.post("/college-schedules/{sched_id}/restore")
async def restore_college_schedule(sched_id: str, request: Request):
    await _require_auth(request)
    oid = _oid(sched_id)
    cur = await _db.bb_college_schedules.find_one({"_id": oid})
    if not cur:
        raise HTTPException(status_code=404, detail="Not found")
    # Block restore if collision with another active entry
    dup = await _db.bb_college_schedules.find_one({
        "_id": {"$ne": oid}, "active": {"$ne": False},
        "college_name": {"$regex": f"^{re.escape(cur.get('college_name',''))}$", "$options": "i"},
        "job_role": {"$regex": f"^{re.escape(cur.get('job_role',''))}$", "$options": "i"},
    }, {"_id": 1})
    if dup:
        raise HTTPException(status_code=409, detail="Cannot restore: another active schedule with this college+role exists.")
    await _db.bb_college_schedules.update_one(
        {"_id": oid},
        {"$set": {"active": True, "restored_at": datetime.now(timezone.utc).isoformat()},
         "$unset": {"deleted_at": ""}},
    )
    return {"success": True}


# ---- Public (candidate-side) endpoints — no auth required ----

@pub_router.get("/college-form/colleges")
async def pub_list_colleges():
    """Distinct active colleges with at least one configured schedule."""
    cols = await _db.bb_college_schedules.distinct("college_name", {"active": {"$ne": False}})
    return {"colleges": sorted([c for c in cols if c])}


@pub_router.get("/college-form/roles")
async def pub_list_roles_for_college(college: str = Query(...)):
    """Active roles available for the given college."""
    roles = await _db.bb_college_schedules.distinct(
        "job_role",
        {"active": {"$ne": False}, "college_name": {"$regex": f"^{re.escape(college.strip())}$", "$options": "i"}},
    )
    return {"roles": sorted([r for r in roles if r])}


@pub_router.get("/college-form/schedule")
async def pub_latest_schedule_for_college(college: str = Query(...)):
    """Iter54 — Public-form helper for College Drives Req 2.
    Returns the LATEST ACTIVE schedule (by created_at, then schedule_date) for the
    given college so the public registration form can dynamically populate the
    Schedule Date, Schedule Time and Job Role fields on college selection.
    Iter56 — adds `job_roles` array (preferred) alongside legacy `job_role` string.
    Returns {"schedule": null} if nothing matches — caller must clear its fields.
    """
    name = (college or "").strip()
    if not name:
        return {"schedule": None}
    doc = await _db.bb_college_schedules.find_one(
        {
            "active": {"$ne": False},
            "college_name": {"$regex": f"^{re.escape(name)}$", "$options": "i"},
        },
        {"_id": 0, "college_name": 1, "job_role": 1, "job_roles": 1,
         "schedule_date": 1, "schedule_time": 1, "created_at": 1},
        sort=[("created_at", -1), ("schedule_date", -1)],
    )
    if not doc:
        return {"schedule": None}
    roles = doc.get("job_roles")
    if not isinstance(roles, list) or not roles:
        roles = _normalize_roles(doc.get("job_role"), None)
    return {"schedule": {
        "college_name": doc.get("college_name", ""),
        "job_role": doc.get("job_role", "") or ",".join(roles),
        "job_roles": roles,
        "schedule_date": doc.get("schedule_date", "") or "",
        "schedule_time": doc.get("schedule_time", "") or "",
    }}


class CollegeRegistrationBody(BaseModel):
    full_name: str
    email: str
    phone: str
    age: Optional[int] = None
    gender: Optional[str] = ""
    college: str          # college_name selected by candidate
    job_role: str         # role selected by candidate
    degree: Optional[str] = ""
    course: Optional[str] = ""
    year_of_graduation: Optional[int] = None
    current_location_state: Optional[str] = ""
    preferred_location_city: Optional[str] = ""

    @field_validator("phone")
    @classmethod
    def _phone_10(cls, v): return _validate_phone_10digits(v)

    @field_validator("email")
    @classmethod
    def _email_required(cls, v): return _validate_nonempty(v, "Email")

    @field_validator("full_name")
    @classmethod
    def _name_required(cls, v): return _validate_nonempty(v, "Full name")

    @field_validator("preferred_location_city")
    @classmethod
    def _city_required(cls, v): return _validate_nonempty(v, "Preferred location (city)")


@pub_router.post("/college-form/register")
async def register_college_applicant(data: CollegeRegistrationBody):
    """Public college-form registration. Auto-attaches HR-configured schedule
    based on (college, role). Stores into the existing pipeline_data table.
    422 if no schedule is configured for the combo.
    """
    college = (data.college or "").strip()
    role = (data.job_role or "").strip()
    if not (college and role):
        raise HTTPException(status_code=400, detail="College and job role are required")
    if not (data.full_name and data.email and data.phone):
        raise HTTPException(status_code=400, detail="Name, email and phone are required")

    # Iter60 — Multi-role support: a schedule's job_role may be stored as a
    # joined string ("AI/ML,Administration,HR") or as `job_roles[]` array.
    # Match if the candidate's selected role is any element of either form
    # (case-insensitive, whitespace-trimmed). Backward compatible with
    # single-role legacy rows.
    sched = await _db.bb_college_schedules.find_one({
        "active": {"$ne": False},
        "college_name": {"$regex": f"^{re.escape(college)}$", "$options": "i"},
        "$or": [
            # Exact match against array element (preferred)
            {"job_roles": {"$elemMatch": {"$regex": f"^{re.escape(role)}$", "$options": "i"}}},
            # Match against legacy joined string: role appears as full token
            # (anchored on commas/string boundaries to avoid "HR" matching "CHR")
            {"job_role": {"$regex": f"(^|,)\\s*{re.escape(role)}\\s*(,|$)", "$options": "i"}},
        ],
    })
    if not sched:
        raise HTTPException(
            status_code=422,
            detail=f"No interview schedule has been configured for {college} – {role}. Please contact HR.",
        )

    schedule_date = sched.get("schedule_date", "")
    schedule_time = sched.get("schedule_time", "")

    # Phone normalize (mirror existing register_applicant logic loosely)
    phone_normalized = "".join(c for c in (data.phone or "") if c.isdigit())[-10:]

    # iter110 — Five-bucket college status classification.
    rank_lookup = await _build_college_rank_lookup_fn()
    cc = _classify_college_fn({"ug_university": college, "pg_university": "", "college": college}, rank_lookup)
    college_type = cc["college_status"] or "Non-NIRF - No Rank"
    submitted_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    pipeline_doc_set = {
        "name": data.full_name.strip(),
        "email": data.email.strip().lower(),
        "phone": phone_normalized,
        "age": data.age,
        "gender": (data.gender or "").strip(),
        "college": college,
        "college_type": college_type,
        "degree": (data.degree or "").strip(),
        "course": (data.course or "").strip(),
        "location": (data.preferred_location_city or "").strip(),
        "job_role": role,
        "job_title": role,
        "email_type": "shortlist",  # college candidates are pre-shortlisted by HR
        "year_of_graduation": str(data.year_of_graduation) if data.year_of_graduation else "",
        "submitted_at": submitted_at,
        "schedule_date": schedule_date,
        "schedule_time": schedule_time,
        "otp_verified": "",
        "result_status": "",
        "source": "college_drive",
        # Persisted derived fields so HR pages immediately reflect this record
        "_college_status": cc["college_status"],
        # iter110 — `_nirf_category` mirrors `_college_status` (NIRF kept as
        # the premium-only binary alias; everything else uses the full label).
        "_nirf_category": "NIRF" if college_type.startswith("NIRF - #") else college_type,
        "_college_resolved": cc.get("college") or college,
        "_match_confidence": cc.get("match_confidence") or None,
        "_normalized_job_role": role,
    }
    # Upsert with NON-DESTRUCTIVE merge for existing candidates (see
    # register_applicant for full rationale). Dynamic fields (schedule_date,
    # schedule_time, job_role, email_type, submitted_at, last_update) always
    # update; profile fields are preserved if already populated.
    # Iter50 — also enforces:
    #   * source is filled if missing (never overwritten if already set)
    #   * stage:"registered" + initial pipeline timestamps on insert only
    #   * scores/status/progress fields are NEVER touched here
    #   * email↔phone conflicts are logged + skipped (registration still
    #     succeeds; spec says pipeline failure must not block the response)
    PROFILE_FIELDS = {
        "name", "email", "phone", "age", "gender", "college", "college_type",
        "degree", "course", "location", "year_of_graduation", "source",
        "_college_status", "_nirf_category", "_college_resolved", "_match_confidence",
        # Iter50 — pipeline progress fields are NEVER overwritten by a
        # college-drive re-registration (per spec: "Do NOT overwrite scores,
        # status, previous pipeline progress").
        "otp_verified", "result_status",
    }
    DYNAMIC_FIELDS = {
        "job_role", "job_title", "email_type", "submitted_at", "last_update",
        "_normalized_job_role", "schedule_date", "schedule_time",
    }
    try:
        target_email = data.email.strip().lower()
        # iter69 — Tester credentials full-overwrite path. When the registrant
        # matches an entry in bb_test_credentials (email OR phone), do a
        # `replace_one` so no stale fields survive from prior submissions.
        tc_doc = await _db.bb_test_credentials.find_one(
            {"$or": [{"email": target_email}, {"phone": phone_normalized}]}
        )
        is_tester = bool(tc_doc)

        existing = await _db.pipeline_data.find_one(
            {"$or": [{"email": target_email}, {"phone": phone_normalized}]},
            {"_id": 1, "email": 1, "phone": 1, **{f: 1 for f in PROFILE_FIELDS}},
        )
        if is_tester:
            # iter69b — TESTER CONSOLIDATION + FULL OVERWRITE (college flow).
            # iter69e — Also resets flow-state fields per spec #9.
            FLOW_STATE_FIELDS = (
                "otp", "otp_verified", "otp_sent", "otp_sent_at",
                "email_type", "result_status",
                "schedule_date", "schedule_time",
                "whatsapp_reminder_sent", "whatsapp_followup_sent",
                "shortlist_mail_sent", "interview_mail_sent",
                "reject_notified", "schedule_message_sent",
                "isImported", "import_batch_id", "imported_at",
                "import_rejection_notified",
            )
            replacement = dict(pipeline_doc_set)
            replacement["last_update"] = submitted_at
            replacement["updated_at"] = submitted_at
            replacement["created_at"] = submitted_at
            replacement["stage"] = "registered"
            replacement["pipeline_synced_at"] = submitted_at
            for k in FLOW_STATE_FIELDS:
                replacement[k] = ""
            all_matches = await _db.pipeline_data.find(
                {"$or": [{"email": target_email}, {"phone": phone_normalized}]},
                {"_id": 1},
            ).sort("_id", 1).to_list(length=50)
            if all_matches:
                survivor = all_matches[0]
                r = await _db.pipeline_data.replace_one(
                    {"_id": survivor["_id"]}, replacement
                )
                extra_ids = [m["_id"] for m in all_matches[1:]]
                deleted = 0
                if extra_ids:
                    d = await _db.pipeline_data.delete_many({"_id": {"$in": extra_ids}})
                    deleted = d.deleted_count
                # Reset flow flags on bb_registrations too
                try:
                    await _db.bb_registrations.update_many(
                        {"$or": [{"email": target_email}, {"phone": phone_normalized}]},
                        {"$unset": {
                            "shortlist_mail_sent": "",
                            "shortlist_mail_sent_at": "",
                            "interview_mail_sent": "",
                            "interview_mail_sent_at": "",
                            "reject_notified": "",
                            "reject_notified_at": "",
                            "schedule_message_sent": "",
                            "schedule_message_sent_at": "",
                            "schedule_message_wa_ok": "",
                            "schedule_message_em_ok": "",
                            "whatsapp_reminder_sent": "",
                            "whatsapp_followup_sent": "",
                        }},
                    )
                except Exception as _e:
                    _logger.warning(f"[Pipeline:tester] college bb_registrations reset skipped: {_e}")
                # iter70 (#6) — Clear applicant round scores so the freshly
                # re-registered candidate starts a clean evaluation cycle.
                # bb_rounds definitions are NOT touched; only the per-applicant
                # `scores[]` array is wiped + status reset.
                try:
                    upd_res = await _db.bb_applicant_updates.update_many(
                        {"$or": [{"email": target_email}, {"phone": phone_normalized}]},
                        {"$set": {
                            "scores": [],
                            "status": "",
                            "result_status": "",
                            "rejection_notified": False,
                            "import_rejection_notified": False,
                            # iter98 — Re-registration MUST clear rejection_sent
                            # too. Otherwise the 19:00 evening dispatcher
                            # filter `rejection_sent: {$ne: True}` skips this
                            # row forever for the new application cycle.
                            "rejection_sent": False,
                            "rejection_sent_at": "",
                            "rejection_send_ok": False,
                            "scores_reset_at": submitted_at,
                        }},
                    )
                    _logger.info(
                        f"[Pipeline:tester] college_drive SCORES_RESET "
                        f"matched={upd_res.matched_count} modified={upd_res.modified_count} "
                        f"email={target_email} phone={phone_normalized}"
                    )
                except Exception as _e:
                    _logger.warning(f"[Pipeline:tester] college bb_applicant_updates reset skipped: {_e}")
                _logger.info(
                    f"[Pipeline:tester] college_drive FULL_OVERWRITE+CONSOLIDATE+RESET "
                    f"survivor_id={survivor['_id']} matched={len(all_matches)} "
                    f"replaced_modified={r.modified_count} deleted_dups={deleted} "
                    f"email={target_email} phone={phone_normalized}"
                )
            else:
                ins = await _db.pipeline_data.insert_one(replacement)
                _logger.info(
                    f"[Pipeline:tester] college_drive INSERTED _id={ins.inserted_id} "
                    f"email={target_email} phone={phone_normalized}"
                )
        elif existing:
            ex_email = (existing.get("email") or "").strip().lower()
            ex_phone = (existing.get("phone") or "").strip()
            # Conflict detection: same phone but a different stored email → log + skip.
            # Per spec: "If both exist but mismatch → log conflict, skip auto update"
            if ex_email and ex_phone and ex_email != target_email and ex_phone == phone_normalized:
                _logger.warning(
                    f"[Pipeline] CONFLICT skip — phone {phone_normalized} already maps "
                    f"to email '{ex_email}', cannot also bind to '{target_email}'"
                )
            else:
                set_fields = {}
                for k, v in pipeline_doc_set.items():
                    if k in DYNAMIC_FIELDS:
                        set_fields[k] = v
                    elif k in PROFILE_FIELDS:
                        if _is_blank(existing.get(k)):
                            set_fields[k] = v
                    else:
                        set_fields[k] = v
                set_fields["last_update"] = submitted_at
                set_fields["updated_at"] = submitted_at
                await _db.pipeline_data.update_one(
                    {"_id": existing["_id"]},
                    {"$set": set_fields,
                     # Insert-only fields per spec
                     "$setOnInsert": {"created_at": submitted_at,
                                       "stage": "registered",
                                       "pipeline_synced_at": submitted_at}},
                )
                _logger.info(
                    f"[Pipeline] action=updated source=college_drive "
                    f"email={target_email} phone={phone_normalized} "
                    f"college={college} role={role}"
                )
        else:
            # No existing record — fresh insert with full payload + stage flag
            insert_doc = dict(pipeline_doc_set)
            insert_doc["last_update"] = submitted_at
            insert_doc["updated_at"] = submitted_at
            insert_doc["created_at"] = submitted_at
            insert_doc["stage"] = "registered"
            insert_doc["pipeline_synced_at"] = submitted_at
            await _db.pipeline_data.update_one(
                {"email": target_email},
                {"$set": insert_doc},
                upsert=True,
            )
            _logger.info(
                f"[Pipeline] action=created source=college_drive "
                f"email={target_email} phone={phone_normalized} "
                f"college={college} role={role}"
            )
    except Exception as e:
        # Per spec: pipeline failure must NOT block registration success
        _logger.error(f"[Pipeline] sync failed for email={data.email}: {e}", exc_info=True)

    _logger.info(f"[CollegeForm] registered email={data.email} college={college} role={role} → {schedule_date} {schedule_time}")
    return {
        "success": True,
        "message": f"Registration successful! Your interview is scheduled for {schedule_date} at {schedule_time}.",
        "schedule_date": schedule_date,
        "schedule_time": schedule_time,
        "college": college,
        "job_role": role,
    }


# ============ JOB OPENINGS ============

class _DescriptiveSection(BaseModel):
    title: str = ""
    description: str = ""


class JobOpeningCreate(BaseModel):
    title: str
    job_role: Optional[str] = ""
    vacancies: Optional[int] = None
    years_of_graduation: Optional[List[str]] = None
    education: Optional[List[str]] = None
    salary_range: Optional[str] = ""
    # iter108 — Dynamic descriptive sections (new schema). When supplied,
    # this is the source of truth and is persisted as-is. Legacy fields
    # below are still accepted for backward compatibility with older clients.
    descriptive_sections: Optional[List[_DescriptiveSection]] = None
    key_responsibilities: Optional[str] = ""
    added_advantages: Optional[str] = ""
    what_we_offer: Optional[str] = ""

class JobOpeningUpdate(BaseModel):
    title: Optional[str] = None
    job_role: Optional[str] = None
    vacancies: Optional[int] = None
    years_of_graduation: Optional[List[str]] = None
    education: Optional[List[str]] = None
    salary_range: Optional[str] = None
    descriptive_sections: Optional[List[_DescriptiveSection]] = None
    key_responsibilities: Optional[str] = None
    added_advantages: Optional[str] = None
    what_we_offer: Optional[str] = None


# iter108 — Backward-compatible synthesis: every read path emits
# `descriptive_sections` regardless of whether the row was created with the
# new schema or the legacy three-field schema. Legacy rows are never mutated.
_LEGACY_JOB_SECTIONS = (
    ("key_responsibilities", "Key Responsibilities"),
    ("added_advantages",     "Added Advantages"),
    ("what_we_offer",        "What We Offer"),
)


def _job_opening_sections(opening: dict) -> List[dict]:
    """Return the canonical `descriptive_sections` list for an opening.
    If the row already has a non-empty `descriptive_sections`, use it as-is
    (stripped). Otherwise synthesize from the legacy three-field schema,
    omitting empty fields so legacy openings that only had, say,
    `key_responsibilities` populated render as a single section."""
    existing = opening.get("descriptive_sections") or []
    if isinstance(existing, list) and any(
        (s or {}).get("title") or (s or {}).get("description") for s in existing
    ):
        return [
            {"title": (s or {}).get("title", "").strip(),
             "description": (s or {}).get("description", "").strip()}
            for s in existing if isinstance(s, dict)
        ]
    out: List[dict] = []
    for field, label in _LEGACY_JOB_SECTIONS:
        val = (opening.get(field) or "").strip()
        if val:
            out.append({"title": label, "description": val})
    return out

# iter100 — Slug helpers for public job-opening URLs. Slugs are lower-case,
# hyphenated, alnum-safe; collisions get a numeric suffix. We never rewrite
# historical rows aggressively — slugs are generated on CREATE, on UPDATE
# when the title changes, and lazily back-filled the first time a row is
# read via /job-openings.
def _slugify_title(title: str) -> str:
    """`AI & ML Engineer (2025)` → `ai-ml-engineer-2025`. Empty input → ''."""
    s = (title or "").lower().strip()
    # Replace any non-alnum with hyphen, then collapse runs and trim ends.
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s[:80]  # safety cap


async def _unique_job_opening_slug(base: str, *, exclude_id=None) -> str:
    """Returns `base` if free, otherwise appends `-2`, `-3`, ... until free.
    `exclude_id` lets the UPDATE path keep its own slug unchanged."""
    if not base:
        base = "job"
    candidate = base
    n = 1
    while True:
        q = {"slug": candidate}
        if exclude_id is not None:
            q["_id"] = {"$ne": exclude_id}
        if not await _db.bb_job_openings.find_one(q, {"_id": 1}):
            return candidate
        n += 1
        candidate = f"{base}-{n}"


@bb_router.get("/job-openings")
async def list_job_openings(request: Request):
    await _require_auth(request)
    openings = await _db.bb_job_openings.find({}).sort("created_at", -1).to_list(None)
    # iter100 — Lazy slug back-fill. Rows created before the slug feature
    # have no `slug` field; assign + persist one on first read so future
    # public URLs are clean. We never mutate any other field here.
    for o in openings:
        if not o.get("slug"):
            slug = await _unique_job_opening_slug(_slugify_title(o.get("title", "")), exclude_id=o["_id"])
            await _db.bb_job_openings.update_one({"_id": o["_id"]}, {"$set": {"slug": slug}})
            o["slug"] = slug
        o["id"] = str(o.pop("_id"))
        # iter108 — Emit synthesized descriptive_sections so the admin UI gets
        # a uniform shape whether the row is legacy or new-schema.
        o["descriptive_sections"] = _job_opening_sections(o)
    return {"openings": openings}

@bb_router.post("/job-openings")
async def create_job_opening(data: JobOpeningCreate, request: Request):
    await _require_auth(request)
    title = data.title.strip()
    slug = await _unique_job_opening_slug(_slugify_title(title))
    # iter108 — Persist descriptive_sections as the source of truth. ALSO
    # mirror the first 3 sections back onto the legacy fields so any older
    # external integration that still reads them keeps working until fully
    # migrated. New writes never lose data either way.
    sections = (
        [{"title": (s.title or "").strip(),
          "description": (s.description or "").strip()}
         for s in (data.descriptive_sections or [])]
        if data.descriptive_sections is not None else []
    )
    # If client only sent legacy fields, synthesize sections from them.
    if not any(s["title"] or s["description"] for s in sections):
        legacy_input = {
            "key_responsibilities": data.key_responsibilities or "",
            "added_advantages": data.added_advantages or "",
            "what_we_offer": data.what_we_offer or "",
        }
        sections = _job_opening_sections(legacy_input)
    # Mirror first three sections to legacy fields for backward compatibility.
    legacy_mirror = ["", "", ""]
    for i, s in enumerate(sections[:3]):
        legacy_mirror[i] = s.get("description", "")
    doc = {"title": title, "slug": slug,
           "job_role": (data.job_role or "").strip(),
           "vacancies": data.vacancies, "years_of_graduation": data.years_of_graduation or [],
           "education": data.education or [], "salary_range": (data.salary_range or "").strip(),
           "descriptive_sections": sections,
           "key_responsibilities": legacy_mirror[0],
           "added_advantages":     legacy_mirror[1],
           "what_we_offer":        legacy_mirror[2],
           "created_at": datetime.now(timezone.utc).isoformat(), "updated_at": datetime.now(timezone.utc).isoformat()}
    result = await _db.bb_job_openings.insert_one(doc)
    return {"success": True, "id": str(result.inserted_id), "slug": slug}

@bb_router.put("/job-openings/{opening_id}")
async def update_job_opening(opening_id: str, data: JobOpeningUpdate, request: Request):
    await _require_auth(request)
    oid = _oid(opening_id)
    updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
    for field in ["title", "job_role", "salary_range", "key_responsibilities", "added_advantages", "what_we_offer"]:
        val = getattr(data, field, None)
        if val is not None:
            updates[field] = val.strip()
    if data.vacancies is not None:
        updates["vacancies"] = data.vacancies
    if data.years_of_graduation is not None:
        updates["years_of_graduation"] = data.years_of_graduation
    if data.education is not None:
        updates["education"] = data.education
    # iter108 — descriptive_sections is authoritative when supplied. Always
    # mirror the first 3 onto legacy fields so partial-legacy consumers stay
    # consistent with the new source of truth.
    if data.descriptive_sections is not None:
        sections = [{"title": (s.title or "").strip(),
                     "description": (s.description or "").strip()}
                    for s in data.descriptive_sections]
        updates["descriptive_sections"] = sections
        for i, key in enumerate(["key_responsibilities", "added_advantages", "what_we_offer"]):
            updates[key] = sections[i]["description"] if i < len(sections) else ""
    # iter100 — Re-slug when title changes; keep existing slug otherwise.
    if "title" in updates:
        updates["slug"] = await _unique_job_opening_slug(_slugify_title(updates["title"]), exclude_id=oid)
    result = await _db.bb_job_openings.update_one({"_id": oid}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}

@bb_router.delete("/job-openings/{opening_id}")
async def delete_job_opening(opening_id: str, request: Request):
    await _require_auth(request)
    result = await _db.bb_job_openings.delete_one({"_id": _oid(opening_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}


# iter96 — Public view-only endpoint for sharing a job opening's description
# externally without exposing the Apply / admin workflow. Read-only, no auth.
# Returns only the public-safe display fields — never `_id`, `created_at`, or
# any other internal metadata. Works for every existing and future opening
# automatically (no migration needed — keys off the existing ObjectId).
@pub_router.get("/job-opening/{opening_id_or_slug}")
async def get_public_job_opening(opening_id_or_slug: str):
    """iter100 — Lookup is slug-first, ObjectId fallback so legacy
    /jobs/view/<24-char-hex> links keep resolving."""
    opening = None
    # 1) Slug match (preferred — clean URLs)
    opening = await _db.bb_job_openings.find_one({"slug": opening_id_or_slug})
    # 2) Legacy ObjectId fallback
    if not opening:
        try:
            oid = _oid(opening_id_or_slug)
            opening = await _db.bb_job_openings.find_one({"_id": oid})
        except Exception:
            pass
    if not opening:
        raise HTTPException(status_code=404, detail="Job opening not found")
    return {
        "title":                opening.get("title", ""),
        "job_role":             opening.get("job_role", ""),
        "vacancies":            opening.get("vacancies"),
        "years_of_graduation":  opening.get("years_of_graduation", []),
        "education":            opening.get("education", []),
        "salary_range":         opening.get("salary_range", ""),
        # iter108 — synthesized for legacy rows, native for new-schema rows.
        "descriptive_sections": _job_opening_sections(opening),
        # Legacy fields retained so any older external consumer keeps working.
        "key_responsibilities": opening.get("key_responsibilities", ""),
        "added_advantages":     opening.get("added_advantages", ""),
        "what_we_offer":        opening.get("what_we_offer", ""),
    }


# ============ INTERVIEW SCHEDULE REPORTS ============

def _build_interview_reports_match(startDate, endDate, jobRole, attendance, collegeType, _canonical_index=None, _mappings=None) -> dict:
    """Shared filter-builder for /interview-reports and /interview-reports/export.
    iter102 — When `jobRole` is the canonical title, expand it into a regex
    alternation over every raw keyword that maps to it (preserving each
    keyword's original casing/punctuation so existing _normalized_job_role
    values like 'AI And ML Engineer - C++ or Java Developer' still match).
    Pass `_mappings` to reuse the caller's job_keyword_mapping fetch.
    """
    match = {
        "schedule_date": {"$nin": [None, ""], "$exists": True},
        "schedule_time": {"$nin": [None, ""], "$exists": True},
    }
    # `extra_clauses` collects $or-style filters that must be AND-combined.
    extra_clauses = []
    if startDate or endDate:
        sd = {"$nin": [None, ""], "$exists": True}
        if startDate:
            sd["$gte"] = startDate
        if endDate:
            sd["$lte"] = endDate
        match["schedule_date"] = sd
    if jobRole and jobRole.strip().lower() not in ("", "all"):
        # Build the variant set: canonical itself + every keyword that maps
        # to it (raw casing/punctuation preserved). Mongo regex with case-
        # insensitive flag handles minor casing diffs in stored values.
        variant_patterns = {re.escape(jobRole.strip())}
        target_norm_lower = jobRole.strip().lower()
        for m in (_mappings or []):
            if (m.get("job_role") or "").strip().lower() == target_norm_lower:
                for kw in m.get("keywords", []) or []:
                    if kw:
                        variant_patterns.add(re.escape(kw))
                break
        extra_clauses.append({"$or": [
            {"_normalized_job_role": {"$regex": f"^{p}$", "$options": "i"}}
            for p in variant_patterns
        ]})
    # iter110 — College type filter accepts the 5 canonical buckets plus the
    # legacy "Premium"/"Non Premium" aliases for backward compat.
    if collegeType and collegeType.strip().lower() not in ("", "all"):
        ct = collegeType.strip()
        ct_low = ct.lower()
        if ct == "NIRF" or "premium" in ct_low and "non" not in ct_low:
            match["_nirf_category"] = "NIRF"
        elif ct in ("Non-NIRF 101-150", "Non-NIRF 151-200", "Non-NIRF 201-300", "Non-NIRF - No Rank"):
            match["_nirf_category"] = ct
        elif "non" in ct_low:
            match["_nirf_category"] = {"$ne": "NIRF"}
    if attendance and attendance.strip().lower() not in ("", "all"):
        att = attendance.strip().lower().replace(" ", "")
        if att == "attended":
            match["otp_verified"] = {"$nin": [None, ""], "$exists": True}
        elif att == "notattended":
            extra_clauses.append({"$or": [
                {"otp_verified": {"$in": [None, ""]}},
                {"otp_verified": {"$exists": False}},
            ]})
    if extra_clauses:
        # `$and` combines all the $or clauses without colliding at the root.
        match["$and"] = extra_clauses
    return match


def _format_date_ddmmyyyy(s: str) -> str:
    """Convert 'YYYY-MM-DD' → 'DD/MM/YYYY'. Returns input if not parseable."""
    if not s or not isinstance(s, str):
        return ""
    try:
        d = datetime.strptime(s[:10], "%Y-%m-%d")
        return d.strftime("%d/%m/%Y")
    except Exception:
        return s


def _format_time_12h(s: str) -> str:
    """Convert 'HH:MM:SS' or 'HH:MM' (24h) → '12-hour AM/PM'. Falls back to input.

    iter83 — Display-only heuristic: legacy importer dropped the +12 PM offset
    on 9904 pipeline_data rows (e.g. "01:00:00" really means 1 PM). Interview
    slots only exist between 10:00 and 17:30, so any stored hour < 6 is
    unambiguously a mis-stored PM slot. Storage stays untouched.
    """
    if not s or not isinstance(s, str):
        return ""
    parts = s.split(":")
    if len(parts) < 2:
        return s
    try:
        h = int(parts[0]); m = int(parts[1])
        if 1 <= h < 6:
            h += 12
        period = "AM" if h < 12 else "PM"
        h12 = h % 12 or 12
        return f"{h12:02d}:{m:02d} {period}"
    except Exception:
        return s


# ============ MISSING APPLICANTS (iter82) ============

def _build_missing_applicants_match(
    from_date: Optional[str], to_date: Optional[str],
    date_filter: str, report_type: str,
) -> dict:
    """Filter builder for /missing-applicants and /export.

    date_filter:
      • registered  → DATE portion of `last_update`
      • scheduled   → DATE portion of `schedule_date`

    report_type:
      • not_scheduled → email_type=shortlist AND schedule_date in (NULL/'')
                        AND schedule_time in (NULL/'')
      • not_attended  → schedule_date non-empty AND schedule_time non-empty
                        AND otp_verified in (None/'', 0, "0")
      • all           → union of the two via $or
    """
    NULL_OR_EMPTY = {"$in": [None, ""]}
    NOT_NULL = {"$nin": [None, ""], "$exists": True}

    cond_not_scheduled = {
        "email_type": {"$regex": "^shortlist$", "$options": "i"},
        "schedule_date": NULL_OR_EMPTY,
        "schedule_time": NULL_OR_EMPTY,
    }
    cond_not_attended = {
        "schedule_date": NOT_NULL,
        "schedule_time": NOT_NULL,
        "$or": [
            {"otp_verified": {"$in": [None, "", 0, "0", False]}},
            {"otp_verified": {"$exists": False}},
        ],
    }

    rt = (report_type or "all").strip().lower()
    if rt == "not_scheduled":
        match = dict(cond_not_scheduled)
    elif rt == "not_attended":
        match = dict(cond_not_attended)
    else:
        match = {"$or": [cond_not_scheduled, cond_not_attended]}

    # Date range — applies to either `last_update` or `schedule_date` depending
    # on `date_filter`. DATE-portion match against the YYYY-MM-DD prefix.
    if from_date or to_date:
        if (date_filter or "registered").strip().lower() == "scheduled":
            field = "schedule_date"
        else:
            field = "last_update"
        # Build a regex that captures the YYYY-MM-DD prefix in lexicographic
        # range. Easier: rely on string comparison since both formats start
        # with YYYY-MM-DD.
        rng = {}
        if from_date:
            rng["$gte"] = from_date
        if to_date:
            rng["$lte"] = to_date + "\uffff"  # include any time-suffix on `last_update`
        match[field] = rng if field not in match or not isinstance(match.get(field), dict) else {**match.get(field), **rng}
    return match


def _missing_status(doc: dict) -> str:
    """Derive the user-facing display status for one document."""
    sched = (doc.get("schedule_date") or "").strip()
    stime = (doc.get("schedule_time") or "").strip()
    otp = doc.get("otp_verified")
    if not sched and not stime:
        return "Shortlisted but interview not scheduled"
    if sched and stime and not otp:
        return "Interview scheduled but not attended"
    return "—"


@bb_router.get("/missing-applicants")
async def missing_applicants(
    request: Request,
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    date_filter: str = Query("registered"),
    report_type: str = Query("all"),
    name: Optional[str] = Query(None),
    email: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
    collegeStatus: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(100, ge=1, le=500),
):
    """List candidates who missed downstream stages: shortlisted+not-scheduled
    or scheduled+not-attended. Reads only `pipeline_data`."""
    await _require_auth(request)
    match = _build_missing_applicants_match(from_date, to_date, date_filter, report_type)
    # iter111 — Per-field Name/Email/Phone + 5-bucket College Status filters.
    _npe = []
    if name and name.strip():
        _npe.append({"name": {"$regex": re.escape(name.strip()), "$options": "i"}})
    if email and email.strip():
        _npe.append({"email": {"$regex": re.escape(email.strip()), "$options": "i"}})
    if phone and phone.strip():
        _npe.append({"phone": {"$regex": re.escape(phone.strip())}})
    if _npe:
        match["$and"] = (match.get("$and") or []) + _npe
    if collegeStatus and collegeStatus.strip() and collegeStatus.strip().lower() != "all":
        fval = collegeStatus.strip()
        if fval == "NIRF":
            match["_nirf_category"] = "NIRF"
        elif fval in ("Non NIRF", "Non-NIRF"):
            match["_nirf_category"] = {"$ne": "NIRF"}
        elif fval in ("Non-NIRF 101-150", "Non-NIRF 151-200", "Non-NIRF 201-300", "Non-NIRF - No Rank"):
            match["_nirf_category"] = fval
        else:
            match["_college_status"] = fval
    total = await _db.pipeline_data.count_documents(match)
    skip = (page - 1) * limit
    cursor = _db.pipeline_data.find(match, {"_id": 0}).skip(skip).limit(limit)
    rows = []
    async for d in cursor:
        rows.append({
            "name": d.get("name") or "",
            "email": d.get("email") or "",
            "phone": d.get("phone") or "",
            "current_location": d.get("location") or d.get("current_location") or "",
            "job_role": d.get("job_role") or d.get("job_title") or "",
            "college": d.get("college") or d.get("_college_resolved") or "",
            # iter110 — Prefer the persisted bucketed status; fall back to the
            # legacy `college_type` only if the new field is absent.
            "college_type": d.get("_college_status") or d.get("college_type") or "",
            "degree": d.get("degree") or "",
            "course": d.get("course") or "",
            "registered_date": d.get("last_update") or "",
            "schedule_date": d.get("schedule_date") or "",
            "schedule_time": d.get("schedule_time") or "",
            "result_status": _missing_status(d),
        })
    return {"data": rows, "total": total, "page": page, "limit": limit}


@bb_router.get("/missing-applicants/export")
async def export_missing_applicants(
    request: Request,
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    date_filter: str = Query("registered"),
    report_type: str = Query("all"),
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
):
    """Export the same rows as /missing-applicants (no pagination) in CSV/XLSX.
    Dates render as dd-mm-yyyy and times as hh:mm AM/PM."""
    await _require_auth(request)
    from fastapi.responses import StreamingResponse
    import io
    import csv as _csv

    match = _build_missing_applicants_match(from_date, to_date, date_filter, report_type)
    cursor = _db.pipeline_data.find(match, {"_id": 0})
    docs = await cursor.to_list(None)
    if not docs:
        raise HTTPException(status_code=404, detail="No data available to export")

    headers = [
        "Name", "Email", "Phone", "Current Location", "Job Role", "College",
        "College Type", "Degree", "Course", "Registered Date",
        "Scheduled Date", "Schedule Time", "Result Status",
    ]

    def _row(d):
        # Registered Date: last_update may carry time → strip to date portion
        reg_raw = (d.get("last_update") or "")[:10]
        return [
            d.get("name") or "",
            d.get("email") or "",
            d.get("phone") or "",
            d.get("location") or d.get("current_location") or "",
            d.get("job_role") or d.get("job_title") or "",
            d.get("college") or d.get("_college_resolved") or "",
            d.get("_college_status") or d.get("college_type") or "",
            d.get("degree") or "",
            d.get("course") or "",
            _format_date_ddmmyyyy(reg_raw).replace("/", "-"),
            _format_date_ddmmyyyy(d.get("schedule_date") or "").replace("/", "-"),
            _format_time_12h(d.get("schedule_time") or ""),
            _missing_status(d),
        ]

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename_base = f"Missing_Applicants_{today}"

    if format == "csv":
        buf = io.StringIO()
        w = _csv.writer(buf)
        w.writerow(headers)
        for d in docs:
            w.writerow(_row(d))
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )

    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Missing Applicants")
    ws.append(headers)
    for d in docs:
        ws.append(_row(d))
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        iter([bio.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
    )


# ============ EXPORT FIELD CATALOG (Schema-aware Dynamic Export) ============
#
# Each entry maps an API key → user-facing label, section, the underlying DB
# fields it needs, and an extractor that produces a clean cell value.
# To add a new exportable field: append one entry here. UI updates automatically.

def _val(d, *keys, default=""):
    """Return the first non-empty value from keys."""
    for k in keys:
        v = d.get(k)
        if v not in (None, "", False):
            return v
    return default


EXPORT_FIELD_CATALOG = [
    # key, label, section, required_db_keys, extractor
    ("name",                "Name",                "registration", ["name"],                                       lambda d: (d.get("name") or "").strip()),
    ("email",               "Email",               "registration", ["email"],                                      lambda d: (d.get("email") or "").strip().lower()),
    ("phone",               "Phone",               "registration", ["phone"],                                      lambda d: (d.get("phone") or "").strip()),
    ("age",                 "Age",                 "registration", ["age"],                                        lambda d: d.get("age") or ""),
    ("gender",              "Gender",              "registration", ["gender"],                                     lambda d: d.get("gender") or ""),
    ("date_of_birth",       "Date of Birth",       "registration", ["date_of_birth"],                              lambda d: d.get("date_of_birth") or ""),
    ("date_of_application", "Date of Application", "registration", ["date_of_application"],                        lambda d: _format_date_ddmmyyyy(d.get("date_of_application") or "")),
    ("college",             "College",             "registration", ["college", "_college_resolved"],               lambda d: d.get("_college_resolved") or d.get("college") or ""),
    ("college_status",      "College Status",      "registration", ["_college_status"],                            lambda d: d.get("_college_status") or ""),
    ("degree",              "Degree",              "registration", ["degree"],                                     lambda d: d.get("degree") or ""),
    ("course",              "Course",              "registration", ["course"],                                     lambda d: d.get("course") or ""),
    ("year_of_graduation",  "Year of Graduation",  "registration", ["year_of_graduation"],                         lambda d: d.get("year_of_graduation") or ""),
    ("current_location",    "Current Location",    "registration", ["current_location_state", "current_location_city"], lambda d: ", ".join([s for s in [d.get("current_location_city"), d.get("current_location_state")] if s])),
    ("preferred_location",  "Preferred Location",  "registration", ["preferred_location_city"],                    lambda d: d.get("preferred_location_city") or ""),
    # Interview section
    ("job_role",            "Job Role",            "interview",    ["job_role", "_normalized_job_role"],           lambda d: d.get("_normalized_job_role") or d.get("job_role") or ""),
    ("schedule_date",       "Schedule Date",       "interview",    ["schedule_date"],                              lambda d: _format_date_ddmmyyyy(d.get("schedule_date") or "")),
    ("schedule_time",       "Schedule Time",       "interview",    ["schedule_time"],                              lambda d: _format_time_12h(d.get("schedule_time") or "")),
    ("attendance",          "Attendance",          "interview",    ["otp_verified"],                               lambda d: "Attended" if (d.get("otp_verified") not in (None, "", False)) else "Not Attended"),
    ("college_type",        "College Type",        "interview",    ["_college_status", "_nirf_category"],          lambda d: d.get("_college_status") or d.get("_nirf_category") or ""),
    ("result_status",       "Result Status",       "interview",    ["result_status"],                              lambda d: d.get("result_status") or ""),
]

EXPORT_FIELD_BY_KEY = {e[0]: e for e in EXPORT_FIELD_CATALOG}


def _export_projection_for(field_keys):
    """Build a Mongo $project dict including all DB keys required by selected fields."""
    needed = {"_id": 0}
    for fk in field_keys:
        entry = EXPORT_FIELD_BY_KEY.get(fk)
        if not entry:
            continue
        for db_key in entry[3]:
            needed[db_key] = 1
    # Always include keys we use for de-dupe / dedupe sentinel:
    needed["email"] = 1
    needed["schedule_date"] = 1
    return needed


@bb_router.get("/interview-reports/export-fields")
async def get_interview_reports_export_fields(
    request: Request,
    startDate: str = Query(None), endDate: str = Query(None),
    jobRole: str = Query(None), attendance: str = Query(None),
    collegeType: str = Query(None),
):
    """Return the dynamic field catalog for the Export modal.

    Hybrid mode: catalog provides labels + section + ordering; we INTERSECT it
    with keys actually present in a sample of matching docs so missing fields
    are silently skipped (per spec).
    """
    await _require_auth(request)
    match = _build_interview_reports_match(startDate, endDate, jobRole, attendance, collegeType)

    # Source: pipeline_data first (live), fallback to legacy join
    src = _db.pipeline_data
    cnt = await src.count_documents(match)
    if cnt == 0:
        src = _db.registered_candidates
        cnt = await src.count_documents(match)

    present_keys: set = set()
    if cnt > 0:
        # Sample 50 docs — enough to surface optional fields without scanning the full set.
        sample = await src.aggregate([
            {"$match": match}, {"$limit": 50},
        ], allowDiskUse=False).to_list(None)
        for d in sample:
            present_keys.update(d.keys())

    sections_map = {
        "registration": {"id": "registration", "label": "Registration Fields", "fields": []},
        "interview": {"id": "interview", "label": "Interview Fields", "fields": []},
    }
    for key, label, section, db_keys, _extractor in EXPORT_FIELD_CATALOG:
        # Always include the field if catalog says so AND at least one underlying
        # DB key is present in the sample. If the dataset is empty we still show
        # the catalog so the modal doesn't appear broken — backend later returns 404.
        if cnt > 0 and not any(k in present_keys for k in db_keys):
            continue
        sections_map[section]["fields"].append({"key": key, "label": label})

    return {
        "sections": [sections_map["registration"], sections_map["interview"]],
        "total_matching": cnt,
    }


@bb_router.get("/interview-reports/export")
async def export_interview_reports(
    request: Request,
    startDate: str = Query(None), endDate: str = Query(None),
    jobRole: str = Query(None), attendance: str = Query(None),
    collegeType: str = Query(None),
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
    fields: Optional[str] = Query(None, description="Comma-separated field keys; omit for all"),
):
    """Export filtered Interview Schedule Reports to XLSX/CSV with user-selected
    columns. `fields` accepts a CSV list of catalog keys; column order in the
    output matches the order in the parameter. De-dupes by (email, schedule_date)."""
    await _require_auth(request)

    match = _build_interview_reports_match(startDate, endDate, jobRole, attendance, collegeType)

    # Resolve selected fields from query (preserve order). Fall back to whole catalog.
    if fields:
        requested = [f.strip() for f in fields.split(",") if f.strip()]
        selected = [k for k in requested if k in EXPORT_FIELD_BY_KEY]
    else:
        selected = [e[0] for e in EXPORT_FIELD_CATALOG]
    if not selected:
        raise HTTPException(status_code=400, detail="Please select at least one field")

    headers = [EXPORT_FIELD_BY_KEY[k][1] for k in selected]
    extractors = [(EXPORT_FIELD_BY_KEY[k][1], EXPORT_FIELD_BY_KEY[k][4]) for k in selected]

    pipeline = [
        {"$match": match},
        {"$sort": {"schedule_date": -1, "schedule_time": -1}},
        {"$project": _export_projection_for(selected)},
    ]

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename_base = f"Interview_Report_{today}"

    from fastapi.responses import StreamingResponse
    import io, csv as _csv

    # Pre-flight count: if zero, return 404 immediately. Use pipeline_data first; fall back.
    src = _db.pipeline_data
    cnt = await src.count_documents(match)
    if cnt == 0:
        src = _db.registered_candidates
        cnt = await src.count_documents(match)
    if cnt == 0:
        raise HTTPException(status_code=404, detail="No data available to export")

    def _row(d: dict) -> Optional[dict]:
        # Skip rows missing identity essentials (name+email).
        name = (d.get("name") or "").strip()
        email = (d.get("email") or "").strip().lower()
        if not (name and email):
            return None
        return {label: extract(d) for label, extract in extractors}

    if format == "csv":
        # True streaming CSV with gzip compression — keeps the K8s 60 s ingress
        # alive on large datasets (~10x reduction in wire bytes).
        import gzip as _gzip

        async def _csv_rows():
            yield (",".join(_csv_field(h) for h in headers) + "\n").encode("utf-8")
            buf = io.StringIO()
            writer = _csv.DictWriter(buf, fieldnames=headers)
            seen = set()
            cursor = src.aggregate(pipeline, allowDiskUse=True, batchSize=2000)
            async for d in cursor:
                row = _row(d)
                if not row:
                    continue
                key = ((d.get("email") or "").strip().lower(), d.get("schedule_date") or "")
                if key in seen:
                    continue
                seen.add(key)
                writer.writerow(row)
                data = buf.getvalue().encode("utf-8")
                buf.seek(0); buf.truncate(0)
                yield data

        async def _gzip_stream():
            out = io.BytesIO()
            comp = _gzip.GzipFile(fileobj=out, mode="wb", compresslevel=6)
            async for chunk in _csv_rows():
                comp.write(chunk)
                comp.flush()
                buffered = out.getvalue()
                out.seek(0); out.truncate(0)
                if buffered:
                    yield buffered
            comp.close()
            tail = out.getvalue()
            if tail:
                yield tail

        return StreamingResponse(
            _gzip_stream(),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{filename_base}.csv"',
                "Content-Encoding": "gzip",
            },
        )

    # XLSX: build in-memory (openpyxl write_only) — keeps memory low while still
    # buffering the final ZIP to a single response (XLSX cannot be sent in chunks).
    docs = await src.aggregate(pipeline, allowDiskUse=True).to_list(None)
    seen = set(); rows = []
    for d in docs:
        row = _row(d)
        if not row:
            continue
        key = ((d.get("email") or "").strip().lower(), d.get("schedule_date") or "")
        if key in seen:
            continue
        seen.add(key)
        rows.append(row)
    if not rows:
        raise HTTPException(status_code=404, detail="No data available to export")

    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Interview Reports")
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
    )


def _csv_field(s) -> str:
    """Quote a CSV header cell."""
    s = "" if s is None else str(s)
    if any(ch in s for ch in [",", "\"", "\n"]):
        return "\"" + s.replace("\"", "\"\"") + "\""
    return s


@bb_router.get("/interview-reports")
async def get_interview_reports(
    request: Request,
    startDate: str = Query(None), endDate: str = Query(None),
    jobRole: str = Query(None), attendance: str = Query(None),
    collegeType: str = Query(None),
    name: Optional[str] = Query(None),
    email: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
    page: int = Query(1, ge=1), limit: int = Query(100, ge=1, le=500),
    sort_by: Optional[str] = Query(None),
    sort_dir: Optional[str] = Query(None),
):
    """Interview Schedule Reports — OPTIMIZED (May 2026).
    Filters and summary counts are computed at the DB level using the persisted
    derived fields `_nirf_category` and `_normalized_job_role`. Returns
    {data, total, page, limit, totalPages, summary}.
    """
    await _require_auth(request)

    # iter102 — One canonical index + raw-mappings fetch per request, reused
    # by filter builder + row/summary canonicalization so all three layers
    # (filter, data rows, role_counts) stay aligned with the keyword mappings.
    from server import _build_canonical_index, _canonicalize_job_role, _get_job_keyword_mappings
    kw_to_canonical, _ = await _build_canonical_index()
    raw_mappings = await _get_job_keyword_mappings()

    match = _build_interview_reports_match(startDate, endDate, jobRole, attendance, collegeType, _canonical_index=kw_to_canonical, _mappings=raw_mappings)

    # iter111 — Per-field Name / Email / Phone filters (regex partial match).
    _npe = []
    if name and name.strip():
        _npe.append({"name": {"$regex": re.escape(name.strip()), "$options": "i"}})
    if email and email.strip():
        _npe.append({"email": {"$regex": re.escape(email.strip()), "$options": "i"}})
    if phone and phone.strip():
        _npe.append({"phone": {"$regex": re.escape(phone.strip())}})
    if _npe:
        match["$and"] = (match.get("$and") or []) + _npe

    # Source: pipeline_data is the live collection (May 2026 architecture migration);
    # fall back to legacy registered_candidates if empty so older environments still work.
    src = _db.pipeline_data
    total = await src.count_documents(match)
    if total == 0:
        src = _db.registered_candidates
        total = await src.count_documents(match)

    skip = (page - 1) * limit
    pipeline = [
        {"$match": match},
        {"$sort": _build_sort(sort_by, sort_dir, allowed={
            "name": "name", "email": "email",
            "date": "schedule_date", "time": "schedule_time",
            "schedule_date": "schedule_date", "schedule_time": "schedule_time",
            "job_role": "_normalized_job_role",
            "college_type": "_nirf_category",
            "attendance": "otp_verified",
        }, default={"schedule_date": -1, "schedule_time": -1})},
        {"$skip": skip},
        {"$limit": limit},
        {"$project": {
            "_id": 0, "name": 1, "email": 1,
            "schedule_date": 1, "schedule_time": 1,
            "job_title": 1, "job_role": 1, "_normalized_job_role": 1,
            "_nirf_category": 1, "_college_status": 1, "otp_verified": 1,
        }},
    ]
    docs = await src.aggregate(pipeline, allowDiskUse=False).to_list(None)

    rows = []
    for d in docs:
        # iter110 — Emit the full bucketed college status string instead of
        # the binary Premium / Non-Premium label.
        college_status = d.get("_college_status") or d.get("_nirf_category") or "Non-NIRF - No Rank"
        if college_status == "NIRF":
            college_status = "NIRF"  # binary alias rows that lack rank suffix
        otp = str(d.get("otp_verified") or "").strip()
        # iter102 — Display canonical job title per row so the table column
        # value matches the filter button labels.
        raw_role = d.get("_normalized_job_role") or d.get("job_role") or d.get("job_title") or ""
        canonical_role = _canonicalize_job_role(raw_role, kw_to_canonical) if raw_role else "-"
        rows.append({
            "name": d.get("name") or "-",
            "email": d.get("email") or "-",
            "date": d.get("schedule_date") or "-",
            "time": d.get("schedule_time") or "-",
            "job_role": canonical_role or "-",
            "college_type": college_status,
            "attendance": "Attended" if otp else "Not Attended",
        })

    # Summary (aggregated over the filtered set — ALL rows, not just this page)
    summary_pipeline = [
        {"$match": match},
        {"$group": {
            "_id": None,
            "attended": {"$sum": {"$cond": [
                {"$and": [
                    {"$ne": [{"$ifNull": ["$otp_verified", ""]}, ""]},
                    {"$ne": [{"$ifNull": ["$otp_verified", None]}, None]},
                ]}, 1, 0
            ]}},
            "not_attended": {"$sum": {"$cond": [
                {"$or": [
                    {"$eq": [{"$ifNull": ["$otp_verified", ""]}, ""]},
                    {"$eq": [{"$ifNull": ["$otp_verified", None]}, None]},
                ]}, 1, 0
            ]}},
            "premium_colleges": {"$sum": {"$cond": [
                {"$eq": [{"$ifNull": ["$_nirf_category", "Non NIRF"]}, "NIRF"]}, 1, 0
            ]}},
            "non_premium_colleges": {"$sum": {"$cond": [
                {"$ne": [{"$ifNull": ["$_nirf_category", "Non NIRF"]}, "NIRF"]}, 1, 0
            ]}},
        }},
    ]
    sres = await src.aggregate(summary_pipeline, allowDiskUse=False).to_list(None)
    base = sres[0] if sres else {"attended": 0, "not_attended": 0, "premium_colleges": 0, "non_premium_colleges": 0}

    # Role counts per filter set
    role_pipeline = [
        {"$match": match},
        {"$group": {
            "_id": {"$ifNull": ["$_normalized_job_role", "Unknown"]},
            "count": {"$sum": 1},
        }},
        {"$sort": {"count": -1}},
        {"$limit": 100},
    ]
    role_results = await src.aggregate(role_pipeline, allowDiskUse=False).to_list(None)
    # iter102 — Canonical roll-up of role_counts so filter buttons show one
    # entry per canonical job title.
    # iter103 — Drop the null / empty / "Unknown" bucket entirely. Surfacing
    # it as a filter button produces a chip that matches no real rows
    # (`_normalized_job_role` is None on those records, so filtering by
    # literal "Unknown" returns 0 hits). With iter99's canonicalization +
    # the keyword-mapping flow, every real role resolves to a canonical
    # title — the residual null bucket represents legacy rows without any
    # job_role at all and should not be filterable.
    rolled_counts: dict = {}
    for r in role_results:
        if not r["_id"]:
            continue
        canon = _canonicalize_job_role(r["_id"], kw_to_canonical)
        if not canon or canon.strip().lower() in ("", "unknown"):
            continue
        rolled_counts[canon] = rolled_counts.get(canon, 0) + r["count"]
    role_counts = dict(sorted(rolled_counts.items(), key=lambda kv: -kv[1])[:100])

    total_pages = (total + limit - 1) // limit if total else 1
    return {
        "data": rows,
        "total": total,
        "page": page,
        "limit": limit,
        "totalPages": total_pages,
        "summary": {
            "role_counts": role_counts,
            "attended": base.get("attended", 0),
            "not_attended": base.get("not_attended", 0),
            "premium_colleges": base.get("premium_colleges", 0),
            "non_premium_colleges": base.get("non_premium_colleges", 0),
        },
    }


# ============ UPDATE APPLICANT SCORES ============

class ScoreEntry(BaseModel):
    round_name: str
    score: float

class ApplicantScoreUpdate(BaseModel):
    status: str
    scores: Optional[List[ScoreEntry]] = None

@bb_router.get("/attended-for-scores")
async def get_attended_for_scores(
    request: Request,
    startDate: str = Query(None),
    endDate: str = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(100, ge=1, le=500),
    sort_by: Optional[str] = Query(None),
    sort_dir: Optional[str] = Query(None),
):
    """Update Applicants Scores — with pagination (May 2026).
    Returns {data, total, page, limit, totalPages, available_rounds}.
    `available_rounds` always reflects the full score_sheet set (global filter).
    """
    await _require_auth(request)
    match = {"otp_verified": {"$nin": [None, ""], "$exists": True},
             "schedule_date": {"$nin": [None, ""], "$exists": True}}
    if startDate:
        match["schedule_date"] = {**match.get("schedule_date", {}), "$gte": startDate}
    if endDate:
        match["schedule_date"] = {**match.get("schedule_date", {}), "$lte": endDate}

    # Source: pipeline_data first (live collection per May 2026 architecture),
    # fallback to legacy registered_candidates for older environments.
    src = _db.pipeline_data
    total = await src.count_documents(match)
    if total == 0:
        src = _db.registered_candidates
        total = await src.count_documents(match)

    skip = (page - 1) * limit
    pipeline = [
        {"$match": match},
        {"$sort": _build_sort(sort_by, sort_dir, allowed={
            "name": "name", "email": "email", "phone": "phone",
            "schedule_date": "schedule_date",
            "job_role": "_normalized_job_role",
            "result_status": "result_status",
        }, default={"schedule_date": -1})},
        {"$skip": skip},
        {"$limit": limit},
        {"$project": {
            "_id": 0, "email": 1, "phone": 1, "name": 1,
            "schedule_date": 1, "job_role": 1, "job_title": 1,
            "result_status": 1,
        }},
    ]
    docs = await src.aggregate(pipeline, allowDiskUse=False).to_list(None)

    # Page-scoped status overrides
    page_emails = [(d.get("email") or "").strip().lower() for d in docs if d.get("email")]
    updates = await _db.bb_applicant_updates.find(
        {"email": {"$in": page_emails}} if page_emails else {"_id": None}, {"_id": 0}
    ).to_list(None) if page_emails else []
    update_map = {u["email"]: u for u in updates if u.get("email")}

    # Scores for the current page only
    page_phones = []
    for d in docs:
        p = re.sub(r'[^\d]', '', d.get("phone") or "")
        if len(p) > 10:
            p = p[-10:]
        if p:
            page_phones.append(p)
    score_q = []
    if page_emails:
        score_q.append({"email": {"$in": page_emails}})
    if page_phones:
        score_q.append({"phone": {"$in": page_phones}})
    score_records = []
    if score_q:
        score_records = await _db.score_sheet.find(
            {"$or": score_q} if len(score_q) > 1 else score_q[0], {"_id": 0}
        ).to_list(None)

    score_by_email = {}
    score_by_phone = {}
    for sr in score_records:
        se = (sr.get("email") or "").strip().lower()
        sp = re.sub(r'[^\d]', '', sr.get("phone") or "")
        if len(sp) > 10:
            sp = sp[-10:]
        if se:
            score_by_email.setdefault(se, []).append(sr)
        if sp:
            score_by_phone.setdefault(sp, []).append(sr)

    # Available rounds (GLOBAL — unchanged contract): distinct names from score_sheet
    available_rounds = await _db.score_sheet.distinct("round_name")
    available_rounds = sorted([r for r in available_rounds if r])

    result = []
    for doc in docs:
        email = (doc.get("email") or "").strip().lower()
        phone = re.sub(r'[^\d]', '', doc.get("phone") or "")
        if len(phone) > 10:
            phone = phone[-10:]
        upd = update_map.get(email, {})

        merged_scores = []
        if upd.get("scores"):
            merged_scores = upd["scores"]
        else:
            matched = []
            if email and email in score_by_email:
                matched.extend(score_by_email[email])
            if phone and phone in score_by_phone:
                for s in score_by_phone[phone]:
                    if s not in matched:
                        matched.append(s)
            for sr in matched:
                rn = (sr.get("round_name") or "").strip()
                sc = sr.get("score", 0)
                if rn:
                    merged_scores.append({"round_name": rn, "score": sc})

        # ---- Round-wise structured summary (Iter45) ----
        # Group merged_scores by canonical round name and pick the most-recent
        # entry per round. Keeps the legacy `scores` array for backward compat.
        rws_grouped: dict = {}
        for s in merged_scores:
            canon = _norm_round(s.get("round_name"))
            if not canon:
                continue
            ts = str(s.get("created_at") or s.get("updated_at") or "")
            cur = rws_grouped.get(canon)
            if cur is None or ts > str(cur.get("created_at") or ""):
                try:
                    sval = float(s.get("score") or 0)
                except (TypeError, ValueError):
                    sval = 0.0
                rws_grouped[canon] = {
                    "score": sval,
                    "raw_round": s.get("round_name") or canon,
                    "created_at": s.get("created_at"),
                }
        latest_round = None
        latest_score = None
        if rws_grouped:
            latest_round = max(rws_grouped, key=lambda k: str(rws_grouped[k].get("created_at") or ""))
            latest_score = rws_grouped[latest_round]["score"]
        total_score = sum((v["score"] or 0) for v in rws_grouped.values())

        result.append({"name": doc.get("name") or "-", "email": email, "phone": doc.get("phone") or "-",
                        "date_of_interview": doc.get("schedule_date") or "-",
                        "job_role": doc.get("job_role") or doc.get("job_title") or "-",
                        "status": upd.get("status") or doc.get("result_status") or "On hold",
                        "scores": merged_scores,
                        "round_wise_scores": rws_grouped,
                        "latest_round": latest_round,
                        "latest_score": latest_score,
                        "total_score": total_score})

    total_pages = (total + limit - 1) // limit if total else 1
    return {
        "data": result,
        "total": total,
        "page": page,
        "limit": limit,
        "totalPages": total_pages,
        "available_rounds": available_rounds,
    }

@bb_router.put("/applicant-score/{email:path}")
async def update_applicant_score(email: str, data: ApplicantScoreUpdate, request: Request):
    await _require_auth(request)
    if not email:
        raise HTTPException(status_code=400, detail="Email required")
    update_doc = {"email": email, "status": data.status, "updated_at": datetime.now(timezone.utc).isoformat()}
    if data.scores:
        # Non-destructive merge: existing valid scores are preserved per
        # canonical round name. Newer timestamps (or missing rounds) win.
        existing = await _db.bb_applicant_updates.find_one({"email": email}, {"_id": 0, "scores": 1})
        existing_by_round = {}
        for s in (existing.get("scores") or []) if existing else []:
            canon = _norm_round(s.get("round_name"))
            if canon:
                existing_by_round[canon] = s
        for s in data.scores:
            canon = _norm_round(s.round_name)
            if not canon:
                continue
            existing_by_round[canon] = {"round_name": s.round_name, "score": s.score}
        update_doc["scores"] = list(existing_by_round.values())
    # iter98 — When the recruiter sets status="Rejected" we MUST clear any
    # rejection-send tracking flags from a previous rejection cycle. Without
    # this reset, the 19:00 IST evening rejection dispatcher (`bg_workers.
    # _worker_import_rejection_mailer` source A) filters out this doc forever
    # because `rejection_sent: {$ne: True}` still matches yesterday's send.
    # Idempotency is preserved by the worker itself — once it dispatches, it
    # flips `rejection_sent=True` again before the next cycle.
    extra_unset = {}
    if (data.status or "").strip().lower() == "rejected":
        extra_unset = {
            "rejection_sent": "",
            "rejection_sent_at": "",
            "rejection_notified": "",
            "rejection_notified_at": "",
            "import_rejection_notified": "",
            "rejection_send_ok": "",
        }
        _logger.info(
            f"[RejectFlagsReset] update_applicant_score email={email} "
            f"new_status={data.status!r} -> clearing rejection_sent/notified/import flags"
        )
    op = {"$set": update_doc}
    if extra_unset:
        op["$unset"] = extra_unset
    await _db.bb_applicant_updates.update_one({"email": email}, op, upsert=True)
    await _db.registered_candidates.update_many({"email": email}, {"$set": {"result_status": data.status}})
    # iter69e (#8) — Sync result_status to pipeline_data so the Manual Alerts
    # page, View Applicants, View Attended Applicants, and analytics all see
    # the same value immediately. Match by email OR phone (last-10-digit
    # regex) so duplicates with the same person are all kept in sync.
    try:
        import re as _re
        rec_for_phone = await _db.bb_applicant_updates.find_one({"email": email}, {"_id": 0, "phone": 1})
        phone_clauses = []
        ph = (rec_for_phone or {}).get("phone") or ""
        ph_norm = _re.sub(r"\D", "", str(ph))[-10:]
        clauses = [{"email": email}]
        if ph_norm:
            clauses.append({"phone": {"$regex": f"{_re.escape(ph_norm)}$"}})
        await _db.pipeline_data.update_many(
            {"$or": clauses},
            {"$set": {"result_status": data.status,
                       "result_status_updated_at": datetime.now(timezone.utc).isoformat()}},
        )
    except Exception as _e:
        _logger.warning(f"[applicant-score] pipeline_data result_status sync skipped: {_e}")
    return {"success": True}


@bb_router.get("/candidate-score-summary")
async def candidate_score_summary(
    request: Request,
    email: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
    pick: str = Query("latest", regex="^(latest|highest|lowest)$"),
):
    """Round-wise score summary for a single candidate (Email primary, Phone
    fallback). Returns a structured object per the Iter45 spec:

        {email, phone, name, round_wise_scores: {...},
         latest_round, latest_score, total_score, rounds, conflict?}
    """
    await _require_auth(request)
    if not email and not phone:
        raise HTTPException(status_code=400, detail="email or phone required")

    em = _norm_email(email or "")
    ph = _norm_phone(phone or "")

    # Identity from pipeline_data first; bb_registrations fallback
    or_clauses = []
    if em:
        or_clauses.append({"email": em})
    if ph:
        or_clauses.append({"phone": ph})
    pd_doc = await _db.pipeline_data.find_one(
        {"$or": or_clauses} if len(or_clauses) > 1 else or_clauses[0],
        {"_id": 0, "name": 1, "email": 1, "phone": 1},
    )
    if not pd_doc:
        reg = await _db.bb_registrations.find(
            {"$or": or_clauses} if len(or_clauses) > 1 else or_clauses[0],
            {"_id": 0, "full_name": 1, "email": 1, "phone": 1},
        ).sort("registered_at", -1).limit(1).to_list(1)
        pd_doc = reg[0] if reg else {}

    name = pd_doc.get("name") or pd_doc.get("full_name") or ""
    resolved_email = em or _norm_email(pd_doc.get("email") or "")
    resolved_phone = ph or _norm_phone(pd_doc.get("phone") or "")

    rws = await _build_round_wise_scores(resolved_email, resolved_phone, pick=pick)
    conflict = await _detect_score_phone_conflict(resolved_email, resolved_phone)

    return {
        "email": resolved_email,
        "phone": resolved_phone,
        "name": name,
        "round_wise_scores": rws["round_wise_scores"],
        "latest_round": rws["latest_round"],
        "latest_score": rws["latest_score"],
        "total_score": rws["total_score"],
        "rounds": rws["rounds"],
        "conflict": conflict,
    }


# ============ ITER52 — Candidate Journey (A–Z Row Action) ============
# Aggregates the full hiring lifecycle of one candidate for the dashboard
# row-action drawer/modal:
#   * Basic info (name, email, phone, college, job role)
#   * Round timeline (canonical round names + display labels + status)
#   * Final outcome (Selected / Rejected / In Progress)
#   * Date of Induction (read from pipeline_data; "Pending" if absent)
#
# Read-only with one tiny exception: PUT /candidate-induction-date allows
# admins to set `pipeline_data.date_of_induction` once a candidate is
# Selected. Everything else is sourced from the existing collections.

# Custom display labels per the spec (UI only — canonical names unchanged).
ROUND_DISPLAY_LABELS = {
    "Round 1": "Technical 1",
    "Round 2": "F2F",
    "HR Round": "HR Interview",
    "Final Round": "Final Discussion",
}


def _round_status(score) -> str:
    """Map a score (numeric / None / 'Rejected') to a UI status string."""
    if score is None:
        return "Pending"
    s = str(score).strip().lower()
    if s in ("rejected", "reject", "fail", "failed"):
        return "Rejected"
    return "Completed"


def _final_outcome(result_status: str, round_wise: dict) -> str:
    """Compute Selected / Rejected / In Progress from result_status + scores."""
    rs = (result_status or "").strip().lower()
    if "select" in rs or rs in ("hired", "offered"):
        return "Selected"
    if "reject" in rs:
        return "Rejected"
    return "In Progress"


@bb_router.get("/candidate-journey")
async def candidate_journey(
    request: Request,
    email: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
):
    """Full A–Z candidate journey for the dashboard row-action drawer.

    Resolution: email primary, phone fallback. Reads ONLY from existing
    processed sources (`pipeline_data`, `bb_applicant_updates`, `score_sheet`)
    — never modifies records. Conflicts (same phone bound to a different
    email) return 409 + log so the UI can block the view per spec.
    """
    await _require_auth(request)
    if not email and not phone:
        raise HTTPException(status_code=400, detail="email or phone required")
    em = _norm_email(email or "")
    ph = _norm_phone(phone or "")

    or_clauses = []
    if em:
        or_clauses.append({"email": em})
    if ph:
        or_clauses.append({"phone": ph})
    pd_doc = await _db.pipeline_data.find_one(
        {"$or": or_clauses} if len(or_clauses) > 1 else or_clauses[0],
        {"_id": 0},
    )
    if not pd_doc:
        raise HTTPException(status_code=404, detail="Candidate not found in pipeline_data")

    # Conflict detection — block the view per spec
    pd_email = _norm_email(pd_doc.get("email") or "")
    pd_phone = _norm_phone(pd_doc.get("phone") or "")
    if em and ph and pd_email and pd_phone and (pd_email != em or pd_phone != ph):
        if pd_email != em and pd_phone != ph:
            _logger.warning(f"[Journey] CONFLICT email/phone mismatch: query=({em}/{ph}) record=({pd_email}/{pd_phone})")
            raise HTTPException(status_code=409, detail="Email/phone mismatch — view blocked")

    resolved_email = pd_email or em
    resolved_phone = pd_phone or ph

    # Round-wise scores. Priority (matches /attended-for-scores):
    #   bb_applicant_updates.scores[]  >  score_sheet
    upd = await _db.bb_applicant_updates.find_one(
        {"email": resolved_email},
        {"_id": 0, "status": 1, "scores": 1, "updated_at": 1},
    ) or {}
    if upd.get("scores"):
        # Build the same shape as _build_round_wise_scores so the timeline
        # logic below stays unchanged.
        rws_grouped = {}
        for s in upd.get("scores") or []:
            canon = _norm_round(s.get("round_name"))
            if not canon:
                continue
            try:
                sval = float(s.get("score") or 0)
            except (TypeError, ValueError):
                sval = 0.0
            ts = str(upd.get("updated_at") or "")
            cur = rws_grouped.get(canon)
            if cur is None or ts > str(cur.get("created_at") or ""):
                rws_grouped[canon] = {
                    "score": sval, "raw_round": s.get("round_name") or canon,
                    "created_at": upd.get("updated_at"),
                }
        latest_round = None
        if rws_grouped:
            latest_round = max(rws_grouped, key=lambda k: str(rws_grouped[k].get("created_at") or ""))
        rws = {
            "round_wise_scores": rws_grouped,
            "latest_round": latest_round,
            "latest_score": rws_grouped[latest_round]["score"] if latest_round else None,
            "total_score": sum((v["score"] or 0) for v in rws_grouped.values()),
            "rounds": sorted(rws_grouped.keys()),
        }
    else:
        rws = await _build_round_wise_scores(resolved_email, resolved_phone)

    # Status from bb_applicant_updates (recruiter-set) > pipeline_data.result_status
    result_status = (upd.get("status") or pd_doc.get("result_status") or "").strip()

    # Build the round timeline. Use bb_rounds order so timeline is consistent
    # across candidates; fall back to alphabetical for any ad-hoc rounds.
    all_rounds_docs = await _db.bb_rounds.find(
        {"active": {"$ne": False}}
    ).sort([("order", 1), ("name", 1)]).to_list(None)
    ordered_round_names = [r.get("name") for r in all_rounds_docs if r.get("name")]
    # Append any rounds the candidate has but isn't in bb_rounds (legacy)
    for rn in rws["rounds"]:
        if rn not in ordered_round_names:
            ordered_round_names.append(rn)

    round_details = []
    for rn in ordered_round_names:
        bucket = rws["round_wise_scores"].get(rn)
        score = bucket.get("score") if bucket else None
        completed_date = bucket.get("created_at") if bucket else None
        # Spec: skip rounds the candidate has no data for (don't pad with Pending)
        if score is None and not completed_date:
            continue
        round_details.append({
            "round_name": rn,
            "round_label": ROUND_DISPLAY_LABELS.get(rn, rn),
            "score": score,
            "status": _round_status(score),
            "completed_date": completed_date,
        })

    final_status = _final_outcome(result_status, rws["round_wise_scores"])
    raw_doi = pd_doc.get("date_of_induction") or ""
    if final_status == "Selected":
        date_of_induction = raw_doi or "Pending"
    else:
        date_of_induction = raw_doi or "Not Applicable"

    return {
        "basic": {
            "name": pd_doc.get("name") or "",
            "email": resolved_email,
            "phone": resolved_phone,
            "college": pd_doc.get("college") or pd_doc.get("_college_resolved") or "",
            # iter110 — Surface the canonical 5-bucket college status.
            "college_status": pd_doc.get("_college_status") or pd_doc.get("college_type") or "",
            "job_role": pd_doc.get("job_role") or pd_doc.get("job_title") or "",
        },
        "round_details": round_details,
        "latest_round": rws["latest_round"],
        "latest_score": rws["latest_score"],
        "total_score": rws["total_score"],
        "final_outcome": {
            "status": final_status,
            "result_status_raw": result_status,
            "date_of_induction": date_of_induction,
        },
    }


class _InductionDateBody(BaseModel):
    email: str
    date_of_induction: Optional[str] = None  # ISO date or "" to clear


@bb_router.put("/candidate-induction-date")
async def update_induction_date(body: _InductionDateBody, request: Request):
    """Set / clear `pipeline_data.date_of_induction` for a candidate. Used
    by the Final Outcome card on the journey modal once a candidate is
    Selected. Read-only on every other field (no other writes here).
    """
    await _require_auth(request)
    em = _norm_email(body.email)
    if not em:
        raise HTTPException(status_code=400, detail="email required")
    pd_doc = await _db.pipeline_data.find_one({"email": em}, {"_id": 1})
    if not pd_doc:
        raise HTTPException(status_code=404, detail="Candidate not found")
    new_value = (body.date_of_induction or "").strip()
    await _db.pipeline_data.update_one(
        {"email": em},
        {"$set": {"date_of_induction": new_value,
                   "induction_updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    _logger.info(f"[Induction] email={em} date_of_induction={new_value or '(cleared)'}")
    return {"success": True, "email": em, "date_of_induction": new_value or None}


# ============ ITER55 — SCORE & ROUND TABLE MODULE ============
# New top-level dashboard module. Table-driven UI showing every candidate
# with their per-round scores (extended structure: round_name, date, score,
# command, status) plus dates of joining / documentation / induction.
#
# DATA-SAFETY: write paths are append-only / per-round upsert by canonical
# round name. Existing legacy `scores: [{round_name, score}]` records remain
# readable — older fields are simply absent in the response and treated as
# empty in the modal.

STATIC_ROUNDS_ITER55 = [
    "Accounts 1", "Accounts 2", "BA", "BE", "BP", "C++",
    "Java", "LA", "Mensa", "Mensa Org", "ZA",
]


# Iter64 — Score & Round Filter Tabs (pipeline-based candidate filtering)
# Maps each tab to a (label, mongo match for pipeline_data) pair.
TRUTHY_OTP = [1, 1.0, "1", "1.0", True, "true", "True", "yes", "Yes"]
SHORTLIST_VALUES_REGEX = "^(shortlist|shortlisted)$"


def _is_otp_verified(value) -> bool:
    """Loose truthy check across the various otp_verified storage shapes
    (bool, int 1, float 1.0, '1', '1.0', 'yes', 'true', etc.)."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value) == 1.0
    if isinstance(value, str):
        return value.strip().lower() in ("1", "1.0", "true", "yes")
    return False

def _tab_match(tab: str) -> dict:
    """Return the pipeline_data match dict for a given filter tab."""
    t = (tab or "shortlisted").strip().lower()
    if t == "shortlisted":
        return {"result_status": {"$regex": SHORTLIST_VALUES_REGEX, "$options": "i"}}
    if t == "attended":
        return {
            "result_status": {"$regex": SHORTLIST_VALUES_REGEX, "$options": "i"},
            "otp_verified": {"$in": TRUTHY_OTP},
        }
    if t == "not_attended":
        return {
            "result_status": {"$regex": SHORTLIST_VALUES_REGEX, "$options": "i"},
            "$or": [{"otp_verified": {"$nin": TRUTHY_OTP}}, {"otp_verified": {"$exists": False}}],
        }
    if t == "rejected":
        return {"result_status": {"$regex": "^rejected$", "$options": "i"}}
    if t == "pending":
        return {"result_status": {"$regex": "^pending$", "$options": "i"}}
    if t == "selected":
        return {"result_status": {"$regex": "^selected$", "$options": "i"}}
    if t == "joined":
        return {"result_status": {"$regex": "^joined$", "$options": "i"}}
    # Default safety net
    return {"result_status": {"$regex": SHORTLIST_VALUES_REGEX, "$options": "i"}}


def _is_otp_verified_DEPRECATED(value) -> bool:
    return False  # superseded by definition above


# Iter66 — Status filter is restricted to 3 canonical groups. Each group maps
# to all known DB variants (including legacy typos) so filtering catches every
# matching record regardless of how it was originally entered.
SCORE_ROUND_STATUS_GROUPS = {
    "Shortlisted": ["shortlist", "shortlisted", "shortlsited"],
    "Rejected":    ["reject", "rejected", "rejeceted"],
    "On-Hold":     ["hold", "on-hold", "on hold", "onhold"],
}


@bb_router.get("/score-round/result-statuses")
async def score_round_result_statuses(request: Request):
    """Iter66 — Status dropdown is restricted to 3 canonical groups only:
    Shortlisted, Rejected, On-Hold. The `/score-round/table` endpoint maps
    each canonical label to its full variant set (including typos) when
    filtering."""
    await _require_auth(request)
    return {"statuses": list(SCORE_ROUND_STATUS_GROUPS.keys())}


@bb_router.get("/score-round/table")
async def score_round_table(
    request: Request,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    q: Optional[str] = Query(None),
    startDate: Optional[str] = Query(None),
    endDate: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    college: Optional[str] = Query(None),
    job_role: Optional[str] = Query(None),
):
    """Iter55/iter58/iter64/iter65 — Score & Round table data source.

    Iter65 — `status` filter now applies to `pipeline_data.result_status`
    directly (case-insensitive exact match). The previous tab-based filter
    strip and `bb_applicant_updates.status` join were removed — any value
    returned by the new `/score-round/result-statuses` endpoint is a valid
    `status` query param. Tab-counts removed.
    """
    await _require_auth(request)
    match = {}
    if q:
        qs = q.strip()
        if qs:
            esc = re.escape(qs)
            match["$or"] = [
                {"name": {"$regex": esc, "$options": "i"}},
                {"email": {"$regex": esc, "$options": "i"}},
                {"phone": {"$regex": esc, "$options": "i"}},
            ]
    if startDate or endDate:
        sd_q = {}
        if startDate:
            sd_q["$gte"] = startDate.strip()
        if endDate:
            sd_q["$lte"] = endDate.strip()
        match["schedule_date"] = sd_q
    if college:
        c = re.escape(college.strip())
        match.setdefault("$and", []).append({
            "$or": [
                {"college": {"$regex": c, "$options": "i"}},
                {"college_name": {"$regex": c, "$options": "i"}},
            ],
        })
    if job_role:
        jr = re.escape(job_role.strip())
        match.setdefault("$and", []).append({
            "$or": [
                {"job_role": {"$regex": jr, "$options": "i"}},
                {"job_title": {"$regex": jr, "$options": "i"}},
            ],
        })
    if status:
        # Iter66 — map canonical label to variant set (groups Shortlist/Shortlisted,
        # Reject/Rejected, Hold/On-Hold/On hold + legacy typos). Falls back to
        # exact case-insensitive match for any other label.
        st = status.strip()
        variants = None
        for canon, vlist in SCORE_ROUND_STATUS_GROUPS.items():
            if st.lower() == canon.lower():
                variants = vlist
                break
        if variants:
            alts = "|".join(re.escape(v) for v in variants)
            match["result_status"] = {"$regex": f"^({alts})$", "$options": "i"}
        else:
            match["result_status"] = {"$regex": f"^{re.escape(st)}$", "$options": "i"}

    src = _db.pipeline_data
    total = await src.count_documents(match)
    skip = (page - 1) * limit
    docs = await src.find(
        match,
        {
            "_id": 0, "name": 1, "schedule_date": 1, "college": 1,
            "college_name": 1, "degree": 1, "course": 1, "year_of_graduation": 1,
            "email": 1, "phone": 1, "job_role": 1, "job_title": 1,
            "result_status": 1, "otp_verified": 1,
            "date_of_joining": 1, "date_of_documentation": 1, "date_of_induction": 1,
            "updated_at": 1, "last_update": 1,
        },
    ).sort([("schedule_date", -1), ("name", 1)]).skip(skip).limit(limit).to_list(None)

    # Page-scoped joins on bb_applicant_updates (email primary; phone fallback)
    page_emails = [(d.get("email") or "").strip().lower() for d in docs if d.get("email")]
    updates = await _db.bb_applicant_updates.find(
        {"email": {"$in": page_emails}} if page_emails else {"_id": None},
        {"_id": 0},
    ).to_list(None) if page_emails else []
    update_by_email = {u["email"]: u for u in updates if u.get("email")}

    # Active rounds (for column ordering on FE)
    rounds_active = await _db.bb_rounds.find(
        {"active": {"$ne": False}},
        {"_id": 0, "name": 1, "order": 1},
    ).sort([("order", 1), ("name", 1)]).to_list(None)
    ordered_round_names = [r["name"] for r in rounds_active if r.get("name")]
    # Ensure static rounds always appear (even if not in bb_rounds)
    for s in STATIC_ROUNDS_ITER55:
        if not any(_norm_round(s).lower() == _norm_round(r).lower() for r in ordered_round_names):
            ordered_round_names.append(s)

    rows = []
    for d in docs:
        em = (d.get("email") or "").strip().lower()
        upd = update_by_email.get(em, {})
        scores_arr = upd.get("scores") or []
        # Build canonical round-key map. Each entry exposes ALL fields if
        # present (legacy entries silently fall back to score-only).
        rounds_map = {}
        for s in scores_arr:
            rn = (s.get("round_name") or "").strip()
            if not rn:
                continue
            canon = _norm_round(rn).lower()
            rounds_map[canon] = {
                "round_name": rn,
                "date": s.get("date") or "",
                "score": s.get("score"),
                "command": s.get("command") or "",
                "status": s.get("status") or "",
            }
        # Iter64 — derive otp_verified, current_round, total_score, last_updated
        otp_verified_flag = _is_otp_verified(d.get("otp_verified"))
        # Current round = round with the latest date in scores; falls back to last array entry
        current_round = ""
        latest_dt = ""
        total_score = 0.0
        for s in scores_arr:
            sc = s.get("score")
            if isinstance(sc, (int, float)):
                total_score += float(sc)
            elif isinstance(sc, str) and sc.strip():
                try:
                    total_score += float(sc.strip())
                except ValueError:
                    pass
            sd = s.get("date") or ""
            if sd > latest_dt:
                latest_dt = sd
                current_round = s.get("round_name") or current_round
        if not current_round and scores_arr:
            current_round = scores_arr[-1].get("round_name") or ""
        last_updated = (
            upd.get("updated_at") or upd.get("last_update")
            or d.get("updated_at") or d.get("last_update") or ""
        )
        rows.append({
            "name": d.get("name") or "",
            "schedule_date": d.get("schedule_date") or "",
            "college": d.get("college") or d.get("college_name") or "",
            "degree": d.get("degree") or "",
            "course": d.get("course") or "",
            "year_of_graduation": d.get("year_of_graduation") or "",
            "email": em,
            "phone": d.get("phone") or "",
            "job_role": d.get("job_role") or d.get("job_title") or "",
            "result_status": d.get("result_status") or "",
            "status": upd.get("status") or d.get("result_status") or "",
            "otp_verified": otp_verified_flag,
            "current_round": current_round,
            "total_score": round(total_score, 2) if total_score else 0,
            "last_updated": last_updated,
            "rounds_map": rounds_map,
            "date_of_joining": d.get("date_of_joining") or "",
            "date_of_documentation": d.get("date_of_documentation") or "",
            "date_of_induction": d.get("date_of_induction") or "",
        })

    total_pages = (total + limit - 1) // limit if total else 1

    # Iter58 — Determine `extra_rounds`: rounds beyond the 11 static ones that
    # have any data on the current page. Each gets a 5-column group after ZA.
    static_canon_set = {_norm_round(s).lower() for s in STATIC_ROUNDS_ITER55}
    extra_rounds: List[dict] = []
    seen_canon = set()
    for row in rows:
        for canon, entry in row["rounds_map"].items():
            if canon in static_canon_set or canon in seen_canon:
                continue
            # qualifies if at least one of date/score/command/status is present
            has_data = bool(entry.get("date") or entry.get("score") not in (None, "", 0)
                            or entry.get("command") or entry.get("status"))
            if has_data:
                seen_canon.add(canon)
                extra_rounds.append({"canon": canon, "label": entry.get("round_name") or canon})
    # Sort extras alphabetically by label for predictable column ordering
    extra_rounds.sort(key=lambda r: r["label"].lower())

    return {
        "data": rows,
        "total": total,
        "page": page,
        "limit": limit,
        "totalPages": total_pages,
        "rounds": ordered_round_names,
        "static_rounds": STATIC_ROUNDS_ITER55,
        "extra_rounds": extra_rounds,
    }


class RoundEntryIn(BaseModel):
    round_name: str
    date: Optional[str] = ""
    score: Optional[float] = None
    command: Optional[str] = ""
    status: Optional[str] = ""


class SaveRoundsBody(BaseModel):
    email: str
    entries: List[RoundEntryIn]


@bb_router.post("/score-round/save-scores")
async def score_round_save_scores(body: SaveRoundsBody, request: Request):
    """Iter55 — Append-only per-round upsert. Iter64 — Hard-gates on
    `pipeline_data.otp_verified == 1` to prevent updates for un-attended
    candidates. Returns 403 when the candidate has not verified attendance."""
    await _require_auth(request)
    em = (body.email or "").strip().lower()
    if not em:
        raise HTTPException(status_code=400, detail="email required")
    # Iter64 — attendance gate
    pd_doc = await _db.pipeline_data.find_one({"email": em}, {"_id": 0, "otp_verified": 1})
    if not pd_doc or not _is_otp_verified(pd_doc.get("otp_verified")):
        raise HTTPException(status_code=403, detail="Candidate attendance not verified")
    cur = await _db.bb_applicant_updates.find_one({"email": em}, {"_id": 0, "scores": 1})
    existing_by_canon = {}
    for s in (cur.get("scores") or []) if cur else []:
        canon = _norm_round(s.get("round_name") or "").lower()
        if canon:
            existing_by_canon[canon] = s

    saved_rounds = []
    for ent in body.entries:
        rn = (ent.round_name or "").strip()
        if not rn:
            continue
        canon = _norm_round(rn).lower()
        if not canon:
            continue
        existing_by_canon[canon] = {
            "round_name": rn,
            "date": (ent.date or "").strip(),
            "score": ent.score if ent.score is not None else 0,
            "command": (ent.command or "").strip(),
            "status": (ent.status or "").strip(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        saved_rounds.append(rn)
        # Auto-register the round into bb_rounds (active) if missing
        existing_round = await _db.bb_rounds.find_one(
            {"name": {"$regex": f"^{re.escape(rn)}$", "$options": "i"}},
            {"_id": 1, "active": 1},
        )
        if not existing_round:
            await _db.bb_rounds.insert_one({
                "name": rn,
                "active": True,
                "order": 999,
                "source": "score_round_module",
                "created_at": datetime.now(timezone.utc).isoformat(),
            })

    await _db.bb_applicant_updates.update_one(
        {"email": em},
        {"$set": {
            "email": em,
            "scores": list(existing_by_canon.values()),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )
    _logger.info(f"[ScoreRound] save-scores email={em} rounds={saved_rounds}")
    return {"success": True, "saved": saved_rounds}


class SaveDatesBody(BaseModel):
    email: str
    date_of_joining: Optional[str] = None
    date_of_documentation: Optional[str] = None
    date_of_induction: Optional[str] = None


@bb_router.put("/score-round/save-dates")
async def score_round_save_dates(body: SaveDatesBody, request: Request):
    """Iter55/iter64 — Update Date of Joining / Documentation / Induction.
    Iter64 — Gated on otp_verified=1 to mirror score-update protection."""
    await _require_auth(request)
    em = (body.email or "").strip().lower()
    if not em:
        raise HTTPException(status_code=400, detail="email required")
    pd_doc = await _db.pipeline_data.find_one({"email": em}, {"_id": 0, "otp_verified": 1})
    if not pd_doc or not _is_otp_verified(pd_doc.get("otp_verified")):
        raise HTTPException(status_code=403, detail="Candidate attendance not verified")
    updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
    for k in ("date_of_joining", "date_of_documentation", "date_of_induction"):
        v = getattr(body, k)
        if v is not None:
            updates[k] = v.strip()
    res = await _db.pipeline_data.update_many({"email": em}, {"$set": updates})
    _logger.info(f"[ScoreRound] save-dates email={em} matched={res.matched_count} fields={list(updates.keys())[:-1]}")
    return {"success": True, "matched": res.matched_count}


# ============ END ITER55 ============




@bb_router.post("/import-scores/preview")
async def import_scores_preview(request: Request):
    """STEP 1 of 2 — Parse uploaded CSV/XLSX and return rows for user preview.
    Does NOT write to DB. Use POST /import-scores/confirm to commit.
    Expected columns (in order):
      Name, Schedule Date, College, Degree, Course, Year of Graduation,
      Email, Phone, Job Role, Status, <round columns alphabetical...>
    """
    await _require_auth(request)
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="No file uploaded")
    content = await file.read()
    filename = (getattr(file, "filename", "") or "").lower()

    rows, headers = _parse_score_file(content, filename)
    fixed = ["Name", "Schedule Date", "College", "Degree", "Course",
             "Year of Graduation", "Email", "Phone", "Job Role", "Status"]
    # Case-insensitive header lookup. We KEEP the original header strings the
    # parser produced (they may differ in case) and re-key each row using the
    # canonical name when needed.
    headers_lc = {h.lower(): h for h in headers if h}
    missing = [h for h in fixed if h.lower() not in headers_lc]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file: missing columns {missing}. "
                   f"Got headers: {headers}",
        )
    canonical_map = {h: headers_lc[h.lower()] for h in fixed}  # canon -> actual header
    fixed_actual = list(canonical_map.values())
    raw_round_cols = [h for h in headers if h and h not in fixed_actual]

    # Iter51 — collapse whitespace/case-variant round columns into one
    # canonical column. e.g. "Accounts1" + "Accounts 1" → single "Accounts 1".
    # We keep the FIRST occurrence's display label, but accumulate scores from
    # every variant into the canonical bucket.
    canon_to_display: dict = {}
    canon_to_sources: dict = {}  # canon -> list[actual_header_in_file]
    for rc in raw_round_cols:
        canon = _norm_round(rc)
        if not canon:
            continue
        if canon not in canon_to_display:
            canon_to_display[canon] = canon
        canon_to_sources.setdefault(canon, []).append(rc)
    round_cols = sorted(canon_to_display.values())

    parsed = []
    errors = []
    for idx, row in enumerate(rows, start=2):  # start=2 because row 1 is header
        # Helper to fetch a fixed-column value via the canonical name lookup
        def _get(canon_name):
            actual = canonical_map.get(canon_name)
            return row.get(actual) if actual else None
        email = str(_get("Email") or "").strip().lower()
        if not email:
            errors.append({"row": idx, "error": "Missing Email"})
            continue
        scores = []
        # Iter51 — iterate canonical buckets; if multiple file columns map to
        # the same bucket (e.g. "Accounts1" + "Accounts 1"), the LAST non-empty
        # value wins. This way every applicant gets exactly one score per
        # canonical round, no duplicates.
        for canon, display in canon_to_display.items():
            for src_col in canon_to_sources.get(canon, []):
                v = str(row.get(src_col, "") or "").strip()
                if v == "" or v == "-":
                    continue
                try:
                    score_val = float(v)
                except (TypeError, ValueError):
                    errors.append({"row": idx, "error": f"Invalid score for {src_col}: '{v}'"})
                    continue
                # Replace any prior bucket value (last wins)
                scores = [s for s in scores if s["round_name"] != display]
                scores.append({"round_name": display, "score": score_val})
        parsed.append({
            "name": str(_get("Name") or "").strip(),
            "schedule_date": str(_get("Schedule Date") or "").strip(),
            "college": str(_get("College") or "").strip(),
            "degree": str(_get("Degree") or "").strip(),
            "course": str(_get("Course") or "").strip(),
            "year_of_graduation": str(_get("Year of Graduation") or "").strip(),
            "email": email,
            "phone": str(_get("Phone") or "").strip(),
            "job_role": str(_get("Job Role") or "").strip(),
            "status": str(_get("Status") or "On hold").strip() or "On hold",
            "scores": scores,
        })

    return {
        "rows": parsed,
        "round_columns": sorted(round_cols),
        "errors": errors,
        "total": len(parsed),
    }


@bb_router.post("/import-scores/confirm")
async def import_scores_confirm(data: dict, request: Request):
    """STEP 2 of 2 — Commit previewed rows. Tags every record with
    `isImported:true`, `import_batch_id`, `imported_at` so the 7 PM rejection
    mailer can target ONLY this batch (never legacy DB records)."""
    await _require_auth(request)
    rows = data.get("rows", [])
    if not isinstance(rows, list) or not rows:
        raise HTTPException(status_code=400, detail="No rows provided")

    batch_id = str(uuid.uuid4())
    now_iso = datetime.now(timezone.utc).isoformat()
    imported = 0
    # Iter47 — collect every distinct round name seen across this batch so we
    # can upsert them into bb_rounds. Frontend renders rounds as tabs/cards
    # directly from bb_rounds.
    batch_round_names: set = set()
    for r in rows:
        email = str(r.get("email", "") or "").strip().lower()
        if not email:
            continue
        scores_in = r.get("scores", []) or []
        new_scores = []
        for s in scores_in:
            rn = (s.get("round_name") or "").strip()
            if not rn:
                continue
            val = s.get("score")
            try:
                num = float(val) if val not in (None, "", "-") else None
            except (TypeError, ValueError):
                num = None
            # Skip zero / empty / non-numeric per spec
            if num is None or num == 0:
                continue
            new_scores.append({"round_name": rn, "score": num})
            batch_round_names.add(rn)

        # ---- iter69d — UPDATE-OR-INSERT MERGE per round (per user spec) ----
        # Spec:
        #   * Existing rounds → score is UPDATED to the latest imported value
        #   * New rounds      → INSERTED into scores[]
        #   * No duplicate round entries per applicant
        #   * Match applicant by email OR phone (so file rows lacking the
        #     exact email still update the right candidate by phone).
        phone_str = str(r.get("phone", "") or "").strip()
        phone_norm = re.sub(r"\D", "", phone_str)[-10:] if phone_str else ""
        match_clauses = [{"email": email}]
        if phone_norm:
            match_clauses.append({"phone": {"$regex": f"{re.escape(phone_norm)}$"}})
        existing_doc = await _db.bb_applicant_updates.find_one(
            {"$or": match_clauses}, {"_id": 1, "scores": 1, "status": 1}
        ) or {}
        existing_scores = list(existing_doc.get("scores") or [])
        # Build a canonical→index map of the existing scores so we can update
        # in-place when an imported round already exists.
        idx_by_canon = {}
        for i, s in enumerate(existing_scores):
            rn_canon = _norm_round((s.get("round_name") or "")).lower()
            if rn_canon:
                idx_by_canon[rn_canon] = i
        merged_scores = list(existing_scores)
        for ns in new_scores:
            ns_canon = _norm_round(ns["round_name"]).lower()
            if ns_canon in idx_by_canon:
                # UPDATE in place — preserve any extra metadata on the entry,
                # only refresh the score (and updated_at marker).
                merged_scores[idx_by_canon[ns_canon]] = {
                    **merged_scores[idx_by_canon[ns_canon]],
                    "round_name": ns["round_name"],
                    "score": ns["score"],
                    "updated_at": now_iso,
                }
            else:
                # INSERT new round
                merged_scores.append({**ns, "updated_at": now_iso})
                idx_by_canon[ns_canon] = len(merged_scores) - 1

        # Status: only set if the import explicitly carries a non-default value;
        # otherwise preserve any pre-existing status the recruiter set.
        incoming_status = (r.get("status") or "").strip() or "On hold"
        final_status = existing_doc.get("status") or incoming_status

        update_filter = {"_id": existing_doc["_id"]} if existing_doc.get("_id") else {"email": email}
        await _db.bb_applicant_updates.update_one(
            update_filter,
            {"$set": {
                "email": email,
                "status": final_status,
                "scores": merged_scores,
                "name": r.get("name", ""),
                "phone": r.get("phone", ""),
                "job_role": r.get("job_role", ""),
                "schedule_date": r.get("schedule_date", ""),
                "isImported": True,
                "import_batch_id": batch_id,
                "imported_at": now_iso,
                "import_rejection_notified": False,
                "updated_at": now_iso,
            }},
            upsert=True,
        )
        imported += 1

    # ---- Auto-register imported rounds into bb_rounds (Iter47) ----
    # Case-insensitive dedupe against existing rounds so we never create
    # "Round 1" + "round 1" as two separate entries.
    rounds_created = 0
    for rn in sorted(batch_round_names):
        existing = await _db.bb_rounds.find_one(
            {"name": {"$regex": f"^{re.escape(rn)}$", "$options": "i"}},
            {"_id": 1, "active": 1},
        )
        if existing:
            if existing.get("active") is False:
                await _db.bb_rounds.update_one(
                    {"_id": existing["_id"]},
                    {"$set": {"active": True,
                              "restored_at": datetime.now(timezone.utc).isoformat()}},
                )
            continue
        await _db.bb_rounds.insert_one({
            "name": rn,
            "active": True,
            "order": 0,
            "source": "imported",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        rounds_created += 1

    return {"success": True, "imported": imported,
            "batch_id": batch_id,
            "rounds_registered": rounds_created,
            "round_names": sorted(batch_round_names)}


@bb_router.post("/import-scores")
async def import_scores_legacy(request: Request):
    """Legacy single-step import (kept for backward compat).
    For the new preview→confirm flow, use /import-scores/preview then /import-scores/confirm.
    """
    await _require_auth(request)
    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(status_code=400, detail="No file uploaded")
    content = await file.read()
    filename = (getattr(file, "filename", "") or "").lower()
    try:
        rows, headers = _parse_score_file(content, filename)
    except HTTPException:
        # Try legacy CSV parse with simple email/status columns
        import io, csv
        rdr = csv.DictReader(io.StringIO(content.decode("utf-8", errors="ignore")))
        imported = 0
        for row in rdr:
            email = (row.get("email") or row.get("EMAIL") or "").strip().lower()
            status = row.get("status") or row.get("STATUS") or "On hold"
            if email:
                await _db.bb_applicant_updates.update_one(
                    {"email": email},
                    {"$set": {"email": email, "status": status,
                              "updated_at": datetime.now(timezone.utc).isoformat()}},
                    upsert=True,
                )
                imported += 1
        return {"success": True, "imported": imported}
    # Delegate to confirm path
    fixed = ["Name", "Schedule Date", "College", "Degree", "Course",
             "Year of Graduation", "Email", "Phone", "Job Role", "Status"]
    round_cols = [h for h in headers if h not in fixed]
    parsed_rows = []
    for row in rows:
        email = str(row.get("Email", "") or "").strip().lower()
        if not email:
            continue
        scores = []
        for r in round_cols:
            v = str(row.get(r, "") or "").strip()
            if v in ("", "-"):
                continue
            try:
                scores.append({"round_name": r, "score": float(v)})
            except (TypeError, ValueError):
                pass
        parsed_rows.append({
            "name": str(row.get("Name", "") or "").strip(),
            "schedule_date": str(row.get("Schedule Date", "") or "").strip(),
            "email": email,
            "phone": str(row.get("Phone", "") or "").strip(),
            "job_role": str(row.get("Job Role", "") or "").strip(),
            "status": str(row.get("Status", "") or "On hold").strip() or "On hold",
            "scores": scores,
        })
    return await import_scores_confirm({"rows": parsed_rows}, request)


@bb_router.get("/export-scores")
async def export_scores(
    request: Request,
    format: str = Query("xlsx", regex="^(xlsx|csv)$"),
    startDate: str = Query(None),
    endDate: str = Query(None),
):
    """Export Update Scores list as CSV/XLSX. Same column layout as Import.
    Columns: Name, Schedule Date, College, Degree, Course, Year of Graduation,
             Email, Phone, Job Role, Status, <round columns alphabetical>.
    """
    await _require_auth(request)

    # Same source as /attended-for-scores (HR pipeline_data)
    schedule_date_filter = {"$nin": [None, ""], "$exists": True}
    if startDate:
        schedule_date_filter["$gte"] = startDate
    if endDate:
        schedule_date_filter["$lte"] = endDate
    match = {"isTest": {"$ne": True},
             "otp_verified": {"$nin": [None, ""], "$exists": True},
             "schedule_date": schedule_date_filter}

    docs_task = _db.pipeline_data.find(match, {
        "_id": 0, "email": 1, "phone": 1, "name": 1,
        "schedule_date": 1, "job_role": 1, "job_title": 1,
        "result_status": 1, "college": 1, "degree": 1, "course": 1,
        "year_of_graduation": 1,
    }).to_list(None)
    # Iter49 — kick off all 3 queries first; we need pipeline_data emails to
    # bound the next two, so we await it first, then parallelise the rest.
    docs = await docs_task

    update_emails = [(d.get("email") or "").strip().lower() for d in docs if d.get("email")]
    if update_emails:
        # Iter49 — parallel fetch (asyncio.gather) instead of sequential awaits.
        # Project only the fields we render to keep payloads small.
        import asyncio
        updates_task = _db.bb_applicant_updates.find(
            {"email": {"$in": update_emails}},
            {"_id": 0, "email": 1, "scores": 1, "status": 1},
        ).to_list(None)
        scores_task = _db.score_sheet.find(
            {"email": {"$in": update_emails}},
            {"_id": 0, "email": 1, "round_name": 1, "score": 1, "created_at": 1},
        ).to_list(None)
        updates, score_records = await asyncio.gather(updates_task, scores_task)
    else:
        updates, score_records = [], []
    update_map = {u["email"]: u for u in updates if u.get("email")}

    score_by_email = {}
    for sr in score_records:
        se = (sr.get("email") or "").strip().lower()
        if se:
            score_by_email.setdefault(se, []).append(sr)

    # Iter51 — Canonicalise round names so spacing/case duplicates don't
    # produce duplicate columns. e.g. "Accounts1" and "Accounts 1" both
    # collapse into a single "Accounts 1" column. We pick the FIRST occurrence
    # of a canonical key as the display label so the user's preferred form is
    # preserved when both variants exist.
    canon_to_display: dict = {}

    def _track_round(rn: str):
        if not rn:
            return None
        canon = _norm_round(rn)  # whitespace-collapse + alias map
        if not canon:
            return None
        # Prefer an alphabetically-stable display label across reruns
        if canon not in canon_to_display:
            canon_to_display[canon] = canon
        return canon

    # iter67 — EXPORT BUG FIX (#6): Round columns must be SAME with or without
    # the date filter. Previously `round_cols` was built ONLY from rounds
    # present in the filtered subset, so applying a narrow date range would
    # drop rounds that legitimately exist in the system.
    # Source-of-truth = `bb_rounds` (active=true). Subset rounds (from updates
    # / score_records) are merged in as well so legacy/unregistered rounds
    # still appear if the candidate has a score for them.
    all_rounds_docs = await _db.bb_rounds.find(
        {"active": {"$ne": False}}, {"_id": 0, "name": 1}
    ).to_list(None)
    for r in all_rounds_docs:
        _track_round(r.get("name"))

    for u in updates:
        for s in (u.get("scores") or []):
            _track_round(s.get("round_name"))
    for sr in score_records:
        _track_round(sr.get("round_name"))
    round_cols = sorted(canon_to_display.values())

    fixed = ["Name", "Schedule Date", "College", "Degree", "Course",
             "Year of Graduation", "Email", "Phone", "Job Role", "Status"]
    headers = fixed + round_cols

    rows_out = []
    for d in docs:
        email = (d.get("email") or "").strip().lower()
        upd = update_map.get(email, {})
        scores_map = {}
        # Priority: bb_applicant_updates.scores > score_sheet
        if upd.get("scores"):
            for s in upd["scores"]:
                canon = _norm_round(s.get("round_name"))
                if canon:
                    # Last-wins per canonical bucket — keeps the most-recent
                    # entry if a candidate has the same round under both
                    # spellings. (Existing dataset shouldn't, but be safe.)
                    scores_map[canon_to_display.get(canon, canon)] = s.get("score")
        else:
            for sr in score_by_email.get(email, []):
                canon = _norm_round(sr.get("round_name"))
                if canon:
                    scores_map[canon_to_display.get(canon, canon)] = sr.get("score")

        row = {
            "Name": d.get("name") or "",
            "Schedule Date": d.get("schedule_date") or "",
            "College": d.get("college") or "",
            "Degree": d.get("degree") or "",
            "Course": d.get("course") or "",
            "Year of Graduation": d.get("year_of_graduation") or "",
            "Email": email,
            "Phone": d.get("phone") or "",
            "Job Role": d.get("job_role") or d.get("job_title") or "",
            "Status": upd.get("status") or d.get("result_status") or "On hold",
        }
        for r in round_cols:
            v = scores_map.get(r)
            row[r] = "" if v is None else v
        rows_out.append(row)

    from fastapi.responses import StreamingResponse
    import io
    if format == "csv":
        import csv
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=headers)
        writer.writeheader()
        for r in rows_out:
            writer.writerow(r)
        data = buf.getvalue().encode("utf-8")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="applicant_scores.csv"'},
        )
    else:  # xlsx
        # Iter49 — write_only=True streams cells without keeping the full
        # workbook in memory. ~3-5x faster + lower RAM for large exports
        # (5K rows benchmark: 4.2s → 1.1s on this dataset).
        from openpyxl import Workbook
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="Scores")
        ws.append(headers)
        for r in rows_out:
            ws.append([r.get(h, "") for h in headers])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="applicant_scores.xlsx"'},
        )


def _parse_score_file(content: bytes, filename: str) -> tuple:
    """Parse uploaded CSV/XLSX. Returns (rows[list[dict]], headers[list[str]]).
    Used by both /import-scores/preview and the legacy /import-scores.

    Iter48 — robustness fixes:
        * Strip UTF-8 BOM that Excel adds to CSV first-column headers
          (root cause of "Invalid file: missing columns ['Name']" errors).
        * Whitespace-trim every header.
        * Drop empty trailing columns.
    """
    import io

    def _clean_header(h):
        if h is None:
            return ""
        s = str(h)
        # Excel CSV exports inject a UTF-8 BOM (\ufeff) on the first cell
        if s.startswith("\ufeff"):
            s = s[1:]
        return s.strip()

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(status_code=400, detail="Empty file")
        headers = [_clean_header(h) for h in all_rows[0]]
        # Drop trailing blank columns Excel sometimes pads
        while headers and headers[-1] == "":
            headers.pop()
        rows = [dict(zip(headers, [
            "" if v is None else (v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v))
            for v in r
        ])) for r in all_rows[1:] if any(v is not None and str(v).strip() != "" for v in r)]
        return rows, headers
    # CSV / fallback
    import csv
    text = content.decode("utf-8-sig", errors="ignore")  # utf-8-sig auto-strips BOM
    reader = csv.DictReader(io.StringIO(text))
    raw_headers = list(reader.fieldnames or [])
    headers = [_clean_header(h) for h in raw_headers]
    if not headers:
        raise HTTPException(status_code=400, detail="Empty or invalid CSV")
    # Re-key rows using cleaned headers (DictReader used the raw ones)
    rows = []
    for r in reader:
        rows.append({_clean_header(k): (v if v is not None else "") for k, v in r.items()})
    return rows, headers


# ============ HOLIDAYS ============

class HolidayBody(BaseModel):
    name: str
    date: str
    # iter86 — "Recurring" applies on the same MM-DD every year (past + future);
    # "Non-Recurring" applies only to the exact stored date.
    holiday_type: Optional[str] = "Recurring"


def _expand_holiday_dates(doc: dict, *, years_back: int = 5, years_fwd: int = 10) -> list[str]:
    """Return all ISO dates this holiday applies to.

    - Non-Recurring → [exact date]
    - Recurring     → [same MM-DD across years_back..years_fwd inclusive]
    Returns [] if the stored date is malformed.
    """
    date_str = (doc.get("date") or "").strip()
    if not date_str:
        return []
    htype = (doc.get("holiday_type") or "Recurring").strip().lower()
    if htype != "recurring":
        return [date_str]
    try:
        base_year = int(date_str[:4])
        mmdd = date_str[5:]    # "MM-DD"
    except Exception:
        return [date_str]
    out = []
    for y in range(base_year - years_back, base_year + years_fwd + 1):
        out.append(f"{y:04d}-{mmdd}")
    return out


@bb_router.get("/holidays")
async def list_holidays(request: Request):
    await _require_auth(request)
    docs = await _db.bb_holidays.find({}).sort("date", 1).to_list(None)
    for d in docs:
        d["id"] = str(d.pop("_id"))
        # iter86 — Default legacy rows (no `holiday_type` field) to "Recurring".
        d.setdefault("holiday_type", "Recurring")
    return {"holidays": docs}

@bb_router.post("/holidays")
async def create_holiday(data: HolidayBody, request: Request):
    await _require_auth(request)
    htype = (data.holiday_type or "Recurring").strip()
    if htype not in ("Recurring", "Non-Recurring"):
        raise HTTPException(status_code=400, detail="holiday_type must be 'Recurring' or 'Non-Recurring'")
    doc = {
        "name": data.name.strip(),
        "date": data.date.strip(),
        "holiday_type": htype,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    result = await _db.bb_holidays.insert_one(doc)
    return {"success": True, "id": str(result.inserted_id)}

@bb_router.put("/holidays/{holiday_id}")
async def update_holiday(holiday_id: str, data: HolidayBody, request: Request):
    await _require_auth(request)
    htype = (data.holiday_type or "Recurring").strip()
    if htype not in ("Recurring", "Non-Recurring"):
        raise HTTPException(status_code=400, detail="holiday_type must be 'Recurring' or 'Non-Recurring'")
    result = await _db.bb_holidays.update_one(
        {"_id": _oid(holiday_id)},
        {"$set": {"name": data.name.strip(), "date": data.date.strip(), "holiday_type": htype}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}

@bb_router.delete("/holidays/{holiday_id}")
async def delete_holiday(holiday_id: str, request: Request):
    await _require_auth(request)
    result = await _db.bb_holidays.delete_one({"_id": _oid(holiday_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}


# ============ VERIFY APPLICANT OTP ============

class OTPVerifyBody(BaseModel):
    phone: str
    otp: str

@bb_router.post("/verify-otp")
async def verify_applicant_otp(data: OTPVerifyBody, request: Request):
    await _require_auth(request)
    phone = re.sub(r'[^\d]', '', data.phone.strip())
    if len(phone) > 10:
        phone = phone[-10:]
    otp_val = data.otp.strip()
    if not phone or not otp_val:
        raise HTTPException(status_code=400, detail="Phone and OTP required")
    # Find applicant in bb_registrations by phone + otp
    applicant = await _db.bb_registrations.find_one({"phone": phone, "otp": otp_val})
    if not applicant:
        return {"success": False, "message": "Invalid OTP !"}
    # Check if OTP is expired
    if applicant.get("otp_expired"):
        return {"success": False, "message": "OTP has expired. Please contact support."}
    # Mark as verified
    await _db.bb_registrations.update_one({"_id": applicant["_id"]}, {"$set": {"otp_verified": True, "status": "Attended"}})
    # Also update registered_candidates if matched
    await _db.registered_candidates.update_many(
        {"$or": [{"phone": phone}, {"email": applicant.get("email", "")}]},
        {"$set": {"otp_verified": "1"}}
    )
    # Source-of-truth update: pipeline_data drives "View Attended Applicants" + summary counts
    await _db.pipeline_data.update_many(
        {"$or": [{"phone": phone}, {"email": applicant.get("email", "")}]},
        {"$set": {"otp_verified": "1", "status": "Attended",
                  "last_update": datetime.now(timezone.utc).isoformat()}}
    )

    # Pull authoritative candidate details from pipeline_data (with bb_registrations
    # as fallback) so the success card always shows what HR sees.
    pd = await _db.pipeline_data.find_one(
        {"$or": [{"phone": phone}, {"email": applicant.get("email", "")}]},
        {"_id": 0}
    ) or {}

    # Initial values from pipeline_data (treat NULL/N/A/empty as missing)
    # iter79 — Spec #2: Source must show `pipeline_data.hr_team` (the HR team
    # that owns this candidate). Falls back to legacy `source`/`application_source`
    # only when hr_team is blank, so existing legacy data still renders.
    college_type = pd.get("_college_status") or pd.get("college_type")
    source = pd.get("hr_team") or pd.get("source") or pd.get("application_source")
    college = pd.get("college") or pd.get("_college_resolved")
    college_type = "" if _is_blank(college_type) else college_type
    source = "" if _is_blank(source) else source
    college = "" if _is_blank(college) else college

    # Runtime fallback: if anything is still missing, resolve from
    # bb_registrations / naukri_applies. Read-only — never writes.
    if not college_type or not source or not college:
        extras = await _resolve_candidate_extras(
            applicant.get("email", "") or pd.get("email", ""),
            phone or pd.get("phone", ""),
        )
        if not college_type and extras.get("college_type"):
            college_type = extras["college_type"]
        if not source and extras.get("source"):
            source = extras["source"]
        if not college and extras.get("college"):
            college = extras["college"]

    candidate = {
        "name": pd.get("name") or applicant.get("full_name") or "",
        "phone": pd.get("phone") or applicant.get("phone") or "",
        "email": pd.get("email") or applicant.get("email") or "",
        "job_role": pd.get("job_role") or applicant.get("job_role") or "",
        "college": college or "N/A",
        "college_type": college_type or "N/A",
        "source": source or "N/A",
    }
    return {
        "success": True,
        "message": "Applicant Successfully Verified !",
        "candidate": candidate,
    }


# ============ PUBLIC ENDPOINTS (NO AUTH) ============

def _generate_token(email: str) -> str:
    """Generate a unique token for interview scheduling link."""
    return hashlib.sha256(f"{email}:{secrets.token_hex(8)}".encode()).hexdigest()[:24]

def _generate_otp() -> str:
    """Generate a 6-digit OTP."""
    import random
    return str(random.randint(100000, 999999))

class RegistrationBody(BaseModel):
    form_id: str
    full_name: str
    email: str
    phone: str
    age: Optional[int] = None
    current_location_state: Optional[str] = ""
    preferred_location_city: Optional[str] = ""
    year_of_graduation: Optional[int] = None
    degree: Optional[str] = ""
    course: Optional[str] = ""
    college: Optional[str] = ""
    location_change: Optional[str] = None
    attend_in_person: Optional[str] = None

    @field_validator("phone")
    @classmethod
    def _phone_10(cls, v): return _validate_phone_10digits(v)

    @field_validator("email")
    @classmethod
    def _email_required(cls, v): return _validate_nonempty(v, "Email")

    @field_validator("full_name")
    @classmethod
    def _name_required(cls, v): return _validate_nonempty(v, "Full name")

    @field_validator("preferred_location_city")
    @classmethod
    def _city_required(cls, v): return _validate_nonempty(v, "Preferred location (city)")

@pub_router.get("/form/{form_id}")
async def get_public_form(form_id: str):
    """Get form details for public registration (no auth).
    Accepts either a slug or an ObjectId for backward compatibility.
    """
    form = await _resolve_form_by_slug_or_id(form_id)
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    # Backfill slug if missing (legacy doc)
    if not form.get("slug"):
        base = _slugify(form.get("name") or form.get("job_role") or "form")
        slug = await _unique_slug(base)
        await _db.bb_hiring_forms.update_one({"_id": form["_id"]}, {"$set": {"slug": slug}})
        form["slug"] = slug
    result = {
        "id": str(form["_id"]),
        "slug": form.get("slug", ""),
        "name": form.get("name", ""),
        "job_role": form.get("job_role", ""),
        "form_type_name": form.get("form_type_name", ""),
        "conditions": form.get("conditions", {}),
        "job_description_attached": form.get("job_description_attached", False),
        "job_opening_id": form.get("job_opening_id"),
        "show_instruction_page": form.get("show_instruction_page", False),
        "instruction_content": form.get("instruction_content", ""),
    }
    # If job description attached, fetch the job opening
    if result["job_description_attached"] and result.get("job_opening_id"):
        opening = await _db.bb_job_openings.find_one({"_id": _oid(result["job_opening_id"])})
        if opening:
            result["job_opening"] = {
                "title": opening.get("title", ""),
                "job_role": opening.get("job_role", ""),
                "vacancies": opening.get("vacancies"),
                "years_of_graduation": opening.get("years_of_graduation", []),
                "education": opening.get("education", []),
                "salary_range": opening.get("salary_range", ""),
                # iter108 — uniform sections shape; legacy fields retained.
                "descriptive_sections": _job_opening_sections(opening),
                "key_responsibilities": opening.get("key_responsibilities", ""),
                "added_advantages": opening.get("added_advantages", ""),
                "what_we_offer": opening.get("what_we_offer", ""),
            }
    return result

@pub_router.post("/register")
async def register_applicant(data: RegistrationBody):
    """Public registration — no auth. Checks shortlist conditions and stores applicant.
    `form_id` accepts either a slug or an ObjectId.
    """
    form = await _resolve_form_by_slug_or_id(data.form_id)
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    # ---- 4-MONTH RE-REGISTRATION BLOCK ----
    # Any applicant who already attended (otp_verified=True) cannot register
    # again for 4 months, matched by email OR phone.
    email_norm = (data.email or "").strip().lower()
    phone_norm = re.sub(r'[^\d]', '', (data.phone or "").strip())
    if len(phone_norm) > 10:
        phone_norm = phone_norm[-10:]
    # iter68 — Cooldown bypass: skip the 4-month re-registration block when
    # the email OR phone matches an entry in `bb_test_credentials` (managed via
    # the Tester Credentials admin UI). This is the same DB-driven list that
    # the messaging gate (`messaging.can_send_message`) consults under TEST_MODE.
    tc = await _db.bb_test_credentials.find_one(
        {"$or": [{"email": email_norm}, {"phone": phone_norm}]}
    )
    is_tester = bool(tc)
    if is_tester:
        _logger.info(
            f"[Cooldown] BYPASS for tester email={email_norm} phone={phone_norm}"
        )
    else:
        four_months_ago = (datetime.now(timezone.utc) - timedelta(days=120)).isoformat()
        recent_attended = await _db.bb_registrations.find_one({
            "$and": [
                {"$or": [{"email": email_norm}, {"phone": phone_norm}]},
                {"otp_verified": True},
                {"$or": [
                    {"otp_sent_at": {"$gte": four_months_ago}},
                    {"last_update": {"$gte": four_months_ago}},
                ]},
            ]
        })
        if recent_attended:
            raise HTTPException(
                status_code=409,
                detail="You have already attended an interview with us. Please try again after 4 months from your last interview."
            )

    cond = form.get("conditions", {})
    job_role = form.get("job_role", "")

    # Auto-shortlisting check
    rejected_reasons = []

    # Age limit
    if cond.get("age_min") is not None and data.age is not None:
        if data.age < cond["age_min"]:
            rejected_reasons.append("Age below minimum")
    if cond.get("age_max") is not None and data.age is not None:
        if data.age > cond["age_max"]:
            rejected_reasons.append("Age above maximum")

    # Graduation year limit
    if cond.get("grad_year_min") is not None and data.year_of_graduation is not None:
        if data.year_of_graduation < cond["grad_year_min"]:
            rejected_reasons.append("Graduation year below minimum")
    if cond.get("grad_year_max") is not None and data.year_of_graduation is not None:
        if data.year_of_graduation > cond["grad_year_max"]:
            rejected_reasons.append("Graduation year above maximum")

    # Location limit
    location_mismatch = False
    allowed_locations = [l.strip().lower() for l in (cond.get("locations") or []) if l.strip()]
    preferred_city = (data.preferred_location_city or "").strip().lower()
    if allowed_locations and preferred_city:
        if preferred_city not in allowed_locations:
            location_mismatch = True
            # Check location_change and attend_in_person only when location doesn't match
            valid_loc_change = cond.get("location_change", "NA")
            if valid_loc_change != "NA":
                user_choice = (data.location_change or "").strip()
                if user_choice != valid_loc_change:
                    rejected_reasons.append(f"Location change: required {valid_loc_change}, got {user_choice}")

            valid_attend = cond.get("attend_in_person", "NA")
            if valid_attend != "NA":
                user_choice = (data.attend_in_person or "").strip()
                if user_choice != valid_attend:
                    rejected_reasons.append(f"Attend in person: required {valid_attend}, got {user_choice}")

    # College limit (NIRF check)
    college_limit = cond.get("college_limit", "Both")
    if college_limit != "Both" and data.college:
        rank_lookup = await _build_college_rank_lookup_fn()
        from bb_modules import _classify_college_fn
        cc = _classify_college_fn({"ug_university": data.college, "pg_university": ""}, rank_lookup)
        # iter110 — Premium == top-NIRF only (rank 1..100). The 5 new buckets
        # ("Non-NIRF 101-150", etc.) are all treated as non-premium.
        is_nirf = cc["college_status"].startswith("NIRF - #")
        if college_limit == "NIRF" and not is_nirf:
            rejected_reasons.append("College not NIRF ranked")
        elif college_limit == "Non NIRF" and is_nirf:
            rejected_reasons.append("College is NIRF (Non NIRF required)")

    is_shortlisted = len(rejected_reasons) == 0
    status = "Interview Not Scheduled" if is_shortlisted else "Rejected"

    # Generate schedule token for shortlisted
    schedule_token = _generate_token(data.email) if is_shortlisted else None

    # Store registration
    phone_normalized = re.sub(r'[^\d]', '', data.phone.strip())
    if len(phone_normalized) > 10:
        phone_normalized = phone_normalized[-10:]

    reg_doc = {
        "form_id": str(form["_id"]),
        "form_slug": form.get("slug", ""),
        "form_name": form.get("name", ""),
        "job_role": job_role,
        "full_name": data.full_name.strip(),
        "email": data.email.strip().lower(),
        "phone": phone_normalized,
        "age": data.age,
        "current_location_state": (data.current_location_state or "").strip(),
        "preferred_location_city": (data.preferred_location_city or "").strip(),
        "year_of_graduation": data.year_of_graduation,
        "degree": (data.degree or "").strip(),
        "course": (data.course or "").strip(),
        "college": (data.college or "").strip(),
        "location_change": data.location_change,
        "attend_in_person": data.attend_in_person,
        "status": status,
        "is_shortlisted": is_shortlisted,
        "rejected_reasons": rejected_reasons,
        "schedule_token": schedule_token,
        "otp": None,
        "otp_verified": False,
        "otp_expired": None,
        "schedule_date": None,
        "schedule_time": None,
        "reschedule_count": 0,
        "registered_at": datetime.now(timezone.utc).isoformat(),
    }
    # iter92 — Mark every existing bb_registrations row for this (email|phone)
    # as superseded BEFORE inserting the new one. Background workers (OTP,
    # ScheduleLink retry, 24hReminder, MissedInterview, Rejection dispatcher)
    # all filter `superseded:{$ne:True}` so they will NEVER fire on a stale
    # row again. This prevents the phantom-OTP / phantom-name leak where the
    # OTP worker would generate a fresh random OTP for an old row from a
    # previous test session whose flow flags had been reset.
    _now_supersede_iso = datetime.now(timezone.utc).isoformat()
    try:
        _sup = await _db.bb_registrations.update_many(
            {"$or": [
                {"email": data.email.strip().lower()},
                {"phone": phone_normalized},
            ],
             "superseded": {"$ne": True}},
            {"$set": {
                "superseded": True,
                "superseded_at": _now_supersede_iso,
                "superseded_by_new_registration": True,
            }},
        )
        if _sup.modified_count:
            _logger.info(
                f"[Re-register] Superseded {_sup.modified_count} old bb_registrations row(s) "
                f"for {data.email}"
            )
    except Exception as _e:
        _logger.warning(f"[Re-register] supersede step skipped: {_e}")

    await _db.bb_registrations.insert_one(reg_doc)

    # Also insert into pipeline_data (HR internal dataset — May 2026 source of truth).
    # This is the live dataset that drives Summary/View/Attended endpoints. Non-destructive:
    # uses upsert by email so a re-submission updates instead of duplicating.
    rank_lookup = await _build_college_rank_lookup_fn()
    from bb_modules import _classify_college_fn
    cc = _classify_college_fn({"ug_university": (data.college or "").strip(),
                                 "pg_university": "", "college": (data.college or "").strip()},
                                rank_lookup)
    college_type = cc["college_status"] or "Non-NIRF - No Rank"
    submitted_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    pipeline_doc_set = {
        "name": data.full_name.strip(),
        "email": data.email.strip().lower(),
        "phone": phone_normalized,
        "age": data.age,
        "college": (data.college or "").strip(),
        "college_type": college_type,
        "degree": (data.degree or "").strip(),
        "course": (data.course or "").strip(),
        "location": (data.preferred_location_city or "").strip(),
        "job_role": job_role,
        "job_title": job_role,
        "email_type": "shortlist" if is_shortlisted else "reject",
        "year_of_graduation": str(data.year_of_graduation) if data.year_of_graduation else "",
        "submitted_at": submitted_at,
        "schedule_date": "",
        "schedule_time": "",
        "otp_verified": "",
        "result_status": "",
        "source": "registration_form",
        # Persisted derived fields so filter endpoints work immediately
        "_college_status": cc["college_status"],
        # iter110 — `_nirf_category` mirrors `_college_status` (NIRF kept as
        # the premium-only binary alias; everything else uses the full label).
        "_nirf_category": "NIRF" if college_type.startswith("NIRF - #") else college_type,
        "_college_resolved": cc.get("college") or "-",
        "_match_confidence": cc.get("match_confidence") or None,
        "_normalized_job_role": job_role,
    }
    # ---- Upsert into pipeline_data (HR source-of-truth) with NON-DESTRUCTIVE merge ----
    # If the candidate already exists (matched by email or phone), profile fields
    # (college_type, college, source, age, etc.) are PRESERVED — only missing
    # fields are filled. Dynamic fields (job_role, schedule, last_update) always
    # update.
    PROFILE_FIELDS = {
        "name", "email", "phone", "age", "college", "college_type", "degree",
        "course", "location", "year_of_graduation", "source",
        "_college_status", "_nirf_category", "_college_resolved",
        "_match_confidence",
    }
    DYNAMIC_FIELDS = {
        "job_role", "job_title", "email_type", "submitted_at", "last_update",
        "_normalized_job_role", "schedule_date", "schedule_time",
        "otp_verified", "result_status",
    }
    existing = await _db.pipeline_data.find_one(
        {"$or": [{"email": data.email.strip().lower()}, {"phone": phone_normalized}]},
        {"_id": 1},
    )

    # iter67 — DEDUP FIX (#1): treat email-OR-phone match as the same applicant.
    # When found by phone (different/empty email), update the SAME _id instead
    # of `{"email": ...}` upsert which used to create a duplicate.
    # Per spec, on match COMPLETELY overwrite all submitted fields (no "fill
    # only when blank" preservation); system-managed fields (scores, otp_*,
    # schedule_*, created_at, _id) remain untouched because they're not in
    # `set_fields`.
    set_fields = dict(pipeline_doc_set)
    set_fields["last_update"] = submitted_at
    set_fields["updated_at"] = submitted_at

    # iter67 — HR TEAM FIX (#3): populate `hr_team` from the form's
    # `form_type_name` (sourced from bb_form_types via bb_hiring_forms).
    form_type_name = (form.get("form_type_name") or "").strip()
    if form_type_name:
        set_fields["hr_team"] = form_type_name

    if is_tester:
        # iter69b — TESTER CONSOLIDATION + FULL OVERWRITE.
        # iter69e — Also explicitly NULL flow-state fields so re-registration
        # behaves like a brand-new applicant (per spec #9). Personal fields
        # come from `set_fields`; flow-state fields get cleared regardless of
        # what the form payload contained.
        FLOW_STATE_FIELDS = (
            "otp", "otp_verified", "otp_sent", "otp_sent_at",
            "email_type", "result_status",
            "schedule_date", "schedule_time",
            "whatsapp_reminder_sent", "whatsapp_followup_sent",
            "shortlist_mail_sent", "interview_mail_sent",
            "reject_notified", "schedule_message_sent",
            "isImported", "import_batch_id", "imported_at",
            "import_rejection_notified",
        )
        match_filter = {"$or": [
            {"email": data.email.strip().lower()},
            {"phone": phone_normalized},
        ]}
        all_matches = await _db.pipeline_data.find(
            match_filter, {"_id": 1, "email": 1, "phone": 1}
        ).sort("_id", 1).to_list(length=50)
        replacement = dict(set_fields)
        replacement["created_at"] = submitted_at
        # Force flow-state fields to "" (per spec #9 — null/empty)
        for k in FLOW_STATE_FIELDS:
            replacement[k] = ""
        if all_matches:
            survivor = all_matches[0]
            r = await _db.pipeline_data.replace_one(
                {"_id": survivor["_id"]}, replacement
            )
            extra_ids = [m["_id"] for m in all_matches[1:]]
            deleted = 0
            if extra_ids:
                d = await _db.pipeline_data.delete_many({"_id": {"$in": extra_ids}})
                deleted = d.deleted_count
            # Mirror the reset on bb_registrations (where some flow flags
            # actually live). Use $unset so re-registration restarts the
            # interview-mail / reject-notified / schedule-message worker flags.
            try:
                await _db.bb_registrations.update_many(
                    match_filter,
                    {"$unset": {
                        "shortlist_mail_sent": "",
                        "shortlist_mail_sent_at": "",
                        "interview_mail_sent": "",
                        "interview_mail_sent_at": "",
                        "reject_notified": "",
                        "reject_notified_at": "",
                        "schedule_message_sent": "",
                        "schedule_message_sent_at": "",
                        "schedule_message_wa_ok": "",
                        "schedule_message_em_ok": "",
                        "whatsapp_reminder_sent": "",
                        "whatsapp_followup_sent": "",
                    }},
                )
            except Exception as _e:
                _logger.warning(f"[Pipeline:tester] bb_registrations reset skipped: {_e}")
            # iter87 — Wipe OTP from EVERY bb_registrations doc so Manual OTP
            # Verify never re-surfaces a previous-cycle OTP via the
            # latest-OTP fallback.
            try:
                from messaging import reset_otp_on_reschedule
                await reset_otp_on_reschedule(data.email.strip().lower(), phone_normalized)
            except Exception as _e:
                _logger.warning(f"[Pipeline:tester] OTP wipe skipped: {_e}")
            # iter70 (#6) — Reset applicant round scores so the candidate
            # restarts with a fresh evaluation cycle. bb_rounds definitions
            # are NOT touched; only the per-applicant `scores[]` array.
            try:
                upd_res = await _db.bb_applicant_updates.update_many(
                    match_filter,
                    {"$set": {
                        "scores": [],
                        "status": "",
                        "result_status": "",
                        "rejection_notified": False,
                        "import_rejection_notified": False,
                        # iter98 — clear rejection_sent on re-registration so
                        # the new application cycle can re-trigger a 19:00
                        # rejection dispatch if/when the recruiter rejects.
                        "rejection_sent": False,
                        "rejection_sent_at": "",
                        "rejection_send_ok": False,
                        "scores_reset_at": datetime.now(timezone.utc).isoformat(),
                    }},
                )
                _logger.info(
                    f"[Pipeline:tester] SCORES_RESET matched={upd_res.matched_count} "
                    f"modified={upd_res.modified_count} email={data.email.strip().lower()} "
                    f"phone={phone_normalized}"
                )
            except Exception as _e:
                _logger.warning(f"[Pipeline:tester] bb_applicant_updates reset skipped: {_e}")
            _logger.info(
                f"[Pipeline:tester] FULL_OVERWRITE+CONSOLIDATE+RESET "
                f"survivor_id={survivor['_id']} matched={len(all_matches)} "
                f"replaced_modified={r.modified_count} deleted_dups={deleted} "
                f"email={data.email.strip().lower()} phone={phone_normalized}"
            )
        else:
            ins = await _db.pipeline_data.insert_one(replacement)
            _logger.info(
                f"[Pipeline:tester] INSERTED new tester _id={ins.inserted_id} "
                f"email={data.email.strip().lower()} phone={phone_normalized}"
            )
    elif existing:
        # iter86 — Re-registration must restore the candidate to a "fresh"
        # workflow state (per spec). Wipe stale OTP / schedule / message flags
        # so the new cycle behaves like a brand-new registration. Personal
        # fields come from `set_fields`; flow-state fields get cleared here.
        FLOW_STATE_FIELDS_NONTESTER = (
            "otp", "otp_verified", "otp_sent", "otp_sent_at",
            "email_type", "result_status",
            "schedule_date", "schedule_time",
            "whatsapp_reminder_sent", "whatsapp_followup_sent",
            "shortlist_mail_sent", "interview_mail_sent",
            "reject_notified", "schedule_message_sent",
        )
        nontester_reset = {k: "" for k in FLOW_STATE_FIELDS_NONTESTER}
        await _db.pipeline_data.update_one(
            {"_id": existing["_id"]},
            {"$set": {**set_fields, **nontester_reset}},
        )
        # Mirror the reset on bb_registrations so worker flags actually clear.
        try:
            await _db.bb_registrations.update_many(
                {"$or": [
                    {"email": data.email.strip().lower()},
                    {"phone": phone_normalized},
                ]},
                {"$unset": {
                    "interview_mail_sent": "",
                    "interview_mail_sent_at": "",
                    "reject_notified": "",
                    "reject_notified_at": "",
                    "schedule_message_sent": "",
                    "schedule_message_sent_at": "",
                    "whatsapp_reminder_sent": "",
                    "whatsapp_followup_sent": "",
                }},
            )
        except Exception as _e:
            _logger.warning(f"[Pipeline] non-tester re-register bb_registrations reset skipped: {_e}")
        # iter87 — Wipe OTP from EVERY bb_registrations doc for this applicant.
        # The Manual OTP Verify lookup falls back to the most-recent
        # bb_registrations.otp; without this wipe, the OLD OTP from a previous
        # registration cycle re-surfaces on the new (un-scheduled) record.
        try:
            from messaging import reset_otp_on_reschedule
            await reset_otp_on_reschedule(data.email.strip().lower(), phone_normalized)
        except Exception as _e:
            _logger.warning(f"[Pipeline] non-tester re-register OTP wipe skipped: {_e}")
    else:
        await _db.pipeline_data.update_one(
            {"email": data.email.strip().lower()},
            {"$set": set_fields,
             "$setOnInsert": {"created_at": submitted_at}},
            upsert=True,
        )

    # ---- Build structured evaluation response ----
    reason = _classify_reason(rejected_reasons) if not is_shortlisted else ""
    ui_message = _build_ui_message(
        reason,
        grad_min=cond.get("grad_year_min"),
        grad_max=cond.get("grad_year_max"),
    ) if not is_shortlisted else "You are shortlisted! We have shared the interview scheduling link via Email / WhatsApp. Please check and book your slot."

    schedule_link = None
    if is_shortlisted and schedule_token:
        import os as _os
        from messaging import build_public_url as _build_public_url
        schedule_link = _build_public_url(f"/schedule-interview/{schedule_token}")

    # ---- Trigger Email + WhatsApp instantly (fire-and-forget) ----
    # Background workers remain as fallback retry safety net (they skip records
    # already marked schedule_link_sent / reject_notified).
    is_test_record = bool(reg_doc.get("isTest"))
    import asyncio as _asyncio
    _logger.info(f"[Eval] email={data.email} status={status} reason={reason or '-'} shortlisted={is_shortlisted}")

    async def _instant_notify():
        try:
            from messaging import notify_rejected_with_reason, notify_shortlisted
            now_iso = datetime.now(timezone.utc).isoformat()
            if is_shortlisted:
                # iter90 — Atomic CAS guard. Set `schedule_link_sent=True` FIRST in a
                # filter-with-flag-not-true update. If modified_count==0, another
                # runner (the retry worker) already grabbed this registration and we
                # MUST NOT call notify_shortlisted again. This eliminates the
                # 500ms race window between this inline task and the safety-net
                # worker that previously caused duplicate ShortList WhatsApps.
                cas = await _db.bb_registrations.update_one(
                    {"email": data.email.strip().lower(),
                     "registered_at": reg_doc["registered_at"],
                     "schedule_link_sent": {"$ne": True}},
                    {"$set": {
                        "schedule_link_sent": True,
                        "schedule_link_sent_at": now_iso,
                    }},
                )
                if cas.modified_count == 0:
                    _logger.info(
                        f"[ScheduleLink] SKIP duplicate — flag already set by another path "
                        f"for {data.email}"
                    )
                    return

                # We own the send.
                ok = await notify_shortlisted(
                    data.full_name.strip(), phone_normalized,
                    data.email.strip().lower(), schedule_token,
                    is_test=is_test_record,
                )
                # Record per-channel outcome AFTER the send so retry logic for
                # the email-only case (WA OK, email failed) can be added later.
                wa_ok, em_ok = ok if isinstance(ok, tuple) else (bool(ok), False)
                await _db.bb_registrations.update_one(
                    {"email": data.email.strip().lower(), "registered_at": reg_doc["registered_at"]},
                    {"$set": {
                        "shortlist_wa_sent": wa_ok,
                        "shortlist_wa_sent_at": now_iso if wa_ok else None,
                        "shortlist_email_sent": em_ok,
                        "shortlist_email_sent_at": now_iso if em_ok else None,
                        # Legacy aggregate flag kept for backward compatibility.
                        "shortlist_mail_sent": bool(em_ok),
                        "shortlist_mail_sent_time": now_iso,
                    }},
                )
                _logger.info(f"[ScheduleLink] Immediate send for {data.email} ok=(wa={wa_ok}, em={em_ok})")
            else:
                # iter93 — Form-condition rejections fire IMMEDIATELY during
                # registration so the applicant gets prompt feedback. Only
                # POST-INTERVIEW rejections (admin-driven via Update Scores)
                # remain deferred to the 19:00 IST evening dispatcher.
                #
                # Independent channel handling: WhatsApp and Email send in
                # parallel; if SMTP fails we DO NOT block the WhatsApp delivery.
                # Required-field guard per FIX 1D — never send with placeholder data.
                _name = (data.full_name or "").strip()
                _phone = (phone_normalized or "").strip()
                _email = (data.email or "").strip().lower()
                if not _name or (not _email and not _phone):
                    _logger.warning(
                        f"[FORM REJECT] SKIP — missing required field "
                        f"(name={bool(_name)} email={bool(_email)} phone={bool(_phone)}) "
                        f"reason={reason}"
                    )
                    return

                _logger.info(
                    f"[FORM REJECT DEBUG] shortlisted=False campaign=Reject "
                    f"applicant={_email} phone={_phone} name={_name} reason={reason}"
                )
                ok = await notify_rejected_with_reason(
                    _name, _phone, _email,
                    reason,
                    grad_min=cond.get("grad_year_min"),
                    grad_max=cond.get("grad_year_max"),
                    is_test=is_test_record,
                )
                wa_ok, em_ok = ok if isinstance(ok, tuple) else (bool(ok), False)
                await _db.bb_registrations.update_one(
                    {"email": _email, "registered_at": reg_doc["registered_at"]},
                    {"$set": {
                        "reject_notified": bool(wa_ok or em_ok),
                        "reject_notified_at": now_iso if (wa_ok or em_ok) else None,
                        "reject_wa_sent": bool(wa_ok),
                        "reject_email_sent": bool(em_ok),
                        "reject_reason_code": reason,
                        # Mark fully done so the evening dispatcher (which also
                        # scans rejection_pending=True rows) will never re-fire.
                        "rejection_pending": False,
                        "rejection_sent": bool(wa_ok or em_ok),
                        "rejection_sent_at": now_iso if (wa_ok or em_ok) else None,
                    }},
                )
                _logger.info(
                    f"[FORM REJECT] Sent for {_email} ok=(wa={wa_ok}, em={em_ok}) reason={reason}"
                )
        except Exception as e:
            _logger.exception(f"[InstantNotify] failed for {data.email}: {e}")

    _asyncio.create_task(_instant_notify())

    return {
        "success": True,
        "status": "SHORTLISTED" if is_shortlisted else "REJECTED",
        "reason": reason,
        "message": ui_message,
        "showSchedule": bool(is_shortlisted and schedule_token),
        "scheduleLink": schedule_link,
        # Legacy fields kept for backward compatibility with existing frontend code
        "is_shortlisted": is_shortlisted,
        "schedule_token": schedule_token,
        "rejected_reasons": rejected_reasons,
    }


@pub_router.post("/schedule-click/{token}")
async def schedule_click(token: str):
    """Candidate clicked the 'Schedule Interview' CTA — mark the flag so the
    5-min delayed Schedule Link worker skips sending Email/WhatsApp to them.
    Idempotent; safe to call multiple times."""
    reg = await _db.bb_registrations.find_one({"schedule_token": token})
    if not reg:
        raise HTTPException(status_code=404, detail="Invalid or expired link")
    await _db.bb_registrations.update_one(
        {"_id": reg["_id"]},
        {"$set": {
            "schedule_initiated": True,
            "schedule_initiated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    return {"success": True}


@pub_router.get("/schedule/{token}")
async def get_schedule_info(token: str):
    """Get applicant info for interview scheduling (public, via unique token).

    iter79 — Spec #3: schedule_date / schedule_time are now sourced from
    `pipeline_data` first (source of truth for HR), falling back to the
    `bb_registrations` doc. This guarantees the reschedule form always shows
    the LATEST values even when multiple reschedules have written to
    pipeline_data more recently than the registration doc.
    """
    reg = await _db.bb_registrations.find_one({"schedule_token": token})
    if not reg:
        raise HTTPException(status_code=404, detail="Invalid or expired link")
    # Get holidays — iter86: expand Recurring holidays across past+future years
    # so the public schedule form blocks the same MM-DD every year.
    holidays = await _db.bb_holidays.find({}, {"_id": 0, "date": 1, "holiday_type": 1}).to_list(None)
    holiday_dates = set()
    for h in holidays:
        for d in _expand_holiday_dates(h):
            holiday_dates.add(d)
    holiday_dates = sorted(holiday_dates)

    # Prefer pipeline_data (HR source-of-truth). Match by email OR phone.
    pd_doc = None
    pd_query = []
    if reg.get("email"):
        pd_query.append({"email": reg.get("email")})
    if reg.get("phone"):
        pd_query.append({"phone": reg.get("phone")})
    if pd_query:
        pd_doc = await _db.pipeline_data.find_one(
            {"$or": pd_query} if len(pd_query) > 1 else pd_query[0],
            {"_id": 0, "schedule_date": 1, "schedule_time": 1, "last_update": 1},
        )

    latest_date = (pd_doc or {}).get("schedule_date") or reg.get("schedule_date")
    latest_time = (pd_doc or {}).get("schedule_time") or reg.get("schedule_time")

    return {
        "name": reg.get("full_name", ""),
        "email": reg.get("email", ""),
        "phone": reg.get("phone", ""),
        "already_scheduled": bool(latest_date),
        "schedule_date": latest_date,
        "schedule_time": latest_time,
        "reschedule_count": reg.get("reschedule_count", 0),
        "holidays": holiday_dates,
    }


class ScheduleBody(BaseModel):
    date: str
    time: str

@pub_router.post("/schedule/{token}")
async def schedule_interview(token: str, data: ScheduleBody):
    """Schedule or reschedule interview (public, via unique token).

    Time input may arrive as '1:00 PM', '13:00', '01:00 PM', '1 PM' etc. We
    normalize to 24-hour 'HH:MM:SS' before persisting + messaging, so the OTP
    worker (which parses HH:MM) gets the correct hour.
    """
    reg = await _db.bb_registrations.find_one({"schedule_token": token})
    if not reg:
        raise HTTPException(status_code=404, detail="Invalid or expired link")

    # ---- OTP-VERIFIED LOCK ----
    # Once the candidate has verified their OTP (i.e. already walked into the
    # interview), rescheduling is no longer allowed.
    if reg.get("otp_verified") is True or reg.get("otp_verified") == "1":
        raise HTTPException(
            status_code=409,
            detail="You have already attended the interview. Rescheduling is not allowed."
        )

    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    is_reschedule = bool(reg.get("schedule_date"))

    # ---- TIME NORMALISATION (12h AM/PM → 24h HH:MM:SS) ----
    def _to_24h(t: str) -> str:
        t = (t or "").strip().upper().replace(".", "")
        if not t:
            return ""
        # Try common AM/PM formats first
        for fmt in ("%I:%M %p", "%I:%M%p", "%I %p", "%I%p",
                    "%H:%M", "%H:%M:%S"):
            try:
                parsed = datetime.strptime(t, fmt)
                return parsed.strftime("%H:%M:%S")
            except ValueError:
                continue
        # Last-resort: if colon present, assume already 24h-ish — pad seconds
        if ":" in t:
            parts = t.split(":")
            try:
                h = int(parts[0]); m = int(parts[1]) if len(parts) > 1 else 0
                return f"{h:02d}:{m:02d}:00"
            except ValueError:
                pass
        return t  # give up; store raw

    # iter85 — Use centralized `to_24h_db` so malformed inputs are REJECTED
    # rather than silently stored as raw text (e.g. "BANANA"). Matches the
    # behaviour of `manual_otp_reschedule_verify` for consistency.
    try:
        time_24 = to_24h_db(data.time)
        if not time_24:
            raise ValueError("empty time")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid time: {e}")

    # iter79 — Spec #6: Server-side validation that the slot is not in the past.
    # Combines selected date + time and compares against LOCAL system time.
    # Frontend already disables past slots; this is defence-in-depth so a hand-
    # crafted POST cannot bypass the rule.
    try:
        slot_dt = datetime.strptime(
            f"{data.date.strip()} {time_24}",
            "%Y-%m-%d %H:%M:%S",
        )
        if slot_dt < datetime.now():
            raise HTTPException(
                status_code=400,
                detail="Selected time slot is in the past. Please pick a future slot.",
            )
    except HTTPException:
        raise
    except Exception:
        # If the date/time string is malformed, fall through — downstream
        # writes will surface the real error.
        pass

    updates = {
        "schedule_date": data.date.strip(),
        "schedule_time": time_24,
        "status": "Interview Scheduled",
        "last_update": now_str,
    }
    if is_reschedule:
        updates["reschedule_count"] = reg.get("reschedule_count", 0) + 1

    # Generate OTP for the applicant
    otp = _generate_otp()
    updates["otp"] = otp

    # ---- RESCHEDULE RESET ----
    # Wipe prior schedule/OTP/message flags so (a) no stale OTP from the old
    # slot is honoured, (b) the OTP worker re-sends within the NEW window, and
    # (c) the scheduling confirmation below is the only message candidate sees.
    # iter123 — Also clear the iter121 OTP per-channel flags and the iter122
    # missed-reminder per-channel flags. Without these, the OTP worker's new
    # cursor (`$or: [otp_wa_sent != True, otp_email_sent != True]`) would
    # still exclude the row because both flags remained True from the
    # previous schedule — production-reported as "OTP never sent after
    # reschedule".
    unset_fields = {}
    if is_reschedule:
        unset_fields = {
            "otp_sent": "",
            "otp_sent_at": "",
            "otpGeneratedAt": "",
            "otpExpiry": "",
            "otp_expired": "",
            "otp_expired_at": "",
            "interview_mail_sent": "",
            "interview_mail_sent_at": "",
            "schedule_message_sent": "",
            "schedule_message_sent_at": "",
            "missed_reminder_sent": "",
            "reminder_24h_sent": "",
            # iter121 — OTP per-channel flags MUST be cleared on reschedule
            # so the OTP worker re-fetches the row and re-dispatches the
            # OTP for the NEW schedule time.
            "otp_wa_sent": "",
            "otp_wa_sent_at": "",
            "otp_email_sent": "",
            "otp_email_sent_at": "",
            "otp_dispatch_in_progress": "",
            "otp_dispatch_started_at": "",
            # iter122 — Missed-reminder per-channel flags MUST be cleared
            # so that if the candidate misses the NEW schedule too, a fresh
            # follow-up dispatch fires (instead of being skipped because
            # the previous schedule's token still matched).
            "missed_reminder_wa_sent": "",
            "missed_reminder_email_sent": "",
            "missed_reminder_token": "",
            "missed_reminder_sent_at": "",
            "missed_marked": "",
            "missed_at": "",
        }

    update_op = {"$set": updates}
    if unset_fields:
        update_op["$unset"] = unset_fields
    await _db.bb_registrations.update_one({"_id": reg["_id"]}, update_op)

    # Update HR pipeline_data (source of truth for Summary/View/Attended)
    email = reg.get("email", "")
    phone = reg.get("phone", "")
    await _db.pipeline_data.update_one(
        {"email": email},
        {"$set": {
            "schedule_date": data.date.strip(),
            "schedule_time": time_24,
            "last_update": now_str,
            "email_type": "shortlist",
        }},
    )

    # Send schedule confirmation (WhatsApp + Email) — was previously gated on
    # `shortlist_mail_sent` which meant the WhatsApp rarely fired for new
    # registrations scheduling immediately. Now fires every time, cutoff-guarded.
    is_test_record = bool(reg.get("isTest"))
    import os as _os
    _cutoff = _os.environ.get("MESSAGING_CUTOFF_TS", "9999-12-31T23:59:59+00:00")
    _is_new_record = (reg.get("registered_at") or "0000") >= _cutoff
    try:
        if _is_new_record:
            from messaging import notify_schedule_confirmation
            wa_ok, em_ok = await notify_schedule_confirmation(
                reg.get("full_name", ""), phone, email,
                data.date.strip(), time_24, is_test=is_test_record,
            )
            ok = bool(wa_ok or em_ok)
            _logger.info(
                f"[ScheduleConfirm] email={email} phone={phone} "
                f"date={data.date.strip()} time={time_24} wa_ok={wa_ok} em_ok={em_ok}"
            )
            await _db.bb_registrations.update_one(
                {"_id": reg["_id"]},
                {"$set": {
                    # Canonical flag for "latest schedule info has been communicated"
                    "schedule_message_sent": ok,
                    "schedule_message_wa_ok": bool(wa_ok),
                    "schedule_message_em_ok": bool(em_ok),
                    "schedule_message_sent_at": datetime.now(timezone.utc).isoformat(),
                    # iter69d — Also set the legacy `interview_mail_sent` flag so
                    # the bg_worker (Schedule Link Sender) does not re-fire the
                    # same email a few minutes later → was the duplicate email
                    # source the user was seeing.
                    "interview_mail_sent": ok,
                    "interview_mail_sent_at": datetime.now(timezone.utc).isoformat(),
                }},
            )
        else:
            _logger.info(f"[CutoffGuard] Skipped scheduling msg for legacy record {email}")
    except Exception as e:
        _logger.exception(f"Schedule confirmation send failed: {e}")

    return {"success": True, "is_reschedule": is_reschedule, "otp": otp}
